#!/usr/bin/env python3
import argparse
import json
import re
import shutil
import subprocess
import time
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


SOURCE_HEADERS = [
    "采集状态",
    "大建页面标题",
    "大建产品类型",
    "大建产地",
    "大建颜色",
    "大建材质",
    "大建填充物",
    "大建适用场景",
    "大建座位个数",
    "大建坐感",
    "大建造型",
    "大建组装长度(in)",
    "大建组装宽度(in)",
    "大建组装高度(in)",
    "大建产品重量(lb)",
    "大建包裹明细",
    "大建最大包裹长(in)",
    "大建最大包裹宽(in)",
    "大建最大包裹高(in)",
    "大建最大包裹重量(lb)",
    "大建总包裹重量(lb)",
    "大建可售库存",
    "大建一件代发物流费($)",
    "大建云送仓物流费($)",
    "大建一件代发时效",
    "大建云送仓时效",
    "大建Seller",
]

AMAZON_HEADERS = [
    "亚马逊SKU",
    "亚马逊Brand Name",
    "亚马逊Product Type",
    "亚马逊Item Name",
    "亚马逊Product Id Type",
    "亚马逊Item Type Keyword",
    "亚马逊Product Description",
    "亚马逊Bullet Point 1",
    "亚马逊Bullet Point 2",
    "亚马逊Bullet Point 3",
    "亚马逊Bullet Point 4",
    "亚马逊Bullet Point 5",
    "亚马逊Country of Origin",
    "亚马逊Dangerous Goods Regulations",
    "亚马逊Color",
    "亚马逊Material",
    "亚马逊Fill Material",
    "亚马逊Upholstery Fabric Type",
    "亚马逊Seating Capacity",
    "亚马逊Sofa Type",
    "亚马逊Required Assembly",
    "亚马逊Item Condition",
    "亚马逊Quantity",
    "亚马逊Item Length",
    "亚马逊Item Width",
    "亚马逊Item Height",
    "亚马逊Item Dimensions Unit",
    "亚马逊Item Weight",
    "亚马逊Item Weight Unit",
    "亚马逊Package Length",
    "亚马逊Package Width",
    "亚马逊Package Height",
    "亚马逊Package Dimensions Unit",
    "亚马逊Package Weight",
    "亚马逊Package Weight Unit",
    "亚马逊Model Number",
    "亚马逊Manufacturer",
]

ADDED_HEADERS = SOURCE_HEADERS + AMAZON_HEADERS + ["待确认事项"]


def run_applescript(script: str) -> str:
    result = subprocess.run(["osascript", "-e", script], capture_output=True, text=True)
    if result.returncode != 0:
        raise RuntimeError(result.stderr.strip() or result.stdout.strip())
    return result.stdout


def chrome_exec(js: str) -> str:
    script = f'tell application "Google Chrome" to execute active tab of front window javascript {json.dumps(js)}'
    return run_applescript(script)


def chrome_open(url: str) -> None:
    script = f'tell application "Google Chrome" to set URL of active tab of front window to {json.dumps(url)}'
    run_applescript(script)


def fetch_page_text(
    url: str,
    item_id: str,
    cache_dir: Path,
    offline_cache_only: bool,
    use_cache: bool,
    wait_seconds: int,
    expected_sku: str = "",
) -> str:
    cache_file = cache_dir / f"{safe_filename(item_id)}.txt"
    if (use_cache or offline_cache_only) and cache_file.exists():
        return cache_file.read_text(encoding="utf-8")
    if offline_cache_only:
        raise RuntimeError(f"missing cache: {cache_file}")

    chrome_open(url)
    last_text = ""
    deadline = time.time() + wait_seconds
    while time.time() < deadline:
        time.sleep(1)
        text = chrome_exec("document.body ? document.body.innerText : ''")
        last_text = text
        has_specs = ("产品规格" in text or "Specification" in text) and ("Item Code:" in text or "Item Code：" in text)
        sku_matches = not expected_sku or expected_sku in text
        if has_specs and sku_matches:
            break
    cache_file.write_text(last_text, encoding="utf-8")
    return last_text


def safe_filename(value: str) -> str:
    return re.sub(r"[^A-Za-z0-9_.-]+", "_", str(value).strip()) or "row"


def next_value(lines, label):
    for idx, line in enumerate(lines):
        if line.strip() in {label, f"{label}:", f"{label}："}:
            for value in lines[idx + 1 : idx + 6]:
                value = value.strip()
                if value:
                    return value
    return ""


def next_any(lines, labels):
    for label in labels:
        value = next_value(lines, label)
        if value:
            return value
    return ""


def search(pattern, text, group=1):
    match = re.search(pattern, text, re.S)
    return match.group(group).strip() if match else ""


def parse_bullets(text):
    if not text:
        return [""] * 5
    pieces = re.split(r"\s*(?:^|\n)\s*\d+\.\s*", str(text).strip())
    bullets = [p.strip().replace("\n", " ") for p in pieces if p.strip()]
    return (bullets + [""] * 5)[:5]


def normalize_sofa_type(title, shape):
    lower = f"{title} {shape}".lower()
    if "sofa bed" in lower or "sleeper" in lower or "convertible" in lower:
        return "Sofa Bed"
    if "chaise" in lower:
        return "Sofa Chaise"
    if "sectional" in lower or "modular" in lower or "l-shaped" in lower or "l shaped" in lower:
        return "Sectional"
    if "loveseat" in lower:
        return "Loveseat"
    if "sofa" in lower or "couch" in lower:
        return "Standard"
    return ""


def is_target_product(local_title, page_title, keywords):
    text = f"{local_title} {page_title}".lower()
    return any(keyword.lower().strip() in text for keyword in keywords if keyword.strip())


def parse_page(text):
    lines = [line.strip() for line in text.splitlines()]
    item_code = next_value(lines, "Item Code")
    product_type = next_any(lines, ["产品类型", "Product Type"])
    title = next_any(lines, ["产品名称", "Product Name"])
    origin = next_any(lines, ["产地", "Place of Origin"])
    color = next_any(lines, ["颜色", "Main Color", "Color"])
    material = next_any(lines, ["材质", "Main Material", "Material"])
    fabric = next_any(lines, ["面料", "Fabric"])
    fill = next_any(lines, ["填充物", "Filler", "Fill Material"])
    scene = next_any(lines, ["适用场景", "Use Case"])
    seats = next_any(lines, ["座位个数", "Seats"])
    feel = next_any(lines, ["坐感", "Seat Plushness"])
    shape = next_any(lines, ["造型", "Sectional Shape", "Shape"])
    item_length = next_any(lines, ["组装长度 (英寸)", "Assembled Length (in.)"])
    item_width = next_any(lines, ["组装宽度 (英寸)", "Assembled Width (in.)"])
    item_height = next_any(lines, ["组装高度 (英寸)", "Assembled Height (in.)"])
    item_weight = next_any(lines, ["产品重量 (磅)", "Product Weight (lbs.)"])

    packages = []
    for match in re.finditer(
        r"子产品\s*\d+\s*:\s*([^\n]+)\n包裹数量:\s*(\d+)\n([\d.]+)\s*\*\s*([\d.]+)\s*\*\s*([\d.]+)\s*英寸\s*([\d.]+)\s*磅",
        text,
    ):
        packages.append(
            {
                "code": match.group(1).strip(),
                "quantity": int(match.group(2)),
                "length": float(match.group(3)),
                "width": float(match.group(4)),
                "height": float(match.group(5)),
                "weight": float(match.group(6)),
            }
        )
    for match in re.finditer(
        r"Sub-item\s*\d+\s*:\s*([^\n]+)\nPackage Quantity:\s*(\d+)\n([\d.]+)\s*\*\s*([\d.]+)\s*\*\s*([\d.]+)\s*in\.\s*([\d.]+)\s*lbs\.",
        text,
    ):
        packages.append(
            {
                "code": match.group(1).strip(),
                "quantity": int(match.group(2)),
                "length": float(match.group(3)),
                "width": float(match.group(4)),
                "height": float(match.group(5)),
                "weight": float(match.group(6)),
            }
        )

    if not packages:
        package_block = search(r"包装尺寸\s*(?:\n外包装图片)?\n(.+?)\n产品特点", text) or search(r"Package Size\s*\n(.+?)(?:\nProduct Features|\n\+1|\nservice@)", text)
        if package_block:
            package_lines = [line.strip() for line in package_block.splitlines()]
            length = next_any(package_lines, ["长度 (英寸)", "Length (in.)"])
            width = next_any(package_lines, ["宽度 (英寸)", "Width (in.)"])
            height = next_any(package_lines, ["高度 (英寸)", "Height (in.)"])
            weight = next_any(package_lines, ["重量 (磅)", "Weight (lbs.)"])
            if length and width and height and weight:
                packages.append(
                    {
                        "code": item_code or "Single Box",
                        "quantity": 1,
                        "length": float(length),
                        "width": float(width),
                        "height": float(height),
                        "weight": float(weight),
                    }
                )

    heaviest = max(packages, key=lambda item: item["weight"], default={})
    package_detail = "; ".join(
        f'{p["code"]}: {p["quantity"]} pkg, {p["length"]} x {p["width"]} x {p["height"]} in, {p["weight"]} lb'
        for p in packages
    )
    total_package_weight = round(sum(p["weight"] * p["quantity"] for p in packages), 2) if packages else ""

    features_block = search(r"产品特点\n(.+?)\n图文描述", text) or search(r"Product Features\n(.+?)(?:\nGraphic Description|\n\+1|\nservice@)", text)
    features = [line.strip() for line in features_block.splitlines() if line.strip()]
    delisted = "产品已下架" in text

    return {
        "title": title,
        "product_type": product_type,
        "item_code": item_code,
        "origin": origin,
        "color": color,
        "material": material,
        "fabric": fabric,
        "fill": fill,
        "scene": scene,
        "seats": seats,
        "feel": feel,
        "shape": shape,
        "item_length": item_length,
        "item_width": item_width,
        "item_height": item_height,
        "item_weight": item_weight,
        "package_detail": package_detail,
        "package_length": heaviest.get("length", ""),
        "package_width": heaviest.get("width", ""),
        "package_height": heaviest.get("height", ""),
        "package_weight": heaviest.get("weight", ""),
        "total_package_weight": total_package_weight,
        "stock": "0" if delisted else search(r"(\d+)\s*可售库存", text),
        "dropship_fee": search(r"一件代发\s*\n预估物流费:\s*\$([\d.]+)", text),
        "cloud_fee": search(r"云送仓\s*\n预估物流费:\s*\$?([\d.]+~\$?[\d.]+)", text).replace("$", ""),
        "dropship_lead": search(r"一件代发发货时效\s*\n([^\n]+)", text),
        "cloud_lead": search(r"云送仓发货时效\s*\n([^\n]+)", text),
        "seller": search(r"加入购物车\s*\n([^\n]+)", text),
        "features": features,
        "delisted": delisted or "This product is no longer available" in text,
    }


def get_cell(ws, headers, row, header):
    if header not in headers:
        return ""
    return ws.cell(row, headers[header]).value or ""


def append_headers(ws):
    existing = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    for header in ADDED_HEADERS:
        if header not in existing:
            ws.cell(1, ws.max_column + 1).value = header
            existing.append(header)
    return {ws.cell(1, col).value: col for col in range(1, ws.max_column + 1) if ws.cell(1, col).value}


def write_values(ws, row, headers, values):
    for header, value in values.items():
        ws.cell(row, headers[header]).value = value


def style_sheet(ws, headers):
    fill = PatternFill("solid", fgColor="D9EAF7")
    for header in ADDED_HEADERS:
        cell = ws.cell(1, headers[header])
        cell.fill = fill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col in range(1, ws.max_column + 1):
        header = ws.cell(1, col).value
        if header in ADDED_HEADERS:
            width = 18
            if header in {"大建页面标题", "亚马逊Item Name", "亚马逊Product Description", "大建包裹明细", "待确认事项"}:
                width = 42
            elif isinstance(header, str) and header.startswith("亚马逊Bullet"):
                width = 36
            ws.column_dimensions[get_column_letter(col)].width = width

    for row in range(2, ws.max_row + 1):
        for header in ADDED_HEADERS:
            ws.cell(row, headers[header]).alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def build_values(args, row_values, parsed, is_target):
    notes = []
    if is_target:
        notes.append(f"品牌按参数填 {args.brand}")
        notes.append(f"未采到UPC/EAN/GTIN，暂填 {args.product_id_type}")
        local_title = str(row_values.get(args.title_column, "") or "")
        if "chenille" in local_title.lower() and parsed["material"] and "chenille" not in parsed["material"].lower():
            notes.append(f"标题含 Chenille，但大建材质为 {parsed['material']}")
        if parsed["delisted"]:
            notes.append("大建页面显示产品已下架，库存按0填")
    else:
        notes.append("非目标关键词商品，不填目标亚马逊字段")

    source_values = {
        "采集状态": "已采集",
        "大建页面标题": parsed["title"],
        "大建产品类型": parsed["product_type"],
        "大建产地": parsed["origin"],
        "大建颜色": parsed["color"],
        "大建材质": parsed["material"],
        "大建填充物": parsed["fill"],
        "大建适用场景": parsed["scene"],
        "大建座位个数": parsed["seats"],
        "大建坐感": parsed["feel"],
        "大建造型": parsed["shape"],
        "大建组装长度(in)": parsed["item_length"],
        "大建组装宽度(in)": parsed["item_width"],
        "大建组装高度(in)": parsed["item_height"],
        "大建产品重量(lb)": parsed["item_weight"],
        "大建包裹明细": parsed["package_detail"],
        "大建最大包裹长(in)": parsed["package_length"],
        "大建最大包裹宽(in)": parsed["package_width"],
        "大建最大包裹高(in)": parsed["package_height"],
        "大建最大包裹重量(lb)": parsed["package_weight"],
        "大建总包裹重量(lb)": parsed["total_package_weight"],
        "大建可售库存": parsed["stock"],
        "大建一件代发物流费($)": parsed["dropship_fee"],
        "大建云送仓物流费($)": parsed["cloud_fee"],
        "大建一件代发时效": parsed["dropship_lead"],
        "大建云送仓时效": parsed["cloud_lead"],
        "大建Seller": parsed["seller"],
        "待确认事项": "；".join(notes),
    }

    amazon_values = {header: "" for header in AMAZON_HEADERS}
    if is_target:
        bullets = parse_bullets(row_values.get(args.bullet_column, ""))
        description = " ".join(parsed["features"][:5]) if parsed["features"] else " ".join(bullets)
        seat_number = search(r"(\d+)", parsed["seats"]) if parsed["seats"] else ""
        material_for_fabric = parsed["fabric"] or parsed["material"]
        title = row_values.get(args.title_column, "") or parsed["title"]
        sku = row_values.get(args.sku_column, "") or parsed["item_code"]
        amazon_values.update(
            {
                "亚马逊SKU": sku,
                "亚马逊Brand Name": args.brand,
                "亚马逊Product Type": args.product_type,
                "亚马逊Item Name": title,
                "亚马逊Product Id Type": args.product_id_type,
                "亚马逊Item Type Keyword": args.item_type_keyword,
                "亚马逊Product Description": description,
                "亚马逊Bullet Point 1": bullets[0],
                "亚马逊Bullet Point 2": bullets[1],
                "亚马逊Bullet Point 3": bullets[2],
                "亚马逊Bullet Point 4": bullets[3],
                "亚马逊Bullet Point 5": bullets[4],
                "亚马逊Country of Origin": parsed["origin"] or args.default_origin,
                "亚马逊Dangerous Goods Regulations": args.dangerous_goods,
                "亚马逊Color": parsed["color"],
                "亚马逊Material": parsed["material"],
                "亚马逊Fill Material": parsed["fill"],
                "亚马逊Upholstery Fabric Type": material_for_fabric,
                "亚马逊Seating Capacity": seat_number,
                "亚马逊Sofa Type": normalize_sofa_type(title, parsed["shape"]),
                "亚马逊Required Assembly": args.required_assembly,
                "亚马逊Item Condition": args.item_condition,
                "亚马逊Quantity": parsed["stock"],
                "亚马逊Item Length": parsed["item_length"],
                "亚马逊Item Width": parsed["item_width"],
                "亚马逊Item Height": parsed["item_height"],
                "亚马逊Item Dimensions Unit": "Inches",
                "亚马逊Item Weight": parsed["item_weight"],
                "亚马逊Item Weight Unit": "Pounds",
                "亚马逊Package Length": parsed["package_length"],
                "亚马逊Package Width": parsed["package_width"],
                "亚马逊Package Height": parsed["package_height"],
                "亚马逊Package Dimensions Unit": "Inches",
                "亚马逊Package Weight": parsed["package_weight"],
                "亚马逊Package Weight Unit": "Pounds",
                "亚马逊Model Number": parsed["item_code"] or sku,
                "亚马逊Manufacturer": parsed["seller"],
            }
        )

    return {**source_values, **amazon_values}


def parse_args():
    parser = argparse.ArgumentParser(description="Collect GIGAB2B product specs from logged-in Chrome and append spreadsheet columns.")
    parser.add_argument("--workbook", required=True, help="Path to .xlsx workbook")
    parser.add_argument("--sheet", default="", help="Sheet name; default active sheet")
    parser.add_argument("--url-column", default="大建云仓URL")
    parser.add_argument("--id-column", default="大建商品id")
    parser.add_argument("--sku-column", default="大建云仓Item Code")
    parser.add_argument("--title-column", default="新商品标题")
    parser.add_argument("--bullet-column", default="新五点描述")
    parser.add_argument("--brand", default="Vindhvisk")
    parser.add_argument("--product-type", default="SOFA")
    parser.add_argument("--product-id-type", default="GTIN Exempt")
    parser.add_argument("--item-type-keyword", default="sofas")
    parser.add_argument("--target-keywords", default="sofa,couch,sectional,loveseat,chaise")
    parser.add_argument("--fill-all-targets", action="store_true", help="Fill Amazon fields for every successfully collected row")
    parser.add_argument("--default-origin", default="China")
    parser.add_argument("--dangerous-goods", default="Not Applicable")
    parser.add_argument("--required-assembly", default="Yes")
    parser.add_argument("--item-condition", default="New")
    parser.add_argument("--cache-dir", default="", help="Directory for page text cache; default next to workbook")
    parser.add_argument("--offline-cache-only", action="store_true", help="Do not use Chrome; only read cached page text")
    parser.add_argument("--use-cache", action="store_true", help="Reuse cached page text when present; default refetches to avoid stale Chrome/page mismatches")
    parser.add_argument("--max-rows", type=int, default=0, help="Limit data rows for testing")
    parser.add_argument("--wait-seconds", type=int, default=30, help="Max wait per page")
    parser.add_argument("--no-backup", action="store_true")
    return parser.parse_args()


def main():
    args = parse_args()
    workbook_path = Path(args.workbook).expanduser().resolve()
    cache_dir = Path(args.cache_dir).expanduser().resolve() if args.cache_dir else workbook_path.with_name(".gigab2b_page_text")
    cache_dir.mkdir(parents=True, exist_ok=True)

    backup = None
    if not args.no_backup:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup = workbook_path.with_name(f"{workbook_path.stem}.backup_before_gigab2b_fill_{timestamp}{workbook_path.suffix}")
        shutil.copy2(workbook_path, backup)

    wb = load_workbook(workbook_path)
    ws = wb[args.sheet] if args.sheet else wb.active
    headers = append_headers(ws)
    required_headers = [args.url_column, args.id_column, args.title_column]
    missing_headers = [header for header in required_headers if header not in headers]
    if missing_headers:
        raise RuntimeError(f"missing required headers: {missing_headers}")

    keywords = [part.strip() for part in args.target_keywords.split(",") if part.strip()]
    rows_attempted = 0
    for row in range(2, ws.max_row + 1):
        if args.max_rows and rows_attempted >= args.max_rows:
            break
        row_values = {header: get_cell(ws, headers, row, header) for header in headers}
        item_id = str(row_values.get(args.id_column, "") or row)
        url = str(row_values.get(args.url_column, "") or "").strip()
        if not url:
            write_values(ws, row, headers, {"采集状态": "缺少URL"})
            continue
        rows_attempted += 1
        try:
            expected_sku = str(row_values.get(args.sku_column, "") or "").strip()
            text = fetch_page_text(url, item_id, cache_dir, args.offline_cache_only, args.use_cache, args.wait_seconds, expected_sku)
            if "产品规格" not in text and "Specification" not in text:
                write_values(ws, row, headers, {"采集状态": "页面未抓到产品规格"})
                print(f"row {row}: {item_id} no specs")
                continue
            parsed = parse_page(text)
            if expected_sku and parsed["item_code"] and parsed["item_code"] != expected_sku:
                raise RuntimeError(f"Item Code mismatch: expected {expected_sku}, got {parsed['item_code']}")
            target = args.fill_all_targets or is_target_product(row_values.get(args.title_column, ""), parsed["title"], keywords)
            values = build_values(args, row_values, parsed, target)
            write_values(ws, row, headers, values)
            print(f"row {row}: {item_id} ok target={target}")
        except Exception as exc:
            write_values(ws, row, headers, {"采集状态": f"采集失败: {exc}"})
            print(f"row {row}: {item_id} failed: {exc}")

    style_sheet(ws, headers)
    wb.save(workbook_path)
    print(f"saved={workbook_path}")
    if backup:
        print(f"backup={backup}")
    print(f"cache_dir={cache_dir}")


if __name__ == "__main__":
    main()
