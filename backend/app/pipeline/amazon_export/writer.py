from __future__ import annotations

import shutil
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.models import Product, ProductData
from app.pipeline.amazon_export.common_fill import build_initial_fill
from app.pipeline.amazon_export.context import AmazonExportContext
from app.pipeline.amazon_export.image_fill import apply_image_fill
from app.pipeline.amazon_export.listing_fill import apply_bullet_fill, apply_listing_fill
from app.pipeline.amazon_export.offer_fill import apply_offer_fill
from app.pipeline.amazon_export.package_fill import apply_package_fill
from app.pipeline.amazon_export.registry import get_strategy
from app.pipeline.amazon_export.validators import build_fill_summary, finalize_warnings, write_fill_values


def build_amazon_template_file(product: Product, pd: ProductData, mapping: dict[str, Any]) -> dict[str, Any]:
    """Build one Amazon import workbook using the template rule layer."""
    from app.pipeline import step10_amazon_template as legacy

    template_path = Path(mapping["template_path"])
    material_dir = Path(pd.material_dir) if pd.material_dir else Path.cwd() / "outputs"
    output_dir = material_dir / "amazon import"
    output_dir.mkdir(parents=True, exist_ok=True)
    output_name = str(mapping.get("output_filename") or "{item_code}_amazon_import.xlsm").format(item_code=pd.item_code)
    output_path = output_dir / output_name
    existing_image_urls = legacy._existing_template_image_urls(pd, mapping)
    shutil.copy2(template_path, output_path)

    wb = load_workbook(output_path, keep_vba=True, data_only=False)
    ws = wb["Template"]
    columns = legacy._index_template_columns(ws)
    data_row = legacy._mapping_data_row(mapping)
    warnings: list[str] = []
    stylesnap_summary = legacy._selected_stylesnap_summary(pd)
    if stylesnap_summary:
        warnings.append(f"Amazon 导入类目参考已选同款候选: {stylesnap_summary}")

    for cell in ws[data_row]:
        cell.value = None

    brand_col = columns.get("brand[marketplace_id=ATVPDKIKX0DER][language_tag=en_US]#1.value")
    if brand_col:
        legacy._remove_column_validations(ws, brand_col, data_row)

    package, package_warnings = legacy._representative_package(pd)
    warnings.extend(package_warnings)
    bullets = legacy._json_loads(pd.listing_bullets, [])
    bullets = bullets if isinstance(bullets, list) else []

    ctx = AmazonExportContext(
        product=product,
        product_data=pd,
        mapping=mapping,
        template_path=template_path,
        output_path=output_path,
        workbook=wb,
        worksheet=ws,
        columns=columns,
        data_row=data_row,
        fields=mapping["dynamic_fields"],
        fill=build_initial_fill(mapping),
        bullets=bullets,
        existing_image_urls=existing_image_urls,
        package=package,
        warnings=warnings,
    )

    apply_listing_fill(ctx)
    apply_offer_fill(ctx)
    get_strategy(mapping)(ctx)
    apply_image_fill(ctx)
    apply_package_fill(ctx)
    apply_bullet_fill(ctx)

    missing_columns = write_fill_values(ctx)
    finalize_warnings(ctx, missing_columns)
    fill_summary = build_fill_summary(ctx, missing_columns)

    wb.save(output_path)

    return {
        "path": str(output_path),
        "warnings": ctx.warnings,
        "uploaded_images": ctx.uploaded_images,
        "fill_summary": fill_summary,
        "filled_fields": fill_summary["filled_count"],
    }
