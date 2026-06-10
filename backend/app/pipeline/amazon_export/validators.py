from __future__ import annotations

import re
from typing import Any

from openpyxl.utils import range_boundaries

from app.pipeline.amazon_export.context import AmazonExportContext


_INDIRECT_SUFFIX_PATTERN = re.compile(
    r'INDIRECT\(IF\(ISNUMBER\(VALUE\(LEFT\(B\d+,1\)\)\),"_",""\)'
    r'&SUBSTITUTE\(SUBSTITUTE\(B\d+,"-","_"\)," ",""\) &"(.+)"\)'
)


def _defined_name_values(ctx: AmazonExportContext, name: str) -> list[str] | None:
    try:
        defined_name = ctx.workbook.defined_names[name]
    except KeyError:
        return None

    values: list[str] = []
    for title, coord in defined_name.destinations:
        if title not in ctx.workbook.sheetnames:
            continue
        min_col, min_row, max_col, max_row = range_boundaries(coord.replace("$", ""))
        sheet = ctx.workbook[title]
        for row in sheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
            for cell in row:
                if cell.value not in (None, ""):
                    values.append(str(cell.value))
    return values


def _dynamic_list_name(ctx: AmazonExportContext, suffix: str) -> str:
    product_type_value = ctx.fill.get("product_type#1.value") or ctx.worksheet.cell(ctx.data_row, 2).value
    product_type = str(product_type_value or "").replace("-", "_").replace(" ", "")
    if product_type[:1].isdigit():
        product_type = "_" + product_type
    return product_type + suffix


def _allowed_list_values(ctx: AmazonExportContext, column: int) -> list[str] | None:
    for validation in ctx.worksheet.data_validations.dataValidation:
        if validation.type != "list":
            continue
        if not any(
            cell_range.min_row <= ctx.data_row <= cell_range.max_row
            and cell_range.min_col <= column <= cell_range.max_col
            for cell_range in validation.cells.ranges
        ):
            continue
        formula = str(validation.formula1 or "")
        match = _INDIRECT_SUFFIX_PATTERN.fullmatch(formula)
        if match:
            return _defined_name_values(ctx, _dynamic_list_name(ctx, match.group(1)))
        if formula.startswith('"') and formula.endswith('"'):
            return formula[1:-1].split(",")
        return None
    return None


def _should_skip_dropdown_validation(attr: str, value: Any, allowed_values: list[str]) -> bool:
    # Amazon's offer price cells expose a "Delete Offer" dropdown, but numeric prices
    # are valid input for the import template and should not be treated as invalid.
    if attr.startswith("purchasable_offer[") and "value_with_tax" in attr:
        return True
    if isinstance(value, (int, float)) and allowed_values == ["Delete Offer (Sell on Amazon)"]:
        return True
    return False


def _value_allowed_for_dropdown(ctx: AmazonExportContext, attr: str, value: Any) -> bool:
    column = ctx.columns.get(attr)
    if not column:
        return True
    allowed_values = _allowed_list_values(ctx, ctx.worksheet[column + str(ctx.data_row)].column)
    if not allowed_values:
        return True
    if _should_skip_dropdown_validation(attr, value, allowed_values):
        return True
    return str(value) in set(allowed_values)


def write_fill_values(ctx: AmazonExportContext) -> list[str]:
    from app.pipeline import step10_amazon_template as legacy

    missing_columns: list[str] = []
    for attr, value in list(ctx.fill.items()):
        if attr not in ctx.columns:
            missing_columns.append(attr)
            continue
        if not _value_allowed_for_dropdown(ctx, attr, value):
            ctx.fill.pop(attr, None)
            label = ctx.worksheet.cell(4, ctx.worksheet[ctx.columns[attr] + str(ctx.data_row)].column).value or attr
            ctx.warnings.append(f"下拉字段 {label} 的值 {value} 不在模板可选项中，已留空。")
            continue
        legacy._set(ctx.worksheet, ctx.columns, ctx.data_row, attr, value)
    return missing_columns


def finalize_warnings(ctx: AmazonExportContext, missing_columns: list[str]) -> None:
    from app.pipeline import step10_amazon_template as legacy

    if missing_columns:
        ctx.warnings.append(f"模板中未找到 {len(missing_columns)} 个预期字段: {', '.join(missing_columns[:5])}")
    ctx.warnings.extend(legacy._listing_template_warnings(ctx.product_data))
    ctx.warnings.extend(legacy._pricing_template_warnings(ctx.product_data))
    ctx.warnings.extend(legacy._inventory_template_warnings(ctx.product_data))
    ctx.warnings.extend(legacy._aplus_template_warnings(ctx.product))
    ctx.warnings.extend(legacy._step6_main_image_warnings(ctx.product))
    ctx.warnings = list(dict.fromkeys(ctx.warnings))


def build_fill_summary(ctx: AmazonExportContext, missing_columns: list[str]) -> dict[str, Any]:
    from app.pipeline import step10_amazon_template as legacy

    return legacy._build_fill_summary(
        ctx.mapping,
        ctx.fill,
        ctx.columns,
        ctx.warnings,
        missing_columns,
        ctx.uploaded_images,
    )
