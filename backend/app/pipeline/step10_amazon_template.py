"""
模块10：Amazon 导入模板 — 将商品数据写入类目专用 Excel 模板

当前支持 Vindhvisk / Sofas & Couches、Vindhvisk / Bicycle，以及儿童骑乘玩具 RIDE_ON_TOY。
模板字段映射保存在 template_mappings/*.json；少数语义字段会先由模型在模板下拉项内做选择。
"""

import json
import logging
import re
import shutil
import asyncio
from datetime import datetime
from pathlib import Path
from types import SimpleNamespace
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import range_boundaries

from app.config import settings
from app.database import async_session
from app.models import Product, ProductData, ProductFile
from app.pipeline.ride_on_category import RIDE_ON_CATEGORY_MARKERS, select_ride_on_category
from app.pipeline.search_terms import normalize_search_terms
from app.services.oss_uploader import oss_configured, upload_private_image
from sqlalchemy import select
from sqlalchemy.orm import selectinload

logger = logging.getLogger(__name__)

DATA_ROW = 8
AMAZON_TEMPLATE_LOGIC_VERSION = "2026-05-24-sofa-seating-chaise-v1"
SINGLE_SEAT_WEIGHT_CAPACITY_LBS = 250
DEFAULT_WEIGHT_CAPACITY_LBS = 500
PIPELINE_DIR = Path(__file__).parent
MAPPING_DIR = PIPELINE_DIR / "template_mappings"
TEMPLATE_DIR = PIPELINE_DIR / "templates"
SOFA_MAPPING = MAPPING_DIR / "vindhvisk_sofa.json"
BICYCLE_MAPPING = MAPPING_DIR / "vindhvisk_bicycle.json"
ANDY_STORAGE_MAPPING = MAPPING_DIR / "andy_storage_furniture.json"
ANDY_SHELF_TABLE_MAPPING = MAPPING_DIR / "andy_shelf_table_cabinet_gate.json"
BICYCLE_LEAF_CATEGORIES = (
    "Kids' Bikes",
    "Cycling",
    "Folding Bikes",
    "Cruiser Bikes",
    "Electric Bicycles",
    "Mountain Bikes",
    "Road Bikes",
)
BRAND_TEMPLATE_MAPPINGS = {
    ("Vindhvisk", "Sofas & Couches"): SOFA_MAPPING,
    **{("Vindhvisk", category): BICYCLE_MAPPING for category in BICYCLE_LEAF_CATEGORIES},
}
GENERIC_TEMPLATE_MAPPINGS = (
    ANDY_SHELF_TABLE_MAPPING,
    ANDY_STORAGE_MAPPING,
)
SHELF_TABLE_LEAF_CATEGORIES = {
    "Bookcases, Cabinets & Shelves",
    "Nightstands",
    "Bookcases",
    "Table & Chair Sets",
    "Furniture-Style Crates",
    "Coffee Tables",
    "Door & Stair Gates",
    "Litter Box Enclosures",
    "Tables",
    "Free Standing Shoe Racks",
    "Storage Benches",
    "Living Room Table Sets",
}
STORAGE_FURNITURE_LEAF_CATEGORIES = {
    "Storage Cabinets",
    "Lidded Storage Bins",
    "Over-the-Toilet Storage",
    "Storage Drawer Units",
    "Toy Chests & Organizers",
    "Armoires & Dressers",
    "Buffets & Sideboards",
    "Chests & Trunks",
    "Dressers",
    "Medicine Cabinets",
    "Step Stools",
    "Storage Boxes",
    "Vanities",
}
RIDE_ON_TOY_MAPPING = MAPPING_DIR / "ride_on_toy.json"
RIDE_ON_TOY_CATEGORY_MARKERS = RIDE_ON_CATEGORY_MARKERS


def _snapshot_model(obj):
    if obj is None:
        return None
    return SimpleNamespace(**{column.name: getattr(obj, column.name) for column in obj.__table__.columns})


def _resolve_template_path(template_path: str | Path) -> Path:
    path = Path(template_path).expanduser()
    if path.is_absolute():
        return path
    return (PIPELINE_DIR / path).resolve()


def _mapping_has_category_marker(mapping_path: Path, category_text: str) -> bool:
    if not mapping_path.exists():
        return False
    try:
        mapping = json.loads(mapping_path.read_text(encoding="utf-8"))
    except Exception:
        return False
    for option in mapping.get("browse_category_options") or []:
        if not isinstance(option, dict):
            continue
        candidates = [
            option.get("node"),
            option.get("path"),
            *(option.get("markers") or []),
        ]
        if any(str(candidate or "").lower() in category_text for candidate in candidates if candidate):
            return True
    return False


def _load_template_mapping(product: Product, pd: ProductData) -> dict:
    key = (product.brand or "", pd.leaf_category or "")
    mapping_path = BRAND_TEMPLATE_MAPPINGS.get(key)
    category_text = " ".join(
        str(item or "")
        for item in (
            pd.leaf_category,
            pd.categories,
            pd.product_type,
            pd.title,
        )
    ).lower()
    if not mapping_path and (product.brand or "") == "Vindhvisk" and _mapping_has_category_marker(BICYCLE_MAPPING, category_text):
        mapping_path = BICYCLE_MAPPING
    if not mapping_path and any(marker in category_text for marker in RIDE_ON_TOY_CATEGORY_MARKERS):
        mapping_path = RIDE_ON_TOY_MAPPING
    if not mapping_path and (product.brand or "") == "Vindhvisk":
        if _mapping_has_category_marker(SOFA_MAPPING, category_text):
            mapping_path = SOFA_MAPPING
    if not mapping_path and pd.leaf_category in SHELF_TABLE_LEAF_CATEGORIES:
        mapping_path = ANDY_SHELF_TABLE_MAPPING
    if not mapping_path and pd.leaf_category in STORAGE_FURNITURE_LEAF_CATEGORIES:
        mapping_path = ANDY_STORAGE_MAPPING
    if not mapping_path:
        for candidate in GENERIC_TEMPLATE_MAPPINGS:
            if _mapping_has_category_marker(candidate, category_text):
                mapping_path = candidate
                break
    if not mapping_path:
        raise ValueError(f"未配置品牌/类目导入模板映射: brand={product.brand or '未知'}, leaf_category={pd.leaf_category or '未知'}")
    with mapping_path.open("r", encoding="utf-8") as f:
        mapping = json.load(f)
    if mapping.get("template_path"):
        mapping["template_path"] = str(_resolve_template_path(mapping["template_path"]))
    return mapping


def _json_loads(value: str | None, fallback: Any) -> Any:
    if not value:
        return fallback
    try:
        return json.loads(value)
    except Exception:
        return fallback


def _compact_template_text(value: Any, limit: int = 1200) -> str:
    text = " ".join(str(value or "").split()).strip()
    if len(text) <= limit:
        return text
    return text[:limit].rstrip(" ,;-") + "..."


def _listing_check_dict(pd: ProductData) -> dict[str, Any]:
    check = _json_loads(pd.listing_check, {})
    return check if isinstance(check, dict) else {}


def _set_listing_check_template_field(pd: ProductData, key: str, value: dict[str, Any]) -> None:
    check = _listing_check_dict(pd)
    fields = check.get("amazon_template_fields")
    if not isinstance(fields, dict):
        fields = {}
    fields[key] = value
    check["amazon_template_fields"] = fields
    pd.listing_check = json.dumps(check, ensure_ascii=False)


SEMANTIC_DROPDOWN_FIELD_KEYS = (
    "target_audience",
    "included_components",
    "recommended_uses_for_product",
    "specific_uses_for_product",
    "room_type",
    "item_shape",
    "container_shape",
    "mounting_type",
    "age_range_description",
    "style",
    "target_gender",
    "theme",
)


def _template_field_key(key: str) -> str:
    return "target_audience_keyword" if key == "target_audience" else key


def _semantic_values_from_listing_check(pd: ProductData, key: str) -> list[str]:
    fields = _listing_check_dict(pd).get("amazon_template_fields")
    if not isinstance(fields, dict):
        return []
    payload = fields.get(_template_field_key(key))
    if not isinstance(payload, dict):
        return []
    values = payload.get("values")
    if not isinstance(values, list):
        return []
    return [str(value).strip() for value in values if str(value or "").strip()][:5]


def _defined_name_values(wb: Any, name: str) -> list[str]:
    try:
        defined_name = wb.defined_names[name]
    except KeyError:
        return []
    values: list[str] = []
    for title, coord in defined_name.destinations:
        if title not in wb.sheetnames:
            continue
        min_col, min_row, max_col, max_row = range_boundaries(coord.replace("$", ""))
        sheet = wb[title]
        for row in sheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
            for cell in row:
                if cell.value not in (None, ""):
                    values.append(str(cell.value))
    return values


def _defined_name_suffix_for_attr(attr: str) -> str:
    return re.sub(r"[\[\]=#]", "", attr)


def _allowed_values_for_template_attr(template_path: Path, product_type: str | None, attr: str) -> list[str]:
    if not product_type:
        return []
    prefix = str(product_type).replace("-", "_").replace(" ", "")
    if prefix[:1].isdigit():
        prefix = "_" + prefix
    expected_name = prefix + _defined_name_suffix_for_attr(attr)
    wb = load_workbook(template_path, keep_vba=True, data_only=False, read_only=False)
    values: list[str] = []
    for value in _defined_name_values(wb, expected_name):
        if value not in values:
            values.append(value)
    return values


def _information_workbook_values(pd: ProductData) -> dict[str, Any]:
    if not pd.material_dir:
        return {}
    material_dir = Path(pd.material_dir).expanduser()
    if not material_dir.is_dir():
        return {}

    candidates = sorted(material_dir.rglob("*productinfo*.xlsx"))
    if not candidates:
        return {}

    for path in candidates:
        try:
            wb = load_workbook(path, data_only=True, read_only=False)
            ws = wb["产品信息"] if "产品信息" in wb.sheetnames else wb.active
            values: dict[str, Any] = {}
            group = None
            for col in range(1, ws.max_column + 1):
                if ws.cell(1, col).value:
                    group = str(ws.cell(1, col).value)
                field = ws.cell(2, col).value or ws.cell(1, col).value
                value = ws.cell(3, col).value
                if field not in (None, "") and value not in (None, ""):
                    field_name = str(field).strip()
                    values[field_name] = value
                    if group:
                        values[f"{group}.{field_name}"] = value
            return values
        except Exception as exc:
            logger.warning(f"[Step10] 读取大健 Information 表格失败: {path}: {exc}")
    return {}


def _number_value(value: Any) -> float | None:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    match = re.search(r"\d+(?:\.\d+)?", str(value).replace(",", ""))
    return float(match.group(0)) if match else None


def _quantity_value(value: Any) -> int:
    number = _number_value(value)
    return max(int(number), 1) if number else 1


def _parse_package_dimensions(text: str | None) -> dict[str, float] | None:
    if not text:
        return None
    nums = [float(item) for item in re.findall(r"\d+(?:\.\d+)?", str(text).replace(",", ""))]
    if len(nums) < 4:
        return None
    return {
        "length": nums[0],
        "width": nums[1],
        "height": nums[2],
        "weight": nums[3],
    }


def _package_dimensions_from_entry(item: Any) -> dict[str, float] | None:
    if not isinstance(item, dict):
        return _parse_package_dimensions(str(item))

    parsed = {
        "length": _number_value(item.get("length")),
        "width": _number_value(item.get("width")),
        "height": _number_value(item.get("height")),
        "weight": _number_value(item.get("weight_value")),
    }
    if not all(parsed.values()):
        parsed = _parse_package_dimensions(item.get("dimensions"))
    if not parsed:
        return None

    quantity = _quantity_value(item.get("qty") or item.get("quantity") or item.get("package_quantity"))
    return {
        "length": parsed["length"] * quantity,
        "width": parsed["width"] * quantity,
        "height": parsed["height"] * quantity,
        "weight": parsed["weight"] * quantity,
    }


def _package_from_information_workbook(pd: ProductData) -> dict[str, float] | None:
    values = _information_workbook_values(pd)
    if not values:
        return None
    keys = {
        "length": "包装尺寸-长度(英寸)",
        "width": "包装尺寸-宽度(英寸)",
        "height": "包装尺寸-高度(英寸)",
        "weight": "包装尺寸-重量(磅)",
    }
    try:
        package = {key: float(str(values[field]).replace(",", "")) for key, field in keys.items() if values.get(field) not in (None, "")}
    except ValueError:
        return None
    return package if set(package) == set(keys) else None


def _representative_package(pd: ProductData) -> tuple[dict[str, float] | None, list[str]]:
    warnings: list[str] = []
    packages = _json_loads(pd.packages, [])
    parsed_packages: list[dict[str, float]] = []
    if isinstance(packages, list):
        for item in packages:
            parsed = _package_dimensions_from_entry(item)
            if parsed:
                parsed_packages.append(parsed)

    if not parsed_packages:
        info_package = _package_from_information_workbook(pd)
        if info_package:
            warnings.append("页面未解析到包裹明细，已使用大健 Information 表格中的包装尺寸/重量。")
            return info_package, warnings
        warnings.append("缺少可解析的外包装尺寸/重量，模板包裹尺寸列未填写。")
        return None, warnings

    if len(parsed_packages) > 1:
        warnings.append("该商品有多个子产品/外包装，Amazon模板只有一组包裹尺寸列；当前按各包裹长宽高和重量分别相加。")

    return {
        "length": round(sum(item["length"] for item in parsed_packages), 2),
        "width": round(sum(item["width"] for item in parsed_packages), 2),
        "height": round(sum(item["height"] for item in parsed_packages), 2),
        "weight": round(sum(item["weight"] for item in parsed_packages), 2),
    }, warnings


def _index_template_columns(ws) -> dict[str, str]:
    columns: dict[str, str] = {}
    for cell in ws[5]:
        if cell.value:
            columns[str(cell.value)] = cell.column_letter
    return columns


def _mapping_data_row(mapping: dict) -> int:
    try:
        data_row = int(mapping.get("data_row") or DATA_ROW)
    except (TypeError, ValueError):
        return DATA_ROW
    return data_row if data_row > 0 else DATA_ROW


def _set(ws, columns: dict[str, str], data_row: int, attr: str, value: Any) -> None:
    if value in (None, ""):
        return
    col = columns.get(attr)
    if col:
        ws[f"{col}{data_row}"] = value


def _nonempty(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, str):
        return bool(value.strip())
    return value != ""


def _flatten_mapping_values(value: Any) -> list[str]:
    if isinstance(value, str):
        return [value]
    if isinstance(value, list):
        result: list[str] = []
        for item in value:
            result.extend(_flatten_mapping_values(item))
        return result
    if isinstance(value, dict):
        result: list[str] = []
        for item in value.values():
            result.extend(_flatten_mapping_values(item))
        return result
    return []


def _critical_template_fields(mapping: dict) -> list[str]:
    fields = mapping.get("dynamic_fields", {})
    keys = (
        "sku",
        "title",
        "brand",
        "description",
        "search_terms",
        "list_price",
        "quantity",
        "price",
        "country_of_origin",
        "shipping_template",
    )
    result: list[str] = []
    for key in keys:
        result.extend(_flatten_mapping_values(fields.get(key)))
    result.extend(_flatten_mapping_values(mapping.get("bullet_fields", [])[:5]))
    image_fields = mapping.get("image_fields", {})
    result.extend(_flatten_mapping_values(image_fields.get("main")))
    package_fields = mapping.get("package_fields", {})
    for key in ("length_value", "width_value", "height_value", "weight_value"):
        result.extend(_flatten_mapping_values(package_fields.get(key)))
    result.extend(_flatten_mapping_values(mapping.get("required_fields", [])))
    return list(dict.fromkeys(field for field in result if field))


def _filled_field_names(fill: dict[str, Any], columns: dict[str, str]) -> set[str]:
    return {attr for attr, value in fill.items() if attr in columns and _nonempty(value)}


def _missing_required_fields(mapping: dict, fill: dict[str, Any], columns: dict[str, str]) -> list[str]:
    return [
        attr
        for attr in _critical_template_fields(mapping)
        if attr in columns and not _nonempty(fill.get(attr))
    ]


def _listing_template_warnings(pd: ProductData) -> list[str]:
    warnings: list[str] = []
    listing_check = _json_loads(pd.listing_check, {})
    issues = listing_check.get("issues") if isinstance(listing_check, dict) else []
    if isinstance(issues, list):
        for issue in issues[:5]:
            if isinstance(issue, dict):
                message = issue.get("message") or issue.get("issue") or issue.get("text")
            else:
                message = str(issue)
            if message:
                warnings.append(f"Listing提醒: {message}")

    primary = (pd.listing_primary_keyword or "").strip().lower()
    title_start = (pd.listing_title or "")[:100].lower()
    if primary and primary not in title_start:
        warnings.append("主关键词未出现在标题前100字符内，可能影响搜索相关性。")
    bullets = _json_loads(pd.listing_bullets, [])
    bullets = bullets if isinstance(bullets, list) else []
    search_terms = normalize_search_terms(
        pd.listing_search_terms,
        visible_copy=" ".join([pd.listing_title or "", *bullets]),
        max_bytes=settings.STEP5_SEARCH_TERMS_MAX_BYTES,
    )[0]
    if not search_terms:
        warnings.append("Search Terms 为空，导入模板缺少后台关键词。")
    elif len(search_terms) < 40:
        warnings.append("Search Terms 偏短，后台关键词覆盖可能不足。")
    return warnings


def _pricing_template_warnings(pd: ProductData) -> list[str]:
    warnings: list[str] = []
    if pd.suggested_price is None:
        warnings.append("建议售价为空，导入模板价格无法确认。")
    if pd.profit is None or pd.profit_rate is None:
        warnings.append("利润数据为空，上传前需要人工确认价格。")
        return warnings
    target_margin_percent = settings.PRICING_TARGET_MARGIN_RATE * 100
    if pd.profit_rate < target_margin_percent:
        warnings.append(f"净利率 {pd.profit_rate:.1f}% 低于系统目标 {target_margin_percent:.1f}%。")
    if pd.profit < settings.PRICING_MIN_PROFIT:
        warnings.append(f"单件利润 ${pd.profit:.2f} 低于系统最低利润 ${settings.PRICING_MIN_PROFIT:.2f}。")
    return warnings


def _inventory_template_warnings(pd: ProductData) -> list[str]:
    stock = _stock_value(pd)
    return [f"Amazon报价数量按采集库存 {stock} 填写。"]


def _aplus_template_warnings(product: Product) -> list[str]:
    if not product.aplus:
        return ["A+规划、脚本和图片均未生成，导入表格只包含 Listing 和图片信息。"]
    warnings: list[str] = []
    if not product.aplus.aplus_plan:
        warnings.append("A+规划未生成，后续上架前需要补 A+ 规划。")
    if product.aplus.aplus_plan and not product.aplus.aplus_scripts:
        warnings.append("A+规划已生成，但 A+脚本未生成，后续可从脚本节点继续。")
    if product.aplus.aplus_scripts and not product.aplus.aplus_images:
        warnings.append("A+脚本已生成，但 A+图片未生成，后续可从出图节点继续。")
    images = _json_loads(product.aplus.aplus_images, [])
    if not isinstance(images, list) or not images:
        return warnings or ["未生成 A+ 图片，后续上架前需要人工补 A+。"]
    done_count = sum(1 for item in images if isinstance(item, dict) and item.get("status", "done") == "done")
    if done_count < 5:
        warnings.append(f"A+ 图片仅完成 {done_count}/5 张，发布前需要人工补齐或确认。")
    if product.aplus.aplus_status and product.aplus.aplus_status not in {"done", "regen_done"}:
        warnings.append(f"A+ 状态为 {product.aplus.aplus_status}，发布前建议复核。")
    return warnings


def _step6_main_image_warnings(product: Product) -> list[str]:
    if not product.images:
        return ["未找到 Step6 图片选择结果。"]
    if product.images.main_image_source == "fallback_substitute":
        return ["主图使用了替代图，不是完全合规白底主图；上传前请人工确认。"]
    return []


def _template_risk_level(warnings: list[str], missing_required: list[str], main_image_url_filled: bool) -> str:
    if missing_required or not main_image_url_filled:
        return "high_risk"
    high_keywords = ("主图上传失败", "缺少Step6选定主图", "OSS 未配置", "利润数据为空", "低于系统最低利润")
    if any(any(keyword in warning for keyword in high_keywords) for warning in warnings):
        return "high_risk"
    return "warning" if warnings else "pass"


def _build_fill_summary(
    mapping: dict,
    fill: dict[str, Any],
    columns: dict[str, str],
    warnings: list[str],
    missing_columns: list[str],
    uploaded_images: list[dict],
) -> dict[str, Any]:
    filled_fields = _filled_field_names(fill, columns)
    missing_required = _missing_required_fields(mapping, fill, columns)
    main_image_field = mapping.get("image_fields", {}).get("main")
    main_image_url_filled = bool(main_image_field and _nonempty(fill.get(main_image_field)))
    image_field_names = set(_flatten_mapping_values(mapping.get("image_fields", {})))
    image_url_count = len([field for field in image_field_names if _nonempty(fill.get(field))])
    return {
        "logic_version": AMAZON_TEMPLATE_LOGIC_VERSION,
        "risk_level": _template_risk_level(warnings, missing_required, main_image_url_filled),
        "filled_count": len(filled_fields),
        "attempted_count": len([value for value in fill.values() if _nonempty(value)]),
        "template_column_count": len(columns),
        "missing_required_count": len(missing_required),
        "missing_required_fields": missing_required[:50],
        "unmapped_count": len(missing_columns),
        "unmapped_fields": missing_columns[:50],
        "image_url_count": image_url_count,
        "main_image_url_filled": main_image_url_filled,
        "uploaded_image_count": len(uploaded_images),
        "warnings_count": len(warnings),
    }


def _is_remote_url(value: str | None) -> bool:
    return bool(value and str(value).strip().lower().startswith(("http://", "https://")))


def _gallery_sources(product: Product) -> list[str]:
    if not product.images:
        return []
    sources = _json_loads(product.images.gallery_images, [])
    if not isinstance(sources, list):
        return []
    return [str(source).strip() for source in sources if str(source or "").strip()]


def _image_field_sequence(mapping: dict) -> list[str]:
    image_fields = mapping.get("image_fields", {})
    return [
        field
        for field in [
            image_fields.get("main"),
            *(image_fields.get("others") or []),
        ]
        if field
    ]


def _existing_template_image_urls(pd: ProductData, mapping: dict) -> dict[str, str]:
    if not pd.amazon_template_path:
        return {}
    path = Path(pd.amazon_template_path).expanduser()
    if not path.is_file():
        return {}
    try:
        wb = load_workbook(path, data_only=True, read_only=True, keep_vba=True)
        ws = wb["Template"]
        columns = _index_template_columns(ws)
        data_row = _mapping_data_row(mapping)
        urls: dict[str, str] = {}
        for field in _image_field_sequence(mapping):
            col = columns.get(field)
            value = ws[f"{col}{data_row}"].value if col else None
            if isinstance(value, str) and value.strip():
                urls[field] = value.strip()
        return urls
    except Exception as exc:
        logger.warning(f"[Step10] 读取既有 Amazon 模板图片URL失败: {path}: {exc}")
        return {}


def _image_health_warnings(product: Product) -> list[str]:
    if not product.images or not product.images.image_analysis:
        return []
    payload = _json_loads(product.images.image_analysis, {})
    if not isinstance(payload, dict):
        return []
    diagnostics = payload.get("selection_diagnostics")
    if not isinstance(diagnostics, dict):
        return []
    health = diagnostics.get("image_health")
    if not isinstance(health, dict):
        return []
    level = health.get("level")
    if level not in {"high_risk", "review_recommended"}:
        return []
    label = health.get("label") or level
    issues = [
        item.get("message")
        for item in (health.get("issues") or [])
        if isinstance(item, dict) and item.get("severity") in {"high", "review"} and item.get("message")
    ]
    if not issues:
        return [f"Step6图片健康等级为 {label}，最终确认前请复核图片。"]
    return [f"Step6图片健康等级为 {label}: {issue}" for issue in issues[:5]]


def _upload_listing_images(
    product: Product,
    pd: ProductData,
    mapping: dict,
    existing_urls: dict[str, str] | None = None,
) -> tuple[dict[str, str], list[str], list[dict]]:
    warnings: list[str] = []
    uploaded: list[dict] = []
    existing_fill: dict[str, str] = dict(existing_urls or {})

    if not product.images or not product.images.main_image_path:
        if not existing_fill:
            warnings.append("缺少Step6选定主图，主图/副图URL未写入导入模板。")
        return existing_fill, warnings, uploaded

    fill: dict[str, str] = dict(existing_fill)
    product_key = pd.item_code or f"product-{product.id}"
    image_sources = [str(product.images.main_image_path).strip()]
    other_fields = mapping.get("image_fields", {}).get("others", [])
    image_sources.extend(_gallery_sources(product)[:len(other_fields)])

    for idx, source in enumerate(image_sources[:1 + len(other_fields)]):
        slot = "main" if idx == 0 else f"other_{idx}"
        field = mapping.get("image_fields", {}).get("main") if idx == 0 else other_fields[idx - 1]
        if not field or not source:
            continue
        if _is_remote_url(source):
            fill[field] = source
            uploaded.append({"slot": slot, "path": source, "url": source, "status": "remote_url"})
            continue
        if field in fill:
            uploaded.append({"slot": slot, "path": source, "url": fill[field], "status": "reused"})
            continue
        if not oss_configured():
            warnings.append(f"OSS 未配置，本地图片未上传 {slot}。")
            continue
        path = Path(source).expanduser()
        try:
            result = upload_private_image(path, product_key, slot)
        except Exception as exc:
            warnings.append(f"图片上传失败 {slot}: {type(exc).__name__}: {exc}")
            continue
        uploaded.append(result)
        fill[field] = result["url"]

    main_field = mapping.get("image_fields", {}).get("main")
    if main_field and main_field not in fill:
        warnings.append("主图上传失败或缺失，导入模板未填写主图URL。")
    if len(fill) < len(image_sources[:9]):
        warnings.append(f"Listing图片URL仅填写 {len(fill)}/{len(image_sources[:9])} 个。")
    warnings.extend(_image_health_warnings(product))
    return fill, warnings, uploaded


def _remove_column_validations(ws, column_letter: str, data_row: int) -> None:
    if not ws.data_validations:
        return
    kept = []
    target_prefix = f"{column_letter}{data_row}:"
    target_cell = f"{column_letter}{data_row}"
    for validation in ws.data_validations.dataValidation:
        sqref = str(validation.sqref)
        if target_prefix in sqref or target_cell in sqref:
            continue
        kept.append(validation)
    ws.data_validations.dataValidation = kept


def _template_values(wb, attr: str) -> list[str]:
    ws = wb["Dropdown Lists"]
    for col in range(1, ws.max_column + 1):
        if ws.cell(3, col).value == attr:
            return [
                str(ws.cell(row, col).value)
                for row in range(4, ws.max_row + 1)
                if ws.cell(row, col).value not in (None, "")
            ]
    return []


def _first_template_value(wb, attr: str, fallback: str | None = None) -> str | None:
    values = _template_values(wb, attr)
    return values[0] if values else fallback


def _shipping_template_for_product(wb, attr: str, mapping: dict, product: Product, warnings: list[str]) -> str | None:
    values = _template_values(wb, attr)
    if not values:
        warnings.append("模板未找到 Shipping Template 下拉值，配送模板未填写。")
        return None

    preferences = mapping.get("shipping_template_by_brand") or {}
    candidates = [
        product.brand,
        settings.DEFAULT_BRAND,
        "Andy店-US",
        "*",
    ]
    for candidate in candidates:
        preferred = preferences.get(str(candidate or ""))
        if not preferred:
            continue
        if preferred in values:
            return preferred
        warnings.append(f"映射指定配送模板 {preferred}，但模板下拉值不存在，已改用 {values[0]}。")
        return values[0]
    return values[0]


def _stock_value(pd: ProductData) -> int | None:
    if pd.stock is not None:
        return pd.stock
    values = _information_workbook_values(pd)
    raw = values.get("可售库存")
    if raw in (None, ""):
        return None
    try:
        return int(float(str(raw).replace(",", "")))
    except ValueError:
        return None


def _offer_quantity(pd: ProductData) -> int:
    stock = _stock_value(pd)
    if stock is None:
        raise ValueError("未解析到可售库存，已停止生成 Amazon 导入模板；请先采集/补充库存后再上传。")
    if stock < 0:
        raise ValueError(f"采集库存为 {stock}，不能导出负数库存。")
    return stock


def _material_value(pd: ProductData) -> str:
    raw = " ".join([pd.material or "", pd.filler or "", pd.description or ""]).lower()
    if "foam" in raw or "all-foam" in raw or "boneless" in raw or "frameless" in raw:
        return "Polyurethane (PU)"
    if re.search(r"\b(?:metal|steel|iron)\s+frame\b|\bframe\s+(?:made\s+of\s+)?(?:metal|steel|iron)\b", raw):
        return "Metal"
    if re.search(r"\b(?:wood|wooden|engineered wood)\s+frame\b|\bframe\s+(?:made\s+of\s+)?(?:wood|wooden|engineered wood)\b", raw):
        return "Engineered Wood"
    return "Polyurethane (PU)"


def _frame_material_value(pd: ProductData) -> str:
    raw = " ".join([pd.material or "", pd.filler or "", pd.description or "", pd.title or ""]).lower()
    def has_positive_frame(pattern: str) -> bool:
        for match in re.finditer(pattern, raw):
            prefix = raw[max(0, match.start() - 45):match.start()]
            if re.search(r"\b(?:no|without|not|free of|unlike)\b", prefix):
                continue
            return True
        return False

    if "all-foam" in raw or "boneless" in raw or "frameless" in raw:
        return "Engineered Wood"
    if has_positive_frame(r"\b(?:metal|steel|iron)\s+frame\b|\bframe\s+(?:made\s+of\s+)?(?:metal|steel|iron)\b"):
        return "Metal"
    if has_positive_frame(r"\bbamboo\s+frame\b|\bframe\s+(?:made\s+of\s+)?bamboo\b"):
        return "Bamboo"
    if has_positive_frame(r"\bplastic\s+frame\b|\bframe\s+(?:made\s+of\s+)?plastic\b"):
        return "Plastic"
    return "Engineered Wood"


def _fabric_value(pd: ProductData) -> str:
    raw = " ".join([pd.material or "", pd.description or "", pd.title or ""]).lower()
    if "corduroy" in raw:
        return "Corduroy"
    if "chenille" in raw:
        return "Chenille"
    if "velvet" in raw:
        return "Velvet"
    if "leather" in raw:
        return "Faux Leather"
    if "flannelette" in raw or "fabric" in raw:
        return "Polyester"
    return "Polyester"


def _description(pd: ProductData) -> str:
    if pd.listing_description:
        return str(pd.listing_description).strip()[:1900]
    parts = []
    if pd.listing_title:
        parts.append(pd.listing_title)
    bullets = _json_loads(pd.listing_bullets, [])
    if isinstance(bullets, list):
        parts.extend(str(item) for item in bullets if item)
    return "\n".join(parts)[:1900]


def _facts_text(pd: ProductData) -> str:
    values = [
        pd.title,
        pd.listing_title,
        pd.color,
        pd.material,
        pd.filler,
        pd.product_type,
        pd.description,
        pd.features,
        pd.variants,
    ]
    info_values = _information_workbook_values(pd)
    values.extend(info_values.get(key) for key in (
        "产品描述",
        "产品特点1",
        "产品特点2",
        "产品特点3",
        "产品特点4",
        "产品特点5",
        "产品特点6",
        "产品特点7",
        "规格描述1",
        "规格描述2",
        "规格描述3",
    ))
    return " ".join(str(value or "") for value in values)


def _included_components(pd: ProductData, seating: int | None) -> str:
    text = " ".join([pd.description or "", pd.title or "", pd.listing_title or ""]).lower()
    parts = []
    if seating:
        parts.append(f"{seating} sofa modules")
    else:
        parts.append("Sofa")

    pillow_match = re.search(r"(\d+)\s+(?:complimentary\s+)?pillows?", text)
    back_pillow_match = re.search(r"(\d+)\s+(?:oversized\s+)?back pillows?", text)
    accent_pillow_match = re.search(r"(\d+)\s+accent cushions?", text)
    if back_pillow_match:
        parts.append(f"{back_pillow_match.group(1)} back pillows")
    if accent_pillow_match:
        parts.append(f"{accent_pillow_match.group(1)} accent cushions")
    if not back_pillow_match and not accent_pillow_match and pillow_match:
        parts.append(f"{pillow_match.group(1)} pillows")
    elif "pillow" in text and not any("pillow" in part for part in parts):
        parts.append("Pillows")

    return ", ".join(parts)


def _seating_capacity(pd: ProductData) -> int | None:
    text = " ".join([pd.title or "", pd.listing_title or "", pd.product_type or "", pd.description or ""])
    match = re.search(r"\b(\d+)\s*[- ]?\s*(?:Seat|Seater|seat|seater)\b", text)
    if match:
        return int(match.group(1))
    variants = _json_loads(pd.variants, [])
    if isinstance(variants, list):
        color = (pd.color or "").lower()
        for item in variants:
            item_text = str(item.get("text") if isinstance(item, dict) else item)
            if color and color not in item_text.lower():
                continue
            match = re.search(r"\b(\d+)\s*[- ]?\s*(?:Seat|Seater|seat|seater)\b", item_text)
            if match:
                return int(match.group(1))
    return None


def _estimated_sofa_seating_capacity(pd: ProductData) -> int | None:
    explicit = _seating_capacity(pd)
    if explicit:
        return explicit

    text = _furniture_match_text(pd)
    if any(marker in text for marker in ("loveseat", "love seat", "2-piece", "2 piece")):
        return 2
    if any(marker in text for marker in ("sofa bed", "futon", "convertible")):
        return 2
    if "sectional" in text or "l shape" in text or "l-shaped" in text or "modular" in text:
        return 3

    longest_side = max((pd.dimension_length or 0), (pd.dimension_width or 0))
    if longest_side >= 96:
        return 3
    if longest_side >= 60:
        return 2
    return None


def _seat_depth(pd: ProductData) -> float | None:
    if not pd.dimension_width:
        return None
    return round(min(max(pd.dimension_width - 12, 20), 30), 1)


def _weight_capacity_maximum(seating: int | None, pd: ProductData | None = None) -> int:
    text = _furniture_match_text(pd).lower() if pd else ""
    sectional_or_modular = any(marker in text for marker in (
        "sectional",
        "modular",
        "l shape",
        "l-shaped",
        "l shaped",
        "l形",
        "模块化",
        "组合沙发",
    ))
    large_sectional = sectional_or_modular and any(marker in text for marker in (
        "large sectional",
        "oversized sectional",
        "大尺寸",
        "5 seater",
        "6 seater",
        "5-seat",
        "6-seat",
    ))

    if seating is not None:
        seats = max(int(seating), 1)
        if sectional_or_modular and seats >= 3:
            return seats * 300
        return seats * SINGLE_SEAT_WEIGHT_CAPACITY_LBS

    if re.search(r"\b(?:single|one|1)[ -]?(?:seat|seater)\b", text) or any(marker in text for marker in (
        "armchair",
        "accent chair",
        "lounge chair",
        "single chair",
        "单人沙发",
        "扶手椅",
        "单人椅",
    )):
        return SINGLE_SEAT_WEIGHT_CAPACITY_LBS
    if re.search(r"\b(?:two|2)[ -]?(?:seat|seater)\b", text) or any(marker in text for marker in (
        "loveseat",
        "love seat",
        "双人沙发",
        "二人沙发",
    )):
        return 500
    if re.search(r"\b(?:three|3)[ -]?(?:seat|seater)\b", text) or "三人沙发" in text:
        return 750
    if large_sectional:
        return 1500
    if sectional_or_modular:
        return 900
    return DEFAULT_WEIGHT_CAPACITY_LBS


def _omit_fields(fill: dict[str, Any], fields: dict[str, str], keys: tuple[str, ...]) -> None:
    for key in keys:
        attr = fields.get(key)
        if attr:
            fill.pop(attr, None)


def _amazon_item_type_keyword(option: dict[str, Any]) -> str:
    return f"{option['path']} ({option['node']})"


def _furniture_match_text(pd: ProductData) -> str:
    text_values = [
        pd.leaf_category,
        pd.categories,
        pd.product_type,
        pd.title,
        pd.listing_title,
        pd.description,
        pd.features,
        pd.variants,
    ]
    return " ".join(str(value or "") for value in text_values).lower()


def _find_browse_option(options: list[Any], node: str) -> dict[str, Any] | None:
    for option in options:
        if isinstance(option, dict) and option.get("node") == node:
            return option
    return None


def _source_category_text(pd: ProductData) -> str:
    return " ".join(str(value or "") for value in (
        pd.leaf_category,
        pd.categories,
        pd.product_type,
    )).lower()


def _source_selected_furniture_option(options: list[Any], pd: ProductData) -> dict[str, Any] | None:
    source_text = _source_category_text(pd)
    if not source_text:
        return None

    high_confidence_nodes = {
        "living-room-chaise-lounges",
    }
    for option in options:
        if not isinstance(option, dict) or option.get("node") not in high_confidence_nodes:
            continue
        if option.get("product_type") == "CHAIR" and _oversized_for_chair_template(pd):
            continue
        candidates = [
            option.get("node"),
            option.get("path"),
            *(option.get("markers") or []),
        ]
        if any(str(candidate or "").lower() in source_text for candidate in candidates if candidate):
            return option
    return None


def _oversized_for_chair_template(pd: ProductData) -> bool:
    longest_side = max((pd.dimension_length or 0), (pd.dimension_width or 0))
    depth_side = min((pd.dimension_length or 0), (pd.dimension_width or 0))
    return longest_side > 40 or depth_side > 35.04


def _looks_like_sofa(pd: ProductData) -> bool:
    text = _furniture_match_text(pd)
    sofa_markers = (
        "sofa",
        "couch",
        "loveseat",
        "love seat",
        "futon",
        "sleeper",
        "settee",
        "sectional",
        "chaise sofa",
        "1-2 seater",
        "2 seater",
        "two seater",
        "多人沙发",
        "双人沙发",
        "沙发",
    )
    if any(marker in text for marker in sofa_markers):
        return True
    # Amazon returned CHAIR dimension warnings for 62.6 x 48.23 in. Sofa-like
    # furniture should not be squeezed into living-room-chairs just because the
    # supplier leaf category says so.
    return _oversized_for_chair_template(pd)


def _preferred_sofa_node(pd: ProductData) -> str:
    text = _furniture_match_text(pd)
    if "futon" in text:
        return "futon-sets"
    if any(marker in text for marker in ("patio loveseat", "outdoor loveseat", "庭院双人沙发")):
        return "patio-loveseats"
    if any(marker in text for marker in ("patio sofa", "outdoor sofa", "户外沙发")):
        return "patio-sofas"
    if any(marker in text for marker in ("children sofa", "kids sofa", "儿童沙发")):
        return "childrens-sofas"
    return "sofas"


def _sofa_type_value(pd: ProductData) -> str:
    text = _furniture_match_text(pd)
    if "sofa bed" in text:
        return "Sofa Bed"
    if "chaise" in text or "lounge chair" in text:
        return "Sofa Chaise"
    if "sleeper" in text:
        return "Sleeper"
    if "futon" in text:
        return "Futon"
    if "loveseat" in text or "love seat" in text:
        return "Loveseat"
    if "sectional" in text:
        return "Sectional"
    if "settee" in text:
        return "Settee"
    if "convertible" in text:
        return "Convertible"
    return "Standard"


def _select_furniture_category_option(mapping: dict, pd: ProductData) -> dict[str, Any] | None:
    options = mapping.get("browse_category_options") or []
    if not isinstance(options, list):
        return None

    source_option = _source_selected_furniture_option(options, pd)
    if source_option:
        return source_option

    if _looks_like_sofa(pd):
        preferred = _find_browse_option(options, _preferred_sofa_node(pd))
        if preferred:
            return preferred

    haystack = _furniture_match_text(pd)
    best: tuple[int, int, dict[str, Any]] | None = None
    for index, option in enumerate(options):
        if not isinstance(option, dict):
            continue
        score = 0
        node = str(option.get("node") or "").lower()
        path = str(option.get("path") or "").lower()
        if node and node in haystack:
            score += 40
        if path and path in haystack:
            score += 40
        for marker in option.get("markers") or []:
            marker_text = str(marker or "").lower().strip()
            if marker_text and marker_text in haystack:
                score += 20 + min(len(marker_text), 30)
        if score and (best is None or score > best[0]):
            best = (score, index, option)

    if best:
        return best[2]

    for option in options:
        if isinstance(option, dict) and option.get("node") == "sofas":
            return option
    return None


def _apply_furniture_category_fill(fill: dict[str, Any], mapping: dict, pd: ProductData) -> tuple[dict[str, Any] | None, list[str]]:
    warnings: list[str] = []
    option = _select_furniture_category_option(mapping, pd)
    if not option:
        warnings.append("SOFA/CHAIR 模板未匹配到细分类目，沿用映射默认类目。")
        return None, warnings

    fill["product_type#1.value"] = option.get("product_type") or "SOFA"
    fill["item_type_keyword[marketplace_id=ATVPDKIKX0DER]#1.value"] = _amazon_item_type_keyword(option)

    if option.get("product_type") == "CHAIR":
        fill.pop("sofa_type[marketplace_id=ATVPDKIKX0DER]#1.value", None)
    else:
        fill["sofa_type[marketplace_id=ATVPDKIKX0DER]#1.value"] = _sofa_type_value(pd)
    return option, warnings


def _option_match_score(option: dict[str, Any], pd: ProductData) -> int:
    source_text = _source_category_text(pd)
    title_text = _furniture_match_text(pd)
    score = 0
    for candidate in [option.get("node"), option.get("path")]:
        candidate_text = str(candidate or "").lower().strip()
        if not candidate_text:
            continue
        if candidate_text in source_text:
            score += 45 + min(len(candidate_text), 45)
        if candidate_text in title_text:
            score += 35 + min(len(candidate_text), 35)

    for marker in option.get("markers") or []:
        marker_text = str(marker or "").lower().strip()
        if not marker_text:
            continue
        if marker_text in source_text:
            score += 45 + min(len(marker_text), 45)
        if marker_text in title_text:
            score += 65 + min(len(marker_text), 45)
    return score


def _select_general_category_option(mapping: dict, pd: ProductData) -> dict[str, Any] | None:
    options = mapping.get("browse_category_options") or []
    if not isinstance(options, list):
        return None

    best: tuple[int, int, dict[str, Any]] | None = None
    for index, option in enumerate(options):
        if not isinstance(option, dict):
            continue
        score = _option_match_score(option, pd)
        if score and (best is None or score > best[0]):
            best = (score, index, option)
    if best:
        return best[2]

    return options[0] if options and isinstance(options[0], dict) else None


def _template_product_type_for_semantic_fields(mapping: dict, pd: ProductData) -> str | None:
    category_type = str(mapping.get("category_type") or "")
    option: dict[str, Any] | None = None
    if category_type in {"shelf_table_cabinet_gate", "home_storage_furniture"}:
        option = _select_general_category_option(mapping, pd)
    elif not category_type and mapping.get("browse_category_options"):
        option = _select_furniture_category_option(mapping, pd)
    if option and option.get("product_type"):
        return str(option["product_type"])
    fixed_values = mapping.get("fixed_values") if isinstance(mapping.get("fixed_values"), dict) else {}
    fixed_product_type = fixed_values.get("product_type#1.value")
    return str(fixed_product_type).strip() if fixed_product_type else None


def _semantic_dropdown_options(mapping: dict, template_path: Path, product_type: str | None) -> dict[str, list[str]]:
    fields = mapping.get("dynamic_fields") if isinstance(mapping.get("dynamic_fields"), dict) else {}
    options: dict[str, list[str]] = {}
    for key in SEMANTIC_DROPDOWN_FIELD_KEYS:
        attrs = _flatten_mapping_values(fields.get(key))
        if key == "target_audience":
            for field_key, field_value in fields.items():
                if str(field_key).startswith("target_audience_"):
                    attrs.extend(_flatten_mapping_values(field_value))
        if key == "theme":
            for field_key, field_value in fields.items():
                if str(field_key).startswith("theme_"):
                    attrs.extend(_flatten_mapping_values(field_value))
        allowed: list[str] = []
        for attr in attrs:
            for value in _allowed_values_for_template_attr(template_path, product_type, attr):
                if value not in allowed:
                    allowed.append(value)
        if allowed:
            options[key] = allowed
    return options


def _semantic_dropdown_prompt(product: Product, pd: ProductData, options: dict[str, list[str]]) -> str:
    bullets = _json_loads(pd.listing_bullets, [])
    bullets = bullets if isinstance(bullets, list) else []
    categories = _json_loads(pd.categories, pd.categories)
    return f"""Analyze Amazon import template semantic dropdown fields for this product.

Rules:
- For each field, choose only from that field's allowed_values list.
- Return [] for a field when none of its allowed values is explicitly supported by the product facts or listing.
- Do not infer generic audience, use, room, shape, component, mounting, style, theme, gender, or age range details unless the selected value is directly supported.
- Do not invent a value outside the allowed list. Values must be exact string matches.
- Output valid JSON only in this shape:
{{
  "fields": {{
    "field_key": {{"values": ["..."], "reason": "short evidence"}}
  }}
}}

Allowed values by field:
{json.dumps(options, ensure_ascii=False, indent=2)}

Product facts:
- Brand: {product.brand}
- Supplier title: {_compact_template_text(pd.title, 500)}
- Listing title: {_compact_template_text(pd.listing_title, 500)}
- Listing bullets: {json.dumps([_compact_template_text(item, 350) for item in bullets[:5]], ensure_ascii=False)}
- Listing description: {_compact_template_text(pd.listing_description, 900)}
- Supplier description: {_compact_template_text(pd.description, 900)}
- Product type: {_compact_template_text(pd.product_type, 200)}
- Category: {_compact_template_text(categories, 500)}
- Leaf category: {_compact_template_text(pd.leaf_category, 200)}
- Features: {_compact_template_text(pd.features, 700)}
- Variants: {_compact_template_text(pd.variants, 700)}
"""


async def ensure_amazon_template_semantic_fields(product: Product, pd: ProductData, mapping: dict, template_path: Path) -> None:
    product_type = _template_product_type_for_semantic_fields(mapping, pd)
    options = _semantic_dropdown_options(mapping, template_path, product_type)
    if not options:
        return

    existing = _listing_check_dict(pd).get("amazon_template_fields")
    if isinstance(existing, dict) and all(
        isinstance(existing.get(_template_field_key(key)), dict)
        and isinstance(existing[_template_field_key(key)].get("values"), list)
        and all(str(value) in allowed_values for value in existing[_template_field_key(key)].get("values", []))
        for key, allowed_values in options.items()
    ):
        return

    try:
        client = settings.get_llm_client()
        response = await client.chat.completions.create(
            model=settings.LLM_MODEL,
            messages=[
                {"role": "system", "content": "You choose Amazon template dropdown values from evidence. Return JSON only."},
                {"role": "user", "content": _semantic_dropdown_prompt(product, pd, options)},
            ],
            temperature=0,
            max_tokens=1200,
            response_format={"type": "json_object"},
        )
        content = response.choices[0].message.content or "{}"
        payload = json.loads(content)
    except Exception as exc:
        logger.warning("[Step10] 模板语义下拉字段模型分析失败 product_id=%s: %s", product.id, exc)
        for key, allowed_values in options.items():
            _set_listing_check_template_field(pd, _template_field_key(key), {
                "values": [],
                "reason": f"Model analysis failed: {type(exc).__name__}: {exc}",
                "source": "llm_error",
                "allowed_values": allowed_values,
                "product_type": product_type,
            })
        return

    fields_payload = payload.get("fields") if isinstance(payload, dict) else {}
    fields_payload = fields_payload if isinstance(fields_payload, dict) else {}
    for key, allowed_values in options.items():
        item = fields_payload.get(key)
        item = item if isinstance(item, dict) else {}
        values = item.get("values")
        values = values if isinstance(values, list) else []
        selected_values: list[str] = []
        for value in values:
            text = str(value or "").strip()
            if text in allowed_values and text not in selected_values:
                selected_values.append(text)
        _set_listing_check_template_field(pd, _template_field_key(key), {
            "values": selected_values[:5],
            "reason": _compact_template_text(item.get("reason") or "", 600),
            "source": "llm",
            "allowed_values": allowed_values,
            "product_type": product_type,
        })


def _field(fill: dict[str, Any], fields: dict[str, Any], key: str, value: Any) -> None:
    target = fields.get(key)
    if not target:
        return
    for field in _flatten_mapping_values(target):
        if _nonempty(value):
            fill[field] = value
            return


def _fields(fill: dict[str, Any], fields: dict[str, Any], key: str, values: list[Any]) -> None:
    targets = _flatten_mapping_values(fields.get(key))
    for field, value in zip(targets, values):
        if _nonempty(value):
            fill[field] = value


def _size_text(pd: ProductData) -> str | None:
    if pd.dimension_length and pd.dimension_width and pd.dimension_height:
        return f'{pd.dimension_length:g}" x {pd.dimension_width:g}" x {pd.dimension_height:g}"'
    return None


def _count_from_text(pd: ProductData, nouns: tuple[str, ...]) -> int | None:
    text = _furniture_match_text(pd)
    pattern = r"\b(\d+)[ -]*(?:" + "|".join(re.escape(noun) for noun in nouns) + r")s?\b"
    match = re.search(pattern, text)
    if match:
        return int(match.group(1))
    words = {
        "one": 1,
        "two": 2,
        "three": 3,
        "four": 4,
        "five": 5,
        "six": 6,
        "seven": 7,
        "eight": 8,
    }
    word_pattern = r"\b(" + "|".join(words) + r")[ -]*(?:" + "|".join(re.escape(noun) for noun in nouns) + r")s?\b"
    match = re.search(word_pattern, text)
    return words.get(match.group(1)) if match else None


def _material_from_product(pd: ProductData) -> str:
    raw = (pd.material or "").strip()
    normalized = raw.lower().replace(".", "")
    if normalized in {"mdf", "medium density fiberboard", "engineered wood"}:
        return "Engineered Wood"
    if "metal" in normalized and "wood" in normalized:
        return "Engineered Wood"
    return raw or _frame_material_value(pd)


def _general_weight_capacity(pd: ProductData) -> int:
    text = _furniture_match_text(pd)
    if any(marker in text for marker in ("bench", "seat", "stool")):
        return 250
    if any(marker in text for marker in ("table", "shelf", "bookcase", "cabinet", "dresser")):
        return 100
    return 50


def _apply_general_template_fill(fill: dict[str, Any], fields: dict[str, Any], mapping: dict, pd: ProductData) -> list[str]:
    warnings: list[str] = []
    option = _select_general_category_option(mapping, pd)
    if option:
        fill["product_type#1.value"] = option.get("product_type") or fill.get("product_type#1.value")
        fill["item_type_keyword[marketplace_id=ATVPDKIKX0DER]#1.value"] = _amazon_item_type_keyword(option)
    else:
        warnings.append("新家具/收纳模板未匹配到细分类目，沿用映射默认类目。")

    material = _material_from_product(pd)
    frame_material = _frame_material_value(pd)
    text = _furniture_match_text(pd)

    _field(fill, fields, "material", material)
    _field(fill, fields, "color", pd.color)
    _field(fill, fields, "size", _size_text(pd))
    _field(fill, fields, "part_number", pd.item_code)
    _field(fill, fields, "number_of_items", 1)
    _field(fill, fields, "number_of_pieces", _count_from_text(pd, ("piece", "pc")) or 1)
    _field(fill, fields, "number_of_packs", 1)
    _fields(fill, fields, "item_shape", _semantic_values_from_listing_check(pd, "item_shape"))
    _fields(fill, fields, "container_shape", _semantic_values_from_listing_check(pd, "container_shape"))
    _field(fill, fields, "unit_count", 1)
    _field(fill, fields, "unit_count_type", "Count")
    _fields(fill, fields, "included_components", _semantic_values_from_listing_check(pd, "included_components"))
    _field(fill, fields, "is_assembly_required", "Yes")
    _field(fill, fields, "assembly_instructions", "Require Assembly")
    _field(fill, fields, "assembly_instructions_description", "Assembly required. Follow the included instructions before use.")
    _field(fill, fields, "frame_material", frame_material)
    _field(fill, fields, "frame_material_structured", frame_material)
    _field(fill, fields, "top_material", material)
    _field(fill, fields, "base_material", frame_material)
    _field(fill, fields, "back_material", material)
    _field(fill, fields, "case_material", material)
    _field(fill, fields, "furniture_finish", pd.color)
    _field(fill, fields, "finish_type", "Painted" if pd.color else None)
    _fields(fill, fields, "mounting_type", _semantic_values_from_listing_check(pd, "mounting_type"))
    _field(fill, fields, "weight_capacity_maximum", _general_weight_capacity(pd))
    _field(fill, fields, "weight_capacity_maximum_unit", "Pounds")
    _field(fill, fields, "maximum_weight_recommendation", _general_weight_capacity(pd))
    _field(fill, fields, "maximum_weight_recommendation_unit", "Pounds")
    _field(fill, fields, "number_of_drawers", _count_from_text(pd, ("drawer",)))
    _field(fill, fields, "number_of_doors", _count_from_text(pd, ("door",)))
    _field(fill, fields, "number_of_shelves", _count_from_text(pd, ("shelf", "shelves", "tier")))
    _field(fill, fields, "number_of_levels", _count_from_text(pd, ("tier", "level")))
    _field(fill, fields, "number_of_racks", _count_from_text(pd, ("rack",)))
    _field(fill, fields, "number_of_steps", _count_from_text(pd, ("step",)) or (2 if "step stool" in text else None))
    _field(fill, fields, "pet_type", "Dog" if "dog" in text else ("Cat" if "cat" in text or "litter" in text else None))
    _field(fill, fields, "dog_breed_size", "Small" if "small" in text else ("Medium" if "medium" in text else None))
    _field(fill, fields, "indoor_outdoor_usage", "Indoor")
    _field(fill, fields, "table_design", "Coffee Table" if "coffee table" in text else ("Dining Table" if "dining" in text else None))
    _fields(fill, fields, "target_audience", _semantic_values_from_listing_check(pd, "target_audience"))

    if pd.dimension_length and pd.dimension_width and pd.dimension_height:
        for prefix in ("item_lwh", "item_dwh", "item_dimensions"):
            _field(fill, fields, f"{prefix}_length_value", pd.dimension_length)
            _field(fill, fields, f"{prefix}_length_unit", "Inches")
            _field(fill, fields, f"{prefix}_width_value", pd.dimension_width)
            _field(fill, fields, f"{prefix}_width_unit", "Inches")
            _field(fill, fields, f"{prefix}_height_value", pd.dimension_height)
            _field(fill, fields, f"{prefix}_height_unit", "Inches")
        _field(fill, fields, "item_width_value", pd.dimension_width)
        _field(fill, fields, "item_width_unit", "Inches")
        _field(fill, fields, "item_length_value", pd.dimension_length)
        _field(fill, fields, "item_length_unit", "Inches")
        _field(fill, fields, "item_depth", f"{pd.dimension_width:g} Inches")
        _field(fill, fields, "base_width_value", pd.dimension_width)
        _field(fill, fields, "base_width_unit", "Inches")
    else:
        warnings.append("商品尺寸缺少长宽高，模板尺寸列只能部分填写。")

    if pd.weight:
        _field(fill, fields, "item_weight_value", pd.weight)
        _field(fill, fields, "item_weight_unit", "Pounds")

    _fields(fill, fields, "room_type", _semantic_values_from_listing_check(pd, "room_type"))
    _fields(fill, fields, "recommended_uses_for_product", _semantic_values_from_listing_check(pd, "recommended_uses_for_product"))
    _fields(fill, fields, "specific_uses_for_product", _semantic_values_from_listing_check(pd, "specific_uses_for_product"))
    return warnings


def _set_sequence_fields(fill: dict[str, Any], field_names: Any, values: list[Any]) -> None:
    fields = _flatten_mapping_values(field_names)
    for field, value in zip(fields, values):
        if _nonempty(value):
            fill[field] = value


def _bicycle_match_text(pd: ProductData) -> str:
    text_values = [
        pd.leaf_category,
        pd.categories,
        pd.product_type,
        pd.title,
        pd.listing_title,
        pd.description,
        pd.features,
        pd.variants,
        _facts_text(pd),
    ]
    return " ".join(str(value or "") for value in text_values).lower()


def _select_bicycle_category_option(mapping: dict, pd: ProductData) -> dict[str, Any] | None:
    options = mapping.get("browse_category_options") or []
    if not isinstance(options, list):
        return None

    source_text = _source_category_text(pd)
    if source_text:
        best: tuple[int, int, dict[str, Any]] | None = None
        for index, option in enumerate(options):
            if not isinstance(option, dict):
                continue
            score = 0
            candidates = [
                option.get("node"),
                option.get("path"),
                *(option.get("markers") or []),
            ]
            for candidate in candidates:
                candidate_text = str(candidate or "").lower().strip()
                if candidate_text and candidate_text in source_text:
                    score += 50 + min(len(candidate_text), 50)
            if score and (best is None or score > best[0]):
                best = (score, index, option)
        if best:
            return best[2]

    text = _bicycle_match_text(pd)
    preferred_nodes: tuple[str, ...]
    if any(marker in text for marker in ("electric bicycle", "electric bike", "e-bike", "ebike", "500w", "750w", "48v", "电动自行车")):
        preferred_nodes = ("electric-bicycles",)
    elif any(marker in text for marker in ("folding bike", "foldable bike", "folding bicycle", "折叠自行车")):
        preferred_nodes = ("folding-bicycles",)
    elif any(marker in text for marker in ("freestyle", "bmx")):
        preferred_nodes = ("freestyle-bmx-bicycles", "bmx-bicycles")
    elif any(marker in text for marker in ("fat tire", "fat tyre", "fat bike", "胖胎")):
        preferred_nodes = ("fat-bicycles",)
    elif any(marker in text for marker in ("kids mountain", "children mountain", "girls mountain", "boys mountain")):
        preferred_nodes = ("childrens-mountain-bicycles",)
    elif "mountain" in text or "mtb" in text:
        preferred_nodes = ("mountain-bicycles",)
    elif "road bike" in text or "road bicycle" in text:
        preferred_nodes = ("road-bicycles",)
    elif any(marker in text for marker in ("cruiser", "city bike", "commuter", "step-through", "comfort bike")):
        preferred_nodes = ("cruiser-bicycles", "comfort-bicycles")
    elif any(marker in text for marker in ("balance bike", "balance bicycle")):
        preferred_nodes = ("childrens-balance-bikes",)
    elif _bicycle_is_child(pd):
        preferred_nodes = ("childrens-bicycles",)
    else:
        preferred_nodes = ("cruiser-bicycles",)

    for node in preferred_nodes:
        option = _find_browse_option(options, node)
        if option:
            return option

    best: tuple[int, int, dict[str, Any]] | None = None
    for index, option in enumerate(options):
        if not isinstance(option, dict):
            continue
        score = 0
        node = str(option.get("node") or "").lower()
        path = str(option.get("path") or "").lower()
        if node and node in text:
            score += 40
        if path and path in text:
            score += 40
        for marker in option.get("markers") or []:
            marker_text = str(marker or "").lower().strip()
            if marker_text and marker_text in text:
                score += 20 + min(len(marker_text), 30)
        if score and (best is None or score > best[0]):
            best = (score, index, option)
    if best:
        return best[2]
    return _find_browse_option(options, "childrens-bicycles")


def _apply_bicycle_category_fill(fill: dict[str, Any], mapping: dict, pd: ProductData) -> tuple[dict[str, Any] | None, list[str]]:
    warnings: list[str] = []
    option = _select_bicycle_category_option(mapping, pd)
    if not option:
        warnings.append("BICYCLE 模板未匹配到细分类目，沿用映射默认儿童自行车类目。")
        return None, warnings

    fill["product_type#1.value"] = option.get("product_type") or "BICYCLE"
    fill["item_type_keyword[marketplace_id=ATVPDKIKX0DER]#1.value"] = _amazon_item_type_keyword(option)
    return option, warnings


def _bicycle_is_child(pd: ProductData) -> bool:
    text = _bicycle_match_text(pd)
    if any(marker in text for marker in ("kids", "kid ", "children", "childrens", "children's", "boys", "girls", "ages 6", "ages 7", "ages 8", "ages 9", "ages 10", "儿童")):
        return True
    return bool(re.search(r"\b(?:12|14|16|18|20)\s*(?:inch|in\\.?|\")\b", text))


def _bicycle_material(pd: ProductData) -> str:
    text = _bicycle_match_text(pd)
    if "aluminum alloy" in text or "aluminium alloy" in text:
        return "Aluminum Alloy"
    if "aluminum" in text or "aluminium" in text:
        return "Aluminum"
    if "high carbon steel" in text:
        return "High Carbon Steel"
    if "carbon steel" in text:
        return "Carbon Steel"
    if "stainless steel" in text:
        return "Stainless Steel"
    if "steel" in text:
        return "Steel"
    if "iron" in text:
        return "Iron"
    return "Steel"


def _bicycle_color_value(pd: ProductData) -> str | None:
    if pd.color:
        return pd.color
    text = _bicycle_match_text(pd)
    colors = (
        "light pink",
        "pink",
        "black",
        "blue",
        "yellow",
        "red",
        "white",
        "green",
        "purple",
        "orange",
        "silver",
        "gray",
        "grey",
    )
    found = [color for color in colors if color in text]
    if not found:
        return None
    return found[0].title().replace("Grey", "Gray")


def _bicycle_color_map(pd: ProductData) -> str | None:
    color = (_bicycle_color_value(pd) or "").lower()
    standard_colors = {
        "black": "Black",
        "blue": "Blue",
        "pink": "Pink",
        "light pink": "Pink",
        "yellow": "Yellow",
        "red": "Red",
        "white": "White",
        "green": "Green",
        "purple": "Purple",
        "orange": "Orange",
        "silver": "Silver",
        "gray": "Gray",
        "grey": "Gray",
    }
    for marker, value in standard_colors.items():
        if marker in color:
            return value
    return None


def _bicycle_wheel_size(pd: ProductData) -> float | int | None:
    text = " ".join(str(value or "") for value in (pd.title, pd.listing_title, pd.product_type)).lower()
    match = re.search(r"\b(12|14|16|18|20|22|24|26|27(?:\\.5)?|27\\.5|29)\s*(?:inch|in\\.?|\")\b", text)
    if not match:
        return None
    value = float(match.group(1))
    return int(value) if value.is_integer() else value


def _bicycle_frame_size(pd: ProductData) -> str:
    wheel_size = _bicycle_wheel_size(pd)
    if isinstance(wheel_size, (int, float)):
        if wheel_size <= 20:
            return "Small"
        if wheel_size <= 24:
            return "Medium"
    return "Large"


def _bicycle_size(pd: ProductData) -> str:
    wheel_size = _bicycle_wheel_size(pd)
    if wheel_size:
        return f"{wheel_size:g} Inch"
    if pd.dimension_length and pd.dimension_width and pd.dimension_height:
        return f'{pd.dimension_length:g}" x {pd.dimension_width:g}" x {pd.dimension_height:g}"'
    return _bicycle_frame_size(pd)


def _bicycle_number_of_speeds(pd: ProductData) -> tuple[int, bool]:
    text = _bicycle_match_text(pd)
    match = re.search(r"\b(\d{1,2})\s*[- ]?speed\b", text)
    if match:
        return max(int(match.group(1)), 1), False
    if "single speed" in text:
        return 1, False
    return 1, True


def _bicycle_brake_styles(pd: ProductData) -> list[str]:
    text = _bicycle_match_text(pd)
    styles: list[str] = []
    if "disc brake" in text or "disc brakes" in text:
        styles.append("Disc")
    if "v-brake" in text or "v brake" in text or "v-brakes" in text:
        styles.append("V Brake")
    if "caliper" in text:
        styles.append("Caliper")
    if "coaster" in text:
        styles.append("Coaster")
    if "drum brake" in text:
        styles.append("Drum")
    if not styles:
        styles.append("V Brake")
    return list(dict.fromkeys(styles))


def _bicycle_suspension_type(pd: ProductData) -> str:
    text = _bicycle_match_text(pd)
    if "full suspension" in text or "dual suspension" in text:
        return "Full"
    if "rear suspension" in text:
        return "Rear"
    if "front suspension" in text or "suspension fork" in text or "fork suspension" in text:
        return "Front"
    return "Rigid"


def _bicycle_bike_type(option: dict[str, Any] | None, pd: ProductData) -> str:
    node = (option or {}).get("node") or ""
    text = _bicycle_match_text(pd)
    if node == "electric-bicycles" or "electric bike" in text or "e-bike" in text or "ebike" in text:
        return "Electric Bike"
    if node in {"childrens-mountain-bicycles", "mountain-bicycles"} or "mountain" in text or "mtb" in text:
        return "Mountain Bike"
    if node in {"folding-bicycles"}:
        return "Folding Bike"
    if node in {"cruiser-bicycles", "comfort-bicycles"} or "cruiser" in text:
        return "Cruiser Bike"
    if node == "road-bicycles":
        return "Road Bike"
    if "bmx" in node or "bmx" in text or "freestyle" in text:
        return "BMX Bike"
    if node == "childrens-balance-bikes":
        return "Balance Bike"
    if node.startswith("childrens"):
        return "Kids Bike"
    return "Bicycle"


def _bicycle_special_features(pd: ProductData) -> list[str]:
    text = _bicycle_match_text(pd)
    features: list[str] = []
    feature_markers = [
        ("fold", "Foldable"),
        ("basket", "Basket"),
        ("rear rack", "Rear Rack"),
        ("fender", "Fenders"),
        ("disc brake", "Disc Brakes"),
        ("suspension", "Suspension"),
        ("chain guard", "Chain Guard"),
        ("phone holder", "Phone Holder"),
        ("fat tire", "Fat Tire"),
        ("removable battery", "Removable Battery"),
    ]
    for marker, label in feature_markers:
        if marker in text:
            features.append(label)
    return list(dict.fromkeys(features))[:5]


def _bicycle_age_bounds(pd: ProductData) -> tuple[int | None, int | None]:
    text = _bicycle_match_text(pd)
    range_match = re.search(r"ages?\s*(\d{1,2})\s*[-~–]\s*(\d{1,2})", text)
    if range_match:
        minimum = int(range_match.group(1))
        maximum = int(range_match.group(2))
        return minimum, maximum
    plus_match = re.search(r"ages?\s*(\d{1,2})\s*\\+", text)
    if plus_match:
        minimum = int(plus_match.group(1))
        return minimum, None
    if _bicycle_is_child(pd):
        return 6, 12
    return None, None


def _bicycle_is_electric(option: dict[str, Any] | None, pd: ProductData) -> bool:
    text = _bicycle_match_text(pd)
    node = (option or {}).get("node")
    return node == "electric-bicycles" or any(marker in text for marker in ("electric bicycle", "electric bike", "e-bike", "ebike", "500w", "750w", "48v"))


def _bicycle_power_number(pd: ProductData, suffix: str) -> float | int | None:
    match = re.search(rf"\b(\d+(?:\.\d+)?)\s*{suffix}\b", _bicycle_match_text(pd))
    if not match:
        return None
    value = float(match.group(1))
    return int(value) if value.is_integer() else value


def _apply_bicycle_fill(fill: dict[str, Any], fields: dict[str, Any], product: Product, pd: ProductData, mapping: dict) -> list[str]:
    warnings: list[str] = []
    category_option, category_warnings = _apply_bicycle_category_fill(fill, mapping, pd)
    warnings.extend(category_warnings)

    material = _bicycle_material(pd)
    color = _bicycle_color_value(pd)
    wheel_size = _bicycle_wheel_size(pd)
    number_of_speeds, defaulted_speeds = _bicycle_number_of_speeds(pd)
    min_age, max_age = _bicycle_age_bounds(pd)
    is_electric = _bicycle_is_electric(category_option, pd)

    if defaulted_speeds:
        warnings.append("未识别自行车变速数量，按单速默认填写；上传前建议核对规格。")
    if not wheel_size:
        warnings.append("未识别车轮尺寸，Wheel Size 暂未填写；上传前建议补车轮尺寸。")
    if not color:
        warnings.append("未识别自行车颜色，Color 暂未填写；上传前建议补颜色。")

    fill.update({
        fields["material"]: material,
        fields["color_map"]: _bicycle_color_map(pd),
        fields["color"]: color,
        fields["size"]: _bicycle_size(pd),
        fields["part_number"]: pd.item_code,
        fields["unit_count"]: 1,
        fields["unit_count_type"]: "Count",
        fields["height_value"]: pd.dimension_height,
        fields["height_unit"]: "Inches",
        fields["length_value"]: pd.dimension_length,
        fields["length_unit"]: "Inches",
        fields["width_value"]: pd.dimension_width,
        fields["width_unit"]: "Inches",
        fields["item_dimensions_length_value"]: pd.dimension_length,
        fields["item_dimensions_length_unit"]: "Inches",
        fields["item_dimensions_width_value"]: pd.dimension_width,
        fields["item_dimensions_width_unit"]: "Inches",
        fields["item_dimensions_height_value"]: pd.dimension_height,
        fields["item_dimensions_height_unit"]: "Inches",
        fields["item_weight_value"]: pd.weight,
        fields["item_weight_unit"]: "Pounds",
        fields["wheel_size_value"]: wheel_size,
        fields["wheel_size_unit"]: "Inches" if wheel_size else None,
        fields["wheel_material"]: "Aluminum Alloy",
        fields["frame_material"]: material,
        fields["frame_material_structured"]: material,
        fields["frame_size"]: _bicycle_frame_size(pd),
        fields["frame_type_1"]: "Step-Through" if "step-through" in _bicycle_match_text(pd) else "Rigid",
        fields["bike_type_1"]: _bicycle_bike_type(category_option, pd),
        fields["power_source"]: "Electric" if is_electric else "Pedal Power",
        fields["suspension_type"]: _bicycle_suspension_type(pd),
        fields["skill_level"]: "Beginner" if _bicycle_is_child(pd) else "Intermediate",
        fields["handlebar_type"]: "Adjustable" if "adjustable" in _bicycle_match_text(pd) else "Flat",
        fields["seat_material"]: "Synthetic Leather",
        fields["wheel_set"]: "Front and Rear Wheels",
        fields["number_of_speeds"]: number_of_speeds,
        fields["bicycle_gear_shifter_type"]: "Trigger" if number_of_speeds > 1 else None,
        fields["age_range_description"]: (_semantic_values_from_listing_check(pd, "age_range_description") or [None])[0],
        fields["style"]: (_semantic_values_from_listing_check(pd, "style") or [None])[0],
    })
    _set_sequence_fields(fill, fields.get("brake_style"), _bicycle_brake_styles(pd))
    _set_sequence_fields(fill, fields.get("specific_uses_for_product"), _semantic_values_from_listing_check(pd, "specific_uses_for_product"))
    _set_sequence_fields(fill, fields.get("included_components"), _semantic_values_from_listing_check(pd, "included_components"))
    _set_sequence_fields(fill, fields.get("special_features"), _bicycle_special_features(pd))

    if min_age is not None:
        fill[fields["min_user_age"]] = min_age
    if max_age is not None:
        fill[fields["max_user_age"]] = max_age

    if is_electric:
        wattage = _bicycle_power_number(pd, "w")
        voltage = _bicycle_power_number(pd, "v")
        battery_capacity = _bicycle_power_number(pd, "ah")
        fill.update({
            fields["power_source"]: "Electric",
            fields["electric_assist_type"]: "Hub Motor",
            fields["wattage"]: wattage,
            fields["wattage_unit"]: "Watts" if wattage else None,
            fields["voltage_value"]: voltage,
            fields["voltage_unit"]: "Volts" if voltage else None,
            fields["batteries_required"]: "Yes",
            fields["batteries_included"]: "Yes",
            fields["battery_cell_composition"]: "Lithium Ion",
            fields["num_batteries_quantity"]: 1,
            fields["num_batteries_type"]: "Lithium Ion",
            fields["lithium_battery_packaging"]: "Batteries contained in equipment",
            fields["lithium_battery_energy_content_unit"]: "Watt Hours",
            fields["contains_battery_or_cell"]: "Battery",
            fields["battery_contains_free_unabsorbed_liquid"]: "No",
            fields["has_replaceable_battery"]: "Yes",
            fields["battery_installation_device_type"]: "Installed in Vehicle",
            fields["battery_capacity"]: battery_capacity,
            fields["battery_capacity_unit"]: "amp hours" if battery_capacity else None,
        })
        if voltage and battery_capacity:
            fill[fields["lithium_battery_energy_content"]] = round(float(voltage) * float(battery_capacity), 2)
        if not wattage:
            warnings.append("电动自行车未识别电机瓦数，Wattage 暂未填写；上传前需核对电机铭牌。")
        if not voltage:
            warnings.append("电动自行车未识别电池电压，Voltage 暂未填写；上传前需核对电池规格。")
        warnings.append("电动自行车电池重量、UL/认证编号、FCC/SDoC 等合规资料未自动填写，发布前需要人工复核。")

    if "tricycle" in _bicycle_match_text(pd) or "trike" in _bicycle_match_text(pd):
        warnings.append("标题包含 Tricycle/Trike，但当前 Amazon 模板按 Bicycle 生成；上传前请确认成人三轮车可放入所选电动自行车/自行车类目。")

    return warnings


def _ride_on_material(pd: ProductData) -> str:
    raw = _facts_text(pd).lower()
    if "metal" in raw or "steel" in raw:
        return "Plastic, Metal"
    if "wood" in raw:
        return "Plastic, Wood"
    return pd.material or "Plastic"


def _ride_on_voltage(pd: ProductData) -> tuple[int, list[str]]:
    warnings: list[str] = []
    raw = _facts_text(pd).lower()
    match = re.search(r"\b(6|12|18|24)\s*v(?:olt)?s?\b", raw)
    if match:
        return int(match.group(1)), warnings
    warnings.append("未识别骑乘玩具电压，按常见 12V 电瓶车默认填写；上传前建议核对电池铭牌/说明书。")
    return 12, warnings


def _ride_on_age_months(pd: ProductData) -> tuple[int, int | None, list[str]]:
    warnings: list[str] = []
    raw = _facts_text(pd).lower()
    range_match = re.search(r"(\d+)\s*[-~–]\s*(\d+)\s*(?:years?|yrs?|岁)", raw)
    if range_match:
        return int(range_match.group(1)) * 12, int(range_match.group(2)) * 12, warnings
    year_match = re.search(r"(?:ages?\s*)?(\d+)\s*\+?\s*(?:years?|yrs?|岁)", raw)
    if year_match:
        return int(year_match.group(1)) * 12, None, warnings
    month_match = re.search(r"(\d+)\s*(?:months?|月)", raw)
    if month_match:
        return int(month_match.group(1)), None, warnings
    warnings.append("未识别建议年龄，按 3 岁以上默认填写玩具合规年龄；上传前建议核对页面/说明书。")
    return 36, None, warnings


def _ride_on_seating_capacity(pd: ProductData) -> int:
    raw = _facts_text(pd).lower()
    if re.search(r"\b(2|two)\s*(?:seat|seater|kids|children)\b", raw):
        return 2
    return 1


def _ride_on_battery_profile(pd: ProductData) -> tuple[dict[str, Any], list[str]]:
    warnings: list[str] = []
    raw = _facts_text(pd).lower()
    voltage, voltage_warnings = _ride_on_voltage(pd)
    warnings.extend(voltage_warnings)

    if "lithium polymer" in raw or "li-polymer" in raw or "lipo" in raw:
        composition = "Lithium Polymer"
    elif "lithium" in raw or "li-ion" in raw or "lion" in raw:
        composition = "Lithium Ion"
    elif "nimh" in raw or "ni-mh" in raw:
        composition = "NiMH"
    elif "nicad" in raw or "ni-cd" in raw:
        composition = "NiCAD"
    elif "alkaline" in raw:
        composition = "Alkaline"
    else:
        composition = "Lead Acid"
        if not any(word in raw for word in ("lead acid", "lead-acid", "sealed lead", "sla")):
            warnings.append("未识别电池化学类型，按骑乘电瓶车常见 Lead Acid 默认填写；如实际为锂电池需改模板。")

    quantity = 1
    quantity_match = re.search(r"\b(\d+)\s*(?:x|pcs?|pieces?)?\s*(?:battery|batteries)\b", raw)
    if quantity_match:
        quantity = max(int(quantity_match.group(1)), 1)

    battery_type = f"{voltage}V" if voltage in (9, 12) else "Nonstandard Battery"
    profile: dict[str, Any] = {
        "voltage": voltage,
        "composition": composition,
        "quantity": quantity,
        "type": battery_type,
    }

    ah_match = re.search(r"(\d+(?:\.\d+)?)\s*ah\b", raw)
    if ah_match:
        profile["energy_content"] = round(voltage * float(ah_match.group(1)), 2)
    return profile, warnings


def _ride_on_size(pd: ProductData) -> str | None:
    if pd.dimension_length and pd.dimension_width and pd.dimension_height:
        return f'{pd.dimension_length:g}" L x {pd.dimension_width:g}" W x {pd.dimension_height:g}" H'
    return None


def _ride_on_assembly(pd: ProductData) -> tuple[str, int | None, str | None]:
    raw = _facts_text(pd).lower()
    if any(phrase in raw for phrase in ("no assembly required", "fully assembled", "ready to use out of box", "ready to use out of the box")):
        return "No", None, None
    return "Yes", 30, "Minutes"


def _apply_ride_on_toy_fill(fill: dict[str, Any], fields: dict[str, str], product: Product, pd: ProductData) -> list[str]:
    warnings: list[str] = []
    age_min, age_max, age_warnings = _ride_on_age_months(pd)
    battery, battery_warnings = _ride_on_battery_profile(pd)
    warnings.extend(age_warnings)
    warnings.extend(battery_warnings)
    voltage = battery["voltage"]

    age_years = max(age_min // 12, 1)
    age_description = f"{age_years} years and up"
    if age_max:
        age_description = f"{age_min // 12}-{age_max // 12} years"
    assembly_required, assembly_time, assembly_time_unit = _ride_on_assembly(pd)

    ships_globally = str(fill.get(fields.get("ships_globally", ""), "No") or "No").strip().lower() == "yes"

    fill.update({
        fields["age_range_description"]: age_description,
        fields["material"]: _ride_on_material(pd),
        fields["color"]: pd.color,
        fields["size"]: _ride_on_size(pd),
        fields["part_number"]: pd.item_code,
        fields["is_assembly_required"]: assembly_required,
        fields["assembly_time"]: assembly_time,
        fields["assembly_time_unit"]: assembly_time_unit,
        fields["seating_capacity"]: _ride_on_seating_capacity(pd),
        fields["voltage_value"]: voltage,
        fields["voltage_unit"]: "Volts",
        fields["unit_count"]: 1,
        fields["unit_count_type"]: "Count",
        fields["manufacturer_min_age"]: age_min,
        fields["manufacturer_max_age"]: age_max,
        fields["height_value"]: pd.dimension_height,
        fields["height_unit"]: "Inches",
        fields["length_value"]: pd.dimension_length,
        fields["length_unit"]: "Inches",
        fields["width_value"]: pd.dimension_width,
        fields["width_unit"]: "Inches",
        fields["item_weight_value"]: pd.weight,
        fields["item_weight_unit"]: "Pounds",
        fields["buyer_age_restrictions"]: "No",
        fields["cpsia_cautionary_statement_1"]: "No Warning Applicable",
        fields["pesticide_marking_type_1"]: "EPA Establishment Number",
        fields["pesticide_registration_status_1"]: "This product is not a pesticide or pesticide device, as defined under the U.S. Federal Insecticide, Fungicide, and Rodenticide Act.",
        fields["has_multiple_battery_powered_components"]: "No",
        fields["contains_battery_or_cell"]: "Battery",
        fields["battery_contains_free_unabsorbed_liquid"]: "No",
        fields["is_battery_non_spillable"]: "Yes" if battery["composition"] in ("Lead Acid", "Wet Alkali") else None,
        fields["non_lithium_battery_packaging"]: "Batteries contained in equipment" if not battery["composition"].startswith("Lithium") else None,
        fields["has_replaceable_battery"]: "Yes",
        fields["battery_installation_device_type"]: "Installed in Vehicle",
        fields["batteries_required"]: "Yes",
        fields["batteries_included"]: "Yes",
        fields["battery_cell_composition"]: battery["composition"],
        fields["num_batteries_quantity"]: battery["quantity"],
        fields["num_batteries_type"]: battery["type"],
    })
    _field(fill, fields, "sub_brand", product.brand)

    if ships_globally:
        fill.update({
            fields["compliance_toy_voltage"]: "Up to 12 Volts" if voltage <= 12 else "More than 12 Volts",
            fields["compliance_toy_type"]: "Cars",
            fields["compliance_recommended_age"]: "Over 3 Years of Age" if age_min >= 36 else "Under 3 Years of Age",
            fields["compliance_operation_mode"]: "Electronic - Battery",
        })
    else:
        warnings.append("RIDE_ON_TOY 默认按美国 FBM 非全球配送生成，已跳过 Ships Globally 条件下才允许的 Toy Compliance 字段。")

    if battery.get("energy_content") and not battery["composition"].startswith("Lithium"):
        fill[fields["non_lithium_battery_energy_content"]] = battery["energy_content"]
        fill[fields["non_lithium_battery_energy_content_unit"]] = "Watt Hours"

    if battery["composition"].startswith("Lithium"):
        fill[fields["lithium_battery_packaging"]] = "Batteries contained in equipment"
        if battery.get("energy_content"):
            fill[fields["lithium_battery_energy_content"]] = battery["energy_content"]
            fill[fields["lithium_battery_energy_content_unit"]] = "Watt Hours"
        warnings.append("识别为锂电池类骑乘玩具，锂电池重量/电芯数量/UN运输信息可能还需人工补充。")

    raw = _facts_text(pd).lower()
    if any(word in raw for word in ("remote", "2.4g", "bluetooth", "wireless")):
        warnings.append("商品疑似含遥控/无线功能，模板未自动填写 FCC ID 或 SDoC 联系资料；如上架报错需补 FCC 合规资料。")
    warnings.append("未自动填写 UL/TUV/Intertek 等监管证书编号或合规媒体 URL；这些字段需要真实证书资料后再补。")

    _set_sequence_fields(fill, [fields.get("target_audience_1"), fields.get("target_audience_2")], _semantic_values_from_listing_check(pd, "target_audience"))
    _fields(fill, fields, "target_gender", _semantic_values_from_listing_check(pd, "target_gender"))
    _set_sequence_fields(fill, [fields.get("theme_1")], _semantic_values_from_listing_check(pd, "theme"))
    _set_sequence_fields(fill, fields.get("included_components"), _semantic_values_from_listing_check(pd, "included_components"))

    return warnings


def _build_amazon_template_file(product: Product, pd: ProductData, mapping: dict) -> dict:
    """同步生成 Excel 文件。调用方应放到线程中执行，避免阻塞 API 事件循环。"""
    from app.pipeline.amazon_export.writer import build_amazon_template_file

    return build_amazon_template_file(product, pd, mapping)


async def run_amazon_template(product_id: int) -> dict:
    """生成 Amazon 类目导入模板。"""
    async with async_session() as db:
        result = await db.execute(
            select(Product)
            .options(
                selectinload(Product.data),
                selectinload(Product.images),
                selectinload(Product.aplus),
                selectinload(Product.catalog_item),
            )
            .where(Product.id == product_id)
        )
        product = result.scalar_one_or_none()
        if not product or not product.data:
            raise ValueError(f"Product {product_id} not found or no data")

        pd = product.data
        mapping = _load_template_mapping(product, pd)
        template_path = Path(mapping["template_path"])
        if not template_path.is_file():
            raise FileNotFoundError(f"模板文件不存在: {template_path}")
        if not pd.item_code:
            raise ValueError("缺少商品Code，无法生成SKU")
        if not pd.listing_title or not pd.listing_bullets:
            raise ValueError("缺少Listing文案，请先执行Step5")
        await ensure_amazon_template_semantic_fields(product, pd, mapping, template_path)
        await db.commit()

        if not product.upc:
            from app.services.upc_pool import UpcPoolEmptyError, ensure_product_upc

            try:
                await ensure_product_upc(db, product)
            except UpcPoolEmptyError as exc:
                raise ValueError(str(exc)) from exc
        if product.catalog_item:
            product.catalog_item.upc = product.upc
        await db.flush()

        product_snapshot = _snapshot_model(product)
        product_snapshot.data = _snapshot_model(product.data)
        product_snapshot.images = _snapshot_model(product.images)
        product_snapshot.aplus = _snapshot_model(product.aplus)
        pd_snapshot = product_snapshot.data

        try:
            template_result = await asyncio.to_thread(_build_amazon_template_file, product_snapshot, pd_snapshot, mapping)
        except Exception:
            await db.rollback()
            raise
        await db.commit()

    output_path = Path(template_result["path"])

    async with async_session() as db:
        result = await db.execute(select(Product).options(selectinload(Product.data)).where(Product.id == product_id))
        product = result.scalar_one_or_none()
        if not product or not product.data:
            raise ValueError(f"Product {product_id} not found or no data")
        pd = product.data
        pd.amazon_template_path = template_result["path"]
        pd.amazon_template_warnings = json.dumps(template_result["warnings"], ensure_ascii=False)
        pd.amazon_template_fill_summary = json.dumps(template_result["fill_summary"], ensure_ascii=False)
        pd.amazon_template_generated_at = datetime.now()
        file_result = await db.execute(
            select(ProductFile).where(
                ProductFile.product_id == product.id,
                ProductFile.file_type == "amazon_import_template",
                ProductFile.path == template_result["path"],
            )
        )
        product_file = file_result.scalar_one_or_none()
        metadata_json = json.dumps({
            "warnings": template_result["warnings"],
            "uploaded_images": template_result["uploaded_images"],
            "fill_summary": template_result["fill_summary"],
            "filled_fields": template_result["filled_fields"],
        }, ensure_ascii=False)
        if product_file:
            product_file.label = "Amazon导入表格"
            product_file.directory = str(output_path.parent)
            product_file.metadata_json = metadata_json
            product_file.updated_at = datetime.now()
        else:
            db.add(ProductFile(
                product_id=product.id,
                file_type="amazon_import_template",
                label="Amazon导入表格",
                path=template_result["path"],
                directory=str(output_path.parent),
                metadata_json=metadata_json,
                created_at=datetime.now(),
                updated_at=datetime.now(),
            ))
        await db.commit()

        logger.info(f"[Step10] Amazon导入模板已生成: {output_path}")
        return template_result
