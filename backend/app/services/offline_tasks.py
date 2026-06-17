import asyncio
import json
import logging
from datetime import datetime
from io import BytesIO
from pathlib import Path
from zipfile import ZipFile

from openpyxl import load_workbook
from sqlalchemy import func, select, update
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.api.schemas import OfflineTaskCatalogExportRequest, OfflineTaskGigaDynamicSyncRequest, OfflineTaskGigaPullRequest
from app.config import settings
from app.database import async_session
from app.models import CatalogProduct, OfflineTask, OfflineTaskStep, Product, ProductAplus, ProductDataSource
from app.models.status import COMPLETED, FAILED, PAUSED, STEP6_CURATING
from app.pipeline.engine import get_step_status, is_running, run_pipeline_tracked
from app.pipeline.step7_aplus_plan import run_aplus_plan
from app.pipeline.step8_aplus_script import run_aplus_script
from app.pipeline.step9_aplus_image import run_aplus_image
from app.services.giga_inventory_sync import GigaInventorySyncOptions, sync_giga_inventory_snapshot
from app.services.giga_image_download_tasks import download_giga_batch_images
from app.services.giga_openapi import GigaSyncOptions, resolve_giga_data_source_context, sync_giga_products
from app.services.giga_price_sync import GigaPriceSyncOptions, sync_giga_price_snapshot
from app.services.oss_uploader import upload_private_file
from app.services.stylesnap_product_tasks import upsert_product_drafts_from_giga_batch
from app.task_planners.product_image_analysis import create_product_image_analysis_runs
from app.task_planners.product_listing import create_product_listing_runs

logger = logging.getLogger(__name__)

_active_offline_tasks: dict[int, asyncio.Task] = {}

TASK_STATUS_ACTIVE = {"pending", "running"}
STEP_STATUS_SUCCESS = "done"
STEP_STATUS_FAILURES = {"failed", "interrupted"}
STEP_STATUS_RESUMABLE = {"pending", "running", "paused", "interrupted"}
STEP_STATUS_CLAIMABLE = {"pending", "interrupted"}
SUPPORTED_TASK_TYPES = {
    "giga_pull",
    "giga_inventory_sync",
    "giga_price_sync",
    "aplus_generate",
    "catalog_export",
    "product_bulk_advance",
}
APLUS_GENERATE_ACTIVE_STATUSES = {"queued", "planning", "scripting", "imaging"}


def _json_dumps(value: object) -> str:
    return json.dumps(value, ensure_ascii=False, default=str)


def _json_loads(value: str | None, default: object | None = None) -> object:
    if not value:
        return {} if default is None else default
    try:
        return json.loads(value)
    except Exception:
        return {} if default is None else default


def _safe_filename_part(value: str | None, fallback: str = "export") -> str:
    raw = (value or fallback).strip() or fallback
    cleaned = "".join(ch if ch.isalnum() or ch in ("-", "_") else "_" for ch in raw)
    return cleaned[:90] or fallback


def _catalog_export_object_key(task_id: int, filename: str) -> str:
    prefix = settings.OSS_EXPORT_UPLOAD_PREFIX.strip().strip("/")
    raw = Path(filename)
    safe_filename = _safe_filename_part(raw.stem, "catalog_export") + (raw.suffix.lower() or ".zip")
    key = f"task_{task_id}/{safe_filename}"
    return f"{prefix}/{key}" if prefix else key


def _catalog_export_result_ready(payload: object) -> bool:
    if not isinstance(payload, dict):
        return False
    file_path = str(payload.get("file_path") or "").strip()
    object_key = str(payload.get("oss_object_key") or "").strip()
    if file_path and Path(file_path).expanduser().is_file():
        return True
    return bool(object_key and payload.get("filename"))


def _catalog_export_row_status(row: dict) -> str:
    status = str(row.get("状态") or "").strip()
    if status == "已导出":
        return "exported"
    if status == "失败":
        return "failed"
    return "skipped"


def _catalog_export_result_rows(report_rows: list[dict]) -> list[dict]:
    rows: list[dict] = []
    for row in report_rows:
        rows.append({
            "catalog_id": row.get("商品资料ID"),
            "product_id": row.get("商品ID"),
            "item_code": row.get("商品Code"),
            "category": row.get("类目"),
            "status": _catalog_export_row_status(row),
            "reason": row.get("原因"),
            "template_file": row.get("模板文件"),
            "output_file": row.get("导出文件"),
        })
    return rows


def _catalog_export_result_payload(
    *,
    category: str,
    categories: list[str],
    template_name: str | None,
    template_path: object,
    catalog_ids: list[int],
    report_rows: list[dict],
    created_at: datetime,
    filename: str | None = None,
    file_path: str | None = None,
    oss_object_key: str | None = None,
    oss_url: str | None = None,
    file_size: int | None = None,
) -> dict:
    rows = _catalog_export_result_rows(report_rows)
    success_count = sum(1 for row in rows if row["status"] == "exported")
    skipped_count = sum(1 for row in rows if row["status"] == "skipped")
    failed_count = sum(1 for row in rows if row["status"] == "failed")
    status = "failed"
    if success_count and (skipped_count or failed_count):
        status = "partial_failed"
    elif success_count:
        status = "done"
    return {
        "status": status,
        "category": category,
        "categories": categories or [category],
        "template_name": template_name or None,
        "template_path": template_path,
        "catalog_product_ids": catalog_ids,
        "requested_count": len(catalog_ids),
        "success_count": success_count,
        "exported_count": success_count,
        "skipped_count": skipped_count,
        "failed_count": failed_count,
        "report_count": len(report_rows),
        "filename": filename,
        "file_path": file_path,
        "oss_object_key": oss_object_key,
        "oss_url": oss_url,
        "file_size": file_size,
        "report_filename": "导出报告.xlsx" if report_rows else None,
        "created_at": created_at.isoformat(),
        "rows": rows,
    }


def _report_rows_from_export_zip(path: Path) -> list[dict]:
    with ZipFile(path) as archive:
        report_name = next((name for name in archive.namelist() if name.lower().endswith(".xlsx")), None)
        if not report_name:
            return []
        workbook = load_workbook(BytesIO(archive.read(report_name)), read_only=True, data_only=True)
        sheet = workbook[workbook.sheetnames[0]]
        rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(value or "").strip() for value in rows[0]]
    report_rows: list[dict] = []
    for row in rows[1:]:
        record = {
            header: row[index] if index < len(row) else None
            for index, header in enumerate(headers)
            if header
        }
        if any(value not in (None, "") for value in record.values()):
            report_rows.append(record)
    return report_rows


def _recover_catalog_export_result_from_file(
    *,
    export_dir: Path,
    category: str,
    categories: list[str],
    template_name: str | None,
    template_path: object,
    catalog_ids: list[int],
) -> dict | None:
    existing_files = sorted(
        export_dir.glob("*.zip"),
        key=lambda path: path.stat().st_mtime if path.exists() else 0,
        reverse=True,
    )
    for path in existing_files:
        try:
            report_rows = _report_rows_from_export_zip(path)
        except Exception:
            logger.warning("[OfflineTask] cannot recover catalog export result from %s", path, exc_info=True)
            continue
        if not report_rows:
            continue
        return _catalog_export_result_payload(
            category=category,
            categories=categories,
            template_name=template_name,
            template_path=template_path,
            catalog_ids=catalog_ids,
            report_rows=report_rows,
            created_at=datetime.fromtimestamp(path.stat().st_mtime),
            filename=path.name,
            file_path=str(path),
            file_size=path.stat().st_size,
        )
    return None


def _product_bulk_advance_row_status(step: OfflineTaskStep) -> tuple[str, str]:
    if step.status == "done":
        return "done", "已推进到待导出"
    if step.status == "running":
        return "running", "任务中心正在推进"
    if step.status == "pending":
        return "queued", "已加入任务中心队列"
    if step.status == "paused":
        return "paused", step.error_message or "任务已挂起"
    if step.status == "interrupted":
        return "interrupted", step.error_message or "任务中心步骤被中断，可重跑"
    if step.status == "failed":
        return "failed", step.error_message or "推进失败"
    return step.status, step.error_message or step.status


def _product_bulk_advance_step_row(step: OfflineTaskStep) -> dict | None:
    payload = _json_loads(step.payload_json, {})
    result = _json_loads(step.result_json, {})
    if not isinstance(payload, dict):
        return None
    product_id = payload.get("product_id")
    if product_id is None:
        return None
    status, reason = _product_bulk_advance_row_status(step)
    row = {
        "product_id": product_id,
        "item_code": payload.get("item_code"),
        "status": result.get("status") if isinstance(result, dict) and result.get("status") else status,
        "reason": result.get("reason") if isinstance(result, dict) and result.get("reason") else reason,
        "current_status": payload.get("current_status"),
        "current_step": payload.get("current_step"),
        "start_step": payload.get("start_step"),
    }
    if isinstance(result, dict):
        for key in ("latest_status", "latest_step", "latest_result", "latest_reason"):
            if key in result:
                row[key] = result[key]
    return row


def _refresh_product_bulk_advance_result(task: OfflineTask, steps: list[OfflineTaskStep]) -> None:
    payload = _json_loads(task.result_json, {})
    if not isinstance(payload, dict):
        payload = {}
    existing_rows = payload.get("rows") if isinstance(payload.get("rows"), list) else []
    rows_by_id = {
        str(row.get("product_id")): dict(row)
        for row in existing_rows
        if isinstance(row, dict) and row.get("product_id") is not None
    }
    for step in steps:
        if step.step_type != "product_bulk_advance_product":
            continue
        row = _product_bulk_advance_step_row(step)
        if not row:
            continue
        existing = rows_by_id.get(str(row["product_id"]), {})
        existing.update(row)
        rows_by_id[str(row["product_id"])] = existing
    rows = list(rows_by_id.values())
    payload["rows"] = rows
    payload["requested_count"] = int(payload.get("requested_count") or len(rows))
    payload["started_count"] = sum(
        1 for row in rows
        if row.get("status") in {"queued", "running", "done", "failed", "interrupted", "paused"}
    )
    payload["skipped_count"] = sum(1 for row in rows if row.get("status") == "skipped")
    payload["failed_count"] = sum(1 for row in rows if row.get("status") in {"failed", "interrupted"})
    payload["done_count"] = sum(1 for row in rows if row.get("status") == "done")
    payload["status"] = task.status
    task.result_json = _json_dumps(payload)


async def _refresh_task_stats(db: AsyncSession, task_id: int) -> OfflineTask | None:
    task = await db.get(OfflineTask, task_id)
    if not task:
        return None
    result = await db.execute(select(OfflineTaskStep).where(OfflineTaskStep.task_id == task_id))
    steps = result.scalars().all()
    total = len(steps)
    success = sum(1 for step in steps if step.status == STEP_STATUS_SUCCESS)
    failed = sum(1 for step in steps if step.status in STEP_STATUS_FAILURES)
    running = sum(1 for step in steps if step.status == "running")
    pending = sum(1 for step in steps if step.status == "pending")
    paused = sum(1 for step in steps if step.status == "paused")
    task.total_steps = total
    task.success_steps = success
    task.failed_steps = failed
    task.running_steps = running
    now = datetime.now()
    task.updated_at = now
    if paused or task.status == "paused":
        task.status = "paused"
        task.finished_at = None
    elif running:
        task.status = "running"
        task.started_at = task.started_at or now
        task.finished_at = None
    elif pending:
        task.status = "pending" if not task.started_at else "running"
        task.finished_at = None
    elif total and success == total:
        task.status = "done"
        task.finished_at = task.finished_at or now
    elif failed and success:
        task.status = "partial_failed"
        task.finished_at = task.finished_at or now
    elif failed:
        task.status = "failed"
        task.finished_at = task.finished_at or now
    else:
        task.status = "pending"
        task.finished_at = None
    if task.task_type == "catalog_export" and task.status == "done":
        payload = _json_loads(task.result_json, {})
        if not isinstance(payload, dict) or not payload:
            for step in steps:
                if step.step_type == "catalog_export_template" and step.status == STEP_STATUS_SUCCESS:
                    payload = _json_loads(step.result_json, {})
                    if isinstance(payload, dict) and payload:
                        break
        if isinstance(payload, dict) and payload.get("status") == "partial_failed":
            task.status = "partial_failed"
    if task.task_type == "product_bulk_advance":
        _refresh_product_bulk_advance_result(task, steps)
    return task


async def _set_step_status(
    db: AsyncSession,
    step: OfflineTaskStep,
    status: str,
    *,
    result: dict | None = None,
    error: str | None = None,
    progress_current: int | None = None,
    progress_total: int | None = None,
) -> None:
    now = datetime.now()
    step.status = status
    step.updated_at = now
    if status == "running":
        step.started_at = step.started_at or now
        step.finished_at = None
        step.error_message = None
    if status in {"done", "failed", "interrupted"}:
        step.finished_at = now
    if result is not None:
        step.result_json = _json_dumps(result)
    if error is not None:
        step.error_message = error
    if progress_current is not None:
        step.progress_current = progress_current
    if progress_total is not None:
        step.progress_total = progress_total
    await _refresh_task_stats(db, step.task_id)
    await db.commit()


async def _claim_offline_step(db: AsyncSession, step_id: int) -> OfflineTaskStep | None:
    now = datetime.now()
    result = await db.execute(
        update(OfflineTaskStep)
        .where(
            OfflineTaskStep.id == step_id,
            OfflineTaskStep.status.in_(tuple(STEP_STATUS_CLAIMABLE)),
        )
        .values(
            status="running",
            started_at=now,
            finished_at=None,
            error_message=None,
            updated_at=now,
        )
    )
    if (result.rowcount or 0) != 1:
        await db.rollback()
        return None
    await db.commit()
    claimed = await db.get(OfflineTaskStep, step_id)
    if claimed:
        await _refresh_task_stats(db, claimed.task_id)
        await db.commit()
    return claimed


async def _ensure_image_step(db: AsyncSession, sync_step: OfflineTaskStep) -> OfflineTaskStep:
    """Legacy helper for old GIGA image download tasks; new pull tasks keep image URLs and download selected images on demand."""
    result = await db.execute(
        select(OfflineTaskStep).where(
            OfflineTaskStep.task_id == sync_step.task_id,
            OfflineTaskStep.step_type == "giga_image_download",
            OfflineTaskStep.batch_id == sync_step.batch_id,
            OfflineTaskStep.data_source_id == sync_step.data_source_id,
        )
    )
    existing = result.scalar_one_or_none()
    if existing:
        existing.status = "pending" if existing.status in STEP_STATUS_FAILURES else existing.status
        existing.error_message = None if existing.status == "pending" else existing.error_message
        existing.updated_at = datetime.now()
        await db.commit()
        return existing

    step = OfflineTaskStep(
        task_id=sync_step.task_id,
        step_type="giga_image_download",
        title=f"下载图片：{sync_step.data_source_name or sync_step.data_source_id}",
        status="pending",
        data_source_id=sync_step.data_source_id,
        data_source_name=sync_step.data_source_name,
        site=sync_step.site,
        batch_id=sync_step.batch_id,
        progress_current=0,
        progress_total=0,
        payload_json=_json_dumps({"batch_id": sync_step.batch_id, "site": sync_step.site}),
        updated_at=datetime.now(),
    )
    db.add(step)
    await _refresh_task_stats(db, sync_step.task_id)
    await db.commit()
    await db.refresh(step)
    return step


async def _run_image_step(db: AsyncSession, step: OfflineTaskStep) -> None:
    await _set_step_status(db, step, "running")

    async def update_progress(counts: dict[str, int]) -> None:
        await db.refresh(step)
        task = await db.get(OfflineTask, step.task_id)
        if task and task.status == "paused":
            raise asyncio.CancelledError()
        step.progress_current = counts.get("done", 0)
        step.progress_total = counts.get("total", 0)
        pending = counts.get("pending", 0)
        failed = counts.get("failed", 0)
        step.result_json = _json_dumps(counts)
        step.error_message = f"图片下载中：待下载 {pending}，失败 {failed}" if pending or failed else None
        step.updated_at = datetime.now()
        await _refresh_task_stats(db, step.task_id)
        await db.commit()

    try:
        result = await download_giga_batch_images(
            batch_id=step.batch_id or "",
            site=step.site or "US",
            data_source_id=step.data_source_id,
            progress_callback=update_progress,
        )
        await upsert_product_drafts_from_giga_batch(
            db,
            batch_id=step.batch_id or "",
            site=step.site or "US",
            data_source_id=step.data_source_id,
        )
        await _set_step_status(
            db,
            step,
            "done" if result.get("failed", 0) == 0 else "failed",
            result=result,
            error=None if result.get("failed", 0) == 0 else f"{result.get('failed', 0)} 张图片下载失败",
            progress_current=result.get("done", 0),
            progress_total=result.get("total", 0),
        )
    except Exception as exc:
        await db.rollback()
        await _set_step_status(db, step, "failed", error=f"{type(exc).__name__}: {exc}")


async def _run_sync_step(db: AsyncSession, step: OfflineTaskStep) -> None:
    await _set_step_status(db, step, "running")
    payload = json.loads(step.payload_json or "{}")

    async def update_progress(live: dict) -> None:
        await db.refresh(step)
        task = await db.get(OfflineTask, step.task_id)
        if task and task.status == "paused":
            raise asyncio.CancelledError()
        existing = _json_loads(step.result_json, {})
        if not isinstance(existing, dict):
            existing = {}
        merged = {**existing, **live, "live": live}
        step.result_json = _json_dumps(merged)
        step.progress_current = int(live.get("progress_current") or live.get("scanned_sku_count") or step.progress_current or 0)
        if live.get("progress_total") is not None:
            step.progress_total = int(live.get("progress_total") or 0)
        step.error_message = live.get("current_message") or step.error_message
        step.updated_at = datetime.now()
        if task:
            task.result_json = _json_dumps(merged)
            task.updated_at = step.updated_at
        await _refresh_task_stats(db, step.task_id)
        await db.commit()

    try:
        result = await sync_giga_products(
            db,
            GigaSyncOptions(
                task_id=f"offline-task-{step.task_id}-step-{step.id}",
                batch_id=step.batch_id or "",
                site=step.site or "US",
                data_source_id=step.data_source_id,
                current_category=payload.get("current_category"),
                page_size=payload.get("page_size") or 200,
                max_pages=payload.get("max_pages"),
                skip_existing=True,
                download_images=False,
                progress_callback=update_progress,
            ),
        )
        await update_progress({
            "current_phase": "materializing_product_drafts",
            "scanned_sku_count": result.raw_sku_count + result.skipped_existing_count,
            "synced_sku_count": result.sku_count,
            "detail_count": result.raw_sku_count,
            "price_count": result.price_count,
            "inventory_count": result.inventory_count,
            "skipped_existing_count": result.skipped_existing_count,
            "item_count": result.item_count,
            "group_count": result.group_count,
            "progress_current": result.raw_sku_count + result.skipped_existing_count,
            "progress_total": result.raw_sku_count + result.skipped_existing_count,
            "current_message": "GIGA snapshot 已完成，正在生成商品工作台草稿",
        })
        draft_result = await upsert_product_drafts_from_giga_batch(
            db,
            batch_id=result.batch_id,
            site=result.site,
            data_source_id=result.data_source_id,
        )
        live_done = {
            "current_phase": "done",
            "scanned_sku_count": result.raw_sku_count + result.skipped_existing_count,
            "synced_sku_count": result.sku_count,
            "detail_count": result.raw_sku_count,
            "price_count": result.price_count,
            "inventory_count": result.inventory_count,
            "image_url_count": result.sku_count,
            "skipped_existing_count": result.skipped_existing_count,
            "item_count": result.item_count,
            "group_count": result.group_count,
            "product_draft_created_count": getattr(draft_result, "created_count", 0),
            "product_draft_updated_count": getattr(draft_result, "updated_count", 0),
            "progress_current": result.raw_sku_count + result.skipped_existing_count,
            "progress_total": result.raw_sku_count + result.skipped_existing_count,
            "current_message": "店铺商品同步完成",
        }
        await _set_step_status(
            db,
            step,
            "done",
            result={**result.__dict__, "product_drafts": draft_result.__dict__, **live_done, "live": live_done},
            progress_current=live_done["progress_current"],
            progress_total=live_done["progress_total"],
        )
        task = await db.get(OfflineTask, step.task_id)
        if task:
            task.result_json = _json_dumps({**result.__dict__, "product_drafts": draft_result.__dict__, **live_done, "live": live_done})
            task.updated_at = datetime.now()
            await db.commit()
    except Exception as exc:
        await db.rollback()
        logger.exception("[OfflineTask] GIGA pull step failed: task=%s step=%s", step.task_id, step.id)
        await _set_step_status(db, step, "failed", error=f"{type(exc).__name__}: {exc}")


async def _run_inventory_sync_step(db: AsyncSession, step: OfflineTaskStep) -> None:
    await _set_step_status(db, step, "running")
    payload = json.loads(step.payload_json or "{}")
    try:
        result = await sync_giga_inventory_snapshot(
            db,
            GigaInventorySyncOptions(
                task_id=f"offline-task-{step.task_id}-step-{step.id}",
                batch_id=step.batch_id or "",
                site=step.site or "US",
                data_source_id=step.data_source_id or 0,
                sku_codes=payload.get("sku_codes") or [],
            ),
        )
        await _set_step_status(
            db,
            step,
            "done" if result.failed_count == 0 else "failed",
            result=result.__dict__,
            error=None if result.failed_count == 0 else f"{result.failed_count} 个 SKU 库存同步失败",
            progress_current=result.success_count,
            progress_total=result.total_skus,
        )
    except Exception as exc:
        await db.rollback()
        logger.exception("[OfflineTask] GIGA inventory sync step failed: task=%s step=%s", step.task_id, step.id)
        await _set_step_status(db, step, "failed", error=f"{type(exc).__name__}: {exc}")


async def _run_price_sync_step(db: AsyncSession, step: OfflineTaskStep) -> None:
    await _set_step_status(db, step, "running")
    payload = json.loads(step.payload_json or "{}")
    try:
        result = await sync_giga_price_snapshot(
            db,
            GigaPriceSyncOptions(
                task_id=f"offline-task-{step.task_id}-step-{step.id}",
                batch_id=step.batch_id or "",
                site=step.site or "US",
                data_source_id=step.data_source_id or 0,
                sku_codes=payload.get("sku_codes") or [],
            ),
        )
        await _set_step_status(
            db,
            step,
            "done" if result.failed_count == 0 else "failed",
            result=result.__dict__,
            error=None if result.failed_count == 0 else f"{result.failed_count} 个 SKU 价格同步失败",
            progress_current=result.success_count,
            progress_total=result.total_skus,
        )
    except Exception as exc:
        await db.rollback()
        logger.exception("[OfflineTask] GIGA price sync step failed: task=%s step=%s", step.task_id, step.id)
        await _set_step_status(db, step, "failed", error=f"{type(exc).__name__}: {exc}")


async def _set_aplus_status(
    product_id: int,
    status: str,
    *,
    error: str | None = None,
    clear_outputs: bool = False,
) -> None:
    async with async_session() as session:
        result = await session.execute(
            select(Product)
            .where(Product.id == product_id)
            .options(
                selectinload(Product.aplus),
                selectinload(Product.catalog_item),
            )
        )
        product = result.scalar_one_or_none()
        if not product:
            return
        if not product.aplus:
            product.aplus = ProductAplus(product_id=product.id)
            session.add(product.aplus)
            await session.flush()
        if clear_outputs:
            product.aplus.aplus_plan = None
            product.aplus.aplus_plan_summary = None
            product.aplus.aplus_scripts = None
            product.aplus.aplus_scripts_summary = None
            product.aplus.aplus_images = None
            product.aplus.aplus_image_count = None
            product.aplus.planned_at = None
            product.aplus.scripted_at = None
            product.aplus.generated_at = None
        product.aplus.aplus_status = status
        if error:
            product.error_message = error
        elif product.error_message and product.error_message.startswith("A+生成"):
            product.error_message = None
        product.updated_at = datetime.now()
        if product.catalog_item:
            product.catalog_item.updated_at = product.updated_at
        await session.commit()


async def _ensure_task_not_paused(db: AsyncSession, task_id: int) -> None:
    task = await db.get(OfflineTask, task_id)
    if task and task.status == "paused":
        raise asyncio.CancelledError()


async def _run_aplus_generate_step(db: AsyncSession, step: OfflineTaskStep) -> None:
    payload = _json_loads(step.payload_json, {})
    product_id = int(payload.get("product_id") or 0) if isinstance(payload, dict) else 0
    force = bool(payload.get("force")) if isinstance(payload, dict) else False
    if product_id <= 0:
        await _set_step_status(db, step, "failed", error="A+生成步骤缺少 product_id")
        return

    try:
        await _set_step_status(db, step, "running", progress_current=0, progress_total=3)
        await _set_aplus_status(product_id, "planning", clear_outputs=force)
        await run_aplus_plan(product_id)
        await _ensure_task_not_paused(db, step.task_id)
        await _set_step_status(
            db,
            step,
            "running",
            result={"stage": "规划完成"},
            progress_current=1,
            progress_total=3,
        )

        await _set_aplus_status(product_id, "scripting")
        await run_aplus_script(product_id)
        await _ensure_task_not_paused(db, step.task_id)
        await _set_step_status(
            db,
            step,
            "running",
            result={"stage": "脚本完成"},
            progress_current=2,
            progress_total=3,
        )

        await _set_aplus_status(product_id, "imaging")
        image_result = await run_aplus_image(product_id)
        await _ensure_task_not_paused(db, step.task_id)
        await _set_aplus_status(product_id, "done")
        await _set_step_status(
            db,
            step,
            "done",
            result={"stage": "出图完成", "image_result": image_result},
            progress_current=3,
            progress_total=3,
        )
    except asyncio.CancelledError:
        raise
    except Exception as exc:
        error = f"A+生成失败: {type(exc).__name__}: {exc}"
        logger.exception("[OfflineTask] A+ generate step failed: task=%s step=%s product=%s", step.task_id, step.id, product_id)
        await _set_aplus_status(product_id, "failed", error=error)
        await _set_step_status(db, step, "failed", error=error)


async def _run_product_bulk_advance_step(db: AsyncSession, step: OfflineTaskStep) -> None:
    payload = _json_loads(step.payload_json, {})
    product_id = int(payload.get("product_id") or 0) if isinstance(payload, dict) else 0
    start_step = int(payload.get("start_step") or 5) if isinstance(payload, dict) else 5
    item_code = payload.get("item_code") if isinstance(payload, dict) else None
    progress_total = 2 if start_step <= 5 else 1
    if product_id <= 0:
        await _set_step_status(db, step, "failed", error="商品推进步骤缺少 product_id")
        return
    if is_running(product_id):
        await _set_step_status(db, step, "failed", error="商品流程正在运行中")
        return

    try:
        if start_step <= 5:
            await _set_step_status(
                db,
                step,
                "running",
                result={
                    "product_id": product_id,
                    "item_code": item_code,
                    "status": "running",
                    "reason": "正在提交新任务中心图片分析",
                },
                progress_current=0,
                progress_total=1,
            )
            async with async_session() as session:
                result = await session.execute(
                    select(Product)
                    .where(Product.id == product_id)
                    .options(selectinload(Product.catalog_item))
                )
                product = result.scalar_one_or_none()
                if not product:
                    await _set_step_status(db, step, "failed", error=f"商品 {product_id} 不存在")
                    return
                product.status = STEP6_CURATING
                product.current_step = 5
                product.error_message = "图片分析已加入新任务中心队列，请到新任务中心查看进度"
                product.updated_at = datetime.now()
                if product.catalog_item:
                    product.catalog_item.status = product.status
                    product.catalog_item.updated_at = product.updated_at
                runs = await create_product_image_analysis_runs(
                    session,
                    [product_id],
                    created_by="product_bulk_advance",
                )
            await _ensure_task_not_paused(db, step.task_id)
            await _set_step_status(
                db,
                step,
                "done",
                result={
                    "product_id": product_id,
                    "item_code": item_code,
                    "status": "queued",
                    "reason": "已提交新任务中心图片分析；完成后从 Step 6 继续 Listing",
                    "latest_status": STEP6_CURATING,
                    "latest_step": 5,
                    "latest_result": "image_analysis_queued",
                    "latest_reason": "图片分析已进入新任务中心",
                    "task_run_ids": [run.id for run in runs],
                },
                progress_current=1,
                progress_total=1,
            )
            return

        if start_step == 6:
            await _set_step_status(
                db,
                step,
                "running",
                result={
                    "product_id": product_id,
                    "item_code": item_code,
                    "status": "running",
                    "reason": "正在提交新任务中心 Listing 生成",
                },
                progress_current=0,
                progress_total=1,
            )
            async with async_session() as session:
                result = await session.execute(
                    select(Product)
                    .where(Product.id == product_id)
                    .options(selectinload(Product.catalog_item))
                )
                product = result.scalar_one_or_none()
                if not product:
                    await _set_step_status(db, step, "failed", error=f"商品 {product_id} 不存在")
                    return
                product.status = get_step_status(6)
                product.current_step = 6
                product.error_message = "Listing 生成已加入新任务中心队列，请到新任务中心查看进度"
                product.updated_at = datetime.now()
                if product.catalog_item:
                    product.catalog_item.status = product.status
                    product.catalog_item.confirmed_at = None
                    product.catalog_item.updated_at = product.updated_at
                runs = await create_product_listing_runs(
                    session,
                    [product_id],
                    created_by="product_bulk_advance",
                )
            await _ensure_task_not_paused(db, step.task_id)
            await _set_step_status(
                db,
                step,
                "done",
                result={
                    "product_id": product_id,
                    "item_code": item_code,
                    "status": "queued",
                    "reason": "已提交新任务中心 Listing 生成",
                    "latest_status": get_step_status(6),
                    "latest_step": 6,
                    "latest_result": "listing_queued",
                    "latest_reason": "Listing 已进入新任务中心",
                    "task_run_ids": [run.id for run in runs],
                },
                progress_current=1,
                progress_total=1,
            )
            return

        await _set_step_status(
            db,
            step,
            "running",
            result={
                "product_id": product_id,
                "item_code": item_code,
                "status": "running",
                "reason": f"正在从 Step {start_step} 推进",
            },
            progress_current=0,
            progress_total=progress_total,
        )
        async with async_session() as session:
            result = await session.execute(
                select(Product)
                .where(Product.id == product_id)
                .options(selectinload(Product.catalog_item))
            )
            product = result.scalar_one_or_none()
            if not product:
                await _set_step_status(db, step, "failed", error=f"商品 {product_id} 不存在")
                return
            product.status = get_step_status(start_step)
            product.current_step = start_step
            product.error_message = None
            product.updated_at = datetime.now()
            if product.catalog_item:
                product.catalog_item.status = product.status
                product.catalog_item.updated_at = product.updated_at
            await session.commit()

        await run_pipeline_tracked(product_id, start_step=start_step)
        await _ensure_task_not_paused(db, step.task_id)

        async with async_session() as session:
            result = await session.execute(
                select(Product)
                .where(Product.id == product_id)
                .options(selectinload(Product.data))
            )
            product = result.scalar_one_or_none()
            if not product:
                await _set_step_status(db, step, "failed", error=f"商品 {product_id} 不存在")
                return
            latest_status = product.status
            latest_step = product.current_step
            if latest_status == COMPLETED and (latest_step or 0) >= 6:
                await _set_step_status(
                    db,
                    step,
                    "done",
                    result={
                        "product_id": product_id,
                        "item_code": item_code or (product.data.item_code if product.data else None),
                        "status": "done",
                        "reason": "已推进到待导出",
                        "latest_status": latest_status,
                        "latest_step": latest_step,
                        "latest_result": "export_ready",
                        "latest_reason": "已到达待导出",
                    },
                    progress_current=progress_total,
                    progress_total=progress_total,
                )
                return
            if latest_status == FAILED:
                error = product.error_message or "商品推进失败"
                await _set_step_status(
                    db,
                    step,
                    "failed",
                    result={
                        "product_id": product_id,
                        "item_code": item_code or (product.data.item_code if product.data else None),
                        "status": "failed",
                        "reason": error,
                        "latest_status": latest_status,
                        "latest_step": latest_step,
                        "latest_result": "failed",
                        "latest_reason": error,
                    },
                    error=error,
                    progress_current=0,
                    progress_total=progress_total,
                )
                return
            if latest_status == PAUSED:
                await _set_step_status(
                    db,
                    step,
                    "interrupted",
                    result={
                        "product_id": product_id,
                        "item_code": item_code or (product.data.item_code if product.data else None),
                        "status": "interrupted",
                        "reason": "商品流程已挂起",
                        "latest_status": latest_status,
                        "latest_step": latest_step,
                        "latest_result": "paused",
                        "latest_reason": "商品流程已挂起",
                    },
                    error="商品流程已挂起",
                    progress_current=0,
                    progress_total=progress_total,
                )
                return
            reason = f"商品未到待导出，当前状态 {latest_status}"
            await _set_step_status(
                db,
                step,
                "failed",
                result={
                    "product_id": product_id,
                    "item_code": item_code or (product.data.item_code if product.data else None),
                    "status": "failed",
                    "reason": reason,
                    "latest_status": latest_status,
                    "latest_step": latest_step,
                    "latest_result": "blocked",
                    "latest_reason": reason,
                },
                error=reason,
                progress_current=0,
                progress_total=progress_total,
            )
    except asyncio.CancelledError:
        await _set_step_status(
            db,
            step,
            "interrupted",
            result={
                "product_id": product_id,
                "item_code": item_code,
                "status": "interrupted",
                "reason": "任务中心步骤被中断，可重跑",
            },
            error="任务中心步骤被中断，可重跑",
            progress_current=0,
            progress_total=progress_total,
        )
        raise
    except Exception as exc:
        error = f"商品推进失败: {type(exc).__name__}: {exc}"
        logger.exception("[OfflineTask] product bulk advance step failed: task=%s step=%s product=%s", step.task_id, step.id, product_id)
        await _set_step_status(
            db,
            step,
            "failed",
            result={
                "product_id": product_id,
                "item_code": item_code,
                "status": "failed",
                "reason": error,
            },
            error=error,
            progress_current=0,
            progress_total=progress_total,
        )


async def _run_catalog_export_step(db: AsyncSession, step: OfflineTaskStep) -> None:
    step_id = step.id
    task_id = step.task_id
    existing_result = _json_loads(step.result_json, {})
    if step.status == STEP_STATUS_SUCCESS and _catalog_export_result_ready(existing_result):
        task = await db.get(OfflineTask, step.task_id)
        if task and isinstance(existing_result, dict):
            if not task.result_json:
                task.result_json = _json_dumps(existing_result)
            if existing_result.get("status") == "partial_failed":
                task.status = "partial_failed"
            if existing_result.get("status") in {"done", "partial_failed"}:
                task.error_message = None
            task.updated_at = datetime.now()
            await db.commit()
        return
    if _catalog_export_result_ready(existing_result):
        progress_total = step.progress_total or len((existing_result if isinstance(existing_result, dict) else {}).get("catalog_product_ids") or [])
        await _set_step_status(
            db,
            step,
            "done",
            result=existing_result if isinstance(existing_result, dict) else None,
            progress_current=progress_total,
            progress_total=progress_total,
        )
        task = await db.get(OfflineTask, step.task_id)
        if task and not task.result_json:
            task.result_json = _json_dumps(existing_result)
            task.updated_at = datetime.now()
            await db.commit()
        return

    payload = _json_loads(step.payload_json, {})
    if not isinstance(payload, dict):
        await _set_step_status(db, step, "failed", error="导出步骤参数格式错误")
        return
    catalog_ids = list(dict.fromkeys(int(item_id) for item_id in payload.get("catalog_product_ids") or []))
    categories = [str(item) for item in (payload.get("categories") or []) if str(item).strip()]
    category = str(payload.get("category") or (categories[0] if categories else "未分类"))
    template_name = str(payload.get("template_name") or "").strip()
    if not catalog_ids:
        await _set_step_status(db, step, "failed", error="导出步骤缺少商品")
        return

    from app.api.products import CatalogExportBuildError, build_catalog_export_zip

    try:
        export_dir = settings.DATA_DIR / "exports" / f"task_{step.task_id}"
        export_dir.mkdir(parents=True, exist_ok=True)
        recovered_result = _recover_catalog_export_result_from_file(
            export_dir=export_dir,
            category=category,
            categories=categories,
            template_name=template_name or None,
            template_path=payload.get("template_path"),
            catalog_ids=catalog_ids,
        )
        if recovered_result:
            await _set_step_status(
                db,
                step,
                "done",
                result=recovered_result,
                progress_current=len(catalog_ids),
                progress_total=len(catalog_ids),
            )
            task = await db.get(OfflineTask, step.task_id)
            if task:
                task.result_json = _json_dumps(recovered_result)
                task.status = recovered_result["status"]
                task.error_message = None if recovered_result["status"] in {"done", "partial_failed"} else task.error_message
                task.updated_at = datetime.now()
                await db.commit()
            return

        await _set_step_status(db, step, "running", progress_current=0, progress_total=len(catalog_ids))
        result = await db.execute(select(CatalogProduct).where(CatalogProduct.id.in_(catalog_ids)))
        catalog_items = result.scalars().all()
        if not catalog_items:
            raise ValueError("导出商品不存在")

        # Reuse the exact same export builder as the direct download endpoint,
        # so task export and manual export cannot drift apart.
        zip_bytes, _filename, report_rows = await build_catalog_export_zip(catalog_items, db)
        target_path = export_dir / f"catalog_export_t{step.task_id}_s{step.id}.zip"
        target_path.write_bytes(zip_bytes)
        uploaded = upload_private_file(target_path, _catalog_export_object_key(step.task_id, target_path.name))
        exported_source_ids = {
            int(row.get("商品ID"))
            for row in report_rows
            if row.get("状态") == "已导出" and row.get("商品ID") is not None
        }
        exported_at = datetime.now()
        for item in catalog_items:
            if item.source_product_id in exported_source_ids:
                item.exported_at = exported_at
                item.export_task_id = step.task_id
                item.export_file_path = str(uploaded.get("url") or target_path)
                item.updated_at = exported_at
        result_payload = _catalog_export_result_payload(
            category=category,
            categories=categories,
            template_name=template_name or None,
            template_path=payload.get("template_path"),
            catalog_ids=catalog_ids,
            report_rows=report_rows,
            created_at=exported_at,
            filename=target_path.name,
            file_path=str(target_path),
            oss_object_key=uploaded.get("object_key"),
            oss_url=uploaded.get("url"),
            file_size=target_path.stat().st_size,
        )
        await _set_step_status(
            db,
            step,
            "done",
            result=result_payload,
            progress_current=len(catalog_ids),
            progress_total=len(catalog_ids),
        )
        task = await db.get(OfflineTask, step.task_id)
        if task:
            task.result_json = _json_dumps(result_payload)
            task.status = result_payload["status"]
            task.error_message = None if result_payload["status"] in {"done", "partial_failed"} else task.error_message
            task.updated_at = datetime.now()
            await db.commit()
    except CatalogExportBuildError as exc:
        await db.rollback()
        step = await db.get(OfflineTaskStep, step_id)
        if not step:
            return
        failed_at = datetime.now()
        result_payload = _catalog_export_result_payload(
            category=category,
            categories=categories,
            template_name=template_name or None,
            template_path=payload.get("template_path"),
            catalog_ids=catalog_ids,
            report_rows=exc.report_rows,
            created_at=failed_at,
        )
        await _set_step_status(
            db,
            step,
            "failed",
            result=result_payload,
            error=exc.message,
            progress_current=0,
            progress_total=len(catalog_ids),
        )
        task = await db.get(OfflineTask, task_id)
        if task:
            task.result_json = _json_dumps(result_payload)
            task.status = "failed"
            task.error_message = exc.message
            task.updated_at = datetime.now()
            await db.commit()
    except Exception as exc:
        await db.rollback()
        step = await db.get(OfflineTaskStep, step_id)
        if not step:
            return
        logger.exception("[OfflineTask] catalog export step failed: task=%s step=%s", task_id, step_id)
        await _set_step_status(db, step, "failed", error=f"{type(exc).__name__}: {exc}")


async def _execute_offline_task(task_id: int, step_ids: list[int] | None = None) -> None:
    try:
        async with async_session() as db:
            query = select(OfflineTaskStep).where(OfflineTaskStep.task_id == task_id)
            if step_ids:
                query = query.where(OfflineTaskStep.id.in_(step_ids))
            else:
                query = query.where(OfflineTaskStep.status.in_(tuple(STEP_STATUS_CLAIMABLE)))
            result = await db.execute(query.order_by(OfflineTaskStep.id.asc()))
            steps = result.scalars().all()
            task = await db.get(OfflineTask, task_id)
            if task and task.status == "paused":
                return
            if task:
                task.status = "running"
                task.started_at = task.started_at or datetime.now()
                task.updated_at = datetime.now()
                await db.commit()
            for step in steps:
                claimed_step = await _claim_offline_step(db, step.id)
                if not claimed_step:
                    continue
                step = claimed_step
                task = await db.get(OfflineTask, task_id)
                if task and task.status == "paused":
                    break
                if step.step_type == "giga_sync":
                    await _run_sync_step(db, step)
                elif step.step_type == "giga_image_download":
                    await _run_image_step(db, step)
                elif step.step_type == "giga_inventory_sync":
                    await _run_inventory_sync_step(db, step)
                elif step.step_type == "giga_price_sync":
                    await _run_price_sync_step(db, step)
                elif step.step_type == "aplus_generate_product":
                    await _run_aplus_generate_step(db, step)
                elif step.step_type == "catalog_export_template":
                    await _run_catalog_export_step(db, step)
                elif step.step_type == "product_bulk_advance_product":
                    await _run_product_bulk_advance_step(db, step)
            await _refresh_task_stats(db, task_id)
            await db.commit()
    finally:
        _active_offline_tasks.pop(task_id, None)


def _schedule_offline_task(task_id: int, step_ids: list[int] | None = None) -> None:
    existing = _active_offline_tasks.get(task_id)
    if existing and not existing.done():
        return
    _active_offline_tasks[task_id] = asyncio.create_task(_execute_offline_task(task_id, step_ids))


def schedule_offline_task(task_id: int, step_ids: list[int] | None = None) -> None:
    _schedule_offline_task(task_id, step_ids)


async def _cancel_active_offline_task(task_id: int) -> None:
    active = _active_offline_tasks.get(task_id)
    if not active or active.done():
        _active_offline_tasks.pop(task_id, None)
        return
    active.cancel()
    try:
        await asyncio.wait_for(active, timeout=5)
    except asyncio.CancelledError:
        pass
    except TimeoutError:
        logger.warning("[OfflineTask] task=%s did not stop within pause timeout", task_id)
    except Exception:
        logger.exception("[OfflineTask] task=%s stopped with error while pausing", task_id)


async def _load_task_detail(db: AsyncSession, task_id: int) -> OfflineTask:
    task = await db.get(OfflineTask, task_id)
    if not task:
        raise ValueError("任务不存在")
    if task.task_type not in SUPPORTED_TASK_TYPES:
        raise ValueError("当前任务类型暂不支持控制")
    return task


async def pause_offline_task(db: AsyncSession, task_id: int) -> OfflineTask:
    task = await _load_task_detail(db, task_id)
    if task.status in {"done", "failed", "partial_failed"}:
        raise ValueError(f"任务已结束，不能挂起，当前状态: {task.status}")
    if task.status == "paused":
        return task

    await _cancel_active_offline_task(task_id)
    result = await db.execute(
        select(OfflineTaskStep).where(
            OfflineTaskStep.task_id == task_id,
            OfflineTaskStep.status.in_(("pending", "running")),
        )
    )
    steps = result.scalars().all()
    now = datetime.now()
    for step in steps:
        step.status = "paused"
        step.error_message = "任务已挂起，恢复后继续执行。"
        step.finished_at = None
        step.updated_at = now
    task.status = "paused"
    task.error_message = "任务已挂起，恢复后会继续未完成步骤。"
    task.running_steps = 0
    task.finished_at = None
    task.updated_at = now
    await db.commit()
    await db.refresh(task)
    return task


async def resume_offline_task(db: AsyncSession, task_id: int) -> OfflineTask:
    task = await _load_task_detail(db, task_id)
    if task.status != "paused":
        raise ValueError(f"只能恢复已挂起任务，当前状态: {task.status}")

    result = await db.execute(
        select(OfflineTaskStep).where(
            OfflineTaskStep.task_id == task_id,
            OfflineTaskStep.status == "paused",
        )
    )
    steps = result.scalars().all()
    if not steps:
        raise ValueError("没有可恢复的挂起步骤")

    now = datetime.now()
    for step in steps:
        step.status = "pending"
        step.error_message = None
        step.finished_at = None
        step.updated_at = now
    task.status = "pending"
    task.error_message = None
    task.finished_at = None
    task.updated_at = now
    await _refresh_task_stats(db, task_id)
    await db.commit()
    _schedule_offline_task(task_id, [step.id for step in steps])
    await db.refresh(task)
    return task


async def _load_enabled_sources(db: AsyncSession, data_source_ids: list[int]) -> tuple[list[int], list[ProductDataSource]]:
    deduped_ids = list(dict.fromkeys(data_source_ids))
    result = await db.execute(
        select(ProductDataSource)
        .where(ProductDataSource.id.in_(deduped_ids), ProductDataSource.enabled == 1)
        .order_by(ProductDataSource.id.asc())
    )
    sources = result.scalars().all()
    found_ids = {source.id for source in sources}
    missing_ids = [source_id for source_id in deduped_ids if source_id not in found_ids]
    if missing_ids:
        raise ValueError(f"店铺不存在或未启用: {', '.join(map(str, missing_ids))}")
    return deduped_ids, sources


async def create_giga_pull_task(
    db: AsyncSession,
    body: OfflineTaskGigaPullRequest,
    *,
    auto_start: bool = True,
) -> OfflineTask:
    data_source_ids, sources = await _load_enabled_sources(db, body.data_source_ids)

    task = OfflineTask(
        task_type="giga_pull",
        title=f"同步店铺商品（{len(sources)} 个店铺）",
        status="pending",
        total_steps=len(sources),
        success_steps=0,
        failed_steps=0,
        running_steps=0,
        payload_json=_json_dumps({
            "data_source_ids": data_source_ids,
            "current_category": body.current_category,
            "page_size": body.page_size,
            "max_pages": body.max_pages,
        }),
        created_at=datetime.now(),
        updated_at=datetime.now(),
    )
    db.add(task)
    await db.flush()
    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    for source in sources:
        context = await resolve_giga_data_source_context(db, source.id, source.site)
        batch_id = f"workbench-giga-t{task.id}-ds{source.id}-{timestamp}"
        db.add(
            OfflineTaskStep(
                task_id=task.id,
                step_type="giga_sync",
                title=f"同步商品：{context.name}",
                status="pending",
                data_source_id=context.id,
                data_source_name=context.name,
                site=context.site,
                batch_id=batch_id,
                progress_current=0,
                progress_total=0,
                payload_json=_json_dumps({
                    "batch_id": batch_id,
                    "site": context.site,
                    "data_source_id": context.id,
                    "current_category": body.current_category,
                    "page_size": body.page_size,
                    "max_pages": body.max_pages,
                    "skip_existing": True,
                }),
                updated_at=datetime.now(),
            )
        )
    await db.commit()
    await db.refresh(task)
    if auto_start:
        _schedule_offline_task(task.id)
    return task


async def create_giga_dynamic_sync_task(
    db: AsyncSession,
    body: OfflineTaskGigaDynamicSyncRequest,
    *,
    kind: str,
    auto_start: bool = True,
) -> OfflineTask:
    if kind not in {"inventory", "price"}:
        raise ValueError(f"不支持的同步类型: {kind}")
    data_source_ids, sources = await _load_enabled_sources(db, body.data_source_ids)
    sku_codes = list(dict.fromkeys(str(sku).strip() for sku in (body.sku_codes or []) if str(sku or "").strip()))
    is_inventory = kind == "inventory"
    task_type = "giga_inventory_sync" if is_inventory else "giga_price_sync"
    step_type = task_type
    title_prefix = "同步大健库存" if is_inventory else "同步大健价格"
    step_prefix = "库存同步" if is_inventory else "价格同步"

    task = OfflineTask(
        task_type=task_type,
        title=f"{title_prefix}（{len(sources)} 个店铺）",
        status="pending",
        total_steps=len(sources),
        success_steps=0,
        failed_steps=0,
        running_steps=0,
        payload_json=_json_dumps({
            "data_source_ids": data_source_ids,
            "sku_codes": sku_codes,
        }),
        created_at=datetime.now(),
        updated_at=datetime.now(),
    )
    db.add(task)
    await db.flush()

    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    for source in sources:
        context = await resolve_giga_data_source_context(db, source.id, source.site)
        batch_kind = "inventory" if is_inventory else "price"
        batch_id = f"giga-{batch_kind}-t{task.id}-ds{context.id}-{timestamp}"
        db.add(
            OfflineTaskStep(
                task_id=task.id,
                step_type=step_type,
                title=f"{step_prefix}：{context.name}",
                status="pending",
                data_source_id=context.id,
                data_source_name=context.name,
                site=context.site,
                batch_id=batch_id,
                progress_current=0,
                progress_total=0,
                payload_json=_json_dumps({
                    "batch_id": batch_id,
                    "site": context.site,
                    "data_source_id": context.id,
                    "sku_codes": sku_codes,
                }),
                updated_at=datetime.now(),
            )
        )
    await db.commit()
    await db.refresh(task)
    if auto_start:
        _schedule_offline_task(task.id)
    return task


async def _active_export_catalog_ids(db: AsyncSession) -> set[int]:
    result = await db.execute(
        select(OfflineTaskStep).where(
            OfflineTaskStep.step_type == "catalog_export_template",
            OfflineTaskStep.status.in_(("pending", "running", "paused")),
        )
    )
    active_ids: set[int] = set()
    for step in result.scalars().all():
        payload = _json_loads(step.payload_json, {})
        if not isinstance(payload, dict):
            continue
        for catalog_id in payload.get("catalog_product_ids") or []:
            try:
                parsed = int(catalog_id)
            except (TypeError, ValueError):
                continue
            if parsed > 0:
                active_ids.add(parsed)
    return active_ids


async def create_catalog_export_tasks(
    db: AsyncSession,
    body: OfflineTaskCatalogExportRequest,
    *,
    auto_start: bool = True,
) -> tuple[list[OfflineTask], list[str]]:
    requested_ids = list(dict.fromkeys(int(item_id) for item_id in body.catalog_product_ids))
    if not requested_ids:
        raise ValueError("请选择要导出的商品")
    if len(requested_ids) > 1000:
        raise ValueError("单次最多导出 1000 个商品")

    result = await db.execute(
        select(CatalogProduct)
        .where(CatalogProduct.id.in_(requested_ids))
        .options(selectinload(CatalogProduct.source_product).selectinload(Product.data))
    )
    catalog_items = result.scalars().all()
    catalog_by_id = {item.id: item for item in catalog_items}
    active_catalog_ids = await _active_export_catalog_ids(db)

    from app.api.products import _catalog_category, _template_status_for_catalog

    errors: list[str] = []
    grouped: dict[str, dict[str, object]] = {}
    for catalog_id in requested_ids:
        catalog = catalog_by_id.get(catalog_id)
        if not catalog:
            errors.append(f"商品资料 {catalog_id} 不存在")
            continue
        product = catalog.source_product
        item_code = (product.data.item_code if product and product.data else None) or catalog.item_code or str(catalog_id)
        if catalog.id in active_catalog_ids:
            errors.append(f"{item_code}: 已在导出任务中")
            continue
        available, template_name, template_path, template_error = _template_status_for_catalog(product, catalog)
        category = _catalog_category(product, catalog)
        if available and template_path:
            template_key = str(Path(str(template_path)).expanduser().resolve())
        else:
            template_key = f"__report_only__:{category}"
        group = grouped.setdefault(
            template_key,
            {
                "items": [],
                "template_name": template_name or ("待报告商品" if not available else None),
                "template_path": template_path,
                "template_error": template_error,
                "categories": [],
                "item_codes": [],
            },
        )
        (group["items"]).append(catalog)
        if category not in group["categories"]:
            (group["categories"]).append(category)
        (group["item_codes"]).append(item_code)

    if not grouped:
        await db.commit()
        return [], errors

    now = datetime.now()
    created_tasks: list[OfflineTask] = []
    requested_count = len(requested_ids)
    for _template_key, group in sorted(grouped.items(), key=lambda pair: str(pair[1].get("template_name") or pair[0])):
        items = group["items"]
        catalog_ids = [item.id for item in items]
        categories = sorted(str(category) for category in (group.get("categories") or []))
        template_name = str(group.get("template_name") or Path(str(group.get("template_path") or "amazon_template")).name)
        category_label = "、".join(categories[:3]) + (f" 等{len(categories)}类目" if len(categories) > 3 else "")
        task = OfflineTask(
            task_type="catalog_export",
            title=f"导出Amazon表格：{template_name}（{len(items)} 个商品）",
            status="pending",
            total_steps=1,
            success_steps=0,
            failed_steps=0,
            running_steps=0,
            payload_json=_json_dumps({
                "categories": categories,
                "catalog_product_ids": catalog_ids,
                "item_codes": group.get("item_codes") or [],
                "template_name": template_name,
                "template_path": group.get("template_path"),
                "requested_catalog_product_ids": requested_ids,
                "requested_count": requested_count,
                "errors": errors,
            }),
            created_at=now,
            updated_at=now,
        )
        db.add(task)
        await db.flush()
        db.add(
            OfflineTaskStep(
                task_id=task.id,
                step_type="catalog_export_template",
                title=f"按模板生成导出文件：{template_name}",
                status="pending",
                progress_current=0,
                progress_total=len(items),
                payload_json=_json_dumps({
                    "categories": categories,
                    "category_label": category_label,
                    "catalog_product_ids": catalog_ids,
                    "item_codes": group.get("item_codes") or [],
                    "template_name": template_name,
                    "template_path": group.get("template_path"),
                }),
                updated_at=now,
            )
        )
        created_tasks.append(task)

    await db.commit()
    for task in created_tasks:
        await db.refresh(task)
        if auto_start:
            _schedule_offline_task(task.id)
    return created_tasks, errors


async def _active_aplus_product_ids(db: AsyncSession) -> set[int]:
    result = await db.execute(
        select(OfflineTaskStep).where(
            OfflineTaskStep.step_type == "aplus_generate_product",
            OfflineTaskStep.status.in_(("pending", "running", "paused")),
        )
    )
    active_ids: set[int] = set()
    for step in result.scalars().all():
        payload = _json_loads(step.payload_json, {})
        if not isinstance(payload, dict):
            continue
        try:
            product_id = int(payload.get("product_id") or 0)
        except (TypeError, ValueError):
            product_id = 0
        if product_id > 0:
            active_ids.add(product_id)
    return active_ids


def _aplus_ready_error(product: Product | None, catalog: CatalogProduct | None) -> str | None:
    if not product:
        return "缺少关联商品，不能生成 A+"
    if not catalog or catalog.confirmed_at is None:
        return "未加入待导出，不能生成 A+"
    if is_running(product.id):
        return "商品主流程仍在运行，不能生成 A+"
    if not product.data or not product.data.listing_title or not product.data.listing_bullets:
        return "Listing 文案未完成，不能生成 A+"
    if not product.images or not product.images.image_analysis:
        return "图片分析未完成，不能生成 A+"
    return None


async def create_aplus_generate_task(
    db: AsyncSession,
    catalog_product_ids: list[int],
    *,
    force: bool = False,
    auto_start: bool = True,
) -> tuple[OfflineTask | None, list[str], list[int]]:
    requested_ids = list(dict.fromkeys(int(item_id) for item_id in catalog_product_ids))
    if not requested_ids:
        raise ValueError("请选择要生成 A+ 的商品")

    result = await db.execute(
        select(CatalogProduct)
        .where(CatalogProduct.id.in_(requested_ids))
        .options(
            selectinload(CatalogProduct.source_product).selectinload(Product.data),
            selectinload(CatalogProduct.source_product).selectinload(Product.images),
            selectinload(CatalogProduct.source_product).selectinload(Product.aplus),
        )
    )
    catalog_items = result.scalars().all()
    catalog_by_id = {item.id: item for item in catalog_items}
    active_product_ids = await _active_aplus_product_ids(db)
    errors: list[str] = []
    eligible: list[tuple[CatalogProduct, Product]] = []
    now = datetime.now()

    for catalog_id in requested_ids:
        catalog = catalog_by_id.get(catalog_id)
        if not catalog:
            errors.append(f"商品资料 {catalog_id} 不存在")
            continue
        product = catalog.source_product
        ready_error = _aplus_ready_error(product, catalog)
        if ready_error:
            errors.append(f"商品资料 {catalog_id}: {ready_error}")
            continue
        if product.id in active_product_ids:
            errors.append(f"商品 {product.id} A+ 已在任务中心生成中")
            continue
        if not product.aplus:
            product.aplus = ProductAplus(product_id=product.id)
            db.add(product.aplus)
            await db.flush()
        if product.aplus.aplus_status in APLUS_GENERATE_ACTIVE_STATUSES:
            errors.append(f"商品 {product.id} A+ 已在生成中")
            continue
        if product.aplus.aplus_status == "done" and not force:
            errors.append(f"商品 {product.id} A+ 已生成，如需重跑请使用强制重跑")
            continue
        product.aplus.aplus_status = "queued"
        if product.error_message and product.error_message.startswith("A+生成"):
            product.error_message = None
        product.updated_at = now
        catalog.updated_at = now
        eligible.append((catalog, product))

    if not eligible:
        await db.commit()
        return None, errors, []

    task = OfflineTask(
        task_type="aplus_generate",
        title=f"A+生成（{len(eligible)} 个商品）",
        status="pending",
        total_steps=len(eligible),
        success_steps=0,
        failed_steps=0,
        running_steps=0,
        payload_json=_json_dumps({
            "catalog_product_ids": [catalog.id for catalog, _ in eligible],
            "requested_catalog_product_ids": requested_ids,
            "force": force,
            "errors": errors,
        }),
        created_at=now,
        updated_at=now,
    )
    db.add(task)
    await db.flush()

    for catalog, product in eligible:
        item_code = product.data.item_code if product.data else None
        title = item_code or f"Product {product.id}"
        db.add(
            OfflineTaskStep(
                task_id=task.id,
                step_type="aplus_generate_product",
                title=f"A+生成：{title}",
                status="pending",
                progress_current=0,
                progress_total=3,
                payload_json=_json_dumps({
                    "catalog_product_id": catalog.id,
                    "product_id": product.id,
                    "item_code": item_code,
                    "force": force,
                }),
                updated_at=now,
            )
        )

    await db.commit()
    await db.refresh(task)
    if auto_start:
        _schedule_offline_task(task.id)
    return task, errors, [product.id for _, product in eligible]


async def rerun_offline_task(db: AsyncSession, task_id: int) -> OfflineTask:
    task = await db.get(OfflineTask, task_id)
    if not task:
        raise ValueError("任务不存在")
    if task.task_type not in SUPPORTED_TASK_TYPES:
        raise ValueError("当前任务类型暂不支持重跑")
    result = await db.execute(
        select(OfflineTaskStep).where(
            OfflineTaskStep.task_id == task_id,
            OfflineTaskStep.status.in_(("failed", "interrupted")),
        )
    )
    steps = result.scalars().all()
    if not steps:
        raise ValueError("没有可重跑的失败步骤")
    for step in steps:
        step.status = "pending"
        step.error_message = None
        step.finished_at = None
        step.updated_at = datetime.now()
    task.status = "pending"
    task.finished_at = None
    task.error_message = None
    await _refresh_task_stats(db, task_id)
    await db.commit()
    _schedule_offline_task(task_id, [step.id for step in steps])
    await db.refresh(task)
    return task


async def recover_offline_tasks() -> None:
    affected_task_ids: set[int] = set()
    async with async_session() as db:
        result = await db.execute(select(OfflineTask).where(OfflineTask.status.in_(("running", "interrupted"))))
        tasks = result.scalars().all()
        for task in tasks:
            affected_task_ids.add(task.id)
            if task.status == "running":
                task.status = "interrupted"
                task.error_message = "服务重启导致任务中断，系统正在恢复执行。"
                task.finished_at = None
            task.updated_at = datetime.now()
        step_result = await db.execute(select(OfflineTaskStep).where(OfflineTaskStep.status.in_(("running", "interrupted"))))
        steps = step_result.scalars().all()
        now = datetime.now()
        for step in steps:
            affected_task_ids.add(step.task_id)
            if step.step_type == "giga_image_download":
                payload = _json_loads(step.result_json, {})
                total = int(payload.get("total") or step.progress_total or 0) if isinstance(payload, dict) else 0
                done = int(payload.get("done") or step.progress_current or 0) if isinstance(payload, dict) else 0
                failed = int(payload.get("failed") or 0) if isinstance(payload, dict) else 0
                pending = int(payload.get("pending") or 0) if isinstance(payload, dict) else 0
                if total > 0 and pending == 0 and done + failed >= total:
                    step.status = "done" if failed == 0 else "failed"
                    step.progress_current = done
                    step.progress_total = total
                    step.error_message = None if failed == 0 else f"{failed} 张图片下载失败"
                    step.finished_at = step.finished_at or now
                    step.updated_at = now
                    continue
            if step.status == "running":
                step.status = "interrupted"
                step.error_message = "服务重启导致步骤中断，系统正在恢复执行。"
                step.finished_at = None
            step.updated_at = now
        for task_id in affected_task_ids:
            await _refresh_task_stats(db, task_id)

        active_aplus_ids = await _active_aplus_product_ids(db)
        orphan_result = await db.execute(
            select(Product)
            .join(ProductAplus)
            .where(ProductAplus.aplus_status.in_(APLUS_GENERATE_ACTIVE_STATUSES))
            .options(selectinload(Product.aplus), selectinload(Product.catalog_item))
        )
        now = datetime.now()
        for product in orphan_result.scalars().all():
            if product.id in active_aplus_ids:
                continue
            if product.aplus:
                product.aplus.aplus_status = "failed"
            product.error_message = "A+生成失败: 旧后台任务已中断，请重新发起 A+生成。"
            product.updated_at = now
            if product.catalog_item:
                product.catalog_item.updated_at = now
        await db.commit()
    for task_id in affected_task_ids:
        _schedule_offline_task(task_id)


async def cancel_active_offline_tasks() -> None:
    tasks = [task for task in _active_offline_tasks.values() if not task.done()]
    for task in tasks:
        task.cancel()
    if tasks:
        await asyncio.gather(*tasks, return_exceptions=True)
    _active_offline_tasks.clear()
