import asyncio
import subprocess
import zipfile
import hashlib
import json
import re
from copy import copy
from io import BytesIO
from pathlib import Path
from types import SimpleNamespace
from typing import Any

from fastapi import APIRouter, BackgroundTasks, Body, Depends, File, Form, HTTPException, Query, UploadFile
from fastapi.responses import FileResponse, StreamingResponse
from openpyxl import Workbook, load_workbook
from sqlalchemy import case, delete, select, func
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import load_only, selectinload
from datetime import datetime

from app.database import get_db
from app.config import settings
from app.models import (
    AplusRegenerateTask,
    AplusUploadBatch,
    AplusUploadItem,
    AsinSyncBatch,
    AsinSyncItem,
    CatalogProduct,
    AmazonListingCapture,
    AmazonStyleSnapCandidate,
    InventorySyncBatch,
    InventorySyncItem,
    GigaItem,
    GigaInventory,
    GigaProductImage,
    GigaSku,
    GigaSyncBatch,
    OfflineTask,
    Product,
    ProductData,
    ProductDataSource,
    ProductImage,
    ProductAplus,
    ProductFile,
    TaskRun,
    TaskStep,
    UpcPoolItem,
)
from app.models.status import COMPLETED, DUPLICATE_SKIPPED, PENDING_REVIEW, SOURCE_UNAVAILABLE, STEP5_DONE, STEP5_LISTING, STEP6_CURATING, STEP6_DONE, STEP_LABELS, STEP_STATUS_MAP
from app.api.schemas import (
    ProductCreate, ProductUpdate, ProductListingImagesUpdate, ProductGigaRefreshRequest, ProductResponse, ProductDetail, ProductImageResponse,
    PaginatedResponse, ProductFileEntry, AplusRegenerateRequest,
    ProductCompetitorReviewDetailResponse, ProductCompetitorReviewQueueResponse,
    ProductImageReviewDetailResponse, ProductImageReviewQueueResponse,
    AsinSyncBatchDetail, AsinSyncBatchResponse, AsinSyncCreateRequest,
    AplusUploadBatchDetail, AplusUploadBatchResponse, AplusUploadCreateRequest, AplusGenerateRequest,
    BulkImportResponse, BulkStartRequest, BulkStartResponse, ProductBulkAdvanceFilterRequest, ProductBulkAdvanceRequest, OfflineTaskResponse, TaskRunResponse,
    PaginatedAsinSyncBatches, PaginatedCatalogProducts, CatalogProductResponse,
    CatalogExportFileResponse, PaginatedCatalogExportFiles,
    CatalogExportByCategoryRequest, CatalogExportCategoriesResponse, CatalogExportCategorySummary,
    CatalogTemplateFileSummary, CatalogTemplateUploadResponse,
    PaginatedAplusUploadBatches, WorkbenchOverview, CatalogAsinUpdateRequest,
    InventorySyncBatchDetail, InventorySyncBatchResponse, InventorySyncCreateRequest,
    PaginatedInventorySyncBatches, PaginatedUpcPoolItems, UpcPoolImportRequest,
    UpcPoolImportResponse, UpcPoolSummary,
)
from app.pipeline.engine import start_pipeline as enqueue_pipeline, cancel_pipeline, get_step_status, is_running
from app.pipeline.step2_pricing import run_pricing
from app.pipeline.step3_keywords import run_keywords
from app.pipeline.step4_category import run_category
from app.pipeline.step5_listing import run_listing
from app.pipeline.step6_image import run_image_analysis
from app.pipeline.ride_on_category import RIDE_ON_CATEGORY_OPTIONS
from app.pipeline.step10_amazon_template import (
    AMAZON_TEMPLATE_LOGIC_VERSION,
    DATA_ROW,
    MAPPING_DIR,
    SEMANTIC_DROPDOWN_FIELD_KEYS,
    _load_template_mapping,
    _offer_quantity,
    _representative_package,
    ensure_amazon_template_semantic_fields,
    run_amazon_template,
)
from app.services.material_assets import (
    IMAGE_EXTENSIONS,
    aplus_folder_summary,
    aplus_image_folder,
    video_folder_summary,
)
from app.services.oss_uploader import download_private_file, sign_private_url, upload_private_file
from app.services.asin_sync import build_sync_item, start_asin_sync_batch
from app.services.inventory_sync import (
    InventorySyncLoginRequired,
    assert_gigab2b_logged_in_for_inventory,
    build_inventory_sync_item,
    start_inventory_sync_batch,
)
from app.services.aplus_upload import build_upload_item, start_aplus_upload_batch
from app.services.aplus_regenerate import create_regenerate_task, retry_latest_regenerate_tasks
from app.services.product_duplicates import (
    extract_gigab2b_item_code_from_url,
    extract_gigab2b_product_id,
    find_duplicate_by_gigab2b_product_id,
    find_duplicate_by_item_code,
)
from app.services.upc_pool import (
    UpcPoolEmptyError,
    add_upcs_to_pool,
    available_upc_count,
    ensure_product_upc,
)
from app.services.giga_openapi import GigaOpenApiError, GigaSyncOptions, sync_giga_products
from app.services.stylesnap_product_tasks import upsert_product_drafts_from_giga_batch
from app.task_planners.aplus_generate import create_aplus_generate_runs
from app.task_planners.product_bulk_advance import create_product_bulk_advance_run
from app.task_planners.product_image_analysis import create_product_image_analysis_runs
from app.task_planners.product_listing import create_product_listing_runs

# Step runners indexed by step number
STEP_RUNNERS = {
    2: run_pricing,
    3: run_keywords,
    4: run_category,
    5: run_image_analysis,
    6: run_listing,
    10: run_amazon_template,
}

router = APIRouter(prefix="/api/products", tags=["products"])


ZIP_EXTENSIONS = {".zip"}
MATERIAL_IMAGE_EXTENSIONS = IMAGE_EXTENSIONS
RUNNING_STATUSES = {
    "step1_collecting",
    "step2_pricing",
    "step3_keywords",
    "step4_category",
    "step5_listing",
    "step6_curating",
}

APLUS_REGEN_ACTIVE_STATUSES = {"regen_queued", "regen_script_running", "regen_image_running"}
COMPETITOR_REVIEW_ERROR_KEYWORDS = ("同款搜索", "StyleSnap", "候选竞品", "参考竞品", "选择竞品")
WORKBENCH_STATUS_KEYS = (
    "select_images",
    "competitor_searching",
    "select_competitor",
    "capture_detail",
    "ready_to_generate",
    "running",
    "interrupted",
    "suspended",
    "manual_review",
    "export_ready",
    "failed",
)
PRODUCT_LIST_WORK_STATUS_KEYS = set(WORKBENCH_STATUS_KEYS) | {"exported"}
AUTO_START_READY_GENERATION_LIMIT = 100
IMAGE_REVIEW_SELECTED_IMAGE_LIMIT = 9
IMAGE_REVIEW_INITIAL_GALLERY_LIMIT = 36
IMAGE_REVIEW_MAX_GALLERY_LIMIT = 200


def _product_task_action_queued_stage(product: Product) -> str | None:
    detail = str(product.error_message or "")
    if "任务中心队列" not in detail:
        return None
    step = int(product.current_step or 0)
    if product.status == STEP6_CURATING and step == 5:
        return "image_analysis"
    if product.status == STEP5_LISTING and step == 6:
        return "listing_generation"
    return None


def _is_stale_running_product(product: Product) -> bool:
    if _product_task_action_queued_stage(product):
        return False
    return product.status in RUNNING_STATUSES and not is_running(product.id)


def _product_data_source_filter(data_source_id: int):
    return Product.source_data_source_id == data_source_id


def normalize_image_path(item: Any) -> str:
    if isinstance(item, str):
        return item.strip()
    if isinstance(item, dict):
        return str(item.get("path") or item.get("local_path") or item.get("image_url") or "").strip()
    return ""


class CatalogExportBuildError(Exception):
    """Raised when a catalog export cannot produce a workbook but has row-level evidence."""

    def __init__(self, message: str, report_rows: list[dict] | None = None):
        super().__init__(message)
        self.message = message
        self.report_rows = report_rows or []


def _catalog_export_exception_status(exc: Exception) -> str:
    message = str(exc)
    business_skip_markers = (
        "已有真实 ASIN",
        "最新 GIGA 库存快照未找到",
        "无可售库存",
        "商品资料不存在",
    )
    if isinstance(exc, ValueError) and any(marker in message for marker in business_skip_markers):
        return "跳过"
    return "失败"


def _aplus_regeneration_running(product: Product) -> bool:
    return bool(product.aplus and product.aplus.aplus_status in APLUS_REGEN_ACTIVE_STATUSES)


def _normalize_listing_image_paths(main_image_path: str | None, gallery_images: list[str] | None) -> tuple[str, list[str]]:
    main_path = str(main_image_path or "").strip()
    if not main_path:
        raise HTTPException(400, "主图不能为空")

    gallery_paths: list[str] = []
    seen_paths = {main_path}
    for item in gallery_images or []:
        path = str(item).strip()
        if path and path not in seen_paths:
            gallery_paths.append(path)
            seen_paths.add(path)
    return main_path, gallery_paths


def _json_loads(value, fallback):
    if not value:
        return fallback
    try:
        return json.loads(value)
    except Exception:
        return fallback


def _compact_error_message(message: str | None) -> str | None:
    if not message:
        return None
    return " ".join(message.strip().splitlines())[:120]


def _is_legacy_giga_browser_collect_error(product: Product) -> bool:
    message = product.error_message or ""
    return (
        (product.current_step or 0) <= 1
        and (
            "大健商品核心信息采集失败" in message
            or "index.php?route=product/product" in message
            or "Step1 浏览器采集已停用" in message
            or "旧浏览器采集已停用" in message
        )
    )


def _legacy_giga_browser_collect_message() -> str:
    return "旧浏览器采集已停用，请通过商品工作台的“同步店铺商品”使用 GIGA OpenAPI 刷新商品源数据。"


def _product_snapshot(product: Product) -> dict:
    if not product.data:
        return {}
    snapshot = _json_loads(product.data.gigab2b_raw_snapshot, {})
    return snapshot if isinstance(snapshot, dict) else {}


ACTIVE_LISTING_CAPTURE_STATUSES = {"queued", "running"}


def _is_competitor_listing_capture_state(product: Product) -> bool:
    return (
        product.status == STEP5_LISTING
        and bool(product.error_message)
        and "竞品" in (product.error_message or "")
        and "抓取中" in (product.error_message or "")
    )


def _is_competitor_search_failed(product: Product) -> bool:
    return bool(
        product.status == "failed"
        and re.search("|".join(re.escape(keyword) for keyword in COMPETITOR_REVIEW_ERROR_KEYWORDS), product.error_message or "", re.I)
    )


def _competitor_search_failed_sql_condition():
    condition = Product.status == "failed"
    keyword_condition = None
    for keyword in COMPETITOR_REVIEW_ERROR_KEYWORDS:
        next_condition = Product.error_message.ilike(f"%{keyword}%")
        keyword_condition = next_condition if keyword_condition is None else keyword_condition | next_condition
    return condition & keyword_condition


async def _product_has_captured_competitor(db: AsyncSession, product: Product) -> bool:
    snapshot = _product_snapshot(product)
    selected = snapshot.get("selected_stylesnap")
    capture = snapshot.get("amazon_listing_capture")
    if _is_competitor_listing_capture_state(product):
        raise HTTPException(400, "不能进入图片分析：竞品详情仍在抓取中，请等待完成")
    if isinstance(capture, dict) and capture.get("status") in ACTIVE_LISTING_CAPTURE_STATUSES:
        raise HTTPException(400, "不能进入图片分析：竞品详情仍在抓取中，请等待完成")
    candidate_id = selected.get("candidate_id") if isinstance(selected, dict) else None
    if candidate_id:
        result = await db.execute(
            select(AmazonListingCapture).where(AmazonListingCapture.selected_candidate_id == int(candidate_id))
        )
        current_capture = result.scalar_one_or_none()
        if current_capture and current_capture.capture_status in ACTIVE_LISTING_CAPTURE_STATUSES:
            raise HTTPException(400, "不能进入图片分析：竞品详情仍在抓取中，请等待完成")
        if (
            current_capture
            and current_capture.capture_status == "captured"
            and current_capture.asin == product.competitor_asin
        ):
            return True
    return bool(
        product.competitor_asin
        and isinstance(selected, dict)
        and isinstance(capture, dict)
        and capture.get("status") == "captured"
        and capture.get("asin") == product.competitor_asin
    )


async def _require_generation_prerequisites(db: AsyncSession, product: Product, start_step: int) -> None:
    """Block a node from starting unless all previous business nodes are complete."""
    if start_step >= 5:
        if not product.images or not product.images.main_image_path:
            raise HTTPException(400, "不能进入图片分析：请先在详情页确认商品主图和 Listing 图片")
        if not product.competitor_asin:
            raise HTTPException(400, "不能进入图片分析：请先从候选竞品中选择一个参考竞品")
        if not await _product_has_captured_competitor(db, product):
            raise HTTPException(400, "不能进入图片分析：请先完成选中竞品的 Amazon Listing 详情抓取")
    if start_step >= 6:
        if not product.images or not product.images.image_analysis:
            raise HTTPException(400, "不能进入 Listing 文案：图片分析节点未完成")
    if start_step >= 7:
        raise HTTPException(400, "A+ 已从商品主流程拆出，请在 A+管理中单独生成")


def _current_task_status(product: Product) -> str:
    step_label = STEP_LABELS.get(product.current_step, f"Step {product.current_step}")
    competitor_capture_message = _compact_error_message(product.error_message)
    if (
        product.status == STEP5_LISTING
        and competitor_capture_message
        and ("竞品详情抓取中" in competitor_capture_message or "竞品 Listing 抓取中" in competitor_capture_message)
    ):
        return competitor_capture_message
    if product.status == "competitor_searching":
        return competitor_capture_message or "竞品搜索中：正在用主图搜索 Amazon 同款"
    if _product_task_action_queued_stage(product):
        return _compact_error_message(product.error_message) or "已加入新任务中心队列"
    if product.status == STEP6_DONE and (product.current_step or 0) <= 5:
        return "图片分析已完成，等待生成 Listing"
    if product.status == STEP5_DONE and (product.current_step or 0) >= 6:
        return "Listing 已生成，等待进入待导出"
    if _is_stale_running_product(product):
        return f"运行状态已中断：{step_label} 未在当前服务中运行，可直接重试当前节点"
    if product.status in RUNNING_STATUSES:
        return f"运行中：{step_label}"
    if product.status == "created":
        if product.current_step <= 0:
            return "待确认商品图片"
        if not product.competitor_asin:
            return "待选择参考竞品"
        return "待启动"
    if product.status == "paused":
        return f"已挂起：{step_label}，不会继续执行后续自动流程"
    if product.status == "pending_review":
        if product.current_step >= 6:
            return "Listing已生成，点击继续后自动进入待导出"
        detail = _compact_error_message(product.error_message)
        return f"待人工处理：{detail or step_label}"
    if product.status == "failed":
        if _is_legacy_giga_browser_collect_error(product):
            return _legacy_giga_browser_collect_message()
        if _is_competitor_search_failed(product):
            detail = _compact_error_message(product.error_message)
            return f"候选竞品搜索失败：{detail or '请重新搜索候选竞品'}"
        detail = _compact_error_message(product.error_message)
        return f"失败：{detail or step_label}"
    if product.status == SOURCE_UNAVAILABLE:
        detail = _compact_error_message(product.error_message)
        return f"原商品下架停止采集：{detail}" if detail else "原商品下架停止采集"
    if product.status == DUPLICATE_SKIPPED:
        detail = _compact_error_message(product.error_message)
        return f"重复商品已跳过：{detail}" if detail else "重复商品已跳过"
    if product.status == "unavailable":
        detail = _compact_error_message(product.error_message)
        return f"商品已下架：{detail}" if detail else "商品已下架"
    if product.status == COMPLETED:
        return "待导出"
    return product.status or "-"


def _workflow_for_step(step: int | None) -> str:
    value = int(step or 0)
    if value <= 0:
        return "image_review"
    if value < 5:
        return "competitor_select"
    if value == 5:
        return "image_analysis"
    if value == 6:
        return "listing_generation"
    return "export"


def _workflow_state(
    product: Product,
    *,
    catalog_exported: bool | None = None,
) -> dict:
    """Single source of truth for the main product workflow shown in the workbench.

    A+ is intentionally excluded. It is a separate workflow and must not override
    the main product status or primary action.
    """
    step = int(product.current_step or 0)
    detail = _compact_error_message(product.error_message)
    exported = bool(catalog_exported)
    if catalog_exported is None:
        catalog_item = getattr(product, "catalog_item", None)
        exported = bool(catalog_item and (catalog_item.exported_at or catalog_item.export_task_id))

    def state(
        stage: str,
        stage_status: str,
        label: str,
        work_status: str,
        *,
        primary_action: str | None = None,
        primary_action_label: str | None = None,
        allowed_actions: list[str] | None = None,
        action_reason: str | None = None,
        color: str = "default",
    ) -> dict:
        actions = ["open_detail", *(allowed_actions or [])]
        if primary_action and primary_action not in actions:
            actions.append(primary_action)
        related_correlation_key = None
        if stage == "image_analysis":
            related_correlation_key = f"product:{product.id}:image_analysis"
        elif stage == "listing_generation":
            related_correlation_key = f"product:{product.id}:listing_generation"
        return {
            "stage": stage,
            "stage_status": stage_status,
            "label": label,
            "work_status": work_status,
            "primary_action": primary_action,
            "primary_action_label": primary_action_label,
            "allowed_actions": list(dict.fromkeys(actions)),
            "action_reason": action_reason or label,
            "color": color,
            "related_task_run_id": None,
            "related_correlation_key": related_correlation_key,
        }

    if product.status == "paused":
        return state(
            _workflow_for_step(step),
            "paused",
            "已挂起",
            "suspended",
            primary_action="resume",
            primary_action_label="继续",
            allowed_actions=["resume", "restart"],
            action_reason=f"已挂起：{STEP_LABELS.get(step, f'Step {step}')}，不会继续执行后续自动流程",
        )

    action_queue_stage = _product_task_action_queued_stage(product)
    if action_queue_stage:
        stage_label = {
            "image_analysis": "图片分析",
            "listing_generation": "Listing 生成",
        }.get(action_queue_stage, "商品任务")
        return state(
            action_queue_stage,
            "queued",
            f"{stage_label}排队中",
            "running",
            primary_action="open_task_center",
            primary_action_label="任务中心",
            allowed_actions=["open_task_center"],
            action_reason=detail or f"{stage_label}已加入任务中心队列",
            color="processing",
        )

    if _is_stale_running_product(product):
        return state(
            _workflow_for_step(step),
            "interrupted",
            "已中断",
            "interrupted",
            primary_action="retry",
            primary_action_label="重试",
            allowed_actions=["retry", "restart"],
            action_reason=f"运行状态已中断：{STEP_LABELS.get(step, f'Step {step}')} 未在当前服务中运行，可直接重试当前节点",
            color="warning",
        )

    if product.status == "competitor_searching":
        queued = "队列" in detail or "排队" in detail
        return state(
            "competitor_search",
            "queued" if queued else "running",
            "候选竞品搜索中",
            "competitor_searching",
            primary_action="open_task_center",
            primary_action_label="任务中心",
            allowed_actions=["open_task_center"],
            action_reason=detail or "正在用主图搜索 Amazon 同款",
            color="processing",
        )

    if product.status == STEP5_LISTING and re.search(r"竞品.*抓取中|Listing.*抓取中", product.error_message or "", re.I):
        queued = "队列" in detail or "排队" in detail
        return state(
            "competitor_capture",
            "queued" if queued else "running",
            "竞品详情抓取中",
            "capture_detail",
            primary_action="open_task_center",
            primary_action_label="任务中心",
            allowed_actions=["open_task_center"],
            action_reason=detail or "正在抓取已选竞品详情",
            color="processing",
        )

    if product.status in RUNNING_STATUSES:
        stage = _workflow_for_step(step)
        stage_label = {
            "image_analysis": "图片分析",
            "listing_generation": "Listing 生成",
            "competitor_select": "商品处理",
            "export": "导出准备",
        }.get(stage, "商品处理")
        queued = "队列" in detail or "排队" in detail or "新任务" in (product.error_message or "")
        return state(
            stage,
            "queued" if queued else "running",
            f"{stage_label}{'排队中' if queued else '中'}",
            "running",
            primary_action="open_task_center",
            primary_action_label="任务中心",
            allowed_actions=["open_task_center"],
            action_reason=detail or f"{stage_label}{'已进入任务中心队列' if queued else '正在执行'}",
            color="processing",
        )

    if product.status == STEP6_DONE and step <= 5:
        return state(
            "listing_generation",
            "queued",
            "Listing 排队中",
            "running",
            primary_action="open_task_center",
            primary_action_label="任务中心",
            allowed_actions=["open_task_center"],
            action_reason="图片分析已完成，等待 Listing 生成任务",
            color="processing",
        )

    if product.status == STEP5_DONE and step >= 6:
        return state(
            "export",
            "pending",
            "待导出",
            "export_ready",
            primary_action="open_export_center",
            primary_action_label="导出",
            allowed_actions=["open_export_center"],
            action_reason="Listing 已生成，等待导出",
            color="success",
        )

    if product.status == COMPLETED and step >= 6:
        if exported:
            return state(
                "export",
                "succeeded",
                "已导出，可重导",
                "exported",
                primary_action="open_export_center",
                primary_action_label="重导",
                allowed_actions=["open_export_center"],
                action_reason="已导出，可在导出中心再次导出",
                color="green",
            )
        return state(
            "export",
            "pending",
            "待导出",
            "export_ready",
            primary_action="open_export_center",
            primary_action_label="导出",
            allowed_actions=["open_export_center"],
            action_reason="Listing 已生成，等待导出",
            color="success",
        )

    if product.status == PENDING_REVIEW:
        return state(
            "manual_review",
            "pending",
            "待人工处理",
            "manual_review",
            primary_action="resume",
            primary_action_label="继续",
            allowed_actions=["resume", "restart"],
            action_reason=detail or "需要人工确认后继续",
            color="warning",
        )

    if product.status == "failed":
        if _is_competitor_search_failed(product):
            return state(
                "competitor_search",
                "failed",
                "候选搜索失败",
                "select_competitor",
                primary_action="open_competitor_review",
                primary_action_label="处理竞品",
                allowed_actions=["open_competitor_review", "restart"],
                action_reason=detail or "候选竞品搜索失败，请重新搜索候选",
                color="error",
            )
        if re.search(r"竞品详情|竞品 Listing|Amazon Listing|选中竞品", product.error_message or "", re.I):
            return state(
                "competitor_capture",
                "failed",
                "竞品详情抓取失败",
                "select_competitor",
                primary_action="open_competitor_review",
                primary_action_label="重新抓取",
                allowed_actions=["open_competitor_review", "restart"],
                action_reason=detail or "竞品详情抓取失败",
                color="error",
            )
        stage = _workflow_for_step(step)
        return state(
            stage,
            "failed",
            f"{STEP_LABELS.get(step, '当前步骤')}失败",
            "failed",
            primary_action="retry" if step > 1 else "open_detail",
            primary_action_label="重试" if step > 1 else "查看",
            allowed_actions=(["retry", "restart"] if step > 1 else ["restart"]),
            action_reason=detail or "当前步骤失败",
            color="error",
        )

    if product.status in {SOURCE_UNAVAILABLE, DUPLICATE_SKIPPED, "unavailable"}:
        label = "原商品不可用" if product.status == SOURCE_UNAVAILABLE else "商品不可用"
        return state(
            "stopped",
            "blocked",
            label,
            "failed",
            primary_action="open_detail",
            primary_action_label="查看",
            action_reason=detail or label,
            color="default",
        )

    if product.status == "created":
        if step <= 0:
            return state(
                "image_review",
                "pending",
                "待确认图片",
                "select_images",
                primary_action="open_image_review",
                primary_action_label="确认图片",
                allowed_actions=["open_image_review"],
                action_reason="需要先确认主图和 Listing 图片",
                color="cyan",
            )
        if not product.competitor_asin:
            return state(
                "competitor_select",
                "pending",
                "待选择竞品",
                "select_competitor",
                primary_action="open_competitor_review",
                primary_action_label="选竞品",
                allowed_actions=["open_competitor_review", "restart"],
                action_reason="需要搜索或选择参考竞品",
                color="purple",
            )
        return state(
            "image_analysis" if step <= 5 else "listing_generation",
            "queued",
            "等待自动入队",
            "ready_to_generate",
            primary_action="open_task_center",
            primary_action_label="任务中心",
            allowed_actions=["open_task_center", "restart"],
            action_reason="前置条件已满足，等待系统自动进入任务中心",
            color="warning",
        )

    return state(
        _workflow_for_step(step),
        "running",
        "处理中",
        "running",
        primary_action="open_task_center",
        primary_action_label="任务中心",
        allowed_actions=["open_task_center"],
        action_reason=product.status or "处理中",
        color="processing",
    )


def _split_category_path(value: str | None) -> list[str]:
    if not value:
        return []
    return [part.strip() for part in re.split(r"\s*>\s*|\s*›\s*", str(value)) if part.strip()]


def _category_option_key(categories: list[str], leaf_category: str | None) -> str:
    return " > ".join(categories) or (leaf_category or "")


def _static_category_options() -> list[dict]:
    options: dict[str, dict] = {}

    for option in RIDE_ON_CATEGORY_OPTIONS:
        categories = option.categories
        key = _category_option_key(categories, option.leaf_category)
        options[key] = {
            "key": key,
            "label": " > ".join(categories),
            "categories": categories,
            "leaf_category": option.leaf_category,
            "source": "ride_on_toy",
        }

    # 项目规则：同一类目 key 冲突时，后导入的映射覆盖前者；
    # 只替换冲突类目，其他类目继续保留。
    for mapping_path in sorted(MAPPING_DIR.glob("*.json")):
        try:
            mapping = json.loads(mapping_path.read_text(encoding="utf-8"))
        except Exception:
            continue
        for item in mapping.get("browse_category_options") or []:
            if not isinstance(item, dict):
                continue
            path = str(item.get("path") or "").strip()
            node = str(item.get("node") or "").strip()
            categories = _split_category_path(path)
            if not categories:
                continue
            leaf = f"{categories[-1]} ({node})" if node else categories[-1]
            option_categories = [*categories[:-1], leaf]
            key = _category_option_key(option_categories, leaf)
            options[key] = {
                "key": key,
                "label": " > ".join(option_categories),
                "categories": option_categories,
                "leaf_category": leaf,
                "source": mapping_path.stem,
            }

    return sorted(options.values(), key=lambda item: item["label"])


IMPORT_TEMPLATE_HEADERS = ["原始数据链接", "竞品ASIN"]
HEADER_ALIASES = {
    "原始数据链接": "gigab2b_url",
    "来源链接": "gigab2b_url",
    "源链接": "gigab2b_url",
    "source_url": "gigab2b_url",
    "大健云仓商品链接": "gigab2b_url",
    "大建云仓商品链接": "gigab2b_url",
    "商品链接": "gigab2b_url",
    "gigab2b_url": "gigab2b_url",
    "url": "gigab2b_url",
    "竞品ASIN": "competitor_asin",
    "竞品 ASIN": "competitor_asin",
    "competitor_asin": "competitor_asin",
    "asin": "competitor_asin",
    "UPC": "upc",
    "upc": "upc",
    "品牌": "brand",
    "brand": "brand",
}


def _is_in_dir(path: Path, root: Path) -> bool:
    try:
        path.resolve().relative_to(root.resolve())
        return True
    except ValueError:
        return False


def _safe_material_dir(product: Product) -> Path:
    material_dir = product.data.material_dir if product.data else None
    if not material_dir:
        raise HTTPException(400, "素材目录不存在")
    path = Path(material_dir).expanduser().resolve()
    if not path.is_dir():
        raise HTTPException(404, f"素材目录不存在: {material_dir}")
    return path


def _summarize_extracted_files(extracted_dir: Path, limit: int = 30) -> list[str]:
    if not extracted_dir.is_dir():
        return []

    priority_words = ("认证", "证书", "certificate", "certification", "cert", "cpsia", "astm", "prop", "test", "report")
    files: list[Path] = []
    for path in extracted_dir.rglob("*"):
        if path.is_file():
            files.append(path)

    files.sort(key=lambda p: (not any(word in p.name.lower() for word in priority_words), str(p).lower()))
    return [str(path.relative_to(extracted_dir)) for path in files[:limit]]


def _find_extracted_dir(material_dir: Path, zip_path: Path) -> Path:
    default_dir = zip_path.with_suffix("")
    if default_dir.is_dir():
        return default_dir

    for path in material_dir.rglob(zip_path.stem):
        if path.is_dir():
            return path

    return default_dir


def _scan_zip_files(material_dir: Path) -> list[ProductFileEntry]:
    entries: list[ProductFileEntry] = []
    for path in sorted(material_dir.rglob("*")):
        if not path.is_file() or path.suffix.lower() not in ZIP_EXTENSIONS:
            continue
        stat = path.stat()
        extracted_dir = _find_extracted_dir(material_dir, path)
        entries.append(ProductFileEntry(
            name=path.name,
            path=str(path),
            size=stat.st_size,
            modified_at=datetime.fromtimestamp(stat.st_mtime),
            extracted_dir=str(extracted_dir),
            extracted_exists=extracted_dir.is_dir(),
            extracted_files=_summarize_extracted_files(extracted_dir),
        ))
    return entries


def _count_material_images(material_dir: Path) -> int:
    count = 0
    for path in material_dir.rglob("*"):
        if not path.is_file() or path.suffix.lower() not in MATERIAL_IMAGE_EXTENSIONS:
            continue
        name = path.name.lower()
        path_text = str(path).lower()
        if "new " in path_text or "image analysis" in path_text or name == "contact_sheet.jpg":
            continue
        count += 1
    return count


def _build_list_item(product: Product) -> dict:
    catalog_exported = product.catalog_item and (product.catalog_item.exported_at or product.catalog_item.export_task_id)
    workflow = _workflow_state(product, catalog_exported=bool(catalog_exported))
    current_task_status = (
        f"已导出，可在导出中心再次导出（任务 #{product.catalog_item.export_task_id}）"
        if catalog_exported and product.catalog_item and product.catalog_item.export_task_id
        else "已导出，可在导出中心再次导出"
        if catalog_exported
        else workflow["action_reason"]
    )
    return {
        "id": product.id,
        "source_url": product.gigab2b_url,
        "source_item_id": product.gigab2b_product_id,
        "gigab2b_url": product.gigab2b_url,
        "gigab2b_product_id": product.gigab2b_product_id,
        "competitor_asin": product.competitor_asin,
        "amazon_asin": product.amazon_asin,
        "asin_sync_status": product.asin_sync_status,
        "asin_synced_at": product.asin_synced_at,
        "asin_sync_error": product.asin_sync_error,
        "amazon_product_status": product.amazon_product_status,
        "amazon_product_status_synced_at": product.amazon_product_status_synced_at,
        "amazon_product_status_error": product.amazon_product_status_error,
        "aplus_upload_status": product.aplus_upload_status,
        "aplus_uploaded_at": product.aplus_uploaded_at,
        "aplus_upload_error": product.aplus_upload_error,
        "aplus_status": product.aplus.aplus_status if product.aplus else None,
        "upc": product.upc,
        "brand": product.brand,
        "source_data_source_id": product.source_data_source_id,
        "source_site": product.source_site,
        "source_batch_id": product.source_batch_id,
        "catalog_exported_at": product.catalog_item.exported_at if product.catalog_item else None,
        "catalog_export_task_id": product.catalog_item.export_task_id if product.catalog_item else None,
        "status": product.status,
        "current_step": product.current_step,
        "current_task_status": current_task_status,
        "workflow": workflow,
        "error_message": product.error_message,
        "created_at": product.created_at,
        "updated_at": product.updated_at,
        "item_code": product.data.item_code if product.data else None,
        "title": product.data.title if product.data else None,
        "leaf_category": product.data.leaf_category if product.data else None,
    }


def _product_workbench_status(product: Product) -> str:
    return _workflow_state(product, catalog_exported=False)["work_status"]


def _product_list_work_status(product: Product) -> str:
    catalog_item = getattr(product, "catalog_item", None)
    catalog_exported = bool(catalog_item and (catalog_item.exported_at or catalog_item.export_task_id))
    return _workflow_state(product, catalog_exported=catalog_exported)["work_status"]


def _template_risk_from_data(pd: ProductData | None) -> tuple[str | None, int | None]:
    if not pd:
        return None, None
    summary = {}
    warnings = []
    try:
        summary = json.loads(pd.amazon_template_fill_summary or "{}")
    except Exception:
        summary = {}
    try:
        warnings = json.loads(pd.amazon_template_warnings or "[]")
    except Exception:
        warnings = []
    risk = summary.get("risk_level") if isinstance(summary, dict) else None
    warning_count = len(warnings) if isinstance(warnings, list) else None
    return risk, warning_count


def _amazon_template_cache_path(pd: ProductData | None) -> Path | None:
    if not pd or not pd.amazon_template_path:
        return None
    source_path = Path(pd.amazon_template_path).expanduser()
    if not source_path.is_file():
        return None
    try:
        summary = json.loads(pd.amazon_template_fill_summary or "{}")
    except Exception:
        summary = {}
    if not isinstance(summary, dict) or summary.get("logic_version") != AMAZON_TEMPLATE_LOGIC_VERSION:
        return None
    return source_path


def _safe_json(value: str | None, fallback):
    if not value:
        return fallback
    try:
        parsed = json.loads(value)
        return parsed if parsed is not None else fallback
    except Exception:
        return fallback


def _shipping_template_preview(product: Product, pd: ProductData, mapping: dict) -> str | None:
    preferences = mapping.get("shipping_template_by_brand") or {}
    for candidate in (product.brand, settings.DEFAULT_BRAND, "Andy店-US", "*"):
        preferred = preferences.get(str(candidate or ""))
        if preferred:
            return preferred
    return None


def _build_amazon_export_preview(product: Product) -> dict | None:
    pd = product.data
    if not pd:
        return None
    try:
        mapping = _load_template_mapping(product, pd)
    except Exception as exc:
        return {"available": False, "error": f"{type(exc).__name__}: {exc}"}

    package, package_warnings = _representative_package(pd)
    try:
        quantity = _offer_quantity(pd)
    except Exception as exc:
        quantity = None
        package_warnings.append(str(exc))

    fields = mapping.get("dynamic_fields") or {}
    return {
        "available": True,
        "logic_version": AMAZON_TEMPLATE_LOGIC_VERSION,
        "template_path": mapping.get("template_path"),
        "output_filename": str(mapping.get("output_filename") or "{item_code}_amazon_import.xlsm").format(item_code=pd.item_code),
        "category": pd.leaf_category or mapping.get("category") or mapping.get("category_type"),
        "field_attributes": {
            "shipping_template": fields.get("shipping_template"),
            "price": fields.get("price"),
            "quantity": fields.get("quantity"),
            "product_weight": fields.get("item_weight_value"),
        },
        "offer": {
            "sku": pd.item_code,
            "brand": product.brand,
            "fulfillment_channel": "Fulfillment by Merchant (Default)",
            "shipping_template": _shipping_template_preview(product, pd, mapping),
            "quantity": quantity,
            "price": pd.suggested_price,
            "list_price": pd.suggested_price,
            "country_of_origin": pd.origin or "China",
        },
        "product_dimensions": {
            "length": pd.dimension_length,
            "width": pd.dimension_width,
            "height": pd.dimension_height,
            "unit": "Inches",
            "size_text": (
                f'{pd.dimension_length:g}" x {pd.dimension_width:g}" x {pd.dimension_height:g}"'
                if pd.dimension_length and pd.dimension_width and pd.dimension_height
                else None
            ),
        },
        "product_weight": {
            "value": pd.weight,
            "unit": "Pounds" if pd.weight else None,
        },
        "package_aggregate": {
            **(package or {}),
            "length_unit": "Inches" if package else None,
            "weight_unit": "Pounds" if package else None,
            "warnings": package_warnings,
        },
        "package_items": _safe_json(pd.packages, []),
        "source_costs": {
            "value_total": pd.value_total,
            "estimated_total": pd.estimated_total,
            "shipping_cost": pd.shipping_cost,
            "shipping_cost_min": pd.shipping_cost_min,
            "shipping_cost_max": pd.shipping_cost_max,
            "suggested_price": pd.suggested_price,
            "cost_total": pd.cost_total,
            "profit": pd.profit,
            "profit_rate": pd.profit_rate,
        },
    }


def _sync_catalog_item(product: Product, db: AsyncSession, confirm: bool = False) -> CatalogProduct:
    pd = product.data
    item = product.catalog_item
    if not item:
        item = CatalogProduct(source_product_id=product.id, gigab2b_url=product.gigab2b_url)
        db.add(item)

    item.gigab2b_url = product.gigab2b_url
    item.gigab2b_product_id = product.gigab2b_product_id
    item.competitor_asin = product.competitor_asin
    item.amazon_asin = product.amazon_asin
    item.asin_sync_status = product.asin_sync_status
    item.asin_synced_at = product.asin_synced_at
    item.asin_sync_error = product.asin_sync_error
    item.amazon_product_status = product.amazon_product_status
    item.amazon_product_status_synced_at = product.amazon_product_status_synced_at
    item.amazon_product_status_error = product.amazon_product_status_error
    item.aplus_upload_status = product.aplus_upload_status
    item.aplus_uploaded_at = product.aplus_uploaded_at
    item.aplus_upload_error = product.aplus_upload_error
    item.upc = product.upc
    item.brand = product.brand
    item.item_code = pd.item_code if pd else None
    item.title = pd.title if pd else None
    item.leaf_category = pd.leaf_category if pd else None
    item.status = product.status
    if confirm and item.confirmed_at is None:
        item.confirmed_at = datetime.now()
    item.updated_at = datetime.now()
    return item


def _safe_oss_key_part(value: str | None, fallback: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "-", value or "").strip("-._")
    return cleaned or fallback


async def _ensure_contact_sheet_oss_urls(product: Product, db: AsyncSession) -> None:
    images = product.images
    if not images or not images.image_analysis:
        return
    try:
        payload = json.loads(images.image_analysis)
    except json.JSONDecodeError:
        return
    if not isinstance(payload, dict):
        return

    sheets = payload.get("contact_sheets")
    if not isinstance(sheets, list) or not sheets:
        return

    changed = False
    product_key = _safe_oss_key_part(product.gigab2b_product_id or product.source_item_id, f"product-{product.id}")
    for index, sheet in enumerate(sheets):
        if not isinstance(sheet, dict):
            continue
        sheet_path = str(sheet.get("sheet_path") or "").strip()
        if not sheet_path or sheet_path.startswith("url_batch:"):
            continue
        if sheet_path.startswith(("http://", "https://")):
            if not sheet.get("display_url"):
                sheet["display_url"] = sheet_path
                changed = True
            continue

        object_key = str(sheet.get("oss_object_key") or "").strip()
        if object_key:
            try:
                signed_url = sign_private_url(object_key)
            except Exception:
                signed_url = str(sheet.get("oss_url") or sheet.get("display_url") or "")
            if signed_url:
                if sheet.get("oss_url") != signed_url or sheet.get("display_url") != signed_url:
                    sheet["oss_url"] = signed_url
                    sheet["display_url"] = signed_url
                    changed = True
            continue

        path = Path(sheet_path).expanduser()
        if not path.is_file():
            continue
        filename = _safe_oss_key_part(path.name, f"contact_sheet_{index + 1:02d}{path.suffix or '.jpg'}")
        object_key = f"image_analysis/{product_key}/contact_sheets/{filename}"
        try:
            uploaded = upload_private_file(path, object_key)
        except Exception:
            continue
        sheet["oss_object_key"] = uploaded.get("object_key")
        sheet["oss_url"] = uploaded.get("url")
        sheet["display_url"] = uploaded.get("url")
        sheet["oss_uploaded_at"] = datetime.now().isoformat()
        changed = True

    if changed:
        images.image_analysis = json.dumps(payload, ensure_ascii=False)
        first_sheet = next((item for item in sheets if isinstance(item, dict)), None)
        if first_sheet:
            images.contact_sheet_path = first_sheet.get("display_url") or first_sheet.get("oss_url") or first_sheet.get("sheet_path")
        images.analyzed_at = images.analyzed_at or datetime.now()
        product.updated_at = datetime.now()
        await db.commit()
        await db.refresh(product)


def _generation_start_step(product: Product) -> int:
    if (product.current_step or 0) >= 6:
        return 6
    if product.images and product.images.image_analysis:
        return 6
    return max(product.current_step or 5, 5)


async def _queue_product_image_analysis(
    db: AsyncSession,
    product: Product,
    *,
    created_by: str,
) -> list[int]:
    await _require_generation_prerequisites(db, product, 5)
    runs = await create_product_image_analysis_runs(db, [product.id], created_by=created_by)
    return [run.id for run in runs]


async def _queue_product_listing_generation(
    db: AsyncSession,
    product: Product,
    *,
    created_by: str,
) -> list[int]:
    await _require_generation_prerequisites(db, product, 6)
    runs = await create_product_listing_runs(db, [product.id], created_by=created_by)
    return [run.id for run in runs]


def _raise_step1_browser_collect_removed() -> None:
    raise HTTPException(
        400,
        "商品中心不再通过浏览器访问大健页面采集商品数据。请先通过店铺/OpenAPI 同步商品，生成商品草稿后再在详情页确认图片和竞品。",
    )


def _workbook_response(wb: Workbook, filename: str) -> StreamingResponse:
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


async def _upc_pool_summary(db: AsyncSession) -> UpcPoolSummary:
    total_result = await db.execute(select(func.count(UpcPoolItem.id)))
    available_result = await db.execute(
        select(func.count(UpcPoolItem.id)).where(UpcPoolItem.status == "available")
    )
    bound_result = await db.execute(
        select(func.count(UpcPoolItem.id)).where(UpcPoolItem.status == "bound")
    )
    return UpcPoolSummary(
        total=total_result.scalar() or 0,
        available=available_result.scalar() or 0,
        bound=bound_result.scalar() or 0,
    )


def _catalog_export_row(product: Product) -> dict:
    pd = product.data
    return {
        "任务ID": product.id,
        "原始数据链接": product.gigab2b_url,
        "来源商品ID": product.gigab2b_product_id,
        "商品Code": pd.item_code if pd else None,
        "竞品ASIN": product.competitor_asin,
        "真实ASIN": product.amazon_asin,
        "ASIN同步状态": product.asin_sync_status,
        "ASIN同步信息": product.asin_sync_error,
        "亚马逊商品状态": product.amazon_product_status,
        "亚马逊商品状态同步时间": product.amazon_product_status_synced_at,
        "亚马逊商品状态同步信息": product.amazon_product_status_error,
        "UPC": product.upc,
        "A+上传状态": product.aplus_upload_status,
        "A+上传信息": product.aplus_upload_error,
        "品牌": product.brand,
        "类目": pd.leaf_category if pd else None,
        "标题": pd.title if pd else None,
        "颜色": pd.color if pd else None,
        "材质": pd.material if pd else None,
        "产品类型": pd.product_type if pd else None,
        "建议售价": pd.suggested_price if pd else None,
        "总成本": pd.cost_total if pd else None,
        "利润": pd.profit if pd else None,
        "利润率": pd.profit_rate if pd else None,
        "Listing标题": pd.listing_title if pd else None,
        "Search Terms": pd.listing_search_terms if pd else None,
    }


def _safe_export_name(value: str | None, fallback: str = "未分类") -> str:
    raw = (value or fallback).strip() or fallback
    return "".join(ch if ch.isalnum() or ch in ("-", "_") else "_" for ch in raw)[:90] or fallback


def _category_template_cache_dir() -> Path:
    path = settings.DATA_DIR / "category_templates"
    path.mkdir(parents=True, exist_ok=True)
    return path


def _category_template_manifest_path() -> Path:
    return _category_template_cache_dir() / "manifest.json"


def _template_file_state_path() -> Path:
    return _category_template_cache_dir() / "template_files.json"


def _load_category_template_manifest() -> dict[str, dict]:
    path = _category_template_manifest_path()
    if not path.is_file():
        return {}
    try:
        loaded = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {}
    return loaded if isinstance(loaded, dict) else {}


def _save_category_template_manifest(manifest: dict[str, dict]) -> None:
    path = _category_template_manifest_path()
    path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2, sort_keys=True), encoding="utf-8")


def _load_template_file_state() -> dict[str, dict]:
    path = _template_file_state_path()
    if not path.is_file():
        return {}
    try:
        loaded = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {}
    return loaded if isinstance(loaded, dict) else {}


def _save_template_file_state(state: dict[str, dict]) -> None:
    path = _template_file_state_path()
    path.write_text(json.dumps(state, ensure_ascii=False, indent=2, sort_keys=True), encoding="utf-8")


def _template_file_id(path: Path | str) -> str:
    raw = str(Path(path).expanduser().resolve())
    return hashlib.sha1(raw.encode("utf-8")).hexdigest()[:16]


def _template_file_no(file_id: str) -> str:
    return f"TPL-{file_id[:8].upper()}"


def _template_file_enabled(path: Path | str) -> bool:
    file_id = _template_file_id(path)
    state = _load_template_file_state().get(file_id)
    if not isinstance(state, dict):
        return True
    return bool(state.get("enabled", True))


def _template_file_object_key(path: Path | str, file_id: str | None = None) -> str:
    resolved = Path(path).expanduser().resolve()
    template_id = file_id or _template_file_id(resolved)
    prefix = settings.OSS_TEMPLATE_UPLOAD_PREFIX.strip().strip("/")
    safe_name = _safe_export_name(resolved.stem, "amazon_template") + (resolved.suffix.lower() or ".xlsm")
    key = f"template_files/{template_id}/{safe_name}"
    return f"{prefix}/{key}" if prefix else key


def _ensure_template_file_oss_metadata(path: Path | str, *, preferred_object_key: str | None = None) -> dict:
    resolved = Path(path).expanduser().resolve()
    file_id = _template_file_id(resolved)
    state = _load_template_file_state()
    current = state.get(file_id) if isinstance(state.get(file_id), dict) else {}
    object_key = str(preferred_object_key or current.get("object_key") or "").strip().lstrip("/")
    if object_key:
        try:
            oss_url = sign_private_url(object_key)
        except Exception:
            oss_url = str(current.get("oss_url") or "")
    else:
        object_key = _template_file_object_key(resolved, file_id)
        uploaded = upload_private_file(resolved, object_key)
        object_key = str(uploaded.get("object_key") or object_key)
        oss_url = str(uploaded.get("url") or "")

    current.update({
        "file_name": resolved.name,
        "template_path": str(resolved),
        "object_key": object_key,
        "oss_url": oss_url,
        "oss_synced_at": datetime.now().isoformat(),
    })
    state[file_id] = current
    _save_template_file_state(state)
    return current


def _category_template_upload_for(category: str) -> dict | None:
    upload = _load_category_template_manifest().get(category)
    return upload if isinstance(upload, dict) else None


def _uploaded_category_template_path(category: str) -> Path | None:
    upload = _category_template_upload_for(category)
    if not upload:
        return None
    cache_path = str(upload.get("cache_path") or "").strip()
    if not cache_path:
        return None
    path = Path(cache_path).expanduser()
    if not path.is_file():
        object_key = str(upload.get("object_key") or "").strip()
        if object_key:
            try:
                download_private_file(object_key, path)
            except Exception:
                return None
    return path if path.is_file() and _template_file_enabled(path) else None


def _catalog_category(product: Product | None, item: CatalogProduct | None) -> str:
    pd = product.data if product else None
    return (pd.leaf_category if pd else None) or (item.leaf_category if item else None) or "未分类"


def _catalog_existing_asin(product: Product | None, item: CatalogProduct | None) -> str:
    return ((product.amazon_asin if product else None) or (item.amazon_asin if item else None) or "").strip()


def _catalog_has_export_file(product: Product | None, item: CatalogProduct | None) -> bool:
    return bool((item.exported_at if item else None) or _catalog_existing_asin(product, item))


def _catalog_export_task_payload(task: OfflineTask) -> dict:
    payload = _json_loads(task.result_json, {})
    if isinstance(payload, dict) and (
        payload.get("filename") or payload.get("file_path") or payload.get("oss_object_key") or payload.get("rows")
    ):
        return payload
    for step in sorted(task.steps, key=lambda item: item.id, reverse=True):
        if step.step_type != "catalog_export_template":
            continue
        step_payload = _json_loads(step.result_json, {})
        if isinstance(step_payload, dict) and (
            step_payload.get("filename")
            or step_payload.get("file_path")
            or step_payload.get("oss_object_key")
            or step_payload.get("rows")
        ):
            return step_payload
    payload = _json_loads(task.payload_json, {})
    return payload if isinstance(payload, dict) else {}


def _catalog_export_datetime(value: object) -> datetime | None:
    if isinstance(value, datetime):
        return value
    if not value:
        return None
    try:
        return datetime.fromisoformat(str(value))
    except ValueError:
        return None


def _catalog_export_file_row(task: OfflineTask) -> CatalogExportFileResponse:
    payload = _catalog_export_task_payload(task)
    rows = payload.get("rows") if isinstance(payload.get("rows"), list) else []
    catalog_ids = [
        int(item_id)
        for item_id in payload.get("catalog_product_ids") or []
        if str(item_id).isdigit()
    ]
    categories = [
        str(category).strip()
        for category in payload.get("categories") or []
        if str(category).strip()
    ]
    if not categories:
        categories = sorted({
            str(row.get("category")).strip()
            for row in rows
            if isinstance(row, dict) and str(row.get("category") or "").strip()
        })
    success_count = int(payload.get("success_count") or payload.get("exported_count") or 0)
    skipped_count = int(payload.get("skipped_count") or 0)
    failed_count = int(payload.get("failed_count") or 0)
    if rows and not (success_count or skipped_count or failed_count):
        success_count = sum(1 for row in rows if isinstance(row, dict) and row.get("status") == "exported")
        skipped_count = sum(1 for row in rows if isinstance(row, dict) and row.get("status") == "skipped")
        failed_count = sum(1 for row in rows if isinstance(row, dict) and row.get("status") == "failed")
    filename = str(payload.get("filename") or "").strip() or None
    file_path = str(payload.get("file_path") or "").strip() or None
    oss_url = str(payload.get("oss_url") or "").strip() or None
    return CatalogExportFileResponse(
        task_id=task.id,
        task_source="offline_task",
        task_status=str(payload.get("status") or task.status),
        title=task.title,
        filename=filename,
        file_path=file_path,
        oss_url=oss_url,
        file_size=payload.get("file_size"),
        exported_at=_catalog_export_datetime(payload.get("created_at")) or task.finished_at or task.updated_at,
        category=str(payload.get("category") or (categories[0] if categories else "")).strip() or None,
        categories=categories,
        category_count=len(categories),
        template_name=str(payload.get("template_name") or "").strip() or None,
        catalog_product_ids=catalog_ids,
        task_product_count=int(payload.get("requested_count") or len(catalog_ids)),
        file_product_count=success_count,
        success_count=success_count,
        exported_count=success_count,
        skipped_count=skipped_count,
        failed_count=failed_count,
        report_count=int(payload.get("report_count") or len(rows)),
        can_download=bool(filename or file_path or payload.get("oss_object_key")),
        created_at=task.created_at,
        finished_at=task.finished_at,
        updated_at=task.updated_at,
    )


def _catalog_export_run_payload(run: TaskRun) -> dict:
    payload = _json_loads(run.summary_json, {})
    if isinstance(payload, dict) and (
        payload.get("filename") or payload.get("file_path") or payload.get("oss_object_key") or payload.get("rows")
    ):
        return payload
    for step in sorted(run.steps, key=lambda item: item.id, reverse=True):
        if step.step_type != "catalog_export_template":
            continue
        step_payload = _json_loads(step.result_json, {})
        if isinstance(step_payload, dict) and (
            step_payload.get("filename")
            or step_payload.get("file_path")
            or step_payload.get("oss_object_key")
            or step_payload.get("rows")
        ):
            return step_payload
    payload = _json_loads(run.payload_json, {})
    return payload if isinstance(payload, dict) else {}


def _catalog_export_file_row_from_run(run: TaskRun) -> CatalogExportFileResponse:
    payload = _catalog_export_run_payload(run)
    rows = payload.get("rows") if isinstance(payload.get("rows"), list) else []
    catalog_ids = [
        int(item_id)
        for item_id in payload.get("catalog_product_ids") or []
        if str(item_id).isdigit()
    ]
    categories = [
        str(category).strip()
        for category in payload.get("categories") or []
        if str(category).strip()
    ]
    if not categories:
        categories = sorted({
            str(row.get("category")).strip()
            for row in rows
            if isinstance(row, dict) and str(row.get("category") or "").strip()
        })
    success_count = int(payload.get("success_count") or payload.get("exported_count") or 0)
    skipped_count = int(payload.get("skipped_count") or 0)
    failed_count = int(payload.get("failed_count") or 0)
    if rows and not (success_count or skipped_count or failed_count):
        success_count = sum(1 for row in rows if isinstance(row, dict) and row.get("status") == "exported")
        skipped_count = sum(1 for row in rows if isinstance(row, dict) and row.get("status") == "skipped")
        failed_count = sum(1 for row in rows if isinstance(row, dict) and row.get("status") == "failed")
    filename = str(payload.get("filename") or "").strip() or None
    file_path = str(payload.get("file_path") or "").strip() or None
    oss_url = str(payload.get("oss_url") or "").strip() or None
    payload_status = str(payload.get("status") or "").strip()
    return CatalogExportFileResponse(
        task_id=run.id,
        task_source="task_run",
        task_status=payload_status or run.status,
        title=run.title,
        filename=filename,
        file_path=file_path,
        oss_url=oss_url,
        file_size=payload.get("file_size"),
        exported_at=_catalog_export_datetime(payload.get("created_at")) or run.finished_at or run.updated_at,
        category=str(payload.get("category") or (categories[0] if categories else "")).strip() or None,
        categories=categories,
        category_count=len(categories),
        template_name=str(payload.get("template_name") or "").strip() or None,
        catalog_product_ids=catalog_ids,
        task_product_count=int(payload.get("requested_count") or len(catalog_ids)),
        file_product_count=success_count,
        success_count=success_count,
        exported_count=success_count,
        skipped_count=skipped_count,
        failed_count=failed_count,
        report_count=int(payload.get("report_count") or len(rows)),
        can_download=bool(filename or file_path or payload.get("oss_object_key")),
        created_at=run.created_at,
        finished_at=run.finished_at,
        updated_at=run.updated_at,
    )


def _template_status_for_catalog(product: Product | None, item: CatalogProduct | None) -> tuple[bool, str | None, str | None, str | None]:
    pd = product.data if product else None
    category = _catalog_category(product, item)
    if not product or not pd:
        return False, None, None, "商品资料不存在"
    try:
        mapping = _load_template_mapping(product, pd)
        uploaded_path = _uploaded_category_template_path(category)
        if uploaded_path:
            return True, uploaded_path.name, str(uploaded_path), None
        template_path = Path(mapping["template_path"]).expanduser()
        if not template_path.is_file():
            return False, template_path.name, str(template_path), f"模板文件不存在: {template_path}"
        if not _template_file_enabled(template_path):
            return False, template_path.name, str(template_path), "模板文件已停用"
        return True, template_path.name, str(template_path), None
    except Exception as exc:
        upload = _category_template_upload_for(category)
        if upload:
            return False, upload.get("filename"), upload.get("cache_path"), f"模板已上传，但类目映射未接入导出: {type(exc).__name__}: {exc}"
        return False, None, None, f"{type(exc).__name__}: {exc}"


def _collect_export_category(
    groups: dict[str, dict],
    product: Product | None,
    item: CatalogProduct,
    *,
    exported: bool,
) -> None:
    category = _catalog_category(product, item)
    group = groups.setdefault(
        category,
        {
            "category": category,
            "count": 0,
            "exportable_count": 0,
            "blocked_count": 0,
            "template_available": False,
            "template_name": None,
            "template_path": None,
            "template_error": None,
            "uploaded_template_name": None,
            "uploaded_template_cache_path": None,
            "uploaded_template_oss_url": None,
            "uploaded_template_object_key": None,
            "uploaded_template_uploaded_at": None,
            "sample_item_codes": [],
            "_template_errors": [],
        },
    )
    upload = _category_template_upload_for(category)
    if upload:
        group["uploaded_template_name"] = upload.get("filename")
        group["uploaded_template_cache_path"] = upload.get("cache_path")
        group["uploaded_template_oss_url"] = upload.get("oss_url")
        group["uploaded_template_object_key"] = upload.get("object_key")
        group["uploaded_template_uploaded_at"] = upload.get("uploaded_at")
    group["count"] += 1
    code = (product.data.item_code if product and product.data else None) or item.item_code
    if code and len(group["sample_item_codes"]) < 5 and code not in group["sample_item_codes"]:
        group["sample_item_codes"].append(code)

    available, template_name, template_path, template_error = _template_status_for_catalog(product, item)
    if available:
        group["template_available"] = True
        group["template_name"] = group["template_name"] or template_name
        group["template_path"] = group["template_path"] or template_path
        if not exported:
            group["exportable_count"] += 1
    else:
        if not exported:
            group["blocked_count"] += 1
        if template_error and template_error not in group["_template_errors"]:
            group["_template_errors"].append(template_error)
            if not group["template_error"]:
                group["template_error"] = template_error


def _export_category_summaries(groups: dict[str, dict]) -> list[CatalogExportCategorySummary]:
    summaries: list[CatalogExportCategorySummary] = []
    for group in groups.values():
        if group.get("_template_errors") and len(group["_template_errors"]) > 1:
            group["template_error"] = "；".join(group["_template_errors"][:3])
        group.pop("_template_errors", None)
        summaries.append(CatalogExportCategorySummary(**group))
    return sorted(
        summaries,
        key=lambda item: (
            not item.template_available,
            -item.exportable_count,
            -item.count,
            item.category,
        ),
    )


def _collect_export_file_category(groups: dict[str, dict], row: CatalogExportFileResponse) -> None:
    categories = row.categories or ([row.category] if row.category else ["未分类"])
    for category in categories:
        group = groups.setdefault(
            category,
            {
                "category": category,
                "count": 0,
                "exportable_count": 0,
                "blocked_count": 0,
                "template_available": True,
                "template_name": row.template_name,
                "template_path": None,
                "template_error": None,
                "uploaded_template_name": None,
                "uploaded_template_cache_path": None,
                "uploaded_template_oss_url": None,
                "uploaded_template_object_key": None,
                "uploaded_template_uploaded_at": None,
                "sample_item_codes": [],
                "_template_errors": [],
            },
        )
        group["count"] += 1
        group["exportable_count"] += 1
        if not group.get("template_name"):
            group["template_name"] = row.template_name


async def _catalog_template_category_samples(db: AsyncSession) -> list[dict[str, Any]]:
    category_expr = func.coalesce(func.nullif(CatalogProduct.leaf_category, ""), "未分类")
    brand_expr = func.coalesce(func.nullif(CatalogProduct.brand, ""), "Vindhvisk")
    category_result = await db.execute(
        select(
            category_expr.label("category"),
            brand_expr.label("brand"),
            func.count(CatalogProduct.id).label("count"),
            func.min(CatalogProduct.item_code).label("sample_item_code"),
        )
        .select_from(CatalogProduct)
        .where(category_expr != "未分类")
        .group_by(category_expr, brand_expr)
        .order_by(category_expr)
    )
    groups: dict[str, dict[str, Any]] = {}
    for category, brand, count, sample_item_code in category_result.all():
        normalized_category = str(category or "未分类").strip() or "未分类"
        group = groups.setdefault(
            normalized_category,
            {
                "category": normalized_category,
                "count": 0,
                "brands": [],
                "sample_item_codes": [],
            },
        )
        group["count"] += int(count or 0)
        normalized_brand = str(brand or "Vindhvisk").strip() or "Vindhvisk"
        if normalized_brand not in group["brands"]:
            group["brands"].append(normalized_brand)
        sample = str(sample_item_code or "").strip()
        if sample and sample not in group["sample_item_codes"] and len(group["sample_item_codes"]) < 5:
            group["sample_item_codes"].append(sample)
    return sorted(groups.values(), key=lambda item: item["category"])


def _category_upload_summary(category: str) -> dict[str, Any]:
    upload = _category_template_upload_for(category)
    if not upload:
        return {}
    return {
        "uploaded_template_name": upload.get("filename"),
        "uploaded_template_cache_path": upload.get("cache_path"),
        "uploaded_template_oss_url": upload.get("oss_url"),
        "uploaded_template_object_key": upload.get("object_key"),
        "uploaded_template_uploaded_at": upload.get("uploaded_at"),
    }


def _catalog_template_status_for_category(
    category: str,
    brands: list[str] | None = None,
) -> tuple[bool, str | None, str | None, str | None]:
    upload = _category_template_upload_for(category)
    if upload:
        cache_path = str(upload.get("cache_path") or "").strip()
        object_key = str(upload.get("object_key") or "").strip()
        if cache_path:
            path = Path(cache_path).expanduser()
            enabled = _template_file_enabled(path)
            if enabled and (path.is_file() or object_key):
                return True, str(upload.get("filename") or path.name), str(path), None
            if not enabled:
                return False, str(upload.get("filename") or path.name), str(path), "模板文件已停用"
            return False, str(upload.get("filename") or path.name), str(path), "模板本地缓存不存在"

    errors: list[str] = []
    for brand in brands or ["Vindhvisk"]:
        product = SimpleNamespace(brand=brand or "Vindhvisk")
        pd = SimpleNamespace(
            leaf_category=category,
            categories=category,
            product_type="",
            title="",
            amazon_stylesnap_selected_asin=None,
            amazon_stylesnap_selected_category_rank=None,
        )
        try:
            mapping = _load_template_mapping(product, pd)
            template_path = Path(mapping["template_path"]).expanduser()
            if not template_path.is_file():
                return False, template_path.name, str(template_path), f"模板文件不存在: {template_path}"
            if not _template_file_enabled(template_path):
                return False, template_path.name, str(template_path), "模板文件已停用"
            return True, template_path.name, str(template_path), None
        except Exception as exc:
            errors.append(f"{type(exc).__name__}: {exc}")
    return False, None, None, errors[0] if errors else "未配置类目导入模板映射"


def _template_file_source(path: Path) -> str:
    cache_root = _category_template_cache_dir().resolve()
    template_root = (Path(__file__).resolve().parents[1] / "pipeline" / "templates").resolve()
    resolved = path.expanduser().resolve()
    if resolved == cache_root or cache_root in resolved.parents:
        return "uploaded"
    if resolved == template_root or template_root in resolved.parents:
        return "builtin"
    return "external"


def _template_file_status(enabled: bool, errors: list[str]) -> str:
    if not enabled:
        return "disabled"
    return "enabled" if not errors else "unmapped"


async def _catalog_template_file_summaries(db: AsyncSession) -> list[CatalogTemplateFileSummary]:
    files: dict[str, dict] = {}
    template_file_state = _load_template_file_state()

    def add_file(
        path: Path,
        category: str,
        *,
        error: str | None = None,
        preferred_object_key: str | None = None,
        preferred_oss_url: str | None = None,
    ) -> None:
        if not path.is_file() and not preferred_object_key:
            return
        resolved = path.expanduser().resolve()
        file_id = _template_file_id(resolved)
        state = template_file_state.get(file_id)
        enabled = bool(state.get("enabled", True)) if isinstance(state, dict) else True
        source = _template_file_source(resolved)
        object_key = str(preferred_object_key or (state or {}).get("object_key") or "").strip() or None
        oss_url = str(preferred_oss_url or (state or {}).get("oss_url") or "").strip() or None
        row = files.setdefault(
            file_id,
            {
                "file_id": file_id,
                "file_no": _template_file_no(file_id),
                "file_name": resolved.name,
                "enabled": enabled,
                "source": source,
                "template_path": str(resolved),
                "oss_object_key": object_key,
                "oss_url": oss_url,
                "support_categories": [],
                "template_errors": [],
                "can_download": bool(path.is_file() or object_key),
                "can_delete": True,
            },
        )
        if object_key:
            row["oss_object_key"] = object_key
        if oss_url:
            row["oss_url"] = oss_url
        if category not in row["support_categories"]:
            row["support_categories"].append(category)
        if error and error not in row["template_errors"]:
            row["template_errors"].append(error)

    for category_row in await _catalog_template_category_samples(db):
        category = category_row["category"]
        upload = _category_template_upload_for(category)
        if upload:
            cache_path = str(upload.get("cache_path") or "").strip()
            if cache_path:
                add_file(
                    Path(cache_path).expanduser(),
                    category,
                    preferred_object_key=upload.get("object_key"),
                    preferred_oss_url=upload.get("oss_url"),
                )
                continue

        available, _template_name, template_path, template_error = _catalog_template_status_for_category(
            category,
            category_row.get("brands") or [],
        )
        if template_path:
            add_file(
                Path(template_path).expanduser(),
                category,
                error=None if available else template_error,
            )

    summaries: list[CatalogTemplateFileSummary] = []
    for row in files.values():
        row["support_categories"] = sorted(row["support_categories"])
        row["file_status"] = _template_file_status(row["enabled"], row["template_errors"])
        summaries.append(CatalogTemplateFileSummary(**row))
    return sorted(
        summaries,
        key=lambda item: (
            item.file_status != "enabled",
            item.source != "uploaded",
            item.file_name,
            item.file_no,
        ),
    )


async def _find_catalog_template_file(db: AsyncSession, file_id: str) -> CatalogTemplateFileSummary:
    for item in await _catalog_template_file_summaries(db):
        if item.file_id == file_id:
            return item
    raise HTTPException(404, "模板文件不存在")


def _copy_row_format(ws, source_row: int, target_row: int) -> None:
    if target_row == source_row:
        return
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
    for col in range(1, ws.max_column + 1):
        source = ws.cell(source_row, col)
        target = ws.cell(target_row, col)
        if source.has_style:
            target._style = copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        if source.alignment:
            target.alignment = copy(source.alignment)
        if source.protection:
            target.protection = copy(source.protection)


def _clear_template_data_rows(ws, row_count: int) -> None:
    max_row = max(ws.max_row, DATA_ROW + row_count + 2)
    for row in range(DATA_ROW, max_row + 1):
        if row > DATA_ROW:
            _copy_row_format(ws, DATA_ROW, row)
        for col in range(1, ws.max_column + 1):
            ws.cell(row, col).value = None


def _template_attribute_columns(ws) -> dict[str, int]:
    return {
        str(cell.value): cell.column
        for cell in ws[5]
        if cell.value not in (None, "")
    }


def _copy_import_data_row(source_path: Path, target_ws, target_row: int) -> None:
    source_wb = load_workbook(source_path, keep_vba=True, data_only=False)
    if "Template" not in source_wb.sheetnames:
        raise ValueError(f"导入表格缺少 Template 工作表: {source_path}")
    source_ws = source_wb["Template"]
    _copy_row_format(target_ws, DATA_ROW, target_row)
    source_columns = _template_attribute_columns(source_ws)
    target_columns = _template_attribute_columns(target_ws)
    for attr, source_col in source_columns.items():
        target_col = target_columns.get(attr)
        if target_col:
            target_ws.cell(target_row, target_col).value = source_ws.cell(DATA_ROW, source_col).value


PRODUCT_ID_TYPE_ATTR = "amzn1.volt.ca.product_id_type"
PRODUCT_ID_VALUE_ATTR = "amzn1.volt.ca.product_id_value"


def _set_import_row_attr(ws, row_number: int, attr: str, value: Any) -> None:
    column = _template_attribute_columns(ws).get(attr)
    if column:
        ws.cell(row_number, column).value = value


def _flatten_template_field_values(value: Any) -> list[str]:
    if isinstance(value, str):
        return [value]
    if isinstance(value, list):
        result: list[str] = []
        for item in value:
            result.extend(_flatten_template_field_values(item))
        return result
    if isinstance(value, dict):
        result: list[str] = []
        for item in value.values():
            result.extend(_flatten_template_field_values(item))
        return result
    return []


def _template_semantic_field_key(key: str) -> str:
    return "target_audience_keyword" if key == "target_audience" else key


def _catalog_export_semantic_values(pd: ProductData, key: str) -> list[str]:
    listing_check = _json_loads(pd.listing_check, {})
    fields = listing_check.get("amazon_template_fields") if isinstance(listing_check, dict) else None
    payload = fields.get(_template_semantic_field_key(key)) if isinstance(fields, dict) else None
    values = payload.get("values") if isinstance(payload, dict) else None
    if not isinstance(values, list):
        return []
    return [str(value).strip() for value in values if str(value or "").strip()][:5]


def _catalog_export_fabric_type_values(pd: ProductData) -> list[str]:
    semantic_values = _catalog_export_semantic_values(pd, "fabric_type")
    if semantic_values:
        return semantic_values[:1]

    material = str(pd.material or "").strip()
    text = " ".join([
        material,
        str(pd.listing_title or ""),
        str(pd.listing_description or ""),
        str(pd.description or ""),
        str(pd.title or ""),
    ]).lower()
    normalized_material = material.replace("+", " and ").replace("/", " and ").strip()
    if "mdf" in text and "particle" in text:
        return ["100% MDF and Particle Board"]
    if "engineered wood" in text:
        return ["100% Engineered Wood"]
    if "wood" in text or "wooden" in text:
        return ["100% Wood"]
    if "metal" in text or "steel" in text or "iron" in text:
        return ["100% Metal"]
    if "polyester" in text or "fabric" in text or "flannelette" in text:
        return ["100% Polyester"]
    if "cotton" in text:
        return ["100% Cotton"]
    if "linen" in text:
        return ["100% Linen"]
    if normalized_material:
        return [f"100% {normalized_material}"[:200]]
    return []


def _apply_catalog_export_row_overrides(ws, row_number: int, product: Product, pd: ProductData, mapping: dict) -> None:
    columns = _template_attribute_columns(ws)
    if product.upc:
        _set_import_row_attr(ws, row_number, PRODUCT_ID_TYPE_ATTR, "UPC")
        _set_import_row_attr(ws, row_number, PRODUCT_ID_VALUE_ATTR, product.upc)

    dynamic_fields = mapping.get("dynamic_fields") if isinstance(mapping.get("dynamic_fields"), dict) else {}
    for key in SEMANTIC_DROPDOWN_FIELD_KEYS:
        attrs = _flatten_template_field_values(dynamic_fields.get(key))
        if key == "target_audience":
            for field_key, field_value in dynamic_fields.items():
                if str(field_key).startswith("target_audience_"):
                    attrs.extend(_flatten_template_field_values(field_value))
        if key == "theme":
            for field_key, field_value in dynamic_fields.items():
                if str(field_key).startswith("theme_"):
                    attrs.extend(_flatten_template_field_values(field_value))
        columns_for_key = [columns[attr] for attr in attrs if attr in columns]
        for column in columns_for_key:
            ws.cell(row_number, column).value = None
        for column, value in zip(columns_for_key, _catalog_export_semantic_values(pd, key)):
            ws.cell(row_number, column).value = value

    fabric_attrs = _flatten_template_field_values(dynamic_fields.get("fabric_type"))
    fabric_attrs.extend(attr for attr in columns if str(attr).startswith("fabric_type["))
    fabric_columns = [columns[attr] for attr in fabric_attrs if attr in columns]
    for column in fabric_columns:
        ws.cell(row_number, column).value = None
    for column, value in zip(fabric_columns, _catalog_export_fabric_type_values(pd)):
        ws.cell(row_number, column).value = value


def _catalog_stock_export_override(ws, mapping: dict, catalog: CatalogProduct | None, stock: int | None = None) -> tuple[int, int] | None:
    stock_value = stock if stock is not None else (catalog.stock if catalog else None)
    if stock_value is None:
        return None
    if stock_value < 0:
        raise ValueError(f"最新 GIGA 库存为 {stock_value}，不能导出负数库存。")
    quantity_attr = (mapping.get("dynamic_fields") or {}).get("quantity")
    if not quantity_attr:
        return None
    quantity_col = _template_attribute_columns(ws).get(str(quantity_attr))
    if not quantity_col:
        raise ValueError("模板未找到 Amazon 数量字段，无法写入最新 GIGA 库存。")
    return quantity_col, stock_value


def _summary_workbook(rows: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "导出报告"
    headers = ["状态", "商品资料ID", "商品ID", "商品Code", "类目", "模板文件", "导出文件", "原因"]
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header) for header in headers])
    for column_cells in ws.columns:
        column_letter = column_cells[0].column_letter
        max_length = max(len(str(cell.value or "")) for cell in column_cells[:80])
        ws.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 60)
    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()


def _save_workbook_bytes(wb: Workbook) -> bytes:
    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()


PRICE_QUANTITY_DATA_ROW = 7
PRICE_QUANTITY_FORMAT_ROW = 6
PRICE_QUANTITY_SKU_ATTR = "contribution_sku#1.value"
PRICE_QUANTITY_FULFILLMENT_ATTR = "fulfillment_availability#1.fulfillment_channel_code"
PRICE_QUANTITY_QUANTITY_ATTR = "fulfillment_availability#1.quantity"
PRICE_QUANTITY_INVENTORY_ALWAYS_AVAILABLE_ATTR = "fulfillment_availability#1.is_inventory_available"
PRICE_QUANTITY_HANDLING_TIME_ATTR = "fulfillment_availability#1.lead_time_to_ship_max_days"
PRICE_QUANTITY_FBM_VALUE = "Fulfillment by Merchant (Default)"


def _inventory_update_report_workbook(rows: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "库存模板导出报告"
    headers = ["状态", "商品资料ID", "任务ID", "商品Code", "真实ASIN", "SKU", "库存", "导出文件", "原因"]
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header) for header in headers])
    for column_cells in ws.columns:
        column_letter = column_cells[0].column_letter
        max_length = max(len(str(cell.value or "")) for cell in column_cells[:80])
        ws.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 60)
    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()


def _clear_price_quantity_data_rows(ws, row_count: int) -> None:
    max_row = max(ws.max_row, PRICE_QUANTITY_DATA_ROW + row_count + 2)
    for row in range(PRICE_QUANTITY_DATA_ROW, max_row + 1):
        _copy_row_format(ws, PRICE_QUANTITY_FORMAT_ROW, row)
        for col in range(1, ws.max_column + 1):
            ws.cell(row, col).value = None


def _catalog_price_quantity_sku(catalog: CatalogProduct) -> str:
    product = catalog.source_product
    product_data = product.data if product and product.data else None
    return str(catalog.item_code or (product_data.item_code if product_data else "") or "").strip()


def _catalog_store_context(catalog: CatalogProduct) -> tuple[str, int | None]:
    product = catalog.source_product
    product_data = product.data if product and product.data else None
    snapshot = _json_loads(product_data.gigab2b_raw_snapshot, {}) if product_data else {}
    if not isinstance(snapshot, dict):
        snapshot = {}
    site = str(snapshot.get("site") or "US").strip().upper()
    data_source_id = snapshot.get("data_source_id")
    try:
        parsed_data_source_id = int(data_source_id) if data_source_id else None
    except (TypeError, ValueError):
        parsed_data_source_id = None
    return site, parsed_data_source_id


async def _latest_giga_inventory_by_catalog_id(
    db: AsyncSession,
    catalog_items: list[CatalogProduct],
) -> dict[int, GigaInventory]:
    grouped: dict[tuple[str, int | None], list[CatalogProduct]] = {}
    for catalog in catalog_items:
        sku = _catalog_price_quantity_sku(catalog)
        if not sku:
            continue
        grouped.setdefault(_catalog_store_context(catalog), []).append(catalog)

    inventory_by_catalog_id: dict[int, GigaInventory] = {}
    for (site, data_source_id), items in grouped.items():
        sku_codes = list(dict.fromkeys(_catalog_price_quantity_sku(item) for item in items if _catalog_price_quantity_sku(item)))
        if not sku_codes:
            continue
        batch_query = (
            select(GigaSyncBatch)
            .where(
                GigaSyncBatch.status == "done",
                GigaSyncBatch.inventory_count > 0,
                GigaSyncBatch.site == site,
            )
            .order_by(GigaSyncBatch.finished_at.is_(None).asc(), GigaSyncBatch.finished_at.desc(), GigaSyncBatch.created_at.desc())
            .limit(1)
        )
        if data_source_id:
            batch_query = batch_query.where(GigaSyncBatch.data_source_id == data_source_id)
        batch_result = await db.execute(batch_query)
        latest_batch = batch_result.scalar_one_or_none()
        if not latest_batch:
            continue
        inventory_query = select(GigaInventory).where(
            GigaInventory.batch_id == latest_batch.batch_id,
            GigaInventory.site == latest_batch.site,
            GigaInventory.sku_code.in_(sku_codes),
        )
        if data_source_id:
            inventory_query = inventory_query.where(GigaInventory.data_source_id == data_source_id)
        inventory_result = await db.execute(inventory_query)
        inventory_by_sku = {row.sku_code: row for row in inventory_result.scalars().all()}
        for item in items:
            sku = _catalog_price_quantity_sku(item)
            inventory = inventory_by_sku.get(sku)
            if inventory:
                inventory_by_catalog_id[item.id] = inventory
    return inventory_by_catalog_id


def _naive_datetime(value: datetime | None) -> datetime | None:
    if value and value.tzinfo:
        return value.replace(tzinfo=None)
    return value


SOURCE_PRODUCT_DATA_FIELDS = {
    "item_code",
    "title",
    "color",
    "material",
    "filler",
    "product_type",
    "dimension_length",
    "dimension_width",
    "dimension_height",
    "weight",
    "packages",
    "value_total",
    "estimated_total",
    "shipping_cost",
    "shipping_cost_min",
    "shipping_cost_max",
    "features",
    "description",
    "variants",
    "gigab2b_raw_snapshot",
    "stock",
    "seller",
    "origin",
    "image_count",
    "material_dir",
    "collected_at",
}


GENERATED_PRODUCT_IMAGE_FIELDS = {
    "contact_sheet_path",
    "image_analysis",
    "image_selling_points",
    "category_style",
    "main_image_summary",
    "analyzed_at",
}


def _reset_product_data(pd: ProductData) -> None:
    for column in ProductData.__table__.columns:
        if column.name in {"id", "product_id"} or column.name in SOURCE_PRODUCT_DATA_FIELDS:
            continue
        setattr(pd, column.name, None)


def _reset_product_images(pi: ProductImage) -> None:
    for column_name in GENERATED_PRODUCT_IMAGE_FIELDS:
        setattr(pi, column_name, None)
    pi.vlm_model = settings.VLM_MODEL


def _strip_competitor_snapshot(snapshot_text: str | None) -> str | None:
    if not snapshot_text:
        return snapshot_text
    try:
        snapshot = json.loads(snapshot_text)
    except Exception:
        return snapshot_text
    if not isinstance(snapshot, dict):
        return snapshot_text
    for key in ("selected_stylesnap", "amazon_listing_capture", "stylesnap_search"):
        snapshot.pop(key, None)
    return json.dumps(snapshot, ensure_ascii=False)


async def _delete_product_competitor_records(db: AsyncSession, product: Product) -> None:
    if not product.data:
        return
    snapshot = {}
    if product.data.gigab2b_raw_snapshot:
        try:
            loaded = json.loads(product.data.gigab2b_raw_snapshot)
            if isinstance(loaded, dict):
                snapshot = loaded
        except Exception:
            snapshot = {}
    batch_id = snapshot.get("batch_id")
    site = str(snapshot.get("site") or "US").strip().upper()
    item_code = product.data.item_code
    representative_sku = snapshot.get("representative_sku") or item_code
    if not batch_id or not item_code:
        return

    candidate_query = select(AmazonStyleSnapCandidate.id).where(
        AmazonStyleSnapCandidate.batch_id == batch_id,
        AmazonStyleSnapCandidate.site == site,
        AmazonStyleSnapCandidate.item_code == item_code,
    )
    if representative_sku:
        candidate_query = candidate_query.where(AmazonStyleSnapCandidate.sku_code == representative_sku)
    result = await db.execute(candidate_query)
    candidate_ids = [row[0] for row in result.all()]
    if not candidate_ids and representative_sku and representative_sku != item_code:
        result = await db.execute(
            select(AmazonStyleSnapCandidate.id).where(
                AmazonStyleSnapCandidate.batch_id == batch_id,
                AmazonStyleSnapCandidate.site == site,
                AmazonStyleSnapCandidate.item_code == item_code,
            )
        )
        candidate_ids = [row[0] for row in result.all()]
    if candidate_ids:
        await db.execute(
            delete(AmazonListingCapture).where(AmazonListingCapture.selected_candidate_id.in_(candidate_ids))
        )
        await db.execute(delete(AmazonStyleSnapCandidate).where(AmazonStyleSnapCandidate.id.in_(candidate_ids)))


async def _giga_image_candidates_for_source(
    db: AsyncSession,
    item_code: str | None,
    raw_snapshot: str | None,
) -> list[dict[str, Any]]:
    if not item_code:
        return []
    snapshot = _json_loads(raw_snapshot, {})
    batch_id = snapshot.get("batch_id") if isinstance(snapshot, dict) else None
    site = str(snapshot.get("site") or "").strip().upper() if isinstance(snapshot, dict) else ""
    data_source_id = snapshot.get("data_source_id") if isinstance(snapshot, dict) else None
    representative_sku = str(snapshot.get("representative_sku") or "").strip() if isinstance(snapshot, dict) else ""

    def candidate_asset_source(image_type: Any) -> str:
        candidate_type = str(image_type or "unknown").strip().lower()
        if candidate_type in {"main", "gallery", "variant_main", "variant_gallery"}:
            return "giga_detail_gallery"
        if candidate_type == "file":
            return "giga_material_package"
        if candidate_type == "brand":
            return "giga_brand_asset"
        return "unknown"

    def classify_candidate_type(image_type: Any, sku_code: Any) -> str:
        candidate_type = str(image_type or "unknown").strip().lower()
        sku_text = str(sku_code or "").strip()
        if representative_sku and sku_text and sku_text != representative_sku and candidate_type in {"main", "gallery"}:
            return f"variant_{candidate_type}"
        return candidate_type

    query = select(GigaProductImage).where(GigaProductImage.item_code == item_code)
    if batch_id:
        query = query.where(GigaProductImage.batch_id == batch_id)
    if site:
        query = query.where(GigaProductImage.site == site)
    if data_source_id:
        try:
            query = query.where(GigaProductImage.data_source_id == int(data_source_id))
        except (TypeError, ValueError):
            pass
    result = await db.execute(
        query.order_by(
            GigaProductImage.sort_order.is_(None).asc(),
            GigaProductImage.sort_order.asc(),
            GigaProductImage.id.asc(),
        )
    )
    candidates: list[dict[str, Any]] = []
    seen: set[str] = set()
    for row in result.scalars().all():
        path = (row.local_path or row.image_url or "").strip()
        if not path or path in seen:
            continue
        seen.add(path)
        image_type = classify_candidate_type(row.image_type, row.sku_code)
        candidates.append({
            "path": path,
            "image_url": row.image_url,
            "local_path": row.local_path,
            "image_type": image_type,
            "sku_code": row.sku_code,
            "representative_sku": representative_sku or None,
            "is_representative_sku": bool(representative_sku and row.sku_code == representative_sku),
            "sort_order": row.sort_order,
            "download_status": row.download_status,
            "asset_source": candidate_asset_source(image_type),
            "source": "giga_product_images",
        })
    if candidates:
        return candidates

    snapshot_images = snapshot.get("giga_listing_images") if isinstance(snapshot, dict) else None
    if not isinstance(snapshot_images, list):
        return []
    for index, item in enumerate(snapshot_images, start=1):
        if not isinstance(item, dict):
            continue
        path = str(item.get("path") or item.get("local_path") or item.get("image_url") or "").strip()
        if not path or path in seen:
            continue
        seen.add(path)
        raw_type = item.get("image_type") or item.get("source") or ("main" if index == 1 else "gallery")
        sku_code = item.get("sku_code")
        image_type = classify_candidate_type(raw_type, sku_code)
        candidates.append({
            "path": path,
            "image_url": item.get("image_url"),
            "local_path": item.get("local_path"),
            "image_type": image_type,
            "sku_code": sku_code,
            "representative_sku": representative_sku or None,
            "is_representative_sku": bool(representative_sku and sku_code == representative_sku),
            "sort_order": item.get("sort_order") or index,
            "download_status": item.get("download_status"),
            "asset_source": candidate_asset_source(image_type),
            "source": item.get("source") or "giga_listing_images",
        })
    return candidates


async def _giga_image_candidates_for_product(db: AsyncSession, product: Product) -> list[dict[str, Any]]:
    if not product.data:
        return []
    return await _giga_image_candidates_for_source(
        db,
        product.data.item_code,
        product.data.gigab2b_raw_snapshot,
    )


async def _gallery_order_for_image_review(
    db: AsyncSession,
    item_code: str | None,
    raw_snapshot: str | None,
    gallery_order: str | None,
) -> str | None:
    existing_order = _json_loads(gallery_order, [])
    if isinstance(existing_order, list) and any(
        isinstance(item, dict) and (item.get("image_type") or item.get("source"))
        for item in existing_order
    ):
        return gallery_order
    if isinstance(existing_order, list) and existing_order:
        structured_order = []
        seen: set[str] = set()
        for index, item in enumerate(existing_order):
            path = normalize_image_path(item)
            if not path or path in seen:
                continue
            seen.add(path)
            structured_order.append({
                "path": path,
                "image_type": "main" if index == 0 else "gallery",
                "source": "saved_gallery_order",
                "sort_order": index + 1,
                "asset_source": "giga_listing",
            })
        if structured_order:
            return json.dumps(structured_order, ensure_ascii=False)
    fallback_images = await _giga_image_candidates_for_source(db, item_code, raw_snapshot)
    if fallback_images:
        return json.dumps(fallback_images, ensure_ascii=False)
    return gallery_order


def _limited_image_review_gallery_order(
    gallery_order: str | None,
    *,
    gallery_images: str | None = None,
    main_image_path: str | None = None,
    limit: int = IMAGE_REVIEW_INITIAL_GALLERY_LIMIT,
) -> tuple[str | None, int | None, int | None]:
    items = _json_loads(gallery_order, [])
    if not isinstance(items, list):
        return gallery_order, None, None
    total = len(items)
    if total <= limit:
        return gallery_order, total, limit

    saved_gallery_images = _json_loads(gallery_images, [])
    if not isinstance(saved_gallery_images, list):
        saved_gallery_images = []
    selected_paths = {
        path
        for path in [main_image_path, *[normalize_image_path(item) for item in saved_gallery_images if item]]
        if path
    }
    seen: set[str] = set()
    selected: list[Any] = []
    main_items: list[Any] = []
    gallery_items: list[Any] = []
    other_items: list[Any] = []

    for index, item in enumerate(items):
        path = normalize_image_path(item)
        if path in seen:
            continue
        seen.add(path)
        item_type = str((item or {}).get("image_type") or (item or {}).get("source") or "").strip().lower() if isinstance(item, dict) else ""
        if path and path in selected_paths:
            selected.append(item)
        elif item_type == "main":
            main_items.append(item)
        elif item_type == "gallery":
            gallery_items.append(item)
        else:
            other_items.append(item)

    limited: list[Any] = []
    limited_paths: set[str] = set()

    def add(entries: list[Any]) -> None:
        for entry in entries:
            if len(limited) >= limit:
                return
            path = normalize_image_path(entry)
            if path and path in limited_paths:
                continue
            if path:
                limited_paths.add(path)
            limited.append(entry)

    add(selected)
    add(main_items)
    add(gallery_items[-IMAGE_REVIEW_SELECTED_IMAGE_LIMIT:])
    add(other_items)
    add(items)
    return json.dumps(limited, ensure_ascii=False), total, limit


async def _ensure_product_detail_gallery_order(db: AsyncSession, product: Product, detail: ProductDetail) -> None:
    if detail.images and detail.images.gallery_order:
        existing_order = _json_loads(detail.images.gallery_order, [])
        if isinstance(existing_order, list) and any(
            isinstance(item, dict) and (item.get("image_type") or item.get("source"))
            for item in existing_order
        ):
            return
    fallback_images = await _giga_image_candidates_for_product(db, product)
    if not fallback_images:
        return
    gallery_order = json.dumps(fallback_images, ensure_ascii=False)
    if detail.images:
        detail.images.gallery_order = gallery_order
    else:
        detail.images = ProductImageResponse(
            id=0,
            product_id=product.id,
            gallery_order=gallery_order,
            vlm_model=settings.VLM_MODEL,
        )


def _snapshot_for_product(product: Product) -> dict:
    snapshot = _json_loads(product.data.gigab2b_raw_snapshot, {}) if product.data else {}
    return snapshot if isinstance(snapshot, dict) else {}


def _normalize_code_list(values: list[str] | tuple[str, ...] | None) -> list[str]:
    return list(dict.fromkeys(str(value).strip() for value in (values or []) if str(value).strip()))


def _item_code_for_product(product: Product, requested_item_code: str | None = None) -> str | None:
    return (
        str(requested_item_code or "").strip()
        or (product.data.item_code if product.data and product.data.item_code else None)
        or extract_gigab2b_item_code_from_url(product.gigab2b_url)
        or extract_gigab2b_product_id(product.gigab2b_url)
        or product.gigab2b_product_id
    )


async def _resolve_giga_refresh_data_source_id(
    db: AsyncSession,
    *,
    requested_data_source_id: int | None,
    snapshot: dict,
    site: str,
) -> int:
    if requested_data_source_id:
        return requested_data_source_id
    snapshot_data_source_id = snapshot.get("data_source_id")
    if snapshot_data_source_id:
        try:
            return int(snapshot_data_source_id)
        except (TypeError, ValueError):
            pass
    query = select(ProductDataSource).where(
        ProductDataSource.platform == "giga",
        ProductDataSource.enabled == 1,
        ProductDataSource.site == site,
    )
    result = await db.execute(query.order_by(ProductDataSource.id.asc()))
    sources = list(result.scalars().all())
    if len(sources) == 1:
        return sources[0].id
    if not sources:
        raise HTTPException(400, f"站点 {site} 没有启用的 GIGA 店铺")
    raise HTTPException(400, "该商品无法自动确定 GIGA 店铺，请在商品工作台选择店铺后同步")


async def _giga_refresh_sku_codes(
    db: AsyncSession,
    *,
    product: Product,
    item_code: str,
    data_source_id: int,
    site: str,
    requested_sku_codes: list[str],
    snapshot: dict,
) -> list[str]:
    requested = _normalize_code_list(requested_sku_codes)
    if requested:
        return requested
    item_snapshot = snapshot.get("item") if isinstance(snapshot.get("item"), dict) else {}
    snapshot_skus = _normalize_code_list(item_snapshot.get("sku_codes") if isinstance(item_snapshot, dict) else [])
    if snapshot_skus:
        return snapshot_skus
    result = await db.execute(
        select(GigaItem.id)
        .where(
            GigaItem.site == site,
            GigaItem.data_source_id == data_source_id,
            GigaItem.item_code == item_code,
        )
        .order_by(GigaItem.updated_at.is_(None).asc(), GigaItem.updated_at.desc(), GigaItem.id.desc())
        .limit(1)
    )
    giga_item_id = result.scalar_one_or_none()
    if not giga_item_id:
        return [item_code]
    result = await db.execute(
        select(GigaSku.sku_code)
        .where(GigaSku.giga_item_id == giga_item_id)
        .order_by(GigaSku.child_sequence.is_(None).asc(), GigaSku.child_sequence.asc(), GigaSku.sku_code.asc())
    )
    db_skus = _normalize_code_list([sku for sku in result.scalars().all()])
    return db_skus or [item_code]


def _reset_product_aplus(pa: ProductAplus) -> None:
    for column in ProductAplus.__table__.columns:
        if column.name in {"id", "product_id"}:
            continue
        if column.name == "llm_model":
            setattr(pa, column.name, settings.LLM_MODEL)
        else:
            setattr(pa, column.name, None)


async def _raise_if_duplicate_gigab2b_url(db: AsyncSession, gigab2b_url: str) -> None:
    gigab2b_product_id = extract_gigab2b_product_id(gigab2b_url)
    duplicate = await find_duplicate_by_gigab2b_product_id(db, gigab2b_product_id)
    if duplicate:
        raise HTTPException(409, f"{duplicate.message}，已跳过创建")
    url_item_code = extract_gigab2b_item_code_from_url(gigab2b_url)
    duplicate = await find_duplicate_by_item_code(db, url_item_code)
    if duplicate:
        raise HTTPException(409, f"{duplicate.message}，已跳过创建")


@router.post("", response_model=ProductResponse, status_code=201)
async def create_product(body: ProductCreate, db: AsyncSession = Depends(get_db)):
    """创建新商品任务"""
    await _raise_if_duplicate_gigab2b_url(db, body.gigab2b_url)
    gigab2b_product_id = extract_gigab2b_product_id(body.gigab2b_url)
    product = Product(
        gigab2b_url=body.gigab2b_url,
        gigab2b_product_id=gigab2b_product_id,
        competitor_asin=body.competitor_asin,
        brand=body.brand,
    )
    db.add(product)
    await db.flush()
    try:
        await ensure_product_upc(db, product)
    except UpcPoolEmptyError as exc:
        raise HTTPException(400, str(exc))
    # 创建关联空子表
    db.add(ProductData(product_id=product.id))
    db.add(ProductImage(product_id=product.id))
    db.add(ProductAplus(product_id=product.id))
    db.add(CatalogProduct(
        source_product_id=product.id,
        gigab2b_url=product.gigab2b_url,
        gigab2b_product_id=product.gigab2b_product_id,
        competitor_asin=product.competitor_asin,
        asin_sync_status=product.asin_sync_status or "not_synced",
        aplus_upload_status=product.aplus_upload_status or "not_uploaded",
        upc=product.upc,
        brand=product.brand,
        status=product.status,
    ))
    await db.commit()
    await db.refresh(product)
    return product


@router.get("/overview", response_model=WorkbenchOverview)
async def get_workbench_overview(
    data_source_id: int | None = Query(None, ge=1),
    db: AsyncSession = Depends(get_db),
):
    """工作台概览：用于顶部快速发现需要处理的任务和商品。"""
    product_query = select(Product).options(
        load_only(
            Product.id,
            Product.status,
            Product.current_step,
            Product.competitor_asin,
            Product.error_message,
            Product.source_data_source_id,
        )
    )
    if data_source_id:
        product_query = product_query.where(_product_data_source_filter(data_source_id))
    product_result = await db.execute(product_query)
    products = product_result.scalars().all()
    status_counts = {key: 0 for key in WORKBENCH_STATUS_KEYS}
    for product in products:
        status_counts[_product_workbench_status(product)] += 1

    exported_condition = (CatalogProduct.exported_at.is_not(None)) | (CatalogProduct.export_task_id.is_not(None))
    export_ready_query = (
        select(
            func.count(Product.id).label("total"),
            func.coalesce(func.sum(case((exported_condition, 1), else_=0)), 0).label("exported"),
        )
        .select_from(Product)
        .join(CatalogProduct, CatalogProduct.source_product_id == Product.id, isouter=True)
        .where(Product.status == COMPLETED, Product.current_step >= 6)
    )
    if data_source_id:
        export_ready_query = export_ready_query.where(_product_data_source_filter(data_source_id))
    export_ready_row = (await db.execute(export_ready_query)).one()
    export_ready_total = int(export_ready_row.total or 0)
    export_ready_exported = int(export_ready_row.exported or 0)
    export_ready_unexported = max(export_ready_total - export_ready_exported, 0)

    asin_not_synced_result = await db.execute(
        select(func.count(CatalogProduct.id)).where(
            CatalogProduct.confirmed_at.is_not(None),
            ((CatalogProduct.amazon_asin.is_(None)) | (CatalogProduct.amazon_asin == "")),
            ((CatalogProduct.asin_sync_status.is_(None)) | (CatalogProduct.asin_sync_status.in_(("not_synced", "skipped")))),
        )
    )
    asin_attention_result = await db.execute(
        select(func.count(CatalogProduct.id)).where(
            CatalogProduct.confirmed_at.is_not(None),
            CatalogProduct.asin_sync_status.in_(("not_found", "multiple_found", "failed")),
        )
    )
    aplus_failed_result = await db.execute(
        select(func.count(CatalogProduct.id)).where(
            CatalogProduct.confirmed_at.is_not(None),
            CatalogProduct.aplus_upload_status == "failed",
        )
    )
    listing_high_risk_result = await db.execute(
        select(func.count(ProductData.id))
        .join(CatalogProduct, CatalogProduct.source_product_id == ProductData.product_id)
        .where(
            CatalogProduct.confirmed_at.is_not(None),
            ProductData.amazon_template_fill_summary.like('%"risk_level"%high_risk%'),
        )
    )
    return WorkbenchOverview(
        total_products=len(products),
        select_images=status_counts["select_images"],
        competitor_searching=status_counts["competitor_searching"],
        select_competitor=status_counts["select_competitor"],
        capture_detail=status_counts["capture_detail"],
        ready_to_generate=status_counts["ready_to_generate"],
        running=status_counts["running"],
        interrupted=status_counts["interrupted"],
        suspended=status_counts["suspended"],
        manual_review=status_counts["manual_review"],
        export_ready=status_counts["export_ready"],
        export_ready_unexported=export_ready_unexported,
        export_ready_exported=export_ready_exported,
        failed=status_counts["failed"],
        running_tasks=status_counts["running"],
        manual_review_tasks=status_counts["manual_review"],
        failed_tasks=status_counts["failed"],
        confirmable_tasks=0,
        asin_not_synced=asin_not_synced_result.scalar() or 0,
        asin_attention=asin_attention_result.scalar() or 0,
        aplus_failed=aplus_failed_result.scalar() or 0,
        listing_high_risk=listing_high_risk_result.scalar() or 0,
    )


@router.get("/category-options")
async def list_category_options(db: AsyncSession = Depends(get_db)):
    """返回人工编辑 Amazon 类目时允许选择的已有类目。"""
    options: dict[str, dict] = {item["key"]: item for item in _static_category_options()}
    result = await db.execute(
        select(ProductData.categories, ProductData.leaf_category)
        .where((ProductData.categories.is_not(None)) | (ProductData.leaf_category.is_not(None)))
    )
    for categories_raw, leaf in result.all():
        categories: list[str] = []
        if categories_raw:
            try:
                parsed = json.loads(categories_raw)
                if isinstance(parsed, list):
                    categories = [str(item).strip() for item in parsed if str(item).strip()]
                else:
                    categories = _split_category_path(str(parsed))
            except Exception:
                categories = _split_category_path(str(categories_raw))
        leaf = str(leaf or "").strip() or (categories[-1] if categories else "")
        if not categories and leaf:
            categories = [leaf]
        if not categories:
            continue
        key = _category_option_key(categories, leaf)
        options[key] = {
            "key": key,
            "label": " > ".join(categories),
            "categories": categories,
            "leaf_category": leaf,
            "source": "history",
        }
    return {"items": sorted(options.values(), key=lambda item: item["label"])}


async def _allowed_category_keys(db: AsyncSession) -> set[str]:
    result = await list_category_options(db)
    return {str(item.get("key") or "") for item in result.get("items", [])}


@router.get("", response_model=PaginatedResponse)
async def list_products(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    status: str | None = None,
    work_status: str | None = Query(None),
    item_id: str | None = None,
    sku_code: str | None = None,
    data_source_id: int | None = Query(None, ge=1),
    competitor_asin: str | None = None,
    upc: str | None = None,
    created_from: datetime | None = None,
    created_to: datetime | None = None,
    db: AsyncSession = Depends(get_db),
):
    """商品任务列表（分页）"""
    if work_status and work_status not in PRODUCT_LIST_WORK_STATUS_KEYS:
        raise HTTPException(400, f"不支持的工作状态筛选：{work_status}")

    query = (
        select(Product)
        .options(
            selectinload(Product.data).load_only(
                ProductData.id,
                ProductData.product_id,
                ProductData.item_code,
                ProductData.title,
                ProductData.leaf_category,
            ),
            selectinload(Product.aplus).load_only(
                ProductAplus.id,
                ProductAplus.product_id,
                ProductAplus.aplus_status,
                ProductAplus.aplus_image_count,
            ),
            selectinload(Product.catalog_item).load_only(
                CatalogProduct.id,
                CatalogProduct.source_product_id,
                CatalogProduct.exported_at,
                CatalogProduct.export_task_id,
            ),
        )
        .order_by(Product.updated_at.is_(None).asc(), Product.updated_at.desc(), Product.created_at.desc())
    )
    count_query = select(func.count(Product.id))

    needs_product_data_join = bool(item_id or sku_code)
    if needs_product_data_join:
        query = query.join(ProductData, ProductData.product_id == Product.id, isouter=True)
        count_query = count_query.join(ProductData, ProductData.product_id == Product.id, isouter=True)

    if status:
        if status == COMPLETED:
            query = query.where(Product.status == COMPLETED, Product.current_step >= 6)
            count_query = count_query.where(Product.status == COMPLETED, Product.current_step >= 6)
        else:
            query = query.where(Product.status == status)
            count_query = count_query.where(Product.status == status)
    if item_id:
        pattern = f"%{item_id.strip()}%"
        query = query.where(
            (Product.gigab2b_product_id.ilike(pattern)) | (ProductData.item_code.ilike(pattern))
        )
        count_query = count_query.where(
            (Product.gigab2b_product_id.ilike(pattern)) | (ProductData.item_code.ilike(pattern))
        )
    if sku_code:
        pattern = f"%{sku_code.strip()}%"
        query = query.where(
            (Product.gigab2b_product_id.ilike(pattern))
            | (Product.source_item_id.ilike(pattern))
            | (ProductData.item_code.ilike(pattern))
            | (ProductData.title.ilike(pattern))
        )
        count_query = count_query.where(
            (Product.gigab2b_product_id.ilike(pattern))
            | (Product.source_item_id.ilike(pattern))
            | (ProductData.item_code.ilike(pattern))
            | (ProductData.title.ilike(pattern))
        )
    if data_source_id:
        query = query.where(_product_data_source_filter(data_source_id))
        count_query = count_query.where(_product_data_source_filter(data_source_id))
    if competitor_asin:
        pattern = f"%{competitor_asin.strip()}%"
        query = query.where(Product.competitor_asin.ilike(pattern))
        count_query = count_query.where(Product.competitor_asin.ilike(pattern))
    if upc:
        pattern = f"%{upc.strip()}%"
        query = query.where(Product.upc.ilike(pattern))
        count_query = count_query.where(Product.upc.ilike(pattern))
    created_from = _naive_datetime(created_from)
    created_to = _naive_datetime(created_to)
    if created_from:
        query = query.where(Product.created_at >= created_from)
        count_query = count_query.where(Product.created_at >= created_from)
    if created_to:
        query = query.where(Product.created_at <= created_to)
        count_query = count_query.where(Product.created_at <= created_to)

    if work_status:
        result = await db.execute(query)
        matched_items = [
            item for item in result.scalars().unique().all()
            if _product_list_work_status(item) == work_status
        ]
        total = len(matched_items)
        start = (page - 1) * page_size
        items = matched_items[start:start + page_size]
        return PaginatedResponse(items=[_build_list_item(item) for item in items], total=total, page=page, page_size=page_size)

    total_result = await db.execute(count_query)
    total = total_result.scalar() or 0

    result = await db.execute(
        query.offset((page - 1) * page_size).limit(page_size)
    )
    items = result.scalars().all()

    return PaginatedResponse(items=[_build_list_item(item) for item in items], total=total, page=page, page_size=page_size)


@router.get("/image-review-queue", response_model=ProductImageReviewQueueResponse)
async def list_product_image_review_queue(
    data_source_id: int | None = Query(None, ge=1),
    limit: int = Query(100, ge=1, le=500),
    db: AsyncSession = Depends(get_db),
):
    """Lightweight queue for the image review page.

    The full product list loads large ORM objects and response fields that this
    workflow does not need. Keep this endpoint narrow so switching into image
    review remains fast.
    """
    base_filter = [
        Product.status == "created",
        ((Product.current_step.is_(None)) | (Product.current_step <= 0)),
    ]
    if data_source_id:
        base_filter.append(_product_data_source_filter(data_source_id))

    total_result = await db.execute(
        select(func.count(Product.id))
        .select_from(Product)
        .where(*base_filter)
    )
    total = total_result.scalar() or 0

    query = (
        select(
            Product.id,
            Product.gigab2b_product_id,
            Product.status,
            Product.current_step,
            Product.created_at,
            Product.updated_at,
            ProductData.item_code,
            ProductData.title,
        )
        .select_from(Product)
        .join(ProductData, ProductData.product_id == Product.id, isouter=True)
        .where(*base_filter)
        .order_by(Product.updated_at.is_(None).asc(), Product.updated_at.desc(), Product.created_at.desc())
        .limit(limit)
    )

    result = await db.execute(query)
    items = [
        {
            "id": row.id,
            "gigab2b_product_id": row.gigab2b_product_id,
            "status": row.status,
            "current_step": row.current_step or 0,
            "current_task_status": "待确认商品图片",
            "item_code": row.item_code,
            "title": row.title,
            "created_at": row.created_at,
            "updated_at": row.updated_at,
        }
        for row in result.all()
    ]
    return ProductImageReviewQueueResponse(items=items, total=total, limit=limit)


@router.get("/image-review-detail/{product_id}", response_model=ProductImageReviewDetailResponse)
async def get_product_image_review_detail(
    product_id: int,
    image_limit: int = Query(IMAGE_REVIEW_INITIAL_GALLERY_LIMIT, ge=12, le=IMAGE_REVIEW_MAX_GALLERY_LIMIT),
    db: AsyncSession = Depends(get_db),
):
    """Lightweight product detail for the image review page."""
    result = await db.execute(
        select(
            Product.id,
            Product.gigab2b_product_id,
            Product.status,
            Product.current_step,
            Product.error_message,
            ProductData.item_code,
            ProductData.title,
            ProductData.gigab2b_raw_snapshot,
            ProductImage.id.label("image_id"),
            ProductImage.main_image_path,
            ProductImage.main_image_source,
            ProductImage.gallery_images,
            ProductImage.gallery_order,
        )
        .select_from(Product)
        .join(ProductData, ProductData.product_id == Product.id, isouter=True)
        .join(ProductImage, ProductImage.product_id == Product.id, isouter=True)
        .where(Product.id == product_id)
    )
    row = result.one_or_none()
    if not row:
        raise HTTPException(404, "Product not found")

    product = Product(
        id=row.id,
        gigab2b_url="",
        gigab2b_product_id=row.gigab2b_product_id,
        status=row.status,
        current_step=row.current_step or 0,
        error_message=row.error_message,
    )
    gallery_order = await _gallery_order_for_image_review(
        db,
        row.item_code,
        row.gigab2b_raw_snapshot,
        row.gallery_order,
    )
    gallery_order, gallery_order_total, gallery_order_limit = _limited_image_review_gallery_order(
        gallery_order,
        gallery_images=row.gallery_images,
        main_image_path=row.main_image_path,
        limit=image_limit,
    )
    images = None
    if row.image_id or row.main_image_path or row.gallery_images or gallery_order:
        images = {
            "id": row.image_id or 0,
            "product_id": row.id,
            "main_image_path": row.main_image_path,
            "main_image_source": row.main_image_source,
            "gallery_images": row.gallery_images,
            "gallery_order": gallery_order,
            "gallery_order_total": gallery_order_total,
            "gallery_order_limit": gallery_order_limit,
        }
    return {
        "id": row.id,
        "source_item_id": row.gigab2b_product_id,
        "gigab2b_product_id": row.gigab2b_product_id,
        "status": row.status,
        "current_step": row.current_step or 0,
        "current_task_status": _current_task_status(product),
        "data": {
            "item_code": row.item_code,
            "title": row.title,
        },
        "images": images,
    }


@router.get("/competitor-review-queue", response_model=ProductCompetitorReviewQueueResponse)
async def list_product_competitor_review_queue(
    data_source_id: int | None = Query(None, ge=1),
    limit: int = Query(100, ge=1, le=500),
    db: AsyncSession = Depends(get_db),
):
    """Lightweight queue for products that need competitor review."""
    created_needs_competitor = (
        (Product.status == "created")
        & (Product.current_step > 0)
        & (Product.competitor_asin.is_(None))
    )
    query = (
        select(
            Product.id,
            Product.gigab2b_product_id,
            Product.competitor_asin,
            Product.status,
            Product.current_step,
            Product.error_message,
            Product.created_at,
            Product.updated_at,
            ProductData.item_code,
            ProductData.title,
            ProductData.leaf_category,
        )
        .select_from(Product)
        .join(ProductData, ProductData.product_id == Product.id, isouter=True)
        .where(
            created_needs_competitor
            | _competitor_search_failed_sql_condition()
        )
        .order_by(Product.updated_at.is_(None).asc(), Product.updated_at.desc(), Product.created_at.desc())
        .limit(limit)
    )
    if data_source_id:
        query = query.where(_product_data_source_filter(data_source_id))

    result = await db.execute(query)
    items = []
    for row in result.all():
        product = Product(
            id=row.id,
            gigab2b_url="",
            gigab2b_product_id=row.gigab2b_product_id,
            competitor_asin=row.competitor_asin,
            status=row.status,
            current_step=row.current_step or 0,
            error_message=row.error_message,
        )
        items.append({
            "id": row.id,
            "source_item_id": row.gigab2b_product_id,
            "gigab2b_product_id": row.gigab2b_product_id,
            "competitor_asin": row.competitor_asin,
            "status": row.status,
            "current_step": row.current_step or 0,
            "current_task_status": _current_task_status(product),
            "error_message": row.error_message,
            "item_code": row.item_code,
            "title": row.title,
            "leaf_category": row.leaf_category,
            "created_at": row.created_at,
            "updated_at": row.updated_at,
        })
    return ProductCompetitorReviewQueueResponse(items=items, total=len(items), limit=limit)


@router.get("/competitor-review-detail/{product_id}", response_model=ProductCompetitorReviewDetailResponse)
async def get_product_competitor_review_detail(
    product_id: int,
    db: AsyncSession = Depends(get_db),
):
    """Lightweight product detail for the competitor review page."""
    result = await db.execute(
        select(
            Product.id,
            Product.gigab2b_product_id,
            Product.competitor_asin,
            Product.status,
            Product.current_step,
            Product.error_message,
            ProductData.item_code,
            ProductData.title,
            ProductData.leaf_category,
            ProductData.gigab2b_raw_snapshot,
            ProductImage.id.label("image_id"),
            ProductImage.main_image_path,
            ProductImage.main_image_source,
        )
        .select_from(Product)
        .join(ProductData, ProductData.product_id == Product.id, isouter=True)
        .join(ProductImage, ProductImage.product_id == Product.id, isouter=True)
        .where(Product.id == product_id)
    )
    row = result.one_or_none()
    if not row:
        raise HTTPException(404, "Product not found")

    product = Product(
        id=row.id,
        gigab2b_url="",
        gigab2b_product_id=row.gigab2b_product_id,
        competitor_asin=row.competitor_asin,
        status=row.status,
        current_step=row.current_step or 0,
        error_message=row.error_message,
    )
    images = None
    if row.image_id or row.main_image_path:
        images = {
            "id": row.image_id or 0,
            "product_id": row.id,
            "main_image_path": row.main_image_path,
            "main_image_source": row.main_image_source,
        }
    return {
        "id": row.id,
        "source_item_id": row.gigab2b_product_id,
        "gigab2b_product_id": row.gigab2b_product_id,
        "competitor_asin": row.competitor_asin,
        "status": row.status,
        "current_step": row.current_step or 0,
        "current_task_status": _current_task_status(product),
        "error_message": row.error_message,
        "leaf_category": row.leaf_category,
        "data": {
            "item_code": row.item_code,
            "title": row.title,
            "gigab2b_raw_snapshot": _compact_gigab2b_snapshot(row.gigab2b_raw_snapshot),
        },
        "images": images,
    }


@router.get("/import/template")
async def download_import_template():
    """下载批量创建任务模板。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "批量导入任务"
    ws.append(IMPORT_TEMPLATE_HEADERS)
    ws.append(["https://www.gigab2b.com/product-detail/example", "B0XXXXXXXX"])
    for width, column in zip((42, 18), ("A", "B")):
        ws.column_dimensions[column].width = width
    return _workbook_response(wb, "fbm_task_import_template.xlsx")


@router.post("/import", response_model=BulkImportResponse)
async def import_products(file: UploadFile = File(...), db: AsyncSession = Depends(get_db)):
    """批量导入任务，并同步创建商品资料记录。"""
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(400, "请上传 xlsx/xlsm Excel 文件")

    try:
        content = await file.read()
        wb = load_workbook(BytesIO(content), data_only=True)
    except Exception:
        raise HTTPException(400, "Excel 文件无法读取")

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(400, "Excel 文件为空")

    header_row = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    index_map: dict[str, int] = {}
    for idx, header in enumerate(header_row):
        normalized = header.replace(" ", "")
        field = HEADER_ALIASES.get(header) or HEADER_ALIASES.get(normalized) or HEADER_ALIASES.get(header.lower())
        if field:
            index_map[field] = idx

    if "gigab2b_url" not in index_map or "competitor_asin" not in index_map:
        raise HTTPException(400, "模板必须包含：原始数据链接、竞品ASIN")

    created_ids: list[int] = []
    errors: list[str] = []
    skipped_details: list[str] = []
    skipped = 0
    seen_product_ids: dict[str, int] = {}
    seen_item_codes: dict[str, int] = {}
    remaining_upcs = await available_upc_count(db)

    for row_number, row in enumerate(rows[1:], start=2):
        values = list(row)
        gigab2b_url = str(values[index_map["gigab2b_url"]]).strip() if index_map["gigab2b_url"] < len(values) and values[index_map["gigab2b_url"]] is not None else ""
        competitor_asin = str(values[index_map["competitor_asin"]]).strip() if index_map["competitor_asin"] < len(values) and values[index_map["competitor_asin"]] is not None else ""
        brand = "Vindhvisk"
        if "brand" in index_map and index_map["brand"] < len(values) and values[index_map["brand"]] is not None:
            brand = str(values[index_map["brand"]]).strip() or brand

        if not any([gigab2b_url, competitor_asin]):
            skipped += 1
            continue
        if not gigab2b_url or not competitor_asin:
            errors.append(f"第 {row_number} 行缺少必填字段")
            continue

        gigab2b_product_id = extract_gigab2b_product_id(gigab2b_url)
        url_item_code = extract_gigab2b_item_code_from_url(gigab2b_url)
        if gigab2b_product_id and gigab2b_product_id in seen_product_ids:
            skipped += 1
            skipped_details.append(f"第 {row_number} 行跳过：大建云仓商品ID {gigab2b_product_id} 已在第 {seen_product_ids[gigab2b_product_id]} 行导入")
            continue
        if url_item_code and url_item_code in seen_item_codes:
            skipped += 1
            skipped_details.append(f"第 {row_number} 行跳过：商品Code {url_item_code} 已在第 {seen_item_codes[url_item_code]} 行导入")
            continue

        duplicate = await find_duplicate_by_gigab2b_product_id(db, gigab2b_product_id)
        if not duplicate:
            duplicate = await find_duplicate_by_item_code(db, url_item_code)
        if duplicate:
            skipped += 1
            skipped_details.append(f"第 {row_number} 行跳过：{duplicate.message}")
            continue

        if remaining_upcs <= 0:
            raise HTTPException(400, f"UPC池子可用UPC不足：已可导入 {len(created_ids)} 行，后续第 {row_number} 行没有可用UPC")

        product = Product(
            gigab2b_url=gigab2b_url,
            gigab2b_product_id=gigab2b_product_id,
            competitor_asin=competitor_asin,
            brand=brand,
        )
        db.add(product)
        await db.flush()
        try:
            await ensure_product_upc(db, product)
        except UpcPoolEmptyError as exc:
            raise HTTPException(400, str(exc))
        remaining_upcs -= 1
        db.add(ProductData(product_id=product.id))
        db.add(ProductImage(product_id=product.id))
        db.add(ProductAplus(product_id=product.id))
        db.add(CatalogProduct(
            source_product_id=product.id,
            gigab2b_url=product.gigab2b_url,
            gigab2b_product_id=gigab2b_product_id,
            competitor_asin=product.competitor_asin,
            asin_sync_status=product.asin_sync_status or "not_synced",
            aplus_upload_status=product.aplus_upload_status or "not_uploaded",
            upc=product.upc,
            brand=product.brand,
            status=product.status,
        ))
        if gigab2b_product_id:
            seen_product_ids[gigab2b_product_id] = row_number
        if url_item_code:
            seen_item_codes[url_item_code] = row_number
        created_ids.append(product.id)

    await db.commit()
    return BulkImportResponse(
        created=len(created_ids),
        skipped=skipped,
        skipped_details=skipped_details,
        errors=errors,
        product_ids=created_ids,
    )


@router.get("/upc-pool", response_model=PaginatedUpcPoolItems)
async def list_upc_pool(
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
    status: str | None = None,
    q: str | None = None,
    db: AsyncSession = Depends(get_db),
):
    """UPC池子列表。已绑定记录只允许同一商品Code/来源商品ID重复确认。"""
    query = select(UpcPoolItem).order_by(UpcPoolItem.id.asc())
    count_query = select(func.count(UpcPoolItem.id))
    if status:
        query = query.where(UpcPoolItem.status == status)
        count_query = count_query.where(UpcPoolItem.status == status)
    if q:
        pattern = f"%{q.strip()}%"
        criteria = (
            (UpcPoolItem.upc.ilike(pattern))
            | (UpcPoolItem.bound_item_code.ilike(pattern))
            | (UpcPoolItem.bound_source_product_id.ilike(pattern))
        )
        query = query.where(criteria)
        count_query = count_query.where(criteria)

    total_result = await db.execute(count_query)
    result = await db.execute(query.offset((page - 1) * page_size).limit(page_size))
    return PaginatedUpcPoolItems(
        items=result.scalars().all(),
        total=total_result.scalar() or 0,
        page=page,
        page_size=page_size,
        summary=await _upc_pool_summary(db),
    )


@router.post("/upc-pool/import", response_model=UpcPoolImportResponse)
async def import_upc_pool(body: UpcPoolImportRequest, db: AsyncSession = Depends(get_db)):
    """批量追加 UPC 到池子，重复 UPC 会保留原记录。"""
    result = await add_upcs_to_pool(db, body.text)
    await db.commit()
    return UpcPoolImportResponse(
        added=result["added"],
        duplicated=result["duplicated"],
        invalid=result["invalid"],
        summary=await _upc_pool_summary(db),
    )


@router.post("/bulk-start", response_model=BulkStartResponse)
async def bulk_start_pipeline(body: BulkStartRequest, db: AsyncSession = Depends(get_db)):
    """批量启动待处理任务。只会启动 created 状态的任务，其余返回跳过原因。"""
    requested_ids = list(dict.fromkeys(body.product_ids))
    if len(requested_ids) > settings.BULK_START_MAX_TASKS:
        raise HTTPException(400, f"单次最多批量启动 {settings.BULK_START_MAX_TASKS} 个任务")
    result = await db.execute(
        select(Product)
        .options(selectinload(Product.data), selectinload(Product.images), selectinload(Product.aplus))
        .where(Product.id.in_(requested_ids))
    )
    products = {product.id: product for product in result.scalars().all()}

    errors: list[str] = []
    started: list[tuple[int, int]] = []
    for product_id in requested_ids:
        product = products.get(product_id)
        if not product:
            errors.append(f"任务 {product_id} 不存在")
            continue
        if product.status != "created":
            errors.append(f"任务 {product_id} 当前状态为 {product.status}，已跳过")
            continue
        if is_running(product.id):
            errors.append(f"任务 {product_id} 已在运行中，已跳过")
            continue
        if (product.current_step or 0) < 5:
            errors.append(f"任务 {product_id} 尚未完成图片/竞品确认，不能启动生成")
            continue
        try:
            await _require_generation_prerequisites(db, product, _generation_start_step(product))
        except HTTPException as exc:
            errors.append(f"任务 {product_id} {exc.detail}")
            continue
        started.append((product.id, _generation_start_step(product)))

    actually_started: list[int] = []
    for product_id, start_step in started:
        product = products.get(product_id)
        if start_step <= 5:
            try:
                if product:
                    await _queue_product_image_analysis(db, product, created_by="bulk_start")
                else:
                    await create_product_image_analysis_runs(db, [product_id], created_by="bulk_start")
                actually_started.append(product_id)
            except ValueError as exc:
                errors.append(f"任务 {product_id} {exc}")
            continue
        if start_step == 6:
            try:
                if product:
                    await _queue_product_listing_generation(db, product, created_by="bulk_start")
                else:
                    await create_product_listing_runs(db, [product_id], created_by="bulk_start")
                actually_started.append(product_id)
            except ValueError as exc:
                errors.append(f"任务 {product_id} {exc}")
            continue
        if enqueue_pipeline(product_id, start_step=start_step):
            actually_started.append(product_id)
            continue
        errors.append(f"任务 {product_id} 后台队列已存在，已跳过")
        if product:
            product.status = "created"
            product.current_step = 0
            product.updated_at = datetime.now()

    if len(actually_started) != len(started):
        await db.commit()

    return BulkStartResponse(
        requested=len(requested_ids),
        started=len(actually_started),
        skipped=len(requested_ids) - len(actually_started),
        errors=errors,
        started_ids=actually_started,
    )


@router.post("/auto-start-ready-generation", response_model=BulkStartResponse)
async def auto_start_ready_generation(
    data_source_id: int | None = Query(None, ge=1),
    limit: int = Query(AUTO_START_READY_GENERATION_LIMIT, ge=1, le=AUTO_START_READY_GENERATION_LIMIT),
    db: AsyncSession = Depends(get_db),
):
    """自动认领已满足前置条件的待生成商品，避免停留在“待启动/待生成”。"""
    query = (
        select(Product)
        .options(
            selectinload(Product.data),
            selectinload(Product.images),
            selectinload(Product.aplus),
            selectinload(Product.catalog_item),
        )
        .where(
            Product.status == "created",
            Product.current_step >= 5,
            Product.competitor_asin.is_not(None),
            Product.competitor_asin != "",
        )
        .order_by(Product.updated_at.asc(), Product.id.asc())
        .limit(limit)
    )
    if data_source_id:
        query = query.where(_product_data_source_filter(data_source_id))

    result = await db.execute(query)
    products = result.scalars().all()
    errors: list[str] = []
    startable: list[tuple[int, int, str, int, str | None]] = []

    for product in products:
        if is_running(product.id):
            errors.append(f"任务 {product.id} 已在运行中，已跳过")
            continue
        start_step = _generation_start_step(product)
        try:
            await _require_generation_prerequisites(db, product, start_step)
        except HTTPException as exc:
            errors.append(f"任务 {product.id} {exc.detail}")
            continue
        startable.append((product.id, start_step, product.status, product.current_step or 0, product.error_message))

    started_ids: list[int] = []
    for product_id, start_step, original_status, original_step, original_error in startable:
        product = next((item for item in products if item.id == product_id), None)
        if start_step <= 5:
            try:
                if product:
                    await _queue_product_image_analysis(db, product, created_by="auto_start_ready_generation")
                else:
                    await create_product_image_analysis_runs(db, [product_id], created_by="auto_start_ready_generation")
                started_ids.append(product_id)
            except ValueError as exc:
                errors.append(f"任务 {product_id} {exc}")
            continue
        if start_step == 6:
            try:
                if product:
                    await _queue_product_listing_generation(db, product, created_by="auto_start_ready_generation")
                else:
                    await create_product_listing_runs(db, [product_id], created_by="auto_start_ready_generation")
                started_ids.append(product_id)
            except ValueError as exc:
                errors.append(f"任务 {product_id} {exc}")
            continue
        if enqueue_pipeline(product_id, start_step=start_step):
            started_ids.append(product_id)
            continue
        errors.append(f"任务 {product_id} 后台队列已存在，已跳过")
        if product:
            product.status = original_status
            product.current_step = original_step
            product.error_message = original_error
            product.updated_at = datetime.now()
            _sync_catalog_item(product, db)

    if len(started_ids) != len(startable):
        await db.commit()

    return BulkStartResponse(
        requested=len(products),
        started=len(started_ids),
        skipped=len(products) - len(started_ids),
        errors=errors,
        started_ids=started_ids,
    )


@router.post("/bulk-advance-task", response_model=TaskRunResponse)
async def create_product_bulk_advance_task(body: ProductBulkAdvanceRequest, db: AsyncSession = Depends(get_db)):
    return await create_product_bulk_advance_run(db, body.product_ids)


@router.post("/bulk-advance-task/by-filter", response_model=TaskRunResponse)
async def create_product_bulk_advance_task_by_filter(body: ProductBulkAdvanceFilterRequest, db: AsyncSession = Depends(get_db)):
    query = (
        select(Product)
        .options(selectinload(Product.data), selectinload(Product.images), selectinload(Product.aplus))
        .order_by(Product.updated_at.desc(), Product.created_at.desc())
        .limit(body.limit)
    )
    needs_product_data_join = bool(body.item_id or body.sku_keyword)
    if needs_product_data_join:
        query = query.join(ProductData, ProductData.product_id == Product.id, isouter=True)

    if body.status:
        if body.status == COMPLETED:
            query = query.where(Product.status == COMPLETED, Product.current_step >= 6)
        else:
            query = query.where(Product.status == body.status)
    if body.item_id:
        pattern = f"%{body.item_id.strip()}%"
        query = query.where(
            (Product.gigab2b_product_id.ilike(pattern)) | (ProductData.item_code.ilike(pattern))
        )
    if body.data_source_id:
        query = query.where(_product_data_source_filter(body.data_source_id))
    if body.competitor_asin:
        query = query.where(Product.competitor_asin.ilike(f"%{body.competitor_asin.strip()}%"))
    if body.upc:
        query = query.where(Product.upc.ilike(f"%{body.upc.strip()}%"))
    created_from = _naive_datetime(body.created_from)
    created_to = _naive_datetime(body.created_to)
    if created_from:
        query = query.where(Product.created_at >= created_from)
    if created_to:
        query = query.where(Product.created_at <= created_to)
    if body.sku_keyword:
        keyword = f"%{body.sku_keyword.strip()}%"
        query = query.where(
            (Product.gigab2b_product_id.ilike(keyword))
            | (ProductData.item_code.ilike(keyword))
            | (ProductData.title.ilike(keyword))
        )

    result = await db.execute(query)
    products = result.scalars().unique().all()
    if body.work_status:
        products = [product for product in products if _product_workbench_status(product) == body.work_status]
    ids = [int(product.id) for product in products]
    if not ids:
        raise HTTPException(400, "当前筛选下没有商品")
    return await create_product_bulk_advance_run(
        db,
        ids,
        payload_extra={
            "filter": body.model_dump(mode="json"),
            "filter_matched_count": len(ids),
        },
        title_suffix="（按筛选）",
    )


@router.get("/catalog", response_model=PaginatedCatalogProducts)
async def list_catalog_products(
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=1000),
    item_id: str | None = None,
    competitor_asin: str | None = None,
    amazon_asin: str | None = None,
    asin_sync_status: str | None = None,
    amazon_product_status: str | None = None,
    aplus_upload_status: str | None = None,
    stock_sync_status: str | None = None,
    template_risk_level: str | None = None,
    upc: str | None = None,
    category: str | None = None,
    export_status: str | None = Query(default=None, pattern="^(pending|exported)$"),
    imported_from: datetime | None = None,
    imported_to: datetime | None = None,
    stock_synced_from: datetime | None = None,
    stock_synced_to: datetime | None = None,
    db: AsyncSession = Depends(get_db),
):
    """独立商品资料列表，最多单页返回 1000 条。"""
    needs_product_join = False
    needs_product_data_join = bool(template_risk_level)
    if needs_product_data_join:
        needs_product_join = True

    query = select(CatalogProduct).order_by(
        CatalogProduct.updated_at.desc(),
        CatalogProduct.imported_at.desc(),
    )
    if needs_product_join or needs_product_data_join:
        query = query.options(
            selectinload(CatalogProduct.source_product).load_only(
            Product.id,
            Product.gigab2b_url,
            Product.gigab2b_product_id,
            Product.competitor_asin,
            Product.amazon_asin,
            Product.asin_sync_status,
            Product.asin_synced_at,
            Product.asin_sync_error,
            Product.amazon_product_status,
            Product.amazon_product_status_synced_at,
            Product.amazon_product_status_error,
            Product.aplus_upload_status,
            Product.aplus_uploaded_at,
            Product.aplus_upload_error,
            Product.upc,
            Product.brand,
            Product.status,
            Product.updated_at,
            ).selectinload(Product.data).load_only(
                ProductData.id,
                ProductData.product_id,
                ProductData.item_code,
                ProductData.title,
                ProductData.leaf_category,
                ProductData.amazon_template_fill_summary,
            ),
            selectinload(CatalogProduct.source_product).selectinload(Product.aplus).load_only(
                ProductAplus.id,
                ProductAplus.product_id,
                ProductAplus.aplus_status,
                ProductAplus.aplus_image_count,
            ),
        )
    count_query = select(func.count(CatalogProduct.id))

    if needs_product_join:
        query = query.join(Product, CatalogProduct.source_product_id == Product.id)
        count_query = count_query.join(Product, CatalogProduct.source_product_id == Product.id)
    if needs_product_data_join:
        query = query.outerjoin(ProductData, ProductData.product_id == Product.id)
        count_query = count_query.outerjoin(ProductData, ProductData.product_id == Product.id)

    query = query.where(CatalogProduct.confirmed_at.is_not(None))
    count_query = count_query.where(CatalogProduct.confirmed_at.is_not(None))

    if item_id:
        pattern = f"%{item_id.strip()}%"
        query = query.where((CatalogProduct.gigab2b_product_id.ilike(pattern)) | (CatalogProduct.item_code.ilike(pattern)))
        count_query = count_query.where((CatalogProduct.gigab2b_product_id.ilike(pattern)) | (CatalogProduct.item_code.ilike(pattern)))
    if competitor_asin:
        pattern = f"%{competitor_asin.strip()}%"
        query = query.where(CatalogProduct.competitor_asin.ilike(pattern))
        count_query = count_query.where(CatalogProduct.competitor_asin.ilike(pattern))
    if amazon_asin:
        pattern = f"%{amazon_asin.strip()}%"
        query = query.where(CatalogProduct.amazon_asin.ilike(pattern))
        count_query = count_query.where(CatalogProduct.amazon_asin.ilike(pattern))
    if asin_sync_status:
        if asin_sync_status == "synced":
            query = query.where((CatalogProduct.amazon_asin.is_not(None)) & (CatalogProduct.amazon_asin != ""))
            count_query = count_query.where((CatalogProduct.amazon_asin.is_not(None)) & (CatalogProduct.amazon_asin != ""))
        elif asin_sync_status == "not_synced":
            query = query.where((CatalogProduct.amazon_asin.is_(None)) | (CatalogProduct.amazon_asin == ""))
            count_query = count_query.where((CatalogProduct.amazon_asin.is_(None)) | (CatalogProduct.amazon_asin == ""))
        elif asin_sync_status == "manual_linked":
            query = query.where(CatalogProduct.asin_sync_status == "manual_linked")
            count_query = count_query.where(CatalogProduct.asin_sync_status == "manual_linked")
        else:
            query = query.where(CatalogProduct.asin_sync_status == asin_sync_status)
            count_query = count_query.where(CatalogProduct.asin_sync_status == asin_sync_status)
    if amazon_product_status:
        if amazon_product_status == "sellable":
            query = query.where((CatalogProduct.amazon_product_status.ilike("%售卖%")) | (CatalogProduct.amazon_product_status.ilike("%在售%")))
            count_query = count_query.where((CatalogProduct.amazon_product_status.ilike("%售卖%")) | (CatalogProduct.amazon_product_status.ilike("%在售%")))
        elif amazon_product_status == "not_synced":
            query = query.where((CatalogProduct.amazon_product_status.is_(None)) | (CatalogProduct.amazon_product_status == ""))
            count_query = count_query.where((CatalogProduct.amazon_product_status.is_(None)) | (CatalogProduct.amazon_product_status == ""))
        else:
            pattern = f"%{amazon_product_status.strip()}%"
            query = query.where(CatalogProduct.amazon_product_status.ilike(pattern))
            count_query = count_query.where(CatalogProduct.amazon_product_status.ilike(pattern))
    if aplus_upload_status:
        query = query.where(CatalogProduct.aplus_upload_status == aplus_upload_status)
        count_query = count_query.where(CatalogProduct.aplus_upload_status == aplus_upload_status)
    if stock_sync_status:
        if stock_sync_status == "not_synced":
            query = query.where((CatalogProduct.stock_sync_status.is_(None)) | (CatalogProduct.stock_sync_status == "not_synced"))
            count_query = count_query.where((CatalogProduct.stock_sync_status.is_(None)) | (CatalogProduct.stock_sync_status == "not_synced"))
        else:
            query = query.where(CatalogProduct.stock_sync_status == stock_sync_status)
            count_query = count_query.where(CatalogProduct.stock_sync_status == stock_sync_status)
    if template_risk_level:
        pattern = f'%"risk_level"%{template_risk_level}%'
        query = query.where(ProductData.amazon_template_fill_summary.like(pattern))
        count_query = count_query.where(ProductData.amazon_template_fill_summary.like(pattern))
    if upc:
        pattern = f"%{upc.strip()}%"
        query = query.where(CatalogProduct.upc.ilike(pattern))
        count_query = count_query.where(CatalogProduct.upc.ilike(pattern))
    if category:
        pattern = f"%{category.strip()}%"
        query = query.where(CatalogProduct.leaf_category.ilike(pattern))
        count_query = count_query.where(CatalogProduct.leaf_category.ilike(pattern))
    if export_status == "pending":
        no_export_file = CatalogProduct.exported_at.is_(None)
        no_asin = (
            (CatalogProduct.amazon_asin.is_(None)) | (CatalogProduct.amazon_asin == "")
        )
        query = query.where(no_export_file & no_asin)
        count_query = count_query.where(no_export_file & no_asin)
    elif export_status == "exported":
        has_export_file = CatalogProduct.exported_at.is_not(None)
        has_asin = (
            (CatalogProduct.amazon_asin.is_not(None)) & (CatalogProduct.amazon_asin != "")
        )
        query = query.where(has_export_file | has_asin)
        count_query = count_query.where(has_export_file | has_asin)
    imported_from = _naive_datetime(imported_from)
    imported_to = _naive_datetime(imported_to)
    if imported_from:
        query = query.where(CatalogProduct.imported_at >= imported_from)
        count_query = count_query.where(CatalogProduct.imported_at >= imported_from)
    if imported_to:
        query = query.where(CatalogProduct.imported_at <= imported_to)
        count_query = count_query.where(CatalogProduct.imported_at <= imported_to)
    stock_synced_from = _naive_datetime(stock_synced_from)
    stock_synced_to = _naive_datetime(stock_synced_to)
    if stock_synced_from:
        query = query.where(CatalogProduct.stock_synced_at >= stock_synced_from)
        count_query = count_query.where(CatalogProduct.stock_synced_at >= stock_synced_from)
    if stock_synced_to:
        query = query.where(CatalogProduct.stock_synced_at <= stock_synced_to)
        count_query = count_query.where(CatalogProduct.stock_synced_at <= stock_synced_to)

    total_result = await db.execute(count_query)
    total = total_result.scalar() or 0
    result = await db.execute(query.offset((page - 1) * page_size).limit(page_size))
    items = result.scalars().all()
    for item in items:
        product = item.__dict__.get("source_product")
        if product:
            pd = product.data
            item.gigab2b_url = product.gigab2b_url
            item.gigab2b_product_id = product.gigab2b_product_id
            item.competitor_asin = product.competitor_asin
            item.amazon_asin = product.amazon_asin
            item.asin_sync_status = product.asin_sync_status
            item.asin_synced_at = product.asin_synced_at
            item.asin_sync_error = product.asin_sync_error
            item.amazon_product_status = product.amazon_product_status
            item.amazon_product_status_synced_at = product.amazon_product_status_synced_at
            item.amazon_product_status_error = product.amazon_product_status_error
            item.aplus_upload_status = product.aplus_upload_status
            item.aplus_uploaded_at = product.aplus_uploaded_at
            item.aplus_upload_error = product.aplus_upload_error
            item.aplus_status = product.aplus.aplus_status if product.aplus else None
            item.aplus_image_count = product.aplus.aplus_image_count if product.aplus else None
            item.upc = product.upc
            item.brand = product.brand
            item.item_code = pd.item_code if pd else None
            item.title = pd.title if pd else None
            item.leaf_category = pd.leaf_category if pd else None
            item.status = product.status
            item.updated_at = product.updated_at or item.updated_at
            risk, warning_count = _template_risk_from_data(pd)
            item.template_risk_level = risk
            item.template_warnings_count = warning_count
    await db.commit()
    return PaginatedCatalogProducts(items=items, total=total, page=page, page_size=page_size)


@router.get("/catalog/export-files", response_model=PaginatedCatalogExportFiles)
async def list_catalog_export_files(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    category: str | None = None,
    db: AsyncSession = Depends(get_db),
):
    """按导出文件/任务维度列出 Amazon 导入表历史记录。"""
    old_result = await db.execute(
        select(OfflineTask)
        .where(OfflineTask.task_type == "catalog_export")
        .where(OfflineTask.status.in_(("done", "partial_failed", "failed")))
        .options(selectinload(OfflineTask.steps))
        .order_by(OfflineTask.id.desc())
    )
    new_result = await db.execute(
        select(TaskRun)
        .where(TaskRun.task_type == "catalog_export")
        .where(TaskRun.status.in_(("succeeded", "failed", "interrupted")))
        .options(selectinload(TaskRun.steps))
        .order_by(TaskRun.id.desc())
    )
    rows = [
        *[_catalog_export_file_row_from_run(run) for run in new_result.scalars().unique().all()],
        *[_catalog_export_file_row(task) for task in old_result.scalars().unique().all()],
    ]
    normalized_category = str(category or "").strip()
    if normalized_category:
        rows = [
            row
            for row in rows
            if normalized_category in row.categories or row.category == normalized_category
        ]
    rows.sort(key=lambda item: item.exported_at or item.updated_at or item.created_at or datetime.min, reverse=True)
    total = len(rows)
    start = (page - 1) * page_size
    return PaginatedCatalogExportFiles(
        items=rows[start:start + page_size],
        total=total,
        page=page,
        page_size=page_size,
    )


@router.get("/catalog/export-categories", response_model=CatalogExportCategoriesResponse)
async def list_catalog_export_categories(db: AsyncSession = Depends(get_db)):
    """按已导出文件/任务维度聚合可选类目。"""
    exported_groups: dict[str, dict] = {}
    task_result = await db.execute(
        select(OfflineTask)
        .where(OfflineTask.task_type == "catalog_export")
        .where(OfflineTask.status.in_(("done", "partial_failed", "failed")))
        .options(selectinload(OfflineTask.steps))
        .order_by(OfflineTask.id.desc())
    )
    for task in task_result.scalars().unique().all():
        _collect_export_file_category(exported_groups, _catalog_export_file_row(task))
    run_result = await db.execute(
        select(TaskRun)
        .where(TaskRun.task_type == "catalog_export")
        .where(TaskRun.status.in_(("succeeded", "failed", "interrupted")))
        .options(selectinload(TaskRun.steps))
        .order_by(TaskRun.id.desc())
    )
    for run in run_result.scalars().unique().all():
        _collect_export_file_category(exported_groups, _catalog_export_file_row_from_run(run))
    return CatalogExportCategoriesResponse(
        pending=[],
        exported=_export_category_summaries(exported_groups),
    )


@router.get("/catalog/template-categories", response_model=list[CatalogExportCategorySummary])
async def list_catalog_template_categories(db: AsyncSession = Depends(get_db)):
    """列出库里所有商品类目，用于类目模板管理。"""
    groups: dict[str, dict] = {}
    for row in await _catalog_template_category_samples(db):
        category = row["category"]
        count = int(row["count"] or 0)
        available, template_name, template_path, template_error = _catalog_template_status_for_category(
            category,
            row.get("brands") or [],
        )
        groups[category] = {
            "category": category,
            "count": count,
            "exportable_count": count if available else 0,
            "blocked_count": 0 if available else count,
            "template_available": available,
            "template_name": template_name,
            "template_path": template_path,
            "template_error": template_error,
            **_category_upload_summary(category),
            "sample_item_codes": row.get("sample_item_codes") or [],
        }
    return _export_category_summaries(groups)


@router.get("/catalog/template-files", response_model=list[CatalogTemplateFileSummary])
async def list_catalog_template_files(db: AsyncSession = Depends(get_db)):
    """按模板文件维度列出当前可管理的 Amazon 导入模板。"""
    return await _catalog_template_file_summaries(db)


@router.get("/catalog/template-files/{file_id}/download")
async def download_catalog_template_file(file_id: str, db: AsyncSession = Depends(get_db)):
    item = await _find_catalog_template_file(db, file_id)
    if not item.template_path:
        raise HTTPException(400, "模板文件路径不存在")
    path = Path(item.template_path).expanduser().resolve()
    allowed_roots = [
        (_category_template_cache_dir()).resolve(),
        (Path(__file__).resolve().parents[1] / "pipeline" / "templates").resolve(),
    ]
    if not any(path == root or root in path.parents for root in allowed_roots):
        raise HTTPException(400, "模板路径非法")
    if not path.is_file():
        state = _load_template_file_state().get(file_id)
        object_key = str((state or {}).get("object_key") or item.oss_object_key or "").strip()
        if object_key:
            try:
                download_private_file(object_key, path)
            except Exception as exc:
                raise HTTPException(404, f"模板文件本地缓存不存在，且从 OSS 下载失败: {type(exc).__name__}: {exc}")
        else:
            raise HTTPException(404, "模板文件不存在")
    return FileResponse(
        path,
        media_type="application/vnd.ms-excel.sheet.macroEnabled.12",
        filename=path.name,
    )


@router.patch("/catalog/template-files/{file_id}/status", response_model=CatalogTemplateFileSummary)
async def update_catalog_template_file_status(
    file_id: str,
    enabled: bool = Body(..., embed=True),
    db: AsyncSession = Depends(get_db),
):
    item = await _find_catalog_template_file(db, file_id)
    state = _load_template_file_state()
    current = state.get(file_id) if isinstance(state.get(file_id), dict) else {}
    current["enabled"] = bool(enabled)
    current["updated_at"] = datetime.now().isoformat()
    current["file_name"] = item.file_name
    current["template_path"] = item.template_path
    state[file_id] = current
    _save_template_file_state(state)
    return await _find_catalog_template_file(db, file_id)


@router.delete("/catalog/template-files/{file_id}", response_model=list[CatalogTemplateFileSummary])
async def delete_catalog_template_file(file_id: str, db: AsyncSession = Depends(get_db)):
    item = await _find_catalog_template_file(db, file_id)
    if not item.template_path:
        raise HTTPException(400, "模板文件路径不存在")

    target_path = Path(item.template_path).expanduser().resolve()
    cache_root = _category_template_cache_dir().resolve()
    template_root = (Path(__file__).resolve().parents[1] / "pipeline" / "templates").resolve()
    allowed_roots = [cache_root, template_root]
    if not any(target_path == root or root in target_path.parents for root in allowed_roots):
        raise HTTPException(400, "模板路径非法")

    manifest = _load_category_template_manifest()
    changed = False
    for category, upload in list(manifest.items()):
        if not isinstance(upload, dict):
            continue
        cache_path = str(upload.get("cache_path") or "").strip()
        if cache_path and _template_file_id(cache_path) == file_id:
            manifest.pop(category, None)
            changed = True
    if changed:
        _save_category_template_manifest(manifest)

    try:
        if target_path.is_file():
            target_path.unlink()
    except OSError as exc:
        raise HTTPException(400, f"删除模板文件失败: {exc}")

    state = _load_template_file_state()
    state.pop(file_id, None)
    _save_template_file_state(state)
    return await _catalog_template_file_summaries(db)


@router.post("/catalog/category-template-upload", response_model=CatalogTemplateUploadResponse)
async def upload_catalog_category_template(
    category: str = Form(..., min_length=1),
    file: UploadFile = File(...),
):
    """缓存并上传类目模板文件。字段映射仍由 template_mappings 控制，上传结果供导出中心追踪模板准备状态。"""
    normalized_category = category.strip()
    if not normalized_category:
        raise HTTPException(400, "类目不能为空")
    filename = Path(file.filename or "").name
    suffix = Path(filename).suffix.lower()
    if suffix not in {".xls", ".xlsx", ".xlsm"}:
        raise HTTPException(400, "只支持上传 .xls / .xlsx / .xlsm 模板文件")

    content = await file.read()
    max_bytes = 50 * 1024 * 1024
    if not content:
        raise HTTPException(400, "上传文件为空")
    if len(content) > max_bytes:
        raise HTTPException(400, "模板文件不能超过 50MB")

    now = datetime.now()
    safe_category = _safe_export_name(normalized_category, "category")
    safe_filename = _safe_export_name(Path(filename).stem, "template") + suffix
    category_dir = _category_template_cache_dir() / safe_category
    category_dir.mkdir(parents=True, exist_ok=True)
    cache_path = category_dir / f"{now.strftime('%Y%m%d_%H%M%S')}_{safe_filename}"
    cache_path.write_bytes(content)

    prefix = settings.OSS_TEMPLATE_UPLOAD_PREFIX.strip().strip("/")
    object_key = f"{prefix}/{safe_category}/{cache_path.name}" if prefix else f"{safe_category}/{cache_path.name}"
    try:
        uploaded = upload_private_file(cache_path, object_key)
    except Exception as exc:
        raise HTTPException(400, f"模板已缓存到本地，但上传 OSS 失败: {type(exc).__name__}: {exc}")

    manifest = _load_category_template_manifest()
    manifest[normalized_category] = {
        "category": normalized_category,
        "filename": filename,
        "cache_path": str(cache_path),
        "object_key": uploaded.get("object_key"),
        "oss_url": uploaded.get("url"),
        "uploaded_at": now.isoformat(),
    }
    _save_category_template_manifest(manifest)
    return CatalogTemplateUploadResponse(
        category=normalized_category,
        filename=filename,
        cache_path=str(cache_path),
        object_key=uploaded.get("object_key"),
        oss_url=uploaded.get("url"),
        uploaded_at=now,
    )


@router.get("/catalog/category-template-download")
async def download_catalog_category_template(
    category: str = Query(..., min_length=1),
    db: AsyncSession = Depends(get_db),
):
    """下载某个类目当前命中的模板文件，优先下载用户上传并缓存的模板。"""
    normalized_category = category.strip()
    if not normalized_category:
        raise HTTPException(400, "类目不能为空")

    upload_path = _uploaded_category_template_path(normalized_category)
    if upload_path:
        path = upload_path.resolve()
    else:
        result = await db.execute(
            select(CatalogProduct)
            .join(Product, CatalogProduct.source_product_id == Product.id, isouter=True)
            .outerjoin(ProductData, ProductData.product_id == Product.id)
            .where(func.coalesce(ProductData.leaf_category, CatalogProduct.leaf_category, "未分类") == normalized_category)
            .options(selectinload(CatalogProduct.source_product).selectinload(Product.data))
            .limit(1)
        )
        item = result.scalar_one_or_none()
        if not item:
            raise HTTPException(404, f"类目不存在: {normalized_category}")
        available, _template_name, template_path, template_error = _template_status_for_catalog(item.source_product, item)
        if not available or not template_path:
            raise HTTPException(400, template_error or "当前类目没有可下载模板")
        path = Path(template_path).expanduser().resolve()

    allowed_roots = [
        (_category_template_cache_dir()).resolve(),
        (Path(__file__).resolve().parents[1] / "pipeline" / "templates").resolve(),
    ]
    if not any(path == root or root in path.parents for root in allowed_roots):
        raise HTTPException(400, "模板路径非法")
    if not path.is_file():
        upload = _category_template_upload_for(normalized_category)
        object_key = str((upload or {}).get("object_key") or "").strip()
        if object_key:
            try:
                download_private_file(object_key, path)
            except Exception as exc:
                raise HTTPException(404, f"模板文件本地缓存不存在，且从 OSS 下载失败: {type(exc).__name__}: {exc}")
        else:
            raise HTTPException(404, "模板文件不存在")
    try:
        _ensure_template_file_oss_metadata(
            path,
            preferred_object_key=(_category_template_upload_for(normalized_category) or {}).get("object_key"),
        )
    except Exception:
        pass
    return FileResponse(
        path,
        media_type="application/vnd.ms-excel.sheet.macroEnabled.12",
        filename=path.name,
    )


async def build_catalog_export_zip(catalog_items: list[CatalogProduct], db: AsyncSession) -> tuple[bytes, str, list[dict]]:
    if not catalog_items:
        raise CatalogExportBuildError("没有可导出的 Amazon 导入表格数据")
    source_ids = [item.source_product_id for item in catalog_items]
    product_result = await db.execute(
        select(Product)
        .options(selectinload(Product.data), selectinload(Product.images), selectinload(Product.aplus))
        .where(Product.id.in_(source_ids))
    )
    products_by_id = {product.id: product for product in product_result.scalars().all()}
    catalog_by_source_id = {item.source_product_id: item for item in catalog_items}
    latest_inventory_by_catalog_id = await _latest_giga_inventory_by_catalog_id(db, catalog_items)

    grouped: dict[str, dict] = {}
    report_rows: list[dict] = []
    for item in catalog_items:
        product = products_by_id.get(item.source_product_id)
        pd = product.data if product else None
        base_report = {
            "商品资料ID": item.id,
            "商品ID": product.id if product else item.source_product_id,
            "商品Code": pd.item_code if pd else item.item_code,
            "类目": (pd.leaf_category if pd else None) or item.leaf_category or "未分类",
        }
        if not product or not pd:
            report_rows.append({**base_report, "状态": "跳过", "原因": "商品资料不存在"})
            continue
        if item.confirmed_at is None:
            report_rows.append({
                **base_report,
                "状态": "跳过",
                "模板文件": None,
                "导出文件": None,
                "原因": "商品还未加入待导出",
            })
            continue
        existing_asin = (product.amazon_asin or item.amazon_asin or "").strip()
        if existing_asin:
            report_rows.append({
                **base_report,
                "状态": "跳过",
                "模板文件": None,
                "导出文件": None,
                "原因": f"已有真实 ASIN {existing_asin}，不能再次导出 Amazon 导入表格",
            })
            continue
        try:
            mapping = _load_template_mapping(product, pd)
            category = pd.leaf_category or mapping.get("category_type") or item.leaf_category or "未分类"
            template_path = _uploaded_category_template_path(str(category)) or Path(mapping["template_path"]).expanduser()
            if not template_path.is_file():
                raise FileNotFoundError(f"模板文件不存在: {template_path}")
            if not _template_file_enabled(template_path):
                raise ValueError(f"模板文件已停用: {template_path.name}")
            key = str(template_path.expanduser().resolve())
            group = grouped.setdefault(key, {
                "template_path": template_path,
                "categories": [],
                "entries": [],
            })
            if str(category) not in group["categories"]:
                group["categories"].append(str(category))
            group["entries"].append({
                "product": product,
                "category": str(category),
                "mapping": mapping,
            })
        except Exception as exc:
            report_rows.append({
                **base_report,
                "状态": _catalog_export_exception_status(exc),
                "模板文件": None,
                "导出文件": None,
                "原因": f"{type(exc).__name__}: {exc}",
            })

    if not grouped:
        if report_rows and all("已有真实 ASIN" in str(row.get("原因") or "") for row in report_rows):
            raise CatalogExportBuildError("选中的商品都已有真实 ASIN，不能再次导出 Amazon 导入表格", report_rows)
        raise CatalogExportBuildError("没有可导出的 Amazon 导入表格数据", report_rows)

    zip_stream = BytesIO()
    with zipfile.ZipFile(zip_stream, mode="w", compression=zipfile.ZIP_DEFLATED) as archive:
        for group in grouped.values():
            entries = group["entries"]
            template_path: Path = group["template_path"]
            categories = sorted(group.get("categories") or [])
            safe_scope = _safe_export_name("_".join(categories[:3]) if len(categories) <= 3 else f"{categories[0]}_and_{len(categories) - 1}_more")
            template_stem = _safe_export_name(template_path.stem, "amazon_template")
            for chunk_index in range(0, len(entries), 500):
                chunk = entries[chunk_index:chunk_index + 500]
                wb = await asyncio.to_thread(load_workbook, template_path, keep_vba=True, data_only=False)
                if "Template" not in wb.sheetnames:
                    raise HTTPException(400, f"模板缺少 Template 工作表: {template_path}")
                ws = wb["Template"]
                _clear_template_data_rows(ws, len(chunk))
                part = chunk_index // 500 + 1
                export_name = f"{template_stem}_{safe_scope}_{part}.xlsm"
                exported_in_workbook = 0

                for offset, entry in enumerate(chunk):
                    product = entry["product"]
                    category = entry["category"]
                    mapping = entry["mapping"]
                    pd = product.data
                    catalog = catalog_by_source_id.get(product.id)
                    row_number = DATA_ROW + offset
                    report_base = {
                        "商品资料ID": catalog.id if catalog else None,
                        "商品ID": product.id,
                        "商品Code": pd.item_code if pd else None,
                        "类目": category,
                        "模板文件": str(template_path),
                        "导出文件": export_name,
                    }
                    try:
                        if not pd:
                            raise ValueError("商品资料不存在")
                        sku = _catalog_price_quantity_sku(catalog) if catalog else ""
                        latest_inventory = latest_inventory_by_catalog_id.get(catalog.id) if catalog else None
                        if sku and not latest_inventory:
                            raise ValueError(f"最新 GIGA 库存快照未找到 SKU {sku}，已停止导出")
                        stock_override = _catalog_stock_export_override(
                            ws,
                            mapping,
                            catalog,
                            latest_inventory.stock_qty if latest_inventory else None,
                        )
                        await ensure_amazon_template_semantic_fields(product, pd, mapping, template_path)
                        await db.commit()
                        source_path = _amazon_template_cache_path(pd)
                        template_result = None
                        if not source_path:
                            template_result = await run_amazon_template(product.id)
                            source_path = Path(template_result["path"]).expanduser()
                            await db.refresh(product)
                            if catalog:
                                await db.refresh(catalog)
                        if not product.upc:
                            try:
                                await ensure_product_upc(db, product)
                            except UpcPoolEmptyError as exc:
                                raise ValueError(str(exc)) from exc
                        if catalog and catalog.upc != product.upc:
                            catalog.upc = product.upc
                            catalog.updated_at = datetime.now()
                        await db.flush()
                        _copy_import_data_row(source_path, ws, row_number)
                        _apply_catalog_export_row_overrides(ws, row_number, product, pd, mapping)
                        if stock_override:
                            quantity_col, stock_quantity = stock_override
                            ws.cell(row_number, quantity_col).value = stock_quantity
                        await db.commit()
                        exported_in_workbook += 1
                        report_rows.append({
                            **report_base,
                            "状态": "已导出",
                            "原因": ("使用已生成表格" if template_result is None else "现场重新生成表格")
                            + (f"，数量按最新 GIGA 库存 {stock_override[1]} 覆盖" if stock_override else ""),
                        })
                    except Exception as exc:
                        await db.rollback()
                        report_rows.append({
                            **report_base,
                            "状态": _catalog_export_exception_status(exc),
                            "原因": f"{type(exc).__name__}: {exc}",
                        })

                if exported_in_workbook:
                    workbook_bytes = await asyncio.to_thread(_save_workbook_bytes, wb)
                    await asyncio.to_thread(archive.writestr, export_name, workbook_bytes)

        report_bytes = await asyncio.to_thread(_summary_workbook, report_rows)
        await asyncio.to_thread(archive.writestr, "导出报告.xlsx", report_bytes)

    if not any(row.get("状态") == "已导出" for row in report_rows):
        raise CatalogExportBuildError("选中的商品没有成功生成 Amazon 导入表格，请查看商品详情中的导入表格检查。", report_rows)

    filename = f"amazon_import_templates_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
    return zip_stream.getvalue(), filename, report_rows


async def _export_catalog_items(catalog_items: list[CatalogProduct], db: AsyncSession) -> StreamingResponse:
    try:
        zip_bytes, filename, _report_rows = await build_catalog_export_zip(catalog_items, db)
    except CatalogExportBuildError as exc:
        raise HTTPException(400, exc.message)
    zip_stream = BytesIO(zip_bytes)
    return StreamingResponse(
        zip_stream,
        media_type="application/zip",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/catalog/export")
async def export_catalog_products(ids: list[int], db: AsyncSession = Depends(get_db)):
    """按 Amazon 类目模板拆分导出可导入 Amazon 的 xlsm，同一类目多个商品写多行。"""
    if not ids:
        raise HTTPException(400, "请选择要导出的商品")
    if len(ids) > 1000:
        raise HTTPException(400, "单次最多导出 1000 个商品")

    catalog_result = await db.execute(select(CatalogProduct).where(CatalogProduct.id.in_(ids)))
    return await _export_catalog_items(catalog_result.scalars().all(), db)


@router.post("/catalog/export-by-category")
async def export_catalog_products_by_category(body: CatalogExportByCategoryRequest, db: AsyncSession = Depends(get_db)):
    """导出指定类目下所有待导出的商品；已有真实 ASIN 的商品不参与。"""
    category = body.category.strip()
    result = await db.execute(
        select(CatalogProduct)
        .join(Product, CatalogProduct.source_product_id == Product.id)
        .outerjoin(ProductData, ProductData.product_id == Product.id)
        .where(CatalogProduct.confirmed_at.is_not(None))
        .where(func.coalesce(ProductData.leaf_category, CatalogProduct.leaf_category, "未分类") == category)
        .order_by(func.coalesce(Product.updated_at, CatalogProduct.updated_at, CatalogProduct.imported_at).desc())
    )
    catalog_items = result.scalars().all()
    if not catalog_items:
        raise HTTPException(400, f"类目「{category}」下没有待导出的商品")
    return await _export_catalog_items(catalog_items, db)


@router.post("/catalog/inventory-template/export")
async def export_inventory_update_template(ids: list[int], db: AsyncSession = Depends(get_db)):
    """导出 Amazon Price & Quantity 库存同步模板。模板按 SKU 更新，仅导出已有真实 ASIN 的商品。"""
    selected_ids = list(dict.fromkeys(ids or []))
    if not selected_ids:
        raise HTTPException(400, "请选择要导出库存同步模板的商品")
    if len(selected_ids) > 5000:
        raise HTTPException(400, "单次最多导出 5000 个商品")

    result = await db.execute(
        select(CatalogProduct)
        .options(selectinload(CatalogProduct.source_product).selectinload(Product.data))
        .where(CatalogProduct.id.in_(selected_ids))
    )
    catalog_by_id = {item.id: item for item in result.scalars().all()}
    latest_inventory_by_catalog_id = await _latest_giga_inventory_by_catalog_id(db, list(catalog_by_id.values()))

    report_rows: list[dict] = []
    export_items: list[tuple[CatalogProduct, int]] = []
    for catalog_id in selected_ids:
        item = catalog_by_id.get(catalog_id)
        product = item.source_product if item else None
        product_data = product.data if product and product.data else None
        sku = _catalog_price_quantity_sku(item) if item else ""
        latest_inventory = latest_inventory_by_catalog_id.get(item.id) if item else None
        latest_stock = latest_inventory.stock_qty if latest_inventory else None
        real_asin = ((item.amazon_asin if item else None) or (product.amazon_asin if product else None) or "").strip()
        base_report = {
            "商品资料ID": catalog_id,
            "任务ID": item.source_product_id if item else None,
            "商品Code": (item.item_code if item else None) or (product_data.item_code if product_data else None),
            "真实ASIN": real_asin or None,
            "SKU": sku or None,
            "库存": latest_stock,
        }
        if not item:
            report_rows.append({**base_report, "状态": "跳过", "原因": "商品资料不存在"})
            continue
        if not item.confirmed_at:
            report_rows.append({**base_report, "状态": "跳过", "原因": "商品还未加入待导出"})
            continue
        if not real_asin:
            report_rows.append({**base_report, "状态": "跳过", "原因": "缺少真实 ASIN，已自动跳过"})
            continue
        if not sku:
            report_rows.append({**base_report, "状态": "跳过", "原因": "缺少 SKU/商品Code，Amazon 库存模板无法定位商品"})
            continue
        if latest_stock is None:
            report_rows.append({**base_report, "状态": "跳过", "原因": "最新 GIGA 库存快照未找到该 SKU"})
            continue
        if latest_stock < 0:
            report_rows.append({**base_report, "状态": "跳过", "原因": f"最新 GIGA 库存为 {latest_stock}，不能导出负数库存"})
            continue
        export_items.append((item, latest_stock))

    if not export_items:
        zip_stream = BytesIO()
        with zipfile.ZipFile(zip_stream, mode="w", compression=zipfile.ZIP_DEFLATED) as archive:
            archive.writestr("库存模板导出报告.xlsx", _inventory_update_report_workbook(report_rows))
        zip_stream.seek(0)
        filename = f"inventory_update_templates_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
        return StreamingResponse(
            zip_stream,
            media_type="application/zip",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    template_path = Path(settings.PRICE_QUANTITY_TEMPLATE_PATH).expanduser()
    if not template_path.is_file():
        raise HTTPException(400, f"库存同步模板不存在: {template_path}")

    wb = load_workbook(template_path, keep_vba=True, data_only=False)
    if "Template" not in wb.sheetnames:
        raise HTTPException(400, f"库存同步模板缺少 Template 工作表: {template_path}")
    ws = wb["Template"]
    columns = _template_attribute_columns(ws)
    required_columns = {
        PRICE_QUANTITY_SKU_ATTR: "SKU",
        PRICE_QUANTITY_FULFILLMENT_ATTR: "Fulfillment Channel Code",
        PRICE_QUANTITY_QUANTITY_ATTR: "Quantity",
    }
    missing = [label for attr, label in required_columns.items() if attr not in columns]
    if missing:
        raise HTTPException(400, f"库存同步模板缺少字段: {', '.join(missing)}")

    _clear_price_quantity_data_rows(ws, len(export_items))
    inventory_always_col = columns.get(PRICE_QUANTITY_INVENTORY_ALWAYS_AVAILABLE_ATTR)
    handling_time_col = columns.get(PRICE_QUANTITY_HANDLING_TIME_ATTR)
    export_name = f"price_quantity_inventory_update_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsm"
    for offset, (item, latest_stock) in enumerate(export_items):
        row_number = PRICE_QUANTITY_DATA_ROW + offset
        sku = _catalog_price_quantity_sku(item)
        product = item.source_product
        real_asin = (item.amazon_asin or (product.amazon_asin if product else None) or "").strip()
        ws.cell(row_number, columns[PRICE_QUANTITY_SKU_ATTR]).value = sku
        ws.cell(row_number, columns[PRICE_QUANTITY_FULFILLMENT_ATTR]).value = PRICE_QUANTITY_FBM_VALUE
        ws.cell(row_number, columns[PRICE_QUANTITY_QUANTITY_ATTR]).value = int(latest_stock or 0)
        if inventory_always_col:
            ws.cell(row_number, inventory_always_col).value = None
        if handling_time_col:
            ws.cell(row_number, handling_time_col).value = 1
        report_rows.append({
            "状态": "已导出",
            "商品资料ID": item.id,
            "任务ID": item.source_product_id,
            "商品Code": item.item_code,
            "真实ASIN": real_asin,
            "SKU": sku,
            "库存": latest_stock,
            "导出文件": export_name,
            "原因": "按 SKU 写入库存；价格列留空，不更新价格；库存来源：最新 GIGA 库存快照",
        })

    workbook_stream = BytesIO()
    wb.save(workbook_stream)
    zip_stream = BytesIO()
    with zipfile.ZipFile(zip_stream, mode="w", compression=zipfile.ZIP_DEFLATED) as archive:
        archive.writestr(export_name, workbook_stream.getvalue())
        archive.writestr("库存模板导出报告.xlsx", _inventory_update_report_workbook(report_rows))
    zip_stream.seek(0)
    filename = f"inventory_update_templates_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
    return StreamingResponse(
        zip_stream,
        media_type="application/zip",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/catalog/inventory-sync", response_model=InventorySyncBatchResponse)
async def create_inventory_sync_batch(body: InventorySyncCreateRequest, db: AsyncSession = Depends(get_db)):
    """创建大建云仓库存同步批次。默认同步全部已确认商品资料。"""
    selected_ids = list(dict.fromkeys(body.catalog_product_ids or []))
    query = (
        select(CatalogProduct)
        .options(selectinload(CatalogProduct.source_product).selectinload(Product.data))
        .where(CatalogProduct.confirmed_at.is_not(None))
        .order_by(CatalogProduct.imported_at.desc())
    )
    if selected_ids:
        query = query.where(CatalogProduct.id.in_(selected_ids))
    result = await db.execute(query)
    catalog_items = result.scalars().all()
    if not catalog_items:
        raise HTTPException(400, "没有找到可同步库存的商品资料")

    if selected_ids:
        found_ids = {item.id for item in catalog_items}
        missing_ids = [item_id for item_id in selected_ids if item_id not in found_ids]
        if missing_ids:
            raise HTTPException(400, f"部分商品不存在或还未加入待导出: {missing_ids[:10]}")

    try:
        await assert_gigab2b_logged_in_for_inventory(catalog_items)
    except InventorySyncLoginRequired as exc:
        raise HTTPException(400, str(exc)) from exc

    batch = InventorySyncBatch(
        status="pending",
        total_count=len(catalog_items),
        created_at=datetime.now(),
    )
    db.add(batch)
    await db.flush()
    for catalog in catalog_items:
        sync_item = build_inventory_sync_item(catalog)
        sync_item.batch_id = batch.id
        db.add(sync_item)
        catalog.stock_sync_status = "skipped" if sync_item.status == "skipped" else "pending"
        catalog.stock_sync_error = sync_item.error_message
        catalog.updated_at = datetime.now()
    await db.commit()
    await db.refresh(batch)

    start_inventory_sync_batch(batch.id)
    return batch


@router.get("/inventory-sync-batches", response_model=PaginatedInventorySyncBatches)
async def list_inventory_sync_batches(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: AsyncSession = Depends(get_db),
):
    total_result = await db.execute(select(func.count(InventorySyncBatch.id)))
    total = total_result.scalar() or 0
    result = await db.execute(
        select(InventorySyncBatch)
        .order_by(InventorySyncBatch.created_at.desc())
        .offset((page - 1) * page_size)
        .limit(page_size)
    )
    return PaginatedInventorySyncBatches(items=result.scalars().all(), total=total, page=page, page_size=page_size)


@router.get("/inventory-sync-batches/{batch_id}", response_model=InventorySyncBatchDetail)
async def get_inventory_sync_batch(batch_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(InventorySyncBatch)
        .options(selectinload(InventorySyncBatch.items))
        .where(InventorySyncBatch.id == batch_id)
    )
    batch = result.scalar_one_or_none()
    if not batch:
        raise HTTPException(404, "Inventory sync batch not found")
    return batch


@router.post("/catalog/{catalog_id}/asin", response_model=CatalogProductResponse)
async def update_catalog_asin(catalog_id: int, body: CatalogAsinUpdateRequest, db: AsyncSession = Depends(get_db)):
    """手动重新关联真实 ASIN，不创建领星同步任务。"""
    asin = body.amazon_asin.strip().upper()
    if not re.fullmatch(r"B0[A-Z0-9]{8}", asin):
        raise HTTPException(400, "ASIN 格式不正确，应为 B0 开头的 10 位编码")

    result = await db.execute(
        select(CatalogProduct)
        .options(selectinload(CatalogProduct.source_product).selectinload(Product.data))
        .where(CatalogProduct.id == catalog_id)
    )
    catalog = result.scalar_one_or_none()
    if not catalog:
        raise HTTPException(404, "商品资料不存在")

    catalog.amazon_asin = asin
    catalog.asin_sync_status = "manual_linked"
    catalog.asin_sync_error = None
    catalog.asin_synced_at = datetime.now()
    catalog.amazon_product_status = None
    catalog.amazon_product_status_synced_at = None
    catalog.amazon_product_status_error = "手动关联 ASIN 后尚未同步亚马逊商品状态"
    catalog.updated_at = datetime.now()

    product = catalog.source_product
    if product:
        product.amazon_asin = asin
        product.asin_sync_status = "manual_linked"
        product.asin_sync_error = None
        product.asin_synced_at = catalog.asin_synced_at
        product.amazon_product_status = None
        product.amazon_product_status_synced_at = None
        product.amazon_product_status_error = catalog.amazon_product_status_error
        product.updated_at = datetime.now()

    await db.commit()
    await db.refresh(catalog)
    return catalog


@router.delete("/catalog/{catalog_id}/asin", response_model=CatalogProductResponse)
async def clear_catalog_asin(catalog_id: int, db: AsyncSession = Depends(get_db)):
    """清除真实 ASIN，让商品回到可同步状态。"""
    result = await db.execute(
        select(CatalogProduct)
        .options(selectinload(CatalogProduct.source_product).selectinload(Product.data))
        .where(CatalogProduct.id == catalog_id)
    )
    catalog = result.scalar_one_or_none()
    if not catalog:
        raise HTTPException(404, "商品资料不存在")

    now = datetime.now()
    catalog.amazon_asin = None
    catalog.asin_sync_status = "not_synced"
    catalog.asin_sync_error = None
    catalog.asin_synced_at = now
    catalog.amazon_product_status = None
    catalog.amazon_product_status_synced_at = None
    catalog.amazon_product_status_error = None
    catalog.updated_at = now

    product = catalog.source_product
    if product:
        product.amazon_asin = None
        product.asin_sync_status = "not_synced"
        product.asin_sync_error = None
        product.asin_synced_at = now
        product.amazon_product_status = None
        product.amazon_product_status_synced_at = None
        product.amazon_product_status_error = None
        product.updated_at = now

    await db.commit()
    await db.refresh(catalog)
    return catalog


@router.post("/catalog/asin-sync", response_model=AsinSyncBatchResponse)
async def create_asin_sync_batch(body: AsinSyncCreateRequest, db: AsyncSession = Depends(get_db)):
    """为选中的商品资料创建 ASIN 同步批次，并在后台按 UPC/商品编码或 MSKU 查询领星 Listing。"""
    result = await db.execute(
        select(CatalogProduct)
        .options(selectinload(CatalogProduct.source_product).selectinload(Product.data))
        .where(CatalogProduct.id.in_(body.catalog_product_ids))
    )
    catalog_items = result.scalars().all()
    if not catalog_items:
        raise HTTPException(400, "没有找到可同步的商品")
    unconfirmed = [item.id for item in catalog_items if item.confirmed_at is None]
    if unconfirmed:
        raise HTTPException(400, f"以下商品还未加入待导出，不能同步 ASIN: {unconfirmed[:10]}")

    found_ids = {item.id for item in catalog_items}
    missing_ids = [item_id for item_id in body.catalog_product_ids if item_id not in found_ids]
    if missing_ids:
        raise HTTPException(400, f"部分商品不存在: {missing_ids[:10]}")

    batch = AsinSyncBatch(
        store=body.store or "Andy店-US",
        status="pending",
        total_count=len(catalog_items),
        created_at=datetime.now(),
    )
    db.add(batch)
    await db.flush()
    for catalog in catalog_items:
        sync_item = build_sync_item(catalog)
        sync_item.batch_id = batch.id
        db.add(sync_item)
        catalog.asin_sync_status = "pending" if sync_item.lookup_code else "skipped"
        catalog.asin_sync_error = sync_item.error_message
        product = catalog.source_product
        if product:
            product.asin_sync_status = catalog.asin_sync_status
            product.asin_sync_error = catalog.asin_sync_error
    await db.commit()
    await db.refresh(batch)

    start_asin_sync_batch(batch.id)
    return batch


@router.get("/asin-sync-batches", response_model=PaginatedAsinSyncBatches)
async def list_asin_sync_batches(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: AsyncSession = Depends(get_db),
):
    total_result = await db.execute(select(func.count(AsinSyncBatch.id)))
    total = total_result.scalar() or 0
    result = await db.execute(
        select(AsinSyncBatch)
        .order_by(AsinSyncBatch.created_at.desc())
        .offset((page - 1) * page_size)
        .limit(page_size)
    )
    return PaginatedAsinSyncBatches(items=result.scalars().all(), total=total, page=page, page_size=page_size)


@router.get("/asin-sync-batches/{batch_id}", response_model=AsinSyncBatchDetail)
async def get_asin_sync_batch(batch_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(AsinSyncBatch)
        .options(selectinload(AsinSyncBatch.items))
        .where(AsinSyncBatch.id == batch_id)
    )
    batch = result.scalar_one_or_none()
    if not batch:
        raise HTTPException(404, "ASIN sync batch not found")
    return batch


@router.post("/catalog/aplus-generate", response_model=BulkStartResponse)
async def create_aplus_generate_batch(body: AplusGenerateRequest, db: AsyncSession = Depends(get_db)):
    """批量生成 A+。A+ 独立于商品主流程，只允许对待导出/已导出的商品执行。"""
    try:
        runs, errors, started_product_ids = await create_aplus_generate_runs(
            db,
            body.catalog_product_ids,
            force=body.force,
            created_by="aplus_generate_batch",
        )
    except ValueError as exc:
        raise HTTPException(400, str(exc))

    return BulkStartResponse(
        requested=len(list(dict.fromkeys(body.catalog_product_ids))),
        started=len(started_product_ids),
        skipped=len(list(dict.fromkeys(body.catalog_product_ids))) - len(started_product_ids),
        errors=errors,
        started_ids=started_product_ids,
    )


@router.post("/{product_id}/aplus/generate")
async def generate_product_aplus(
    product_id: int,
    force: bool = Query(False),
    db: AsyncSession = Depends(get_db),
):
    """为单个已加入导出中心的商品后台生成 A+。"""
    result = await db.execute(
        select(Product)
        .options(
            selectinload(Product.data),
            selectinload(Product.images),
            selectinload(Product.aplus),
            selectinload(Product.catalog_item),
        )
        .where(Product.id == product_id)
    )
    product = result.scalar_one_or_none()
    if not product:
        raise HTTPException(404, "Product not found")
    if not product.catalog_item:
        raise HTTPException(400, "只有待导出/已导出的商品可以生成 A+")
    try:
        runs, errors, started_product_ids = await create_aplus_generate_runs(
            db,
            [product.catalog_item.id],
            force=force,
            created_by="product_aplus_generate",
        )
    except ValueError as exc:
        raise HTTPException(400, str(exc))
    if not started_product_ids:
        raise HTTPException(400, "；".join(errors) or "A+ 生成任务未创建")
    task_run_id = runs[0].id if runs else None
    return {"status": "queued", "product_id": product.id, "task_id": task_run_id, "task_run_id": task_run_id}


@router.post("/catalog/aplus-upload", response_model=AplusUploadBatchResponse)
async def create_aplus_upload_batch(body: AplusUploadCreateRequest, db: AsyncSession = Depends(get_db)):
    """为选中的商品资料创建领星 A+ 上传批次。默认自动提交审批。"""
    result = await db.execute(
        select(CatalogProduct)
        .options(selectinload(CatalogProduct.source_product).selectinload(Product.data))
        .where(CatalogProduct.id.in_(body.catalog_product_ids))
    )
    catalog_items = result.scalars().all()
    if not catalog_items:
        raise HTTPException(400, "没有找到可上传 A+ 的商品")
    unconfirmed = [item.id for item in catalog_items if item.confirmed_at is None]
    if unconfirmed:
        raise HTTPException(400, f"以下商品还未加入待导出，不能上传 A+: {unconfirmed[:10]}")

    batch = AplusUploadBatch(
        store=body.store or "Andy店-US",
        submit_for_approval=1 if body.submit_for_approval else 0,
        status="pending",
        total_count=len(catalog_items),
        created_at=datetime.now(),
    )
    db.add(batch)
    await db.flush()
    for catalog in catalog_items:
        upload_item = build_upload_item(catalog)
        upload_item.batch_id = batch.id
        product = catalog.source_product
        if upload_item.status == "skipped":
            catalog.aplus_upload_status = "skipped"
            catalog.aplus_upload_error = upload_item.error_message
            catalog.aplus_uploaded_at = datetime.now()
            if product:
                product.aplus_upload_status = "skipped"
                product.aplus_upload_error = upload_item.error_message
                product.aplus_uploaded_at = datetime.now()
        else:
            catalog.aplus_upload_status = "pending"
            catalog.aplus_upload_error = None
            if product:
                product.aplus_upload_status = "pending"
                product.aplus_upload_error = None
        db.add(upload_item)
    await db.commit()
    await db.refresh(batch)
    start_aplus_upload_batch(batch.id)
    return batch


@router.get("/aplus-upload-batches", response_model=PaginatedAplusUploadBatches)
async def list_aplus_upload_batches(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: AsyncSession = Depends(get_db),
):
    total_result = await db.execute(select(func.count(AplusUploadBatch.id)))
    total = total_result.scalar() or 0
    result = await db.execute(
        select(AplusUploadBatch)
        .order_by(AplusUploadBatch.created_at.desc())
        .offset((page - 1) * page_size)
        .limit(page_size)
    )
    return PaginatedAplusUploadBatches(items=result.scalars().all(), total=total, page=page, page_size=page_size)


@router.get("/aplus-upload-batches/{batch_id}", response_model=AplusUploadBatchDetail)
async def get_aplus_upload_batch(batch_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(AplusUploadBatch)
        .options(selectinload(AplusUploadBatch.items))
        .where(AplusUploadBatch.id == batch_id)
    )
    batch = result.scalar_one_or_none()
    if not batch:
        raise HTTPException(404, "A+ upload batch not found")
    return batch


@router.post("/{product_id}/confirm", response_model=ProductResponse)
async def confirm_product(product_id: int, db: AsyncSession = Depends(get_db)):
    """人工确认商品生成结果，并同步进入待导出列表。"""
    result = await db.execute(
        select(Product)
        .options(selectinload(Product.data), selectinload(Product.aplus), selectinload(Product.catalog_item))
        .where(Product.id == product_id)
    )
    product = result.scalar_one_or_none()
    if not product:
        raise HTTPException(404, "Product not found")
    if is_running(product.id):
        raise HTTPException(400, "任务还在运行中，完成后再确认")
    if product.current_step < 6:
        raise HTTPException(400, "Listing 内容还没有生成完成")
    if not product.data or not product.data.listing_title or not product.data.listing_bullets:
        raise HTTPException(400, "Listing 标题和五点还没有生成完成")

    product.status = COMPLETED
    product.current_step = 6
    product.error_message = None
    product.updated_at = datetime.now()
    _sync_catalog_item(product, db, confirm=True)
    await db.commit()
    await db.refresh(product)
    return product


def _compact_gigab2b_snapshot(value: str | None) -> str | None:
    snapshot = _json_loads(value, None)
    if not isinstance(snapshot, dict):
        return value
    capture = snapshot.get("amazon_listing_capture")
    compact_capture = None
    if isinstance(capture, dict):
        compact_capture = {
            key: capture.get(key)
            for key in (
                "asin",
                "status",
                "capture_status",
                "capture_error",
                "title",
                "brand",
                "categories",
                "leaf_category",
                "category_rank",
                "url",
            )
            if capture.get(key) is not None
        }
    keep_keys = (
        "batch_id",
        "site",
        "data_source_id",
        "data_source_name",
        "representative_sku",
        "selected_stylesnap",
        "stylesnap_search",
        "item",
    )
    compact = {key: snapshot.get(key) for key in keep_keys if snapshot.get(key) is not None}
    if compact_capture:
        compact["amazon_listing_capture"] = compact_capture
    return json.dumps(compact, ensure_ascii=False)


def _compact_product_detail(detail: ProductDetail) -> ProductDetail:
    if detail.data:
        detail.data.gigab2b_raw_snapshot = _compact_gigab2b_snapshot(detail.data.gigab2b_raw_snapshot)
    if detail.images:
        detail.images.image_analysis = None
    if detail.aplus:
        detail.aplus.aplus_plan = None
        detail.aplus.aplus_scripts = None
        detail.aplus.aplus_images = None
    detail.zip_files = []
    detail.generated_files = []
    detail.video_folder = None
    detail.aplus_folder = None
    detail.amazon_export_preview = None
    return detail


@router.get("/{product_id}", response_model=ProductDetail)
async def get_product(
    product_id: int,
    compact: bool = Query(False, description="返回首屏轻量详情，跳过大字段和本地文件扫描"),
    db: AsyncSession = Depends(get_db),
):
    """商品详情（含子表数据）"""
    result = await db.execute(
        select(Product)
        .options(
            selectinload(Product.data),
            selectinload(Product.images),
            selectinload(Product.aplus),
            selectinload(Product.files),
            selectinload(Product.catalog_item),
        )
        .where(Product.id == product_id)
    )
    product = result.scalar_one_or_none()
    if not product:
        raise HTTPException(404, "Product not found")

    detail = ProductDetail.model_validate(product)
    await _ensure_product_detail_gallery_order(db, product, detail)
    catalog_exported = bool(product.catalog_item and (product.catalog_item.exported_at or product.catalog_item.export_task_id))
    detail.workflow = _workflow_state(product, catalog_exported=catalog_exported)
    if compact:
        detail.current_task_status = detail.workflow["action_reason"]
        return _compact_product_detail(detail)

    detail.current_task_status = detail.workflow["action_reason"]
    detail.amazon_export_preview = _build_amazon_export_preview(product)
    if product.data and product.data.material_dir:
        material_dir = Path(product.data.material_dir).expanduser()
        if material_dir.is_dir():
            detail.zip_files = _scan_zip_files(material_dir)
            if detail.data:
                detail.data.image_count = _count_material_images(material_dir)
            detail.video_folder = video_folder_summary(material_dir)
            detail.aplus_folder = aplus_folder_summary(aplus_image_folder(material_dir))
    detail.generated_files = sorted(product.files or [], key=lambda item: item.created_at or datetime.min, reverse=True)
    return detail


@router.post("/{product_id}/files/open")
async def open_product_file(
    product_id: int,
    path: str | None = None,
    directory: bool = False,
    db: AsyncSession = Depends(get_db),
):
    """在 Finder 中打开素材目录或指定文件/文件夹"""
    result = await db.execute(
        select(Product)
        .options(selectinload(Product.data))
        .where(Product.id == product_id)
    )
    product = result.scalar_one_or_none()
    if not product:
        raise HTTPException(404, "Product not found")

    material_dir = _safe_material_dir(product)
    target = Path(path).expanduser().resolve() if path else material_dir
    if not _is_in_dir(target, material_dir):
        raise HTTPException(403, "只能打开当前商品素材目录内的文件")
    if not target.exists():
        raise HTTPException(404, f"文件不存在: {target}")
    if directory and target.is_file():
        target = target.parent

    subprocess.Popen(["open", str(target)])
    return {"status": "ok", "path": str(target)}


@router.post("/{product_id}/files/extract")
async def extract_product_zip(
    product_id: int,
    path: str,
    db: AsyncSession = Depends(get_db),
):
    """解压素材目录内的 zip 文件"""
    result = await db.execute(
        select(Product)
        .options(selectinload(Product.data))
        .where(Product.id == product_id)
    )
    product = result.scalar_one_or_none()
    if not product:
        raise HTTPException(404, "Product not found")

    material_dir = _safe_material_dir(product)
    zip_path = Path(path).expanduser().resolve()
    if not _is_in_dir(zip_path, material_dir):
        raise HTTPException(403, "只能解压当前商品素材目录内的压缩包")
    if not zip_path.is_file() or zip_path.suffix.lower() not in ZIP_EXTENSIONS:
        raise HTTPException(400, "请选择 zip 压缩包")

    output_dir = zip_path.with_suffix("")
    output_dir.mkdir(parents=True, exist_ok=True)

    try:
        with zipfile.ZipFile(zip_path) as archive:
            for member in archive.infolist():
                member_path = (output_dir / member.filename).resolve()
                if not _is_in_dir(member_path, output_dir):
                    raise HTTPException(400, "压缩包包含不安全路径")
            archive.extractall(output_dir)
    except zipfile.BadZipFile:
        raise HTTPException(400, "压缩包无法读取")

    return {
        "status": "ok",
        "extracted_dir": str(output_dir),
        "files": _summarize_extracted_files(output_dir),
    }


@router.patch("/{product_id}", response_model=ProductResponse)
async def update_product(
    product_id: int,
    body: ProductUpdate,
    db: AsyncSession = Depends(get_db),
):
    """更新商品任务"""
    result = await db.execute(
        select(Product)
        .options(selectinload(Product.data), selectinload(Product.images), selectinload(Product.catalog_item))
        .where(Product.id == product_id)
    )
    product = result.scalar_one_or_none()
    if not product:
        raise HTTPException(404, "Product not found")

    update_data = body.model_dump(exclude_unset=True)
    if "upc" in update_data:
        requested_upc = str(update_data.pop("upc") or "").strip() or None
        current_upc = str(product.upc or "").strip() or None
        if requested_upc != current_upc:
            raise HTTPException(400, "UPC由UPC池子绑定后不可手动修改")
    categories_value = update_data.pop("categories", None)
    leaf_category_value = update_data.pop("leaf_category", None)
    main_image_path_value = update_data.pop("main_image_path", None)
    gallery_images_value = update_data.pop("gallery_images", None)
    product_data_updates = {
        key: update_data.pop(key)
        for key in list(update_data.keys())
        if key in {
            "listing_title",
            "listing_bullets",
            "listing_description",
            "listing_search_terms",
            "listing_title_zh",
            "listing_bullets_zh",
            "listing_description_zh",
            "listing_search_terms_zh",
            "listing_primary_keyword",
        }
    }
    for key, value in update_data.items():
        setattr(product, key, value)
    if categories_value is not None or leaf_category_value is not None or product_data_updates:
        if not product.data:
            product.data = ProductData(product_id=product.id)
            db.add(product.data)
        if categories_value is not None:
            if isinstance(categories_value, list):
                categories = [str(item).strip() for item in categories_value if str(item).strip()]
            else:
                categories = [
                    part.strip()
                    for part in re.split(r"\s*>\s*|\s*›\s*", str(categories_value))
                    if part.strip()
                ]
            category_key = _category_option_key(categories, leaf_category_value or (categories[-1] if categories else None))
            if category_key and category_key not in await _allowed_category_keys(db):
                raise HTTPException(400, "Amazon 类目只能从已有类目列表中选择")
            product.data.categories = json.dumps(categories, ensure_ascii=False) if categories else None
            if categories and leaf_category_value is None:
                product.data.leaf_category = categories[-1]
        if leaf_category_value is not None:
            product.data.leaf_category = leaf_category_value.strip() or None
        for key, value in product_data_updates.items():
            if key in {"listing_bullets", "listing_bullets_zh"}:
                if isinstance(value, list):
                    normalized = [" ".join(str(item).split()).strip() for item in value if str(item).strip()]
                else:
                    normalized = [" ".join(line.split()).strip() for line in str(value or "").splitlines() if line.strip()]
                setattr(product.data, key, json.dumps(normalized, ensure_ascii=False))
            else:
                setattr(product.data, key, str(value).strip() if value is not None else None)
    if main_image_path_value is not None or gallery_images_value is not None:
        if not product.images:
            product.images = ProductImage(product_id=product.id)
            db.add(product.images)
        gallery_input = gallery_images_value if gallery_images_value is not None else json.loads(product.images.gallery_images or "[]")
        if not isinstance(gallery_input, list):
            raise HTTPException(400, "副图列表格式不正确")
        main_path, gallery_paths = _normalize_listing_image_paths(
            main_image_path_value if main_image_path_value is not None else product.images.main_image_path,
            gallery_input,
        )
        product.images.main_image_path = main_path
        product.images.main_image_source = "manual_selected"
        product.images.gallery_images = json.dumps(gallery_paths, ensure_ascii=False)
    product.updated_at = datetime.now()
    _sync_catalog_item(product, db)
    await db.commit()
    await db.refresh(product)
    return product


@router.put("/{product_id}/listing-images", response_model=ProductImageResponse)
async def update_product_listing_images(
    product_id: int,
    body: ProductListingImagesUpdate,
    background_tasks: BackgroundTasks,
    db: AsyncSession = Depends(get_db),
):
    """保存商品主图/副图选择。"""
    result = await db.execute(
        select(Product)
        .options(selectinload(Product.images), selectinload(Product.data), selectinload(Product.catalog_item))
        .where(Product.id == product_id)
    )
    product = result.scalar_one_or_none()
    if not product:
        raise HTTPException(404, "Product not found")

    main_path, gallery_paths = _normalize_listing_image_paths(body.main_image_path, body.gallery_images)
    old_main_path = product.images.main_image_path if product.images else None
    main_image_changed = bool(old_main_path and old_main_path != main_path)
    if not product.images:
        product.images = ProductImage(product_id=product.id)
        db.add(product.images)

    if main_image_changed:
        await _delete_product_competitor_records(db, product)
        if product.data:
            product.data.gigab2b_raw_snapshot = _strip_competitor_snapshot(product.data.gigab2b_raw_snapshot)
        product.competitor_asin = None

    product.images.main_image_path = main_path
    product.images.main_image_source = "manual_selected"
    product.images.gallery_images = json.dumps(gallery_paths, ensure_ascii=False)
    now = datetime.now()
    snapshot = _json_loads(product.data.gigab2b_raw_snapshot, {}) if product.data else {}
    if not isinstance(snapshot, dict):
        snapshot = {}
    batch_id = snapshot.get("batch_id")
    site = str(snapshot.get("site") or "US").strip().upper()
    item_code = product.data.item_code if product.data else None
    representative_sku = snapshot.get("representative_sku") or item_code
    existing_candidate_count = 0
    if batch_id and item_code:
        candidate_count_query = select(func.count(AmazonStyleSnapCandidate.id)).where(
            AmazonStyleSnapCandidate.batch_id == batch_id,
            AmazonStyleSnapCandidate.site == site,
            AmazonStyleSnapCandidate.item_code == item_code,
        )
        if representative_sku:
            candidate_count_query = candidate_count_query.where(AmazonStyleSnapCandidate.sku_code == representative_sku)
        existing_candidate_count = int((await db.execute(candidate_count_query)).scalar_one() or 0)
    should_search_candidates = bool(
        (product.status == "created" or _is_competitor_search_failed(product))
        and not product.competitor_asin
        and batch_id
        and item_code
        and existing_candidate_count == 0
    )
    if should_search_candidates:
        product.status = "competitor_searching"
        product.current_step = 2
        product.error_message = "商品图片已确认：正在用主图自动搜索 Amazon 候选竞品"
        if product.data:
            product.data.gigab2b_raw_snapshot = json.dumps(
                {
                    **snapshot,
                    "batch_id": batch_id,
                    "site": site,
                    "representative_sku": representative_sku or item_code,
                    "stylesnap_search": {
                        "status": "running",
                        "started_at": now.isoformat(),
                        "source_image_path": main_path,
                        "append": False,
                        "previous_count": 0,
                        "auto_started": True,
                    },
                },
                ensure_ascii=False,
            )
    elif product.status == "created" and product.current_step <= 0:
        product.current_step = 1
        product.error_message = None
    product.updated_at = now
    if product.catalog_item:
        product.catalog_item.status = product.status
        product.catalog_item.competitor_asin = product.competitor_asin
        product.catalog_item.updated_at = now
    await db.commit()
    await db.refresh(product.images)
    if should_search_candidates:
        from app.api.amazon_stylesnap import _run_product_competitor_search_background

        background_tasks.add_task(_run_product_competitor_search_background, product.id)
    return product.images


@router.delete("/{product_id}", status_code=204)
async def delete_product(product_id: int, db: AsyncSession = Depends(get_db)):
    """删除商品任务"""
    result = await db.execute(select(Product).options(selectinload(Product.catalog_item)).where(Product.id == product_id))
    product = result.scalar_one_or_none()
    if not product:
        raise HTTPException(404, "Product not found")
    cancel_pipeline(product_id)
    catalog_id = product.catalog_item.id if product.catalog_item else None
    await db.execute(delete(AplusRegenerateTask).where(AplusRegenerateTask.product_id == product_id))
    await db.execute(delete(AsinSyncItem).where(AsinSyncItem.product_id == product_id))
    await db.execute(delete(AplusUploadItem).where(AplusUploadItem.product_id == product_id))
    await db.execute(delete(InventorySyncItem).where(InventorySyncItem.product_id == product_id))
    if catalog_id is not None:
        await db.execute(delete(AsinSyncItem).where(AsinSyncItem.catalog_product_id == catalog_id))
        await db.execute(delete(AplusUploadItem).where(AplusUploadItem.catalog_product_id == catalog_id))
        await db.execute(delete(InventorySyncItem).where(InventorySyncItem.catalog_product_id == catalog_id))
    await db.delete(product)
    await db.commit()


@router.post("/{product_id}/start", response_model=ProductResponse)
async def start_pipeline(product_id: int, db: AsyncSession = Depends(get_db)):
    """旧入口已禁用：商品中心不再触发浏览器 Step1 采集。"""
    result = await db.execute(
        select(Product)
        .options(
            selectinload(Product.data),
            selectinload(Product.images),
            selectinload(Product.aplus),
            selectinload(Product.catalog_item),
        )
        .where(Product.id == product_id)
    )
    product = result.scalar_one_or_none()
    if not product:
        raise HTTPException(404, "Product not found")
    _raise_step1_browser_collect_removed()


@router.post("/{product_id}/restart", response_model=ProductResponse)
async def restart_pipeline(
    product_id: int,
    background_tasks: BackgroundTasks,
    db: AsyncSession = Depends(get_db),
):
    """重新开始商品流程：保留已选图片，清空竞品和生成结果；有主图时重新搜索候选竞品。"""
    result = await db.execute(
        select(Product)
        .options(
            selectinload(Product.data),
            selectinload(Product.images),
            selectinload(Product.aplus),
            selectinload(Product.files),
            selectinload(Product.catalog_item),
        )
        .where(Product.id == product_id)
    )
    product = result.scalar_one_or_none()
    if not product:
        raise HTTPException(404, "Product not found")
    if is_running(product.id):
        raise HTTPException(400, "任务正在运行中，请先挂起后再重新开始")
    try:
        await ensure_product_upc(db, product)
    except UpcPoolEmptyError as exc:
        raise HTTPException(400, str(exc))

    if product.data:
        _reset_product_data(product.data)
        product.data.gigab2b_raw_snapshot = _strip_competitor_snapshot(product.data.gigab2b_raw_snapshot)
    else:
        db.add(ProductData(product_id=product.id))

    if product.images:
        _reset_product_images(product.images)
    else:
        db.add(ProductImage(product_id=product.id))

    if product.aplus:
        _reset_product_aplus(product.aplus)
    else:
        db.add(ProductAplus(product_id=product.id))

    for file_record in list(product.files or []):
        await db.delete(file_record)

    await _delete_product_competitor_records(db, product)

    snapshot = _json_loads(product.data.gigab2b_raw_snapshot, {}) if product.data else {}
    if not isinstance(snapshot, dict):
        snapshot = {}
    batch_id = snapshot.get("batch_id")
    site = str(snapshot.get("site") or "US").strip().upper()
    item_code = product.data.item_code if product.data else None
    representative_sku = snapshot.get("representative_sku") or item_code
    has_confirmed_main_image = bool(product.images and product.images.main_image_path)
    can_search_competitors = bool(has_confirmed_main_image and batch_id and item_code)
    now = datetime.now()

    product.competitor_asin = None
    product.aplus_upload_status = "not_uploaded"
    product.aplus_uploaded_at = None
    product.aplus_upload_error = None
    product.updated_at = now
    if can_search_competitors:
        product.status = "competitor_searching"
        product.current_step = 2
        product.error_message = "重新开始流程：正在用已确认主图重新搜索 Amazon 候选竞品"
        if product.data:
            product.data.gigab2b_raw_snapshot = json.dumps(
                {
                    **snapshot,
                    "batch_id": batch_id,
                    "site": site,
                    "representative_sku": representative_sku or item_code,
                    "stylesnap_search": {
                        "status": "running",
                        "started_at": now.isoformat(),
                        "source_image_path": product.images.main_image_path,
                        "append": False,
                        "previous_count": 0,
                        "restart": True,
                    },
                },
                ensure_ascii=False,
            )
    else:
        product.status = "created"
        product.current_step = 1 if has_confirmed_main_image else 0
        product.error_message = None if has_confirmed_main_image else "待确认商品图片"
    if product.catalog_item:
        product.catalog_item.confirmed_at = None
        product.catalog_item.upc = product.upc
        product.catalog_item.competitor_asin = None
        product.catalog_item.aplus_upload_status = product.aplus_upload_status
        product.catalog_item.aplus_uploaded_at = None
        product.catalog_item.aplus_upload_error = None
        product.catalog_item.status = product.status
        product.catalog_item.updated_at = now
    await db.commit()
    await db.refresh(product)

    if can_search_competitors:
        from app.api.amazon_stylesnap import _run_product_competitor_search_background

        background_tasks.add_task(_run_product_competitor_search_background, product.id)

    return product


@router.post("/{product_id}/refresh-giga", response_model=ProductResponse)
async def refresh_product_from_giga_openapi(
    product_id: int,
    body: ProductGigaRefreshRequest | None = None,
    db: AsyncSession = Depends(get_db),
):
    """用 GIGA OpenAPI 重新拉取指定商品源数据，并刷新当前 Product 草稿。"""
    body = body or ProductGigaRefreshRequest()
    result = await db.execute(
        select(Product)
        .options(selectinload(Product.data), selectinload(Product.images), selectinload(Product.aplus), selectinload(Product.catalog_item))
        .where(Product.id == product_id)
    )
    product = result.scalar_one_or_none()
    if not product:
        raise HTTPException(404, "Product not found")
    if is_running(product.id):
        raise HTTPException(400, "商品流程正在运行中，请先挂起后再重新拉取")

    snapshot = _snapshot_for_product(product)
    site = str(snapshot.get("site") or "US").strip().upper()
    item_code = _item_code_for_product(product, body.item_code)
    if not item_code:
        raise HTTPException(400, "无法确定大健 item_code，请传 item_code 后重新拉取")
    data_source_id = await _resolve_giga_refresh_data_source_id(
        db,
        requested_data_source_id=body.data_source_id,
        snapshot=snapshot,
        site=site,
    )
    sku_codes = await _giga_refresh_sku_codes(
        db,
        product=product,
        item_code=item_code,
        data_source_id=data_source_id,
        site=site,
        requested_sku_codes=body.sku_codes,
        snapshot=snapshot,
    )

    if not product.data:
        product.data = ProductData(product_id=product.id, item_code=item_code)
        db.add(product.data)
        await db.flush()
    else:
        product.data.item_code = product.data.item_code or item_code
    product.gigab2b_product_id = product.gigab2b_product_id or item_code
    product.gigab2b_url = product.gigab2b_url or f"https://www.gigab2b.com/product-detail/{item_code}"
    await db.commit()

    safe_item = re.sub(r"[^A-Za-z0-9_.-]+", "_", item_code).strip("._") or str(product.id)
    batch_id = f"product-refresh-p{product.id}-ds{data_source_id}-{safe_item}-{datetime.now().strftime('%Y%m%d-%H%M%S')}"
    try:
        sync_result = await sync_giga_products(
            db,
            GigaSyncOptions(
                task_id=f"product-refresh-{product.id}",
                batch_id=batch_id,
                site=site,
                data_source_id=data_source_id,
                page_size=max(len(sku_codes), 1),
                max_pages=1,
                skip_existing=False,
                download_images=False,
                sku_codes=tuple(sku_codes),
            ),
        )
        draft_result = await upsert_product_drafts_from_giga_batch(
            db,
            batch_id=sync_result.batch_id,
            site=sync_result.site,
            data_source_id=sync_result.data_source_id,
        )
        if product.id not in draft_result.product_ids:
            await db.refresh(product)
            product.updated_at = datetime.now()
            await db.commit()
    except GigaOpenApiError as exc:
        raise HTTPException(400, str(exc)) from exc
    except Exception as exc:
        raise HTTPException(502, f"GIGA 指定商品重新拉取失败: {type(exc).__name__}: {exc}") from exc

    refreshed = await db.execute(
        select(Product)
        .options(selectinload(Product.data), selectinload(Product.images), selectinload(Product.aplus))
        .where(Product.id == product_id)
    )
    refreshed_product = refreshed.scalar_one_or_none()
    if not refreshed_product:
        raise HTTPException(404, "Product not found after refresh")
    refreshed_product.current_task_status = _current_task_status(refreshed_product)
    return refreshed_product


@router.post("/{product_id}/retry", response_model=ProductResponse)
async def retry_step(product_id: int, db: AsyncSession = Depends(get_db)):
    """重试当前失败的步骤"""
    result = await db.execute(
        select(Product)
        .options(selectinload(Product.data), selectinload(Product.images), selectinload(Product.aplus), selectinload(Product.catalog_item))
        .where(Product.id == product_id)
    )
    product = result.scalar_one_or_none()
    if not product:
        raise HTTPException(404, "Product not found")
    if product.status != "failed":
        raise HTTPException(400, "Can only retry failed tasks")

    step = product.current_step
    if (step or 0) <= 1:
        _raise_step1_browser_collect_removed()
    if (step or 0) >= 5:
        await _require_generation_prerequisites(db, product, step)
        if step == 5:
            await _queue_product_image_analysis(db, product, created_by="retry_step")
            refreshed = await db.execute(
                select(Product)
                .options(selectinload(Product.data), selectinload(Product.images), selectinload(Product.aplus), selectinload(Product.catalog_item))
                .where(Product.id == product_id)
            )
            queued_product = refreshed.scalar_one_or_none()
            if not queued_product:
                raise HTTPException(404, "Product not found")
            queued_product.current_task_status = _current_task_status(queued_product)
            return queued_product
        if step == 6 and product.catalog_item:
            product.catalog_item.confirmed_at = None
            product.catalog_item.updated_at = datetime.now()
        if step == 6:
            await _queue_product_listing_generation(db, product, created_by="retry_step")
        else:
            await create_product_bulk_advance_run(
                db,
                [product.id],
                payload_extra={"source": "retry_step", "start_step": step},
                title_suffix=f"（商品 #{product.id}）",
                start_step_override=step,
            )
        refreshed = await db.execute(
            select(Product)
            .options(selectinload(Product.data), selectinload(Product.images), selectinload(Product.aplus), selectinload(Product.catalog_item))
            .where(Product.id == product_id)
        )
        queued_product = refreshed.scalar_one_or_none()
        if not queued_product:
            raise HTTPException(404, "Product not found")
        queued_product.current_task_status = _current_task_status(queued_product)
        return queued_product
    product.status = STEP_STATUS_MAP.get(step, "created")
    product.error_message = None
    product.updated_at = datetime.now()
    await db.commit()
    await db.refresh(product)
    
    # 重新触发Pipeline引擎
    enqueue_pipeline(product.id, start_step=step)
    return product


@router.post("/{product_id}/run-from-step", response_model=ProductResponse)
async def run_product_from_step(
    product_id: int,
    start_step: int = Query(5, ge=5, le=6),
    db: AsyncSession = Depends(get_db),
):
    """从指定商品节点启动后续生成流程。用于商品工作台，不依赖旧 StyleSnap 任务入口。"""
    result = await db.execute(
        select(Product)
        .options(selectinload(Product.data), selectinload(Product.images), selectinload(Product.aplus), selectinload(Product.catalog_item))
        .where(Product.id == product_id)
    )
    product = result.scalar_one_or_none()
    if not product:
        raise HTTPException(404, "Product not found")
    if is_running(product.id):
        raise HTTPException(400, "商品流程正在运行中")
    if product.status in {"source_unavailable", "unavailable"}:
        raise HTTPException(400, f"当前状态不能启动生成: {product.status}")
    if start_step == 5 and product.images and product.images.image_analysis:
        start_step = 6
    if product.status == "completed" and start_step != 6:
        raise HTTPException(400, f"当前状态不能从该节点启动生成: {product.status}")
    if product.status == "paused":
        raise HTTPException(400, "商品已挂起，请先点击继续")
    await _require_generation_prerequisites(db, product, start_step)

    if start_step == 5:
        await _queue_product_image_analysis(db, product, created_by="run_from_step")
        refreshed = await db.execute(
            select(Product)
            .options(selectinload(Product.data), selectinload(Product.images), selectinload(Product.aplus), selectinload(Product.catalog_item))
            .where(Product.id == product_id)
        )
        queued_product = refreshed.scalar_one_or_none()
        if not queued_product:
            raise HTTPException(404, "Product not found")
        queued_product.current_task_status = _current_task_status(queued_product)
        return queued_product

    if start_step == 6:
        await _queue_product_listing_generation(db, product, created_by="run_from_step")
    else:
        await db.commit()
        await create_product_bulk_advance_run(
            db,
            [product.id],
            payload_extra={"source": "run_from_step", "start_step": start_step},
            title_suffix=f"（商品 #{product.id}）",
            start_step_override=start_step,
        )
    refreshed = await db.execute(
        select(Product)
        .options(selectinload(Product.data), selectinload(Product.images), selectinload(Product.aplus))
        .where(Product.id == product_id)
    )
    queued_product = refreshed.scalar_one_or_none()
    if not queued_product:
        raise HTTPException(404, "Product not found")
    queued_product.current_task_status = _current_task_status(queued_product)
    return queued_product


@router.post("/{product_id}/resume", response_model=ProductResponse)
async def resume_pipeline(product_id: int, db: AsyncSession = Depends(get_db)):
    """从挂起时记录的当前步骤继续 Pipeline。"""
    result = await db.execute(
        select(Product)
        .options(selectinload(Product.data), selectinload(Product.images), selectinload(Product.aplus), selectinload(Product.catalog_item))
        .where(Product.id == product_id)
    )
    product = result.scalar_one_or_none()
    if not product:
        raise HTTPException(404, "Product not found")
    if product.status not in {"paused", PENDING_REVIEW}:
        raise HTTPException(400, f"只能继续已挂起或待人工处理的任务，当前状态: {product.status}")
    if is_running(product.id):
        raise HTTPException(400, "任务已经在运行中")

    step = product.current_step or 1
    if step < 1:
        step = 1
    if step <= 1:
        _raise_step1_browser_collect_removed()
    if step > 6:
        raise HTTPException(400, f"当前步骤无效，无法继续: {step}")
    if product.status == PENDING_REVIEW and step >= 6:
        product.status = COMPLETED
        product.current_step = 6
        product.error_message = None
        product.updated_at = datetime.now()
        _sync_catalog_item(product, db, confirm=True)
        await db.commit()
        await db.refresh(product)
        return product
    if step >= 5:
        await _require_generation_prerequisites(db, product, step)
    if step == 5:
        await _queue_product_image_analysis(db, product, created_by="resume_pipeline")
        refreshed = await db.execute(
            select(Product)
            .options(selectinload(Product.data), selectinload(Product.images), selectinload(Product.aplus), selectinload(Product.catalog_item))
            .where(Product.id == product_id)
        )
        queued_product = refreshed.scalar_one_or_none()
        if not queued_product:
            raise HTTPException(404, "Product not found")
        queued_product.current_task_status = _current_task_status(queued_product)
        return queued_product
    if step == 6:
        await _queue_product_listing_generation(db, product, created_by="resume_pipeline")
        refreshed = await db.execute(
            select(Product)
            .options(selectinload(Product.data), selectinload(Product.images), selectinload(Product.aplus), selectinload(Product.catalog_item))
            .where(Product.id == product_id)
        )
        queued_product = refreshed.scalar_one_or_none()
        if not queued_product:
            raise HTTPException(404, "Product not found")
        queued_product.current_task_status = _current_task_status(queued_product)
        return queued_product

    product.status = STEP_STATUS_MAP.get(step, "created")
    product.error_message = None
    product.updated_at = datetime.now()
    await db.commit()
    await db.refresh(product)

    enqueue_pipeline(product.id, start_step=step)
    return product


@router.post("/{product_id}/step/{step}")
async def run_single_step(product_id: int, step: int, db: AsyncSession = Depends(get_db)):
    """单独执行某个步骤（调试/重试用）"""
    if step == 1:
        _raise_step1_browser_collect_removed()
    if step not in STEP_RUNNERS:
        raise HTTPException(400, f"Invalid step: {step}. 主流程只支持 Step2-6；A+ 请在 A+管理中生成。")

    result = await db.execute(
        select(Product)
        .options(selectinload(Product.data), selectinload(Product.images), selectinload(Product.aplus), selectinload(Product.catalog_item))
        .where(Product.id == product_id)
    )
    product = result.scalar_one_or_none()
    if not product:
        raise HTTPException(404, "Product not found")
    if is_running(product.id):
        raise HTTPException(400, "商品流程正在运行中，不能单独执行节点")
    if step >= 5:
        await _require_generation_prerequisites(db, product, step)
    if step == 5:
        task_run_ids = await _queue_product_image_analysis(db, product, created_by="single_step")
        return {"status": "queued", "step": step, "task_run_ids": task_run_ids}
    if step == 6:
        task_run_ids = await _queue_product_listing_generation(db, product, created_by="single_step")
        return {"status": "queued", "step": step, "task_run_ids": task_run_ids}

    product.status = STEP_STATUS_MAP.get(step, "created")
    product.current_step = step
    product.error_message = None
    product.updated_at = datetime.now()
    _sync_catalog_item(product, db)
    await db.commit()

    try:
        runner = STEP_RUNNERS[step]
        data = await runner(product_id)
        await db.refresh(product)
        product.status = COMPLETED if step == 6 else get_step_status(step, "done")
        product.current_step = step
        product.error_message = None
        product.updated_at = datetime.now()
        _sync_catalog_item(product, db, confirm=(step == 6))
        await db.commit()
        return {"status": "ok", "step": step, "data": data}
    except Exception as e:
        await db.refresh(product)
        product.status = "failed"
        product.current_step = step
        product.error_message = f"{type(e).__name__}: {e}"
        product.updated_at = datetime.now()
        _sync_catalog_item(product, db)
        await db.commit()
        raise HTTPException(500, f"Step {step} failed: {e}")


@router.post("/{product_id}/aplus/regenerate")
async def regenerate_aplus_module(product_id: int, body: AplusRegenerateRequest, db: AsyncSession = Depends(get_db)):
    """提交单个 A+ 模块重新生成任务，写入数据库队列后后台执行。"""
    result = await db.execute(
        select(Product)
        .options(selectinload(Product.aplus))
        .where(Product.id == product_id)
    )
    product = result.scalar_one_or_none()
    if not product:
        raise HTTPException(404, "Product not found")
    if not product.aplus or not product.aplus.aplus_plan or not product.aplus.aplus_scripts:
        raise HTTPException(400, "未找到A+规划/脚本，请先执行Step7/Step8")
    try:
        scripts_data = json.loads(product.aplus.aplus_scripts)
    except json.JSONDecodeError:
        raise HTTPException(400, "A+脚本数据损坏")
    scripts = scripts_data.get("scripts") if isinstance(scripts_data, dict) else None
    if not isinstance(scripts, list) or not any(item.get("module_position") == body.module_position for item in scripts if isinstance(item, dict)):
        raise HTTPException(400, f"未找到模块 {body.module_position} 的A+脚本")

    task = await create_regenerate_task(product_id, body.module_position, body.reason.strip())
    return {
        "status": task.status,
        "message": "已提交后台重新生成" if task.status == "queued" else "该模块正在后台重新生成",
        "module_position": body.module_position,
        "task_id": task.id,
    }


@router.post("/{product_id}/aplus/regenerate/retry")
async def retry_aplus_regenerate_tasks(product_id: int, db: AsyncSession = Depends(get_db)):
    """重试该商品最新一批失败/中断的 A+ 重新生图任务。"""
    result = await db.execute(
        select(Product)
        .options(selectinload(Product.aplus))
        .where(Product.id == product_id)
    )
    product = result.scalar_one_or_none()
    if not product:
        raise HTTPException(404, "Product not found")
    if not product.aplus or not product.aplus.aplus_scripts:
        raise HTTPException(400, "未找到A+脚本，不能重试重新生图")

    tasks = await retry_latest_regenerate_tasks(product_id)
    if not tasks:
        raise HTTPException(400, "没有找到可重试的 A+ 重新生图任务，请在对应模块手动重新生成")
    return {
        "status": "queued",
        "message": f"已重新排队 {len(tasks)} 个 A+ 重新生图任务",
        "task_ids": [task.id for task in tasks],
        "module_positions": [task.module_position for task in tasks],
    }


@router.post("/{product_id}/pause", response_model=ProductResponse)
async def pause_pipeline(product_id: int, db: AsyncSession = Depends(get_db)):
    """挂起 Pipeline：停止当前后台任务，并禁止后续自动流程继续跑，直到用户点击继续。"""
    result = await db.execute(select(Product).where(Product.id == product_id))
    product = result.scalar_one_or_none()
    if not product:
        raise HTTPException(404, "Product not found")
    product.status = "paused"
    product.updated_at = datetime.now()
    await db.commit()
    await db.refresh(product)
    
    # 取消后台任务
    cancel_pipeline(product.id)
    return product
