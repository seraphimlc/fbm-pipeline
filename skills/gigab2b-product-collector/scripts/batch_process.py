#!/usr/bin/env python3
"""大健云仓批量处理"""

import openpyxl
import os
import re
import time
import glob
import shutil
import zipfile
import subprocess
import json
import sys

# 导入extract_info和download_materials的功能
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from extract_info import extract_id, get_info, format_attributes, format_features, ensure_js_file
from download_materials import download_materials


def batch_process(excel_path, url_column=6, output_dir=None, start_row=2, end_row=None):
    """批量处理大健云仓商品

    Args:
        excel_path: Excel表格路径
        url_column: URL所在列（默认第6列）
        output_dir: 输出目录（默认同级目录）
        start_row: 开始行（默认第2行）
        end_row: 结束行（默认到最后一行）
    """
    # 确保JS文件存在
    ensure_js_file()

    # 读取表格
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    # 输出目录
    if not output_dir:
        output_dir = os.path.dirname(excel_path)

    # 结束行
    if not end_row:
        end_row = ws.max_row + 1

    total = end_row - start_row
    processed = 0
    failed = []

    for row in range(start_row, end_row):
        try:
            print(f"\n{'='*50}")
            print(f"处理第 {row-start_row+1}/{total} 个商品")

            # 读取URL
            url = ws.cell(row, url_column).value

            if not url:
                print(f"  ⚠️ 跳过：无URL")
                continue

            # 提取商品ID
            product_id = extract_id(url)

            if not product_id:
                print(f"  ❌ 失败：无法提取商品ID")
                failed.append((row, '无法提取商品ID'))
                continue

            # 创建商品文件夹
            product_dir = os.path.join(output_dir, product_id)
            os.makedirs(product_dir, exist_ok=True)

            # 写入商品ID
            ws.cell(row, 1).value = product_id
            print(f"  ✅ 商品ID: {product_id}")

            # 获取商品信息
            info = get_info(product_id)

            if not info:
                print(f"  ❌ 失败：无法获取页面信息")
                failed.append((row, '无法获取页面信息'))
                ws.cell(row, 3).value = '2-获取大建信息失败'
                wb.save(excel_path)
                continue

            # 下载素材包（自动检测所有可用类型）
            dl_results = download_materials(product_id, product_dir)

            # 汇总图片数量
            image_count = sum(r.get('image_count', 0) for r in dl_results.values() if r.get('success'))

            # 写入数据
            # Item Code（第4列）
            ws.cell(row, 4).value = info.get('itemCode')

            # 成本（第9列）
            if info.get('estimatedTotal'):
                ws.cell(row, 9).value = float(info.get('estimatedTotal'))

            # 货值（第10列）
            if info.get('valueTotal'):
                ws.cell(row, 10).value = float(info.get('valueTotal'))

            # 图片数量（第8列）
            ws.cell(row, 8).value = image_count

            # 标题（第13列）
            ws.cell(row, 13).value = info.get('title')

            # 五点描述（第14列）
            features = info.get('features', [])
            if features:
                ws.cell(row, 14).value = format_features(features)

            # 商品属性（第15列）- 包含颜色、材质、尺寸、重量、包装
            ws.cell(row, 15).value = format_attributes(info)

            # 打印关键信息
            print(f"  ✅ Item Code: {info.get('itemCode')}")
            print(f"  ✅ 颜色: {info.get('color')}")
            print(f"  ✅ 材质: {info.get('material')}")
            dims = info.get('dimensions', {})
            if any(dims.values()):
                print(f"  ✅ 尺寸: {dims.get('length')}×{dims.get('width')}×{dims.get('height')} in")
            print(f"  ✅ 重量: {info.get('weight')} lbs")
            print(f"  ✅ 成本: ${info.get('estimatedTotal')}")
            print(f"  ✅ 图片: {image_count}张")

            # 更新工作节点
            ws.cell(row, 3).value = '2-获取大建信息完成'

            # 保存
            wb.save(excel_path)

            processed += 1
            print(f"  🎉 完成: {processed}/{total}")

            # 每5个商品汇报一次
            if processed % 5 == 0:
                print(f"\n📊 进度汇报: 已完成 {processed}/{total} 个商品")

            # 间隔3秒
            time.sleep(3)

        except Exception as e:
            print(f"  ❌ 异常: {str(e)}")
            failed.append((row, str(e)))
            wb.save(excel_path)
            continue

    print(f"\n{'='*50}")
    print(f"🎉 批量处理完成!")
    print(f"  成功: {processed}/{total}")
    print(f"  失败: {len(failed)}")
    if failed:
        print(f"  失败列表: {failed}")

def main():
    if len(sys.argv) < 2:
        print("用法:")
        print("  batch_process.py <excel_path> [--url-column <n>] [--output-dir <dir>] [--start-row <n>] [--end-row <n>]")
        sys.exit(1)

    excel_path = sys.argv[1]
    url_column = 6
    output_dir = None
    start_row = 2
    end_row = None

    # 解析参数
    i = 2
    while i < len(sys.argv):
        if sys.argv[i] == '--url-column' and i + 1 < len(sys.argv):
            url_column = int(sys.argv[i + 1])
            i += 2
        elif sys.argv[i] == '--output-dir' and i + 1 < len(sys.argv):
            output_dir = sys.argv[i + 1]
            i += 2
        elif sys.argv[i] == '--start-row' and i + 1 < len(sys.argv):
            start_row = int(sys.argv[i + 1])
            i += 2
        elif sys.argv[i] == '--end-row' and i + 1 < len(sys.argv):
            end_row = int(sys.argv[i + 1])
            i += 2
        else:
            i += 1

    batch_process(excel_path, url_column, output_dir, start_row, end_row)

if __name__ == '__main__':
    main()