#!/usr/bin/env python3
"""批量下载素材包 v6

增强功能：
1. 检测下架/404页面，跳过并标记
2. 下载失败跳过并标记
3. 更新Excel"处理备注"列
4. 已有素材包的商品可跳过
"""

import openpyxl
import os
import sys
import time
import subprocess
import json
import glob

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from download_materials import download_materials, navigate_to_product
from extract_info import extract_id

EXCEL_PATH = '/Users/jiyuhang/Documents/F/亚马逊工作目录/亚马逊商品/05-03/商品表格.xlsx'
OUTPUT_DIR = '/Users/jiyuhang/Documents/F/亚马逊工作目录/亚马逊商品/05-03'
STATUS_COL = 32  # 处理备注列


def run_js_file(js_code):
    js_path = '/tmp/gigab2b_check.js'
    with open(js_path, 'w') as f:
        f.write(js_code)
    osascript = '''tell application "Google Chrome"
    tell active tab of front window
        set jsFile to do shell script "cat /tmp/gigab2b_check.js"
        set result to execute javascript jsFile
        return result
    end tell
end tell'''
    result = subprocess.run(['osascript', '-e', osascript], capture_output=True, text=True)
    return result.stdout.strip()


def check_page_status():
    """检测当前页面状态

    Returns:
        str: 'ok' | 'off_shelf' | 'error'
    """
    js = '''(function() {
    var text = document.body.innerText || '';
    var title = document.title || '';
    
    // 检测404
    if (title.includes('404') || text.includes('Page Not Found') || text.includes('页面未找到')) {
        return 'off_shelf';
    }
    
    // 检测下架提示
    var offShelfKeywords = ['已下架', '商品已下架', '该商品已下架', 'product has been removed', 
                            'no longer available', '商品不存在', '产品已下架',
                            '该商品已失效', '商品已失效', '已售罄'];
    for (var i = 0; i < offShelfKeywords.length; i++) {
        if (text.includes(offShelfKeywords[i])) {
            return 'off_shelf';
        }
    }
    
    // 检测是否有商品详情区域（正常页面标志）
    var hasProductDetail = document.querySelector('.product-info') || 
                          document.querySelector('.product-detail') ||
                          document.querySelector('[class*="product"]') ||
                          document.querySelector('.item-code');
    
    // 检测是否有下载按钮
    var hasDownloadBtn = false;
    var btns = document.querySelectorAll('button');
    for (var j = 0; j < btns.length; j++) {
        if (btns[j].innerText.includes('下载素材包') || btns[j].innerText.includes('下载')) {
            hasDownloadBtn = true;
            break;
        }
    }
    
    if (hasProductDetail || hasDownloadBtn) {
        return 'ok';
    }
    
    // 页面可能还在加载
    if (text.length < 100) {
        return 'error';
    }
    
    return 'ok';
})()'''
    result = run_js_file(js)
    if result == 'missing value':
        return 'error'
    return result if result in ('ok', 'off_shelf', 'error') else 'ok'


def has_existing_files(product_id):
    """检查商品文件夹是否已有素材"""
    product_dir = os.path.join(OUTPUT_DIR, str(product_id))
    if not os.path.exists(product_dir):
        return False
    # 检查是否有图片
    images = glob.glob(os.path.join(product_dir, '**/*.jpg'), recursive=True)
    images += glob.glob(os.path.join(product_dir, '**/*.png'), recursive=True)
    return len(images) > 0


def update_excel_status(row, status):
    """更新Excel处理备注列"""
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active
    ws.cell(row, STATUS_COL).value = status
    wb.save(EXCEL_PATH)


def batch_download_with_status(start_row=2, end_row=None, skip_existing=True):
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active

    if not end_row:
        end_row = ws.max_row + 1

    total = end_row - start_row
    success_count = 0
    off_shelf_count = 0
    failed_count = 0
    skipped_count = 0
    results = []

    for row in range(start_row, end_row):
        url = ws.cell(row, 6).value  # URL列
        product_id = ws.cell(row, 1).value  # 商品ID列

        if not url and not product_id:
            print(f"  ⚠️ Row {row}: 无URL和商品ID，跳过")
            update_excel_status(row, '无URL')
            failed_count += 1
            continue

        if url:
            extracted_id = extract_id(url)
            if extracted_id:
                product_id = extracted_id

        if not product_id:
            print(f"  ⚠️ Row {row}: 无法提取商品ID，跳过")
            update_excel_status(row, '无法提取ID')
            failed_count += 1
            continue

        idx = row - start_row + 1
        print(f"\n{'='*50}")
        print(f"📦 [{idx}/{total}] 商品 {product_id} (Row {row})")

        # 跳过已有素材的
        if skip_existing and has_existing_files(product_id):
            print(f"  ⏭️ 已有素材，跳过")
            skipped_count += 1
            update_excel_status(row, '已有素材')
            continue

        # 跳转到商品页面
        print(f"  🌐 跳转到商品页面...")
        nav_ok = navigate_to_product(product_id)
        time.sleep(2)

        if not nav_ok:
            print(f"  ❌ 页面加载失败")
            update_excel_status(row, '页面加载失败')
            failed_count += 1
            results.append((product_id, '页面加载失败'))
            continue

        # 检测页面状态
        page_status = check_page_status()
        if page_status == 'off_shelf':
            print(f"  🚫 商品已下架")
            update_excel_status(row, '已下架')
            off_shelf_count += 1
            results.append((product_id, '已下架'))
            continue
        elif page_status == 'error':
            # 重试一次
            print(f"  ⚠️ 页面状态异常，重试...")
            time.sleep(3)
            page_status = check_page_status()
            if page_status == 'off_shelf':
                print(f"  🚫 商品已下架")
                update_excel_status(row, '已下架')
                off_shelf_count += 1
                results.append((product_id, '已下架'))
                continue
            elif page_status == 'error':
                print(f"  ❌ 页面异常")
                update_excel_status(row, '页面异常')
                failed_count += 1
                results.append((product_id, '页面异常'))
                continue

        # 下载素材包
        product_dir = os.path.join(OUTPUT_DIR, str(product_id))
        try:
            dl_results = download_materials(product_id, product_dir, extract=True)
            any_success = any(r['success'] for r in dl_results.values()) if dl_results else False
            if any_success:
                total_images = sum(r.get('image_count', 0) for r in dl_results.values() if r.get('success'))
                print(f"  ✅ 下载成功，共 {total_images} 张图片")
                update_excel_status(row, f'下载成功({total_images}张)')
                success_count += 1
            else:
                print(f"  ❌ 下载失败（无素材包）")
                update_excel_status(row, '下载失败-无素材包')
                failed_count += 1
                results.append((product_id, '下载失败'))
        except Exception as e:
            print(f"  ❌ 异常: {e}")
            update_excel_status(row, f'下载异常: {str(e)[:30]}')
            failed_count += 1
            results.append((product_id, str(e)))

        # 间隔
        time.sleep(2)

    # 汇总
    print(f"\n{'='*60}")
    print(f"🎉 批量下载完成!")
    print(f"  ✅ 成功: {success_count}")
    print(f"  ⏭️ 跳过(已有): {skipped_count}")
    print(f"  🚫 下架: {off_shelf_count}")
    print(f"  ❌ 失败: {failed_count}")
    print(f"  📊 总计: {total}")
    if results:
        print(f"\n  异常商品:")
        for pid, reason in results:
            print(f"    - {pid}: {reason}")


def main():
    start_row = 2
    end_row = None
    no_skip = False

    i = 1
    while i < len(sys.argv):
        if sys.argv[i] == '--start-row' and i + 1 < len(sys.argv):
            start_row = int(sys.argv[i + 1])
            i += 2
        elif sys.argv[i] == '--end-row' and i + 1 < len(sys.argv):
            end_row = int(sys.argv[i + 1])
            i += 2
        elif sys.argv[i] == '--no-skip':
            no_skip = True
            i += 1
        else:
            i += 1

    batch_download_with_status(
        start_row=start_row,
        end_row=end_row,
        skip_existing=not no_skip
    )


if __name__ == '__main__':
    main()
