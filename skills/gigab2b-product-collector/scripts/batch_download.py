#!/usr/bin/env python3
"""批量重新下载素材包（跳过信息提取）

只做：打开页面 → 下载素材包 → 解压到商品文件夹
不修改Excel，不提取商品信息。
"""

import openpyxl
import os
import sys
import time
import subprocess
import json

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from download_materials import download_materials
from extract_info import ensure_js_file, extract_id


def navigate_to_product(product_id):
    """用AppleScript让Chrome跳转到商品页面"""
    url = f"https://www.gigab2b.com/index.php?route=product/product&product_id={product_id}"
    osascript = f'''tell application "Google Chrome"
    tell active tab of front window
        set URL to "{url}"
    end tell
end tell'''
    subprocess.run(['osascript', '-e', osascript], capture_output=True, text=True)
    # 等待页面加载
    time.sleep(3)
    for _ in range(10):
        osascript = '''tell application "Google Chrome"
    tell active tab of front window
        execute javascript "(document.querySelector('button') !== null).toString()"
    end tell
end tell'''
        result = subprocess.run(['osascript', '-e', osascript], capture_output=True, text=True)
        if result.stdout.strip() == 'true':
            return True
        time.sleep(1)
    return False


def batch_download(excel_path, url_column=6, output_dir=None, start_row=2, end_row=None):
    """批量下载素材包

    Args:
        excel_path: Excel表格路径
        url_column: URL所在列
        output_dir: 输出目录
        start_row: 开始行
        end_row: 结束行
    """
    ensure_js_file()

    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    if not output_dir:
        output_dir = os.path.dirname(excel_path)

    if not end_row:
        end_row = ws.max_row + 1

    total = end_row - start_row
    success_count = 0
    failed = []

    for row in range(start_row, end_row):
        url = ws.cell(row, url_column).value
        if not url:
            print(f"  ⚠️ Row {row}: 无URL，跳过")
            continue

        product_id = extract_id(url)
        if not product_id:
            product_id = ws.cell(row, 1).value
            if not product_id:
                print(f"  ⚠️ Row {row}: 无法提取商品ID，跳过")
                continue

        print(f"\n{'='*50}")
        print(f"📦 [{row-start_row+1}/{total}] 商品 {product_id}")

        # 跳转到商品页面
        print(f"  🌐 跳转到商品页面...")
        nav_ok = navigate_to_product(product_id)
        if not nav_ok:
            print(f"  ❌ 页面加载失败")
            failed.append((product_id, '页面加载失败'))
            continue

        # 下载素材包（v4: 自动检测弹出框选项，逐个下载）
        product_dir = os.path.join(output_dir, str(product_id))
        try:
            results = download_materials(product_id, product_dir, extract=True)
            any_success = any(r['success'] for r in results.values()) if results else False
            if any_success:
                success_count += 1
            else:
                failed.append((product_id, '下载失败'))
        except Exception as e:
            print(f"  ❌ 异常: {e}")
            failed.append((product_id, str(e)))

        # 间隔2秒
        time.sleep(2)

    print(f"\n{'='*50}")
    print(f"🎉 批量下载完成!")
    print(f"  成功: {success_count}/{total}")
    print(f"  失败: {len(failed)}")
    if failed:
        print(f"  失败列表:")
        for pid, reason in failed:
            print(f"    - {pid}: {reason}")


def main():
    if len(sys.argv) < 2:
        print("用法:")
        print("  batch_download.py <excel_path> [--start-row <n>] [--end-row <n>]")
        print("")
        print("只下载素材包，不提取商品信息，不修改Excel")
        sys.exit(1)

    excel_path = sys.argv[1]
    start_row = 2
    end_row = None

    i = 2
    while i < len(sys.argv):
        if sys.argv[i] == '--start-row' and i + 1 < len(sys.argv):
            start_row = int(sys.argv[i + 1])
            i += 2
        elif sys.argv[i] == '--end-row' and i + 1 < len(sys.argv):
            end_row = int(sys.argv[i + 1])
            i += 2
        else:
            i += 1

    batch_download(excel_path, start_row=start_row, end_row=end_row)


if __name__ == '__main__':
    main()
