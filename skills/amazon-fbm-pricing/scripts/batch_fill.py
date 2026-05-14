#!/usr/bin/env python3
"""Amazon FBM 售价利润批量填充"""

import openpyxl
import sys
import os

def calc_price_with_min_profit(T, G, min_profit=10):
    """计算建议售价和预期利润

    Args:
        T: 预估总额含运费（大建云仓成本）
        G: 货值总计
        min_profit: 最低利润（默认$10）

    Returns:
        (建议售价, 预期利润)
    """
    if not T or not G or T <= 0 or G <= 0:
        return None, None

    # 综合成本
    C = T + 9 - 0.06 * G

    # 公式一：利润 = 5% × P
    P1 = C / 0.635

    # 公式二：利润 = $10
    P2 = (C + min_profit) / 0.685

    # 取较大值
    P = max(P1, P2)

    # 验证利润
    profit = P * 0.685 - T - 9 + 0.06 * G

    # 确保利润不低于最低值
    if profit < min_profit:
        profit = min_profit

    return round(P, 2), round(profit, 2)

def batch_fill(excel_path, cost_col=9, value_col=10, price_col=11, profit_col=12, start_row=2, end_row=None):
    """批量填充售价和利润

    Args:
        excel_path: Excel表格路径
        cost_col: 成本列（默认第9列）
        value_col: 货值列（默认第10列）
        price_col: 售价列（默认第11列）
        profit_col: 利润列（默认第12列）
        start_row: 开始行（默认第2行）
        end_row: 结束行（默认到最后一行）
    """
    # 读取表格
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    # 结束行
    if not end_row:
        end_row = ws.max_row + 1

    processed = 0
    skipped = 0

    print(f"开始批量计算售价和利润...")
    print(f"范围: 行{start_row} - 行{end_row-1}")

    for row in range(start_row, end_row):
        # 读取成本和货值
        T = ws.cell(row, cost_col).value
        G = ws.cell(row, value_col).value

        if not T or not G:
            skipped += 1
            continue

        try:
            T = float(T)
            G = float(G)

            # 计算售价和利润
            P, profit = calc_price_with_min_profit(T, G)

            if P and profit:
                # 写入表格
                ws.cell(row, price_col).value = P
                ws.cell(row, profit_col).value = profit

                processed += 1
                print(f"  行{row}: 成本=${T:.2f}, 货值=${G:.2f} → 售价=${P:.2f}, 利润=${profit:.2f}")

        except (ValueError, TypeError):
            skipped += 1
            continue

    # 保存
    wb.save(excel_path)

    print(f"\n✅ 批量计算完成!")
    print(f"  处理: {processed} 个商品")
    print(f"  跳过: {skipped} 个商品（缺少成本或货值）")

def main():
    if len(sys.argv) < 2:
        print("用法:")
        print("  batch_fill.py <excel_path> [--cost-col <n>] [--value-col <n>] [--price-col <n>] [--profit-col <n>] [--start-row <n>] [--end-row <n>]")
        print("\n默认列映射:")
        print("  成本列: 9 (大建云仓成本包含邮费)")
        print("  货值列: 10 (货值总计)")
        print("  售价列: 11 (建议亚马逊售价)")
        print("  利润列: 12 (预期利润)")
        sys.exit(1)

    excel_path = sys.argv[1]
    cost_col = 9
    value_col = 10
    price_col = 11
    profit_col = 12
    start_row = 2
    end_row = None

    # 解析参数
    i = 2
    while i < len(sys.argv):
        if sys.argv[i] == '--cost-col' and i + 1 < len(sys.argv):
            cost_col = int(sys.argv[i + 1])
            i += 2
        elif sys.argv[i] == '--value-col' and i + 1 < len(sys.argv):
            value_col = int(sys.argv[i + 1])
            i += 2
        elif sys.argv[i] == '--price-col' and i + 1 < len(sys.argv):
            price_col = int(sys.argv[i + 1])
            i += 2
        elif sys.argv[i] == '--profit-col' and i + 1 < len(sys.argv):
            profit_col = int(sys.argv[i + 1])
            i += 2
        elif sys.argv[i] == '--start-row' and i + 1 < len(sys.argv):
            start_row = int(sys.argv[i + 1])
            i += 2
        elif sys.argv[i] == '--end-row' and i + 1 < len(sys.argv):
            end_row = int(sys.argv[i + 1])
            i += 2
        else:
            i += 1

    batch_fill(excel_path, cost_col, value_col, price_col, profit_col, start_row, end_row)

if __name__ == '__main__':
    main()