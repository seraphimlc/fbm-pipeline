#!/usr/bin/env python3
"""卖家精灵关键词反查 - 直接API导出（0浏览器 0多模态）

通过直接调用卖家精灵后端API导出关键词反查Excel，无需打开浏览器。

API发现过程：
  从卖家精灵JS源码 chunk-01ccfd8e 中逆向出导出API：
  - 同步导出: POST /v3/api/relation/ta/export-keyword-new
  - 异步导出: POST /v3/api/relation/async-export-keyword-new
  注意：前端拦截器会将 export-keyword 改写为 export-keyword-new

Commands:
  export         导出单个ASIN的关键词Excel
  export-batch   批量导出多个ASIN
  extract        从导出的Excel提取关键词列表
  token          从Chrome获取当前Cookie/Token
"""

import argparse
import json
import re
import sys
import time
from pathlib import Path

try:
    import openpyxl
    from openpyxl import load_workbook
except ImportError:
    print("Error: openpyxl not installed. Run: uv pip install openpyxl", file=sys.stderr)
    sys.exit(1)

try:
    import urllib.request
    import urllib.error
except ImportError:
    pass

ASIN_RE = re.compile(r"\bB0[A-Z0-9]{8}\b", re.I)

# 卖家精灵API配置
BASE_URL = "https://www.sellersprite.com"
EXPORT_API = "/v3/api/relation/ta/export-keyword-new"
ASYNC_EXPORT_API = "/v3/api/relation/async-export-keyword-new"
MONTHLY_API = "/v3/api/relation/ta/monthly"

# 默认输出目录
DEFAULT_OUTPUT_DIR = Path("~/Documents/F/亚马逊工作目录/亚马逊商品").expanduser()


def get_token_from_chrome():
    """从Chrome获取卖家精灵的Cookie（需要macOS + AppleScript权限）"""
    import subprocess
    script = '''
tell application "Google Chrome"
    tell active tab of front window
        return execute javascript "document.cookie"
    end tell
end tell
'''
    try:
        result = subprocess.run(
            ["osascript", "-e", script],
            capture_output=True, text=True, timeout=10
        )
        if result.returncode != 0:
            return None, "AppleScript执行失败，请确认Chrome允许JS from Apple Events"
        cookie_str = result.stdout.strip()
        if not cookie_str or "Sprite-X-Token" not in cookie_str:
            return None, "未找到Sprite-X-Token，请确认已在Chrome登录卖家精灵"
        return cookie_str, None
    except Exception as e:
        return None, str(e)


def api_request(url, cookie_str, data=None, method="GET"):
    """发送HTTP请求到卖家精灵API（使用curl，更可靠）"""
    import subprocess
    
    cmd = ["curl", "-s", "-w", "\n%{http_code}", "-o", "-"]
    
    if method == "POST":
        cmd.extend(["-X", "POST"])
    
    cmd.extend([
        "-H", f"Cookie: {cookie_str}",
        "-H", "Content-Type: application/json;charset=UTF-8",
        "-H", "Accept: application/json, text/plain, */*",
        "-H", "Referer: https://www.sellersprite.com/v3/keyword-reverse",
    ])
    
    if data:
        cmd.extend(["-d", json.dumps(data)])
    
    cmd.append(url)
    
    try:
        result = subprocess.run(cmd, capture_output=True, timeout=60)
        # 最后一行是HTTP状态码
        output = result.stdout
        parts = output.rsplit(b"\n", 1)
        if len(parts) == 2:
            status_code = int(parts[1].strip())
            body = parts[0]
        else:
            status_code = 0
            body = output
        
        if status_code == 401:
            print("❌ 认证失败：Token已过期，请重新获取", file=sys.stderr)
            raise Exception("401 Unauthorized")
        elif status_code == 403:
            print("❌ 权限不足：可能需要VIP账号", file=sys.stderr)
            raise Exception("403 Forbidden")
        elif status_code == 429:
            print("❌ 请求过于频繁，请稍后重试", file=sys.stderr)
            raise Exception("429 Too Many Requests")
        elif status_code != 200:
            print(f"❌ HTTP错误 {status_code}", file=sys.stderr)
            raise Exception(f"HTTP {status_code}")
        
        return body
    except subprocess.TimeoutExpired:
        print("❌ 请求超时", file=sys.stderr)
        raise
    except Exception as e:
        if "401" in str(e) or "403" in str(e) or "429" in str(e):
            raise
        print(f"❌ 请求失败: {e}", file=sys.stderr)
        raise


def export_asin(asin, cookie_str, market=1, export_variations=False, export_gk_images=False, output_dir=None):
    """导出单个ASIN的关键词反查Excel
    
    Args:
        asin: Amazon ASIN (如 B0GMWKDNBC)
        cookie_str: 卖家精灵Cookie字符串
        market: 市场ID (1=美国, 2=英国, 3=德国, 4=法国, 5=意大利, 6=西班牙, 7=日本)
        export_variations: 是否导出变体数据
        export_gk_images: 是否导出广告图片
        output_dir: 输出目录
    
    Returns:
        dict: {success, path, asin, rows, keywords_count}
    """
    asin = asin.upper().strip()
    if not ASIN_RE.match(asin):
        return {"success": False, "error": f"无效ASIN: {asin}"}
    
    # 构造API URL
    url = (
        f"{BASE_URL}{EXPORT_API}"
        f"?market={market}"
        f"&exportVariations={'true' if export_variations else 'false'}"
        f"&exportGkImages={'true' if export_gk_images else 'false'}"
    )
    
    print(f"🔍 导出 ASIN: {asin} (market={market})")
    
    try:
        body = api_request(url, cookie_str, data={"asin": asin}, method="POST")
    except Exception as e:
        return {"success": False, "asin": asin, "error": str(e)}
    
    # 生成文件名（curl无法获取content-disposition，用固定格式）
    filename = f"ReverseASIN-US-{asin}-Keywords.xlsx"
    
    # 保存文件
    if output_dir is None:
        output_dir = DEFAULT_OUTPUT_DIR
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    
    output_path = output_dir / filename
    with open(output_path, "wb") as f:
        f.write(body)
    
    # 解析Excel获取统计
    try:
        wb = load_workbook(output_path, read_only=True)
        ws = wb.active
        row_count = ws.max_row - 1  # 减去表头
        wb.close()
    except Exception:
        row_count = -1
    
    print(f"✅ 导出成功: {output_path.name} ({row_count}条关键词)")
    
    return {
        "success": True,
        "asin": asin,
        "path": str(output_path),
        "filename": filename,
        "keywords_count": row_count
    }


def export_batch(asins, cookie_str, market=1, interval=3, output_dir=None, **kwargs):
    """批量导出多个ASIN的关键词
    
    Args:
        asins: ASIN列表
        cookie_str: Cookie字符串
        market: 市场ID
        interval: 每个ASIN之间的间隔秒数
        output_dir: 输出目录
    
    Returns:
        dict: {total, success, failed, results}
    """
    results = []
    for i, asin in enumerate(asins):
        asin = asin.upper().strip()
        if not ASIN_RE.match(asin):
            results.append({"success": False, "asin": asin, "error": "无效ASIN"})
            continue
        
        print(f"\n[{i+1}/{len(asins)}] 处理 {asin}")
        result = export_asin(asin, cookie_str, market=market, output_dir=output_dir, **kwargs)
        results.append(result)
        
        # 间隔控制，避免触发频率限制
        if i < len(asins) - 1 and result.get("success"):
            print(f"⏳ 等待 {interval} 秒...")
            time.sleep(interval)
    
    success = sum(1 for r in results if r.get("success"))
    failed = sum(1 for r in results if not r.get("success"))
    
    summary = {
        "total": len(asins),
        "success": success,
        "failed": failed,
        "results": results
    }
    print(f"\n📊 批量导出完成: 成功 {success}/{len(asins)}")
    return summary


def extract_keywords(export_path, limit=20, column=None):
    """从导出的Excel提取关键词列表"""
    export_path = Path(export_path)
    if not export_path.exists():
        print(f"❌ 文件不存在: {export_path}", file=sys.stderr)
        return []
    
    wb = load_workbook(export_path, read_only=True, data_only=True)
    ws = wb.active
    
    # 找关键词列
    headers = {}
    for idx, cell in enumerate(ws[1], start=1):
        if cell.value:
            headers[str(cell.value).strip()] = idx
    
    if column:
        col = int(column) if str(column).isdigit() else headers.get(column)
        if not col:
            print(f"❌ 列不存在: {column}", file=sys.stderr)
            wb.close()
            return []
    else:
        col = headers.get("关键词", 1)
    
    keywords = []
    for row in range(2, ws.max_row + 1):
        kw = str(ws.cell(row, col).value or "").strip()
        if kw and kw != "关键词":
            keywords.append(kw)
        if len(keywords) >= limit:
            break
    
    wb.close()
    
    result = {
        "source": str(export_path),
        "column": col,
        "count": len(keywords),
        "keywords": keywords
    }
    print(json.dumps(result, ensure_ascii=False, indent=2))
    return keywords


def main():
    parser = argparse.ArgumentParser(
        description="卖家精灵关键词反查 - 直接API导出（0浏览器 0多模态）"
    )
    sub = parser.add_subparsers(dest="cmd", required=True)

    # export
    p = sub.add_parser("export", help="导出单个ASIN的关键词Excel")
    p.add_argument("asin", help="Amazon ASIN (如 B0GMWKDNBC)")
    p.add_argument("--cookie", help="Cookie字符串（不传则自动从Chrome获取）")
    p.add_argument("--market", type=int, default=1, help="市场ID: 1=美国(默认) 2=英国 3=德国 4=法国 5=意大利 6=西班牙 7=日本")
    p.add_argument("--output-dir", help="输出目录（默认: ~/Documents/F/亚马逊工作目录/亚马逊商品）")
    p.add_argument("--variations", action="store_true", help="导出变体数据")
    p.add_argument("--gk-images", action="store_true", help="导出广告图片")

    # export-batch
    p = sub.add_parser("export-batch", help="批量导出多个ASIN")
    p.add_argument("asins", nargs="+", help="ASIN列表")
    p.add_argument("--cookie", help="Cookie字符串")
    p.add_argument("--market", type=int, default=1)
    p.add_argument("--interval", type=int, default=3, help="每个ASIN间隔秒数（默认3）")
    p.add_argument("--output-dir", help="输出目录")

    # extract
    p = sub.add_parser("extract", help="从导出Excel提取关键词列表")
    p.add_argument("export_file", help="导出的Excel文件路径")
    p.add_argument("--limit", type=int, default=20, help="提取关键词数量（默认20）")
    p.add_argument("--column", help="关键词列名或列号")

    # token
    p = sub.add_parser("token", help="从Chrome获取当前Cookie/Token")

    args = parser.parse_args()

    if args.cmd == "token":
        cookie, err = get_token_from_chrome()
        if err:
            print(f"❌ {err}")
            sys.exit(1)
        # 提取关键token
        tokens = {}
        for part in cookie.split("; "):
            if "=" in part:
                k, v = part.split("=", 1)
                if k in ("Sprite-X-Token", "rank-login-user", "ecookie", "current_guest"):
                    tokens[k] = v[:50] + "..." if len(v) > 50 else v
        print(json.dumps({
            "status": "ok",
            "has_token": "Sprite-X-Token" in cookie,
            "tokens": tokens,
            "cookie_length": len(cookie)
        }, ensure_ascii=False, indent=2))

    elif args.cmd == "export":
        cookie = args.cookie
        if not cookie:
            cookie, err = get_token_from_chrome()
            if err:
                print(f"❌ {err}")
                sys.exit(1)
        result = export_asin(
            args.asin, cookie,
            market=args.market,
            export_variations=args.variations,
            export_gk_images=args.gk_images,
            output_dir=args.output_dir
        )
        print(json.dumps(result, ensure_ascii=False, indent=2))

    elif args.cmd == "export-batch":
        cookie = args.cookie
        if not cookie:
            cookie, err = get_token_from_chrome()
            if err:
                print(f"❌ {err}")
                sys.exit(1)
        result = export_batch(
            args.asins, cookie,
            market=args.market,
            interval=args.interval,
            output_dir=args.output_dir
        )
        print(json.dumps(result, ensure_ascii=False, indent=2))

    elif args.cmd == "extract":
        extract_keywords(args.export_file, limit=args.limit, column=args.column)


if __name__ == "__main__":
    main()
