#!/usr/bin/env python3
"""Excel handler - 读写 Amazon 商品表格和卖家精灵导出文件

Commands:
  inspect        检查表格结构（ASIN列、关键词列、行状态）
  pending        列出关键词为空的 ASIN 行
  keywords       从卖家精灵导出文件提取关键词
  write          写回关键词到指定行
  write-downloads 批量扫描 Downloads 写回 Excel
"""

import argparse
import re
import json
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment
except ImportError:
    print("Error: openpyxl not installed. Run: uv pip install openpyxl", file=sys.stderr)
    sys.exit(1)

ASIN_RE = re.compile(r"\bB0[A-Z0-9]{8}\b")
ASIN_HEADERS = {"竞品ASIN", "ASIN", "asin", "Asin"}
KEYWORD_HEADERS = ("竞品关键词前20", "新的关键词", "关键词")
REVERSE_EXPORT_RE = re.compile(r"ReverseASIN-US-(B0[A-Z0-9]{8})", re.I)


def norm(value):
    return str(value).strip() if value is not None else ""


def header_map(ws):
    return {norm(cell.value): idx for idx, cell in enumerate(ws[1], start=1) if norm(cell.value)}


def find_asin_col(ws):
    headers = header_map(ws)
    for name in ASIN_HEADERS:
        if name in headers:
            return headers[name]
    for col in range(1, ws.max_column + 1):
        for row in range(2, min(ws.max_row, 20) + 1):
            if ASIN_RE.search(norm(ws.cell(row, col).value)):
                return col
    raise SystemExit("No ASIN column found")


def find_keyword_col(ws, requested=None):
    headers = header_map(ws)
    if requested:
        if requested.isdigit():
            return int(requested)
        if requested in headers:
            return headers[requested]
        raise SystemExit(f"Keyword column not found: {requested}")
    for name in KEYWORD_HEADERS:
        if name in headers:
            return headers[name]
    raise SystemExit("No keyword target column found")


def inspect(path, sheet=None):
    """检查 Excel 表格结构"""
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet else wb.active
    asin_col = find_asin_col(ws)
    keyword_col = find_keyword_col(ws)
    rows = []
    for row in range(2, ws.max_row + 1):
        text = norm(ws.cell(row, asin_col).value)
        match = ASIN_RE.search(text)
        if match:
            existing = norm(ws.cell(row, keyword_col).value)
            rows.append({
                "row": row,
                "asin": match.group(0),
                "status": "filled" if existing else "empty",
                "keywords_preview": existing[:50] if existing else None
            })
    result = {
        "sheet": ws.title,
        "asin_col": asin_col,
        "keyword_col": keyword_col,
        "total_asins": len(rows),
        "empty_asins": sum(1 for r in rows if r["status"] == "empty"),
        "rows": rows[:30]
    }
    print(json.dumps(result, ensure_ascii=False, indent=2))
    wb.close()


def pending(path, sheet=None, keyword_column=None):
    """列出关键词为空的 ASIN 行"""
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet else wb.active
    asin_col = find_asin_col(ws)
    keyword_col = find_keyword_col(ws, keyword_column)
    rows = []
    for row in range(2, ws.max_row + 1):
        match = ASIN_RE.search(norm(ws.cell(row, asin_col).value))
        if match and not norm(ws.cell(row, keyword_col).value):
            rows.append({"row": row, "asin": match.group(0)})
    result = {
        "sheet": ws.title,
        "asin_col": asin_col,
        "keyword_col": keyword_col,
        "pending_count": len(rows),
        "pending": rows
    }
    print(json.dumps(result, ensure_ascii=False, indent=2))
    wb.close()


def export_keywords(path, limit=20, column=None):
    """从卖家精灵导出的 Excel 提取关键词"""
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    headers = header_map(ws)

    if column:
        col = int(column) if str(column).isdigit() else headers.get(column)
        if not col:
            raise SystemExit(f"Export keyword column not found: {column}")
    else:
        col = headers.get("关键词", 1)

    keywords = []
    for row in range(2, ws.max_row + 1):
        kw = norm(ws.cell(row, col).value)
        if kw and kw != "关键词":
            keywords.append(kw)
        if len(keywords) >= limit:
            break
    wb.close()

    result = {
        "source": str(path),
        "column": col,
        "count": len(keywords),
        "keywords": keywords
    }
    print(json.dumps(result, ensure_ascii=False, indent=2))


def read_export_keywords(path, limit=20, column=None):
    """内部函数：从导出文件读取关键词列表（返回list，不打印）"""
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    headers = header_map(ws)
    if column:
        col = int(column) if str(column).isdigit() else headers.get(column)
        if not col:
            return []
    else:
        col = headers.get("关键词", 1)
    keywords = []
    for row in range(2, ws.max_row + 1):
        kw = norm(ws.cell(row, col).value)
        if kw and kw != "关键词":
            keywords.append(kw)
        if len(keywords) >= limit:
            break
    wb.close()
    return keywords


def write_keywords(path, row, keywords, sheet=None, keyword_column=None):
    """写回关键词到指定行"""
    wb = load_workbook(path)
    ws = wb[sheet] if sheet else wb.active
    col = find_keyword_col(ws, keyword_column)
    cell = ws.cell(row=row, column=col)
    cell.value = "\n".join(keywords)
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    wb.save(path)
    wb.close()
    result = {"success": True, "row": row, "col": col, "count": len(keywords)}
    print(json.dumps(result, ensure_ascii=False, indent=2))


def write_downloads(path, downloads="~/Downloads", sheet=None, keyword_column=None, limit=20, export_column=None):
    """批量扫描 Downloads 目录的导出文件，自动匹配 ASIN 写回 Excel"""
    downloads = Path(downloads).expanduser()
    kw_by_asin = {}
    for export_path in sorted(
        downloads.glob("ReverseASIN-US-*.xlsx"),
        key=lambda p: p.stat().st_mtime,
        reverse=True
    ):
        match = REVERSE_EXPORT_RE.search(export_path.name)
        if not match:
            continue
        asin = match.group(1).upper()
        if asin in kw_by_asin:
            continue  # 保留最新的
        keywords = read_export_keywords(export_path, limit=limit, column=export_column)
        if keywords:
            kw_by_asin[asin] = keywords

    wb = load_workbook(path)
    ws = wb[sheet] if sheet else wb.active
    asin_col = find_asin_col(ws)
    keyword_col = find_keyword_col(ws, keyword_column)
    written = []
    skipped = []
    for row in range(2, ws.max_row + 1):
        match = ASIN_RE.search(norm(ws.cell(row, asin_col).value))
        if not match:
            continue
        asin = match.group(0).upper()
        cell = ws.cell(row, keyword_col)
        if norm(cell.value):
            skipped.append({"row": row, "asin": asin, "reason": "already_filled"})
            continue
        if asin not in kw_by_asin:
            skipped.append({"row": row, "asin": asin, "reason": "no_export_file"})
            continue
        cell.value = "\n".join(kw_by_asin[asin])
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        written.append({"row": row, "asin": asin, "count": len(kw_by_asin[asin])})

    wb.save(path)
    wb.close()
    result = {
        "exports_found": len(kw_by_asin),
        "written_count": len(written),
        "skipped_count": len(skipped),
        "written": written,
        "skipped": skipped[:10]  # 只显示前10个跳过的
    }
    print(json.dumps(result, ensure_ascii=False, indent=2))


def main():
    parser = argparse.ArgumentParser(description="Excel handler for SellerSprite keywords")
    sub = parser.add_subparsers(dest="cmd", required=True)

    p = sub.add_parser("inspect", help="检查表格结构")
    p.add_argument("workbook")
    p.add_argument("--sheet")

    p = sub.add_parser("pending", help="列出待处理的ASIN行")
    p.add_argument("workbook")
    p.add_argument("--sheet")
    p.add_argument("--keyword-column")

    p = sub.add_parser("keywords", help="从导出文件提取关键词")
    p.add_argument("export_workbook")
    p.add_argument("--limit", type=int, default=20)
    p.add_argument("--column")

    p = sub.add_parser("write", help="写回关键词到指定行")
    p.add_argument("workbook")
    p.add_argument("--row", type=int, required=True)
    p.add_argument("--keywords", nargs="+", required=True)
    p.add_argument("--sheet")
    p.add_argument("--keyword-column")

    p = sub.add_parser("write-downloads", help="批量从Downloads写回Excel")
    p.add_argument("workbook")
    p.add_argument("--downloads", default="~/Downloads")
    p.add_argument("--sheet")
    p.add_argument("--keyword-column")
    p.add_argument("--limit", type=int, default=20)
    p.add_argument("--export-column")

    args = parser.parse_args()
    if args.cmd == "inspect":
        inspect(args.workbook, args.sheet)
    elif args.cmd == "pending":
        pending(args.workbook, args.sheet, args.keyword_column)
    elif args.cmd == "keywords":
        export_keywords(args.export_workbook, args.limit, args.column)
    elif args.cmd == "write":
        write_keywords(args.workbook, args.row, args.keywords, args.sheet, args.keyword_column)
    elif args.cmd == "write-downloads":
        write_downloads(args.workbook, args.downloads, args.sheet, args.keyword_column, args.limit, args.export_column)


if __name__ == "__main__":
    main()
