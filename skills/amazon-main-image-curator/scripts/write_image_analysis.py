#!/usr/bin/env python3
"""Export image review decisions to product folders and write lightweight Excel links."""

from __future__ import annotations

import argparse
import datetime as dt
import json
import shutil
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover
    raise SystemExit("Missing dependency: openpyxl. Install with: python3 -m pip install openpyxl") from exc

ANALYSIS_FOLDER_COL = "图片分析文件夹"
A_PLUS_INPUT_COL = "A+图片规划输入文件"
SUMMARY_COL = "图片卖点摘要"
SELECTION_SUMMARY_COL = "主图选择摘要"
WORKFLOW_NODE_COL = "当前工作节点"
REMARK_COL = "处理备注"


def compact_json(value) -> str:
    return json.dumps(value, ensure_ascii=False, separators=(",", ":"))


def load_decisions(path: Path) -> dict:
    data = json.loads(path.read_text(encoding="utf-8"))
    if isinstance(data, list):
        return {"products": data}
    return data


def by_product_key(items: list[dict]) -> dict[str, dict]:
    out = {}
    for item in items:
        for key in [item.get("product_key"), item.get("folder_id"), str(item.get("row_number", ""))]:
            if key:
                out[str(key)] = item
    return out


def ensure_col(ws, header: str) -> int:
    headers = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column + 1)]
    for i, value in enumerate(headers, start=1):
        if value == header:
            return i
    col = ws.max_column + 1
    ws.cell(row=1, column=col, value=header)
    return col


def review_record(review: dict, selected_by_path: dict[str, dict]) -> dict:
    selection = selected_by_path.get(str(Path(review.get("path", "")).resolve())) if review.get("path") else None
    return {
        "image_id": review.get("image_id", ""),
        "filename": review.get("filename", ""),
        "path": review.get("path", ""),
        "image_type": review.get("image_type", ""),
        "selling_point": review.get("visible_selling_point", ""),
        "conversion_role": review.get("conversion_role", ""),
        "risk_flags": review.get("risk_flags", []),
        "slot01_score": review.get("slot01_score"),
        "gallery_score": review.get("gallery_score"),
        "selected": bool(selection),
        "selected_slot": selection.get("slot") if selection else "",
        "reason": review.get("decision_reason", "") or (selection.get("selection_reason") if selection else ""),
        "multimodal_result": review.get("multimodal_result", {}),
        "contact_sheet_evidence": review.get("contact_sheet_evidence", {}),
    }


def compact_review_for_aplus(review: dict) -> dict:
    mm = review.get("multimodal_result", {}) or {}
    return {
        "image_id": review.get("image_id", ""),
        "filename": review.get("filename", ""),
        "image_type": review.get("image_type", ""),
        "visible_selling_point": review.get("visible_selling_point", ""),
        "conversion_role": review.get("conversion_role", ""),
        "a_plus_reference_value": mm.get("A+ reference value") or mm.get("a_plus_reference_value") or "",
        "visual_summary": mm.get("visual_summary", ""),
        "product_angle": mm.get("product_angle", ""),
        "product_state": mm.get("product_state", ""),
        "scale_cues": mm.get("scale_cues", ""),
        "material_texture": mm.get("material_texture", ""),
        "scene_type": mm.get("scene_type", ""),
        "risk_flags": review.get("risk_flags", []),
        "uncertainty": mm.get("uncertainty", ""),
    }


def summary_text(reviews: list[dict], selections: list[dict]) -> str:
    selected = [f"{s.get('slot')}:{s.get('filename')}({s.get('selection_reason', '')})" for s in selections]
    risks = sorted({flag for r in reviews for flag in r.get("risk_flags", []) if flag})
    parts = []
    if selected:
        parts.append("已选 " + "；".join(selected[:9]))
    if risks:
        parts.append("风险：" + "、".join(risks[:8]))
    return "。".join(parts)[:1000]


def md_list(items: list) -> str:
    if not items:
        return "无"
    if isinstance(items, str):
        return items or "无"
    return "、".join(str(x) for x in items if x) or "无"


def write_product_outputs(product: dict, decision: dict, selected_by_path: dict[str, dict]) -> dict:
    product_folder = Path(product.get("matched_folder") or "").expanduser()
    if not product_folder.exists():
        return {"status": "missing_matched_folder"}

    out_dir = product_folder / "image analysis"
    out_dir.mkdir(parents=True, exist_ok=True)

    reviews = decision.get("image_reviews", [])
    selections = sorted(decision.get("gallery_selection", []), key=lambda x: str(x.get("slot", "")))
    review_rows = [review_record(r, selected_by_path) for r in reviews]
    aplus_reviews = [compact_review_for_aplus(r) for r in reviews]

    data = {
        "product_key": decision.get("product_key", ""),
        "folder_id": decision.get("folder_id", ""),
        "row_number": decision.get("row_number", ""),
        "listing_intent": decision.get("listing_intent", {}),
        "image_reviews": review_rows,
        "gallery_selection": selections,
        "missing_recommended_images": decision.get("missing_recommended_images", []),
    }
    (out_dir / "image_selling_points.json").write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")

    lines = [
        "# 图片卖点分析",
        "",
        f"- 商品ID：{decision.get('folder_id') or decision.get('product_key')}",
        f"- Excel行号：{decision.get('row_number', '')}",
        "",
        "## Listing Intent",
    ]
    intent = decision.get("listing_intent", {}) or {}
    for key in ["product_type", "variant", "buyer_problem"]:
        lines.append(f"- {key}: {intent.get(key, '')}")
    lines.extend(["", "## 每张图片判断"])
    for row in review_rows:
        lines.extend(
            [
                "",
                f"### {row.get('image_id')} {row.get('filename')}",
                f"- 图片类型：{row.get('image_type', '')}",
                f"- 可见卖点：{row.get('selling_point', '')}",
                f"- 转化角色：{row.get('conversion_role', '')}",
                f"- 风险标记：{md_list(row.get('risk_flags', []))}",
                f"- 选择状态：{row.get('selected_slot') or '未选'}",
                f"- 判断原因：{row.get('reason', '')}",
                f"- 视觉摘要：{(row.get('multimodal_result') or {}).get('visual_summary', '')}",
            ]
        )
    (out_dir / "image_selling_points.md").write_text("\n".join(lines).strip() + "\n", encoding="utf-8")

    gallery_lines = [
        "# 主图/副图选择",
        "",
        f"- 商品ID：{decision.get('folder_id') or decision.get('product_key')}",
        "",
        "## 已选 01-09",
    ]
    for sel in selections:
        gallery_lines.extend(
            [
                "",
                f"### {str(sel.get('slot', '')).zfill(2)} {sel.get('filename', '')}",
                f"- 原图路径：{sel.get('path', '')}",
                f"- 卖点：{md_list(sel.get('selected_selling_points', []))}",
                f"- 分数：{sel.get('score', '')}",
                f"- 选择原因：{sel.get('selection_reason', '')}",
            ]
        )
    if decision.get("missing_recommended_images"):
        gallery_lines.extend(["", "## 建议补拍/补生成图片"])
        for item in decision.get("missing_recommended_images", []):
            gallery_lines.append(f"- {item}")
    (out_dir / "gallery_selection.md").write_text("\n".join(gallery_lines).strip() + "\n", encoding="utf-8")

    aplus = {
        "product_key": decision.get("product_key", ""),
        "folder_id": decision.get("folder_id", ""),
        "listing_intent": decision.get("listing_intent", {}),
        "reference_image_selling_points": aplus_reviews,
        "selected_gallery_coverage": selections,
        "missing_recommended_images": decision.get("missing_recommended_images", []),
        "notes_for_aplus_planner": [
            "Use reference_image_selling_points as product-form and feature evidence.",
            "Use selected_gallery_coverage as existing main/secondary image coverage; avoid repeating it unless A+ adds stronger scenario proof.",
        ],
    }
    (out_dir / "aplus_planner_input.json").write_text(json.dumps(aplus, ensure_ascii=False, indent=2), encoding="utf-8")

    aplus_lines = [
        "# A+ 图片规划输入",
        "",
        "## 参考图卖点",
    ]
    for item in aplus_reviews:
        aplus_lines.append(
            f"- {item.get('image_id')} {item.get('filename')}: {item.get('visible_selling_point') or item.get('visual_summary')}；"
            f"角度/形态：{item.get('product_angle') or item.get('product_state')}；"
            f"A+参考价值：{item.get('a_plus_reference_value')}"
        )
    aplus_lines.extend(["", "## 主图/副图已表达内容"])
    for sel in selections:
        aplus_lines.append(
            f"- {str(sel.get('slot', '')).zfill(2)} {sel.get('filename', '')}: "
            f"{md_list(sel.get('selected_selling_points', []))}；{sel.get('selection_reason', '')}"
        )
    if decision.get("missing_recommended_images"):
        aplus_lines.extend(["", "## 仍缺少的图片证明"])
        for item in decision.get("missing_recommended_images", []):
            aplus_lines.append(f"- {item}")
    (out_dir / "aplus_planner_input.md").write_text("\n".join(aplus_lines).strip() + "\n", encoding="utf-8")

    return {
        "status": "written",
        "analysis_folder": str(out_dir),
        "aplus_input": str(out_dir / "aplus_planner_input.md"),
        "summary": summary_text(reviews, selections),
        "selection_summary": "；".join(
            f"{s.get('slot')}:{s.get('filename')}({s.get('selection_reason', '')})" for s in selections
        )[:1000],
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", required=True)
    parser.add_argument("--index", required=True)
    parser.add_argument("--decisions", required=True)
    parser.add_argument("--output", default=None)
    parser.add_argument("--in-place", action="store_true")
    args = parser.parse_args()

    workbook_path = Path(args.workbook).expanduser().resolve()
    index = json.loads(Path(args.index).expanduser().resolve().read_text(encoding="utf-8"))
    decisions = load_decisions(Path(args.decisions).expanduser().resolve())
    if len(index.get("products", [])) != 1 or len(decisions.get("products", [])) != 1:
        raise SystemExit("Single-product red line: write_image_analysis requires exactly one indexed product and one decision product.")
    product_index = by_product_key(index.get("products", []))
    decision_index = by_product_key(decisions.get("products", []))

    if args.in_place:
        backup = workbook_path.with_name(f"{workbook_path.stem}_backup_{dt.datetime.now():%Y%m%d_%H%M%S}{workbook_path.suffix}")
        shutil.copy2(workbook_path, backup)
        output_path = workbook_path
    else:
        output_path = (
            Path(args.output).expanduser().resolve()
            if args.output
            else workbook_path.with_name(f"{workbook_path.stem}_image_analysis_{dt.datetime.now():%Y%m%d_%H%M%S}{workbook_path.suffix}")
        )

    wb = load_workbook(workbook_path)
    source_sheet_name = index.get("products", [{}])[0].get("sheet_name") if index.get("products") else None
    ws = wb[source_sheet_name] if source_sheet_name in wb.sheetnames else wb.active
    analysis_folder_col = ensure_col(ws, ANALYSIS_FOLDER_COL)
    aplus_input_col = ensure_col(ws, A_PLUS_INPUT_COL)
    summary_col = ensure_col(ws, SUMMARY_COL)
    selection_summary_col = ensure_col(ws, SELECTION_SUMMARY_COL)
    workflow_node_col = ensure_col(ws, WORKFLOW_NODE_COL)
    remark_col = ensure_col(ws, REMARK_COL)

    for pkey, product in product_index.items():
        decision = decision_index.get(pkey)
        if not decision:
            row_number = int(product.get("row_number"))
            ws.cell(row=row_number, column=remark_col, value="图片分析失败：未找到该商品的 decisions 记录")
            continue
        row_number = int(product.get("row_number") or decision.get("row_number"))
        selections = sorted(decision.get("gallery_selection", []), key=lambda x: str(x.get("slot", "")))
        selected_by_path = {}
        for item in selections:
            path = item.get("path")
            if path:
                selected_by_path[str(Path(path).resolve())] = item
        result = write_product_outputs(product, decision, selected_by_path)
        if result.get("status") == "written":
            ws.cell(row=row_number, column=analysis_folder_col, value=result.get("analysis_folder", ""))
            ws.cell(row=row_number, column=aplus_input_col, value=result.get("aplus_input", ""))
            ws.cell(row=row_number, column=summary_col, value=result.get("summary", ""))
            ws.cell(row=row_number, column=selection_summary_col, value=result.get("selection_summary", ""))
            ws.cell(row=row_number, column=workflow_node_col, value="6-图片分析")
            ws.cell(row=row_number, column=remark_col, value="")
        else:
            ws.cell(row=row_number, column=remark_col, value=f"图片分析失败：{result.get('status', 'unknown_error')}")

    wb.save(output_path)
    print(f"Wrote {output_path}")


if __name__ == "__main__":
    main()
