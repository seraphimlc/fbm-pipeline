#!/usr/bin/env python3
"""Scan one Amazon product workbook row and match it to one supplier image folder."""

from __future__ import annotations

import argparse
import datetime as dt
import hashlib
import json
import os
import re
from pathlib import Path
from urllib.parse import unquote, urlparse

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover
    raise SystemExit("Missing dependency: openpyxl. Install with: python3 -m pip install openpyxl") from exc

try:
    from PIL import Image
except ImportError:  # pragma: no cover
    Image = None

IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".webp", ".bmp", ".tif", ".tiff", ".gif"}
SKIP_DIR_NAMES = {"new main image", "main image", "contact_sheets", "multimodal_upload_cache"}


def norm(value):
    return "" if value is None else str(value).strip()


def header_score(header: str, candidates: list[str]) -> int:
    h = header.lower()
    return max((1 for c in candidates if c.lower() in h), default=0)


def find_header(headers: list[str], candidates: list[str]) -> int | None:
    for i, header in enumerate(headers):
        if header_score(header, candidates):
            return i
    return None


def extract_folder_id(value: str) -> str:
    text = norm(value)
    if not text:
        return ""
    parsed = urlparse(text)
    if parsed.path:
        parts = [unquote(p) for p in parsed.path.split("/") if p]
        if parts:
            return re.sub(r"\.[A-Za-z0-9]+$", "", parts[-1]).strip()
    matches = re.findall(r"[A-Za-z0-9][A-Za-z0-9_-]{2,}", text)
    return matches[-1] if matches else text


def file_hash(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def image_dimensions(path: Path) -> dict:
    if Image is None:
        return {"width": None, "height": None}
    try:
        with Image.open(path) as im:
            return {"width": im.width, "height": im.height}
    except Exception:
        return {"width": None, "height": None}


def list_images(folder: Path) -> list[dict]:
    rows = []
    for root, dirs, files in os.walk(folder):
        dirs[:] = [d for d in dirs if d not in SKIP_DIR_NAMES and not d.startswith(".")]
        for name in sorted(files):
            path = Path(root) / name
            if path.suffix.lower() not in IMAGE_EXTS:
                continue
            stat = path.stat()
            rows.append(
                {
                    "image_id": f"#{len(rows) + 1:02d}",
                    "filename": path.name,
                    "path": str(path.resolve()),
                    "relative_path": str(path.relative_to(folder)),
                    "extension": path.suffix.lower(),
                    "bytes": stat.st_size,
                    "sha256": file_hash(path),
                    "dimensions": image_dimensions(path),
                }
            )
    return rows


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", required=True)
    parser.add_argument("--image-root", default=None)
    parser.add_argument("--sheet", default=None)
    parser.add_argument("--output-dir", default=None)
    parser.add_argument("--row-number", type=int, default=None, help="Excel row number to process. Required unless --folder-id is provided.")
    parser.add_argument("--folder-id", default=None, help="Single product folder ID to process. Required unless --row-number is provided.")
    args = parser.parse_args()

    if (args.row_number is None) == (args.folder_id is None):
        raise SystemExit("Single-product red line: provide exactly one of --row-number or --folder-id.")

    workbook_path = Path(args.workbook).expanduser().resolve()
    if not workbook_path.exists():
        raise SystemExit(f"Workbook not found: {workbook_path}")
    image_root = Path(args.image_root).expanduser().resolve() if args.image_root else workbook_path.parent
    if not image_root.exists():
        raise SystemExit(f"Image root not found: {image_root}")

    run_dir = (
        Path(args.output_dir).expanduser().resolve()
        if args.output_dir
        else workbook_path.parent / "amazon_image_curator_runs" / f"run_{dt.datetime.now():%Y%m%d_%H%M%S}"
    )
    run_dir.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(workbook_path, data_only=True)
    ws = wb[args.sheet] if args.sheet else wb.active
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [norm(v) or f"column_{i + 1}" for i, v in enumerate(header_row)]
    url_idx = find_header(headers, ["大件云仓", "云仓", "warehouse", "图片链接", "url", "link", "链接", "ID", "id"])
    title_idx = find_header(headers, ["标题", "title", "商品名", "product"])
    bullet_idxs = [i for i, h in enumerate(headers) if any(k in h.lower() for k in ["bullet", "卖点", "要点"])]

    products = []
    for excel_row_number, values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(norm(v) for v in values):
            continue
        if args.row_number is not None and excel_row_number != args.row_number:
            continue
        row = {headers[i]: norm(values[i]) if i < len(values) else "" for i in range(len(headers))}
        source_value = norm(values[url_idx]) if url_idx is not None and url_idx < len(values) else ""
        folder_id = extract_folder_id(source_value)
        if args.folder_id is not None and folder_id != args.folder_id:
            continue
        matched_folder = image_root / folder_id if folder_id else None
        matched_folder = matched_folder if matched_folder and matched_folder.exists() and matched_folder.is_dir() else None
        images = list_images(matched_folder) if matched_folder else []
        products.append(
            {
                "product_key": folder_id or f"row_{excel_row_number}",
                "row_number": excel_row_number,
                "sheet_name": ws.title,
                "folder_id": folder_id,
                "warehouse_source": source_value,
                "title": norm(values[title_idx]) if title_idx is not None and title_idx < len(values) else "",
                "bullets": [norm(values[i]) for i in bullet_idxs if i < len(values) and norm(values[i])],
                "row": row,
                "match_status": "matched" if matched_folder else "missing",
                "matched_folder": str(matched_folder.resolve()) if matched_folder else "",
                "candidate_folders": [str((image_root / folder_id).resolve())] if folder_id else [],
                "images": images,
            }
        )
        break

    if len(products) != 1:
        target = f"row {args.row_number}" if args.row_number is not None else f"folder_id {args.folder_id}"
        raise SystemExit(f"Single-product red line: no matching workbook row found for {target}.")

    index = {
        "created_at": dt.datetime.now().isoformat(timespec="seconds"),
        "workbook": str(workbook_path),
        "image_root": str(image_root),
        "run_dir": str(run_dir),
        "products": products,
    }
    index_path = run_dir / "product_index.json"
    index_path.write_text(json.dumps(index, ensure_ascii=False, indent=2), encoding="utf-8")

    template = {
        "created_at": dt.datetime.now().isoformat(timespec="seconds"),
        "source_index": str(index_path),
        "products": [
            {
                "product_key": p["product_key"],
                "row_number": p["row_number"],
                "folder_id": p["folder_id"],
                "listing_intent": {
                    "product_type": "",
                    "variant": "",
                    "buyer_problem": "",
                    "visible_claims": [],
                    "caution_claims": [],
                },
                "image_reviews": [
                    {
                        "image_id": img["image_id"],
                        "filename": img["filename"],
                        "path": img["path"],
                        "multimodal_result": {},
                        "contact_sheet_evidence": {},
                        "image_type": "",
                        "visible_selling_point": "",
                        "matched_title_bullet_evidence": "",
                        "conversion_role": "",
                        "risk_flags": [],
                        "slot01_score": None,
                        "gallery_score": None,
                        "decision_reason": "",
                    }
                    for img in p["images"]
                ],
                "gallery_selection": [],
                "missing_recommended_images": [],
            }
            for p in products
        ],
    }
    template_path = run_dir / "analysis_decisions_template.json"
    template_path.write_text(json.dumps(template, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"Wrote {index_path}")
    print(f"Wrote {template_path}")
    product = products[0]
    print(f"Single product: {product['product_key']}; match_status: {product['match_status']}; images: {len(product['images'])}")


if __name__ == "__main__":
    main()
