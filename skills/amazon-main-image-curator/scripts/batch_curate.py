#!/usr/bin/env python3
"""
Batch Image Curator - Analyze all products in Excel using the configured GPT multimodal API.
"""

import json
import os
import sys
import time
import base64
import urllib.request
import urllib.error
import openpyxl
from pathlib import Path
from PIL import Image

# Config
EXCEL_PATH = "/Users/jiyuhang/Documents/F/亚马逊工作目录/亚马逊商品/05-03/商品表格.xlsx"
IMAGE_ROOT = "/Users/jiyuhang/Documents/F/亚马逊工作目录/亚马逊商品/05-03"
API_URL = os.environ.get("LLM_API_BASE", "https://sub2api.127space.com/v1").rstrip("/") + "/chat/completions"
API_KEY = os.environ.get("LLM_API_KEY", "YOUR_API_KEY_HERE")
MODEL = os.environ.get("LLM_MODEL", "gpt-5.5")
CONTACT_SHEET_SIZE = (1600, 1600)
CONTACT_SHEET_QUALITY = 80
MAX_RETRIES = 3
RETRY_DELAY = 10

# Excel column mapping
COL_PRODUCT_ID = 1
COL_TITLE = 17
COL_BULLETS = 18
COL_STATUS = 20
COL_GALLERY_COUNT = 21
COL_MAIN_CANDIDATE = 22
COL_MAIN_SCORE = 23
COL_MAIN_GAP = 24
COL_MAIN_PATH = 25
COL_ANALYSIS_FILE = 26
COL_SELLING_FILE = 27
COL_SELLING_SUMMARY = 28
COL_RUN_DIR = 29
COL_APLUS_INPUT = 30
COL_SELECTION_SUMMARY = 31


def find_product_images(product_dir):
    """Find all product images in directory and subdirectories."""
    images = []
    for root, dirs, files in os.walk(product_dir):
        # Skip analysis and output dirs
        dirs[:] = [d for d in dirs if d not in ('gallery_selected', 'new main image', 'amazon_image_curator_runs')]
        for f in sorted(files):
            if f.lower().endswith(('.jpg', '.jpeg', '.png', '.webp')):
                images.append(os.path.join(root, f))
    return images


def build_contact_sheet(image_paths, output_path, tiles_per_row=3, tile_size=400):
    """Build a contact sheet from image paths."""
    if not image_paths:
        return None
    
    n = len(image_paths)
    rows = (n + tiles_per_row - 1) // tiles_per_row
    
    canvas_w = tiles_per_row * tile_size
    canvas_h = rows * tile_size
    
    canvas = Image.new('RGB', (canvas_w, canvas_h), (240, 240, 240))
    
    from PIL import ImageDraw, ImageFont
    draw = ImageDraw.Draw(canvas)
    
    for idx, img_path in enumerate(image_paths):
        row = idx // tiles_per_row
        col = idx % tiles_per_row
        x = col * tile_size
        y = row * tile_size
        
        try:
            img = Image.open(img_path).convert('RGB')
            img.thumbnail((tile_size - 4, tile_size - 20), Image.LANCZOS)
            
            # Center tile
            offset_x = x + (tile_size - img.width) // 2
            offset_y = y + 2
            canvas.paste(img, (offset_x, offset_y))
            
            # Label
            label = f"#{idx+1:02d}"
            draw.text((x + 4, y + tile_size - 18), label, fill=(0, 0, 0))
        except Exception as e:
            draw.text((x + 4, y + 4), f"#{idx+1:02d} ERR", fill=(255, 0, 0))
    
    canvas.save(output_path, quality=CONTACT_SHEET_QUALITY)
    return output_path


def call_vlm_api(image_path, prompt):
    """Call the configured multimodal API with an image and prompt."""
    # Compress image
    img = Image.open(image_path).convert('RGB')
    img.thumbnail(CONTACT_SHEET_SIZE, Image.LANCZOS)
    
    import io
    buf = io.BytesIO()
    img.save(buf, format='JPEG', quality=CONTACT_SHEET_QUALITY)
    b64 = base64.b64encode(buf.getvalue()).decode()
    
    payload = {
        'model': MODEL,
        'messages': [{
            'role': 'user',
            'content': [
                {'type': 'image_url', 'image_url': {'url': f'data:image/jpeg;base64,{b64}'}},
                {'type': 'text', 'text': prompt}
            ]
        }],
        'max_tokens': 2000,
        'temperature': 0.1
    }
    
    for attempt in range(MAX_RETRIES):
        try:
            req = urllib.request.Request(
                API_URL,
                data=json.dumps(payload).encode(),
                headers={
                    'Authorization': f'Bearer {API_KEY}',
                    'Content-Type': 'application/json'
                }
            )
            with urllib.request.urlopen(req, timeout=180) as resp:
                result = json.loads(resp.read())
                return result['choices'][0]['message']['content']
        except Exception as e:
            if attempt < MAX_RETRIES - 1:
                print(f"  ⚠️ API error (attempt {attempt+1}): {e}, retrying...")
                time.sleep(RETRY_DELAY * (attempt + 1))
            else:
                print(f"  ❌ API failed after {MAX_RETRIES} attempts: {e}")
                return None


def analyze_product(product_id, product_dir, title, bullets):
    """Full analysis pipeline for one product."""
    print(f"\n{'='*60}")
    print(f"📦 Analyzing: {product_id}")
    print(f"{'='*60}")
    
    # Step 1: Find images
    images = find_product_images(product_dir)
    if not images:
        print(f"  ⚠️ No images found for {product_id}")
        return None
    
    print(f"  📷 Found {len(images)} images")
    
    # Step 2: Build contact sheets (max 9 per sheet)
    run_dir = os.path.join(IMAGE_ROOT, "amazon_image_curator_runs", f"run_{time.strftime('%Y%m%d_%H%M%S')}")
    os.makedirs(run_dir, exist_ok=True)
    
    sheets = []
    for i in range(0, len(images), 9):
        batch = images[i:i+9]
        sheet_path = os.path.join(run_dir, f"{product_id}_sheet_{i//9+1:02d}.jpg")
        build_contact_sheet(batch, sheet_path)
        sheets.append((sheet_path, batch))
        print(f"  📋 Contact sheet {len(sheets)}: {len(batch)} images")
    
    # Step 3: Analyze each sheet with VLM
    all_reviews = []
    for sheet_idx, (sheet_path, batch) in enumerate(sheets):
        tile_start = sheet_idx * 9 + 1
        tile_end = tile_start + len(batch) - 1
        
        prompt = f"""You are an Amazon senior operator reviewing a contact sheet of product images. Tiles are labeled #{tile_start:02d}-#{tile_end:02d}.

Product Title: {title}

Key Bullets:
{bullets}

For EACH tile, provide a JSON object with these fields:
- image_id: tile label (e.g. "#{tile_start:02d}")
- filename: just the filename from the path
- image_type: one of [white_background_product, lifestyle, detail, dimension, infographic, unboxing, variant_comparison]
- visible_selling_point: primary selling point shown (e.g. product_identity, material_or_texture, usage_scenario, dimension_information, full_set, construction_detail)
- conversion_role: one of [click, confirm_product, show_use, prove_material, explain_size, no_useful_role]
- risk_flags: array of issues (e.g. non-white_bg, text_overlay, wrong_variant, near_duplicate, cropped_product, watermark)
- slot01_eligible: true only if white background, no text/logo/watermark, full product visible, correct variant
- slot01_score: 0-100 (hard reject = non-white bg/text/wrong variant caps at 59)
- gallery_score: 0-100
- visual_summary: one sentence description

Output as a JSON array. No markdown fences."""

        print(f"  🔍 Analyzing sheet {sheet_idx+1}...")
        result = call_vlm_api(sheet_path, prompt)
        
        if result:
            # Parse JSON from response
            try:
                # Try to extract JSON from response
                text = result.strip()
                if text.startswith('```'):
                    text = text.split('\n', 1)[1].rsplit('```', 1)[0].strip()
                
                reviews = json.loads(text)
                if isinstance(reviews, list):
                    # Add filenames from batch
                    for j, rev in enumerate(reviews):
                        if j < len(batch):
                            rev['filepath'] = batch[j]
                            rev['filename'] = os.path.basename(batch[j])
                    all_reviews.extend(reviews)
                    print(f"  ✅ Got {len(reviews)} reviews")
                else:
                    print(f"  ⚠️ Unexpected response format, using raw text")
                    # Create basic reviews from batch
                    for j, img_path in enumerate(batch):
                        all_reviews.append({
                            'image_id': f'#{tile_start+j:02d}',
                            'filepath': img_path,
                            'filename': os.path.basename(img_path),
                            'image_type': 'unknown',
                            'slot01_score': 50,
                            'gallery_score': 50,
                            'risk_flags': [],
                            'visual_summary': 'VLM parse failed'
                        })
            except json.JSONDecodeError:
                print(f"  ⚠️ JSON parse failed, creating basic reviews")
                for j, img_path in enumerate(batch):
                    all_reviews.append({
                        'image_id': f'#{tile_start+j:02d}',
                        'filepath': img_path,
                        'filename': os.path.basename(img_path),
                        'image_type': 'unknown',
                        'slot01_score': 50,
                        'gallery_score': 50,
                        'risk_flags': [],
                        'visual_summary': 'VLM response parse failed'
                    })
        else:
            print(f"  ❌ API call failed for sheet {sheet_idx+1}")
            for j, img_path in enumerate(batch):
                all_reviews.append({
                    'image_id': f'#{tile_start+j:02d}',
                    'filepath': img_path,
                    'filename': os.path.basename(img_path),
                    'image_type': 'unknown',
                    'slot01_score': 50,
                    'gallery_score': 50,
                    'risk_flags': ['api_failed'],
                    'visual_summary': 'API call failed'
                })
    
    # Step 4: Select gallery (up to 9 images)
    gallery = select_gallery(all_reviews)
    
    # Step 5: Save results
    save_results(product_id, product_dir, run_dir, all_reviews, gallery, title)
    
    return {
        'product_id': product_id,
        'run_dir': run_dir,
        'total_images': len(images),
        'gallery_count': len(gallery),
        'reviews': all_reviews,
        'gallery': gallery
    }


def select_gallery(reviews):
    """Select up to 9 gallery images from reviews."""
    # Sort by gallery_score descending
    sorted_revs = sorted(reviews, key=lambda r: r.get('gallery_score', 0), reverse=True)
    
    # Pick slot 01 first (highest slot01_score)
    slot01_candidates = [r for r in sorted_revs if r.get('slot01_eligible', False)]
    if not slot01_candidates:
        slot01_candidates = [r for r in sorted_revs if r.get('slot01_score', 0) >= 60]
    
    gallery = []
    used_selling_points = set()
    used_types = set()
    
    # Slot 01: best main image candidate
    if slot01_candidates:
        best = max(slot01_candidates, key=lambda r: r.get('slot01_score', 0))
        gallery.append({'slot': '01', **best})
        used_selling_points.add(best.get('visible_selling_point', ''))
        used_types.add(best.get('image_type', ''))
    
    # Slots 02-09: diversify by selling point and type
    remaining = [r for r in sorted_revs if r not in [g for g in gallery]]
    
    for r in remaining:
        if len(gallery) >= 9:
            break
        
        sp = r.get('visible_selling_point', '')
        it = r.get('image_type', '')
        score = r.get('gallery_score', 0)
        
        # Skip low scores
        if score < 40:
            continue
        
        # Prefer diversity but allow some overlap for important types
        priority = 0
        if sp not in used_selling_points:
            priority += 20
        if it not in used_types:
            priority += 10
        
        effective_score = score + priority
        r['_effective_score'] = effective_score
    
    # Re-sort by effective score
    remaining.sort(key=lambda r: r.get('_effective_score', r.get('gallery_score', 0)), reverse=True)
    
    for r in remaining:
        if len(gallery) >= 9:
            break
        slot_num = len(gallery) + 1
        gallery.append({'slot': f'{slot_num:02d}', **r})
        used_selling_points.add(r.get('visible_selling_point', ''))
        used_types.add(r.get('image_type', ''))
    
    return gallery


def save_results(product_id, product_dir, run_dir, reviews, gallery, title):
    """Save analysis results to files."""
    # Save full analysis JSON
    analysis = {
        'product_id': product_id,
        'title': title,
        'image_reviews': reviews,
        'gallery_selection': gallery
    }
    
    analysis_path = os.path.join(run_dir, f'{product_id}_analysis.json')
    with open(analysis_path, 'w', encoding='utf-8') as f:
        json.dump(analysis, f, ensure_ascii=False, indent=2)
    
    # Copy gallery images to 'new main image' folder
    gallery_dir = os.path.join(product_dir, 'new main image')
    if os.path.exists(gallery_dir) and os.listdir(gallery_dir):
        # Backup existing folder
        import time
        backup_dir = os.path.join(product_dir, f'new main image_{time.strftime("%Y%m%d_%H%M%S")}')
        os.rename(gallery_dir, backup_dir)
    os.makedirs(gallery_dir, exist_ok=True)
    
    for g in gallery:
        slot = g['slot']
        src = g.get('filepath', '')
        if src and os.path.exists(src):
            ext = os.path.splitext(src)[1]
            dst = os.path.join(gallery_dir, f'{slot}{ext}')
            import shutil
            shutil.copy2(src, dst)
    
    # Write markdown analysis
    md_path = os.path.join(product_dir, 'image_analysis.md')
    with open(md_path, 'w', encoding='utf-8') as f:
        f.write(f'# Image Analysis: {product_id}\n\n')
        f.write(f'## Product\n{title}\n\n')
        f.write(f'## Gallery Selection ({len(gallery)} images)\n\n')
        f.write('| Slot | Image | Type | Selling Point | Gallery Score |\n')
        f.write('|------|-------|------|---------------|---------------|\n')
        for g in gallery:
            f.write(f'| {g["slot"]} | {g.get("filename","?")} | {g.get("image_type","?")} | {g.get("visible_selling_point","?")} | {g.get("gallery_score",0)} |\n')
        
        # Excluded
        included_files = {g.get('filename') for g in gallery}
        excluded = [r for r in reviews if r.get('filename') not in included_files]
        if excluded:
            f.write(f'\n## Excluded Images ({len(excluded)})\n\n')
            f.write('| Image | Type | Score | Reason |\n')
            f.write('|-------|------|-------|--------|\n')
            for r in excluded:
                flags = ', '.join(r.get('risk_flags', []))
                f.write(f'| {r.get("filename","?")} | {r.get("image_type","?")} | {r.get("gallery_score",0)} | {flags or "lower priority"} |\n')
    
    print(f"  💾 Saved: {analysis_path}")
    print(f"  💾 Saved: {md_path}")
    print(f"  📁 Gallery: {gallery_dir} ({len(gallery)} images)")


def update_excel(row, result):
    """Update Excel row with analysis results."""
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active
    
    if result is None:
        ws.cell(row, COL_STATUS).value = '⚠️ 无图片'
        wb.save(EXCEL_PATH)
        return
    
    gallery = result['gallery']
    main_img = next((g for g in gallery if g['slot'] == '01'), None)
    
    ws.cell(row, COL_STATUS).value = '✅ 已完成'
    ws.cell(row, COL_GALLERY_COUNT).value = len(gallery)
    
    if main_img:
        ws.cell(row, COL_MAIN_CANDIDATE).value = main_img.get('filename', '')
        ws.cell(row, COL_MAIN_SCORE).value = main_img.get('slot01_score', 0)
        ws.cell(row, COL_MAIN_PATH).value = f"new main image/01{os.path.splitext(main_img.get('filename',''))[1]}"
    
    # Check for main image gap
    has_white_bg = any(g.get('image_type') == 'white_background_product' for g in gallery)
    ws.cell(row, COL_MAIN_GAP).value = '无缺口' if has_white_bg else '⚠️ 无白底图'
    
    ws.cell(row, COL_ANALYSIS_FILE).value = 'image_analysis.md'
    ws.cell(row, COL_SELLING_FILE).value = 'image_analysis.md'
    
    selling_points = list({g.get('visible_selling_point', '') for g in gallery if g.get('visible_selling_point')})
    ws.cell(row, COL_SELLING_SUMMARY).value = ', '.join(selling_points)
    ws.cell(row, COL_RUN_DIR).value = os.path.relpath(result['run_dir'], IMAGE_ROOT)
    
    # Selection summary
    summary_parts = [f"Slot{g['slot']}={g.get('filename','?')}({g.get('gallery_score',0)})" for g in gallery[:4]]
    ws.cell(row, COL_SELECTION_SUMMARY).value = ', '.join(summary_parts)
    
    wb.save(EXCEL_PATH)


def main():
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument('--start-row', type=int, default=2)
    parser.add_argument('--end-row', type=int, default=None)
    parser.add_argument('--product-id', type=str, default=None)
    parser.add_argument('--skip-done', action='store_true', default=True)
    parser.add_argument('--no-skip-done', action='store_false', dest='skip_done')
    args = parser.parse_args()
    
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active
    
    end_row = args.end_row or ws.max_row
    
    # Collect products to process
    products = []
    for row in range(args.start_row, end_row + 1):
        pid = str(ws.cell(row, COL_PRODUCT_ID).value).strip()
        if not pid or pid == 'None':
            continue
        
        status = ws.cell(row, COL_STATUS).value
        if args.skip_done and status and '✅' in str(status):
            print(f"  ⏭️ Row {row}: {pid} already done")
            continue
        
        title = ws.cell(row, COL_TITLE).value or ''
        bullets = ws.cell(row, COL_BULLETS).value or ''
        
        product_dir = os.path.join(IMAGE_ROOT, pid)
        if not os.path.isdir(product_dir):
            print(f"  ⏭️ Row {row}: {pid} directory not found")
            continue
        
        if args.product_id and pid != args.product_id:
            continue
        
        products.append((row, pid, product_dir, str(title), str(bullets)))
    
    print(f"\n🚀 Batch Image Curator")
    print(f"   Products to process: {len(products)}")
    print(f"   Start time: {time.strftime('%H:%M:%S')}")
    
    success = 0
    failed = 0
    
    for i, (row, pid, pdir, title, bullets) in enumerate(products, 1):
        print(f"\n[{i}/{len(products)}] Processing {pid}...")
        
        try:
            result = analyze_product(pid, pdir, title, bullets)
            update_excel(row, result)
            if result:
                success += 1
                print(f"  ✅ Done: {result['gallery_count']} gallery images selected")
            else:
                failed += 1
                print(f"  ⚠️ No images found")
        except Exception as e:
            failed += 1
            print(f"  ❌ Error: {e}")
            import traceback
            traceback.print_exc()
        
        # Rate limit: small delay between products
        if i < len(products):
            time.sleep(2)
    
    print(f"\n{'='*60}")
    print(f"🏁 Batch Complete!")
    print(f"   ✅ Success: {success}")
    print(f"   ❌ Failed: {failed}")
    print(f"   End time: {time.strftime('%H:%M:%S')}")


if __name__ == '__main__':
    main()
