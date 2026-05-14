#!/usr/bin/env python3
import argparse
import base64
import json
import math
import subprocess
import sys
import time
from io import BytesIO
from pathlib import Path

from PIL import Image


DEFAULT_CONFIG = Path.home() / ".openclaw/workspace/skills/gpt-image-async/scripts/config.json"


TEXT_MODE_RULES = {
    "none": "",
    "no-text": (
        "\n\nCritical rendering rule: do not generate any headline, body copy, icons, pictograms, labels, "
        "badges, seals, QR codes, URLs, brand names, store names, wordmarks, logos, UI graphics, or decorative text inside the image. "
        "Leave clean negative space for a designer to add copy later. The product and scene must communicate the idea without text."
    ),
    "minimal": (
        "\n\nCritical rendering rule: controlled Amazon A+ text is allowed. You may render a short headline, "
        "1-4 short selling-point phrases, and simple clean feature icons when they improve readability. "
        "Do not render any brand name, store name, brand wordmark, brand logo, Amazon logo, certification badge, warranty seal, star rating, review quote, "
        "prices, discounts, QR codes, URLs, dense paragraphs, or tiny unreadable labels. "
        "If any earlier instruction says no icons or no text, interpret it as no fake badges, no clutter, and no dense unreadable text; "
        "simple A+ feature icons and short truthful copy are allowed, but brand text and brand marks are not allowed."
    ),
}


def load_config(path: Path) -> dict:
    return json.loads(path.read_text(encoding="utf-8"))


def normalize_size(size: str) -> str:
    cleaned = size.lower().replace(" ", "")
    if "x" not in cleaned:
        raise ValueError(f"Invalid size: {size}")
    width, height = [int(x) for x in cleaned.split("x", 1)]
    width = math.ceil(width / 16) * 16
    height = math.ceil(height / 16) * 16
    if max(width, height) > 3840:
        raise ValueError(f"Size exceeds provider limit: {width}x{height}")
    if max(width, height) / min(width, height) > 3:
        raise ValueError(f"Aspect ratio exceeds provider limit: {width}x{height}")
    return f"{width}x{height}"


def parse_size(size: str) -> tuple[int, int]:
    cleaned = size.lower().replace(" ", "")
    if "x" not in cleaned:
        raise ValueError(f"Invalid size: {size}")
    width, height = [int(x) for x in cleaned.split("x", 1)]
    return width, height


def load_script(script_json: Path, image_no: int) -> dict:
    data = json.loads(script_json.read_text(encoding="utf-8"))
    for script in data.get("scripts", []):
        if int(script.get("image_no")) == image_no:
            return data, script
    raise ValueError(f"Image script {image_no} not found in {script_json}")


def validate_refs(script: dict) -> list[Path]:
    refs = script.get("reference_images") or []
    if not refs:
        raise ValueError("No reference_images found. This generator refuses text-only generation.")
    paths = []
    missing = []
    for ref in refs:
        path = Path(ref.get("path", "")).expanduser()
        if not path.exists():
            missing.append(str(path))
        else:
            paths.append(path)
    if missing:
        raise FileNotFoundError("Missing reference image(s): " + "; ".join(missing))
    return paths


def reference_data_url(path: Path, max_side: int, quality: int) -> str:
    with Image.open(path) as image:
        image = image.convert("RGB")
        image.thumbnail((max_side, max_side))
        buffer = BytesIO()
        image.save(buffer, format="JPEG", quality=quality, optimize=True)
    return "data:image/jpeg;base64," + base64.b64encode(buffer.getvalue()).decode("ascii")


def submit_generation(
    config: dict,
    prompt: str,
    ref_paths: list[Path],
    aspect_ratio: str,
    quality: str,
    n: int,
    response_format: str,
    ref_max_side: int,
    ref_quality: int,
    timeout: int,
) -> dict:
    url = f"{config['base_url'].rstrip('/')}/v1/images/generations"
    payload = {
        "model": config["model"],
        "prompt": prompt,
        "aspect_ratio": aspect_ratio,
        "quality": quality,
        "n": n,
        "response_format": response_format,
        "image": [reference_data_url(path, ref_max_side, ref_quality) for path in ref_paths],
    }
    cmd = [
        "curl",
        "-sS",
        "-X",
        "POST",
        url,
        "-H",
        f"Authorization: Bearer {config['api_key']}",
        "-H",
        "Content-Type: application/json",
        "--max-time",
        str(timeout),
        "--data-binary",
        "@-",
    ]
    result = subprocess.run(cmd, input=json.dumps(payload, ensure_ascii=False), capture_output=True, text=True)
    if result.returncode != 0:
        raise RuntimeError(f"curl generation failed: {result.stderr or result.stdout}")
    try:
        data = json.loads(result.stdout)
    except json.JSONDecodeError as exc:
        raise RuntimeError(f"Provider returned non-JSON response: {result.stdout[:1000]}") from exc
    if data.get("error"):
        raise RuntimeError(f"Provider error: {json.dumps(data, ensure_ascii=False)}")
    return data


def submit_edit(config: dict, prompt: str, ref_paths: list[Path], size: str, n: int) -> dict:
    url = f"{config['base_url'].rstrip('/')}/v1/images/edits?async=true"
    cmd = [
        "curl",
        "-s",
        "-X",
        "POST",
        url,
        "-H",
        f"Authorization: Bearer {config['api_key']}",
        "-F",
        f"model={config['model']}",
        "-F",
        f"prompt={prompt}",
        "-F",
        f"size={size}",
        "-F",
        f"n={n}",
        "--max-time",
        "120",
    ]
    for path in ref_paths:
        cmd.extend(["-F", f"image=@{path}"])
    result = subprocess.run(cmd, capture_output=True, text=True)
    if result.returncode != 0:
        raise RuntimeError(f"curl submit failed: {result.stderr}")
    try:
        data = json.loads(result.stdout)
    except json.JSONDecodeError as exc:
        raise RuntimeError(f"Provider returned non-JSON response: {result.stdout[:1000]}") from exc
    if data.get("error"):
        raise RuntimeError(f"Provider error: {json.dumps(data, ensure_ascii=False)}")
    if not data.get("task_id"):
        raise RuntimeError(f"Provider response missing task_id: {json.dumps(data, ensure_ascii=False)}")
    return data


def query_task(config: dict, task_id: str) -> dict:
    url = f"{config['base_url'].rstrip('/')}/v1/images/tasks/{task_id}"
    cmd = [
        "curl",
        "-s",
        "-X",
        "GET",
        url,
        "-H",
        f"Authorization: Bearer {config['api_key']}",
        "-H",
        "Content-Type: application/json",
        "--max-time",
        "60",
    ]
    result = subprocess.run(cmd, capture_output=True, text=True)
    if result.returncode != 0:
        raise RuntimeError(f"curl query failed: {result.stderr}")
    return json.loads(result.stdout)


def extract_urls(result: dict) -> list[str]:
    urls = []
    data = result.get("data")
    if isinstance(data, list):
        urls.extend([img.get("url") for img in data if isinstance(img, dict) and img.get("url")])
    if isinstance(data, dict):
        nested = data.get("data", {}).get("data", [])
        if isinstance(nested, list):
            urls.extend([img.get("url") for img in nested if isinstance(img, dict) and img.get("url")])
    return urls


def extract_b64_images(result: dict) -> list[str]:
    data = result.get("data")
    if not isinstance(data, list):
        return []
    return [img.get("b64_json") for img in data if isinstance(img, dict) and img.get("b64_json")]


def download(url: str, output: Path) -> None:
    output.parent.mkdir(parents=True, exist_ok=True)
    result = subprocess.run(["curl", "-sL", url, "-o", str(output)], capture_output=True, text=True)
    if result.returncode != 0:
        raise RuntimeError(f"download failed: {result.stderr}")
    if not output.exists() or output.stat().st_size == 0:
        raise RuntimeError(f"download produced empty file: {output}")


def center_crop_to(image_path: Path, target_size: str) -> Path:
    target_width, target_height = parse_size(target_size)
    with Image.open(image_path) as image:
        width, height = image.size
        if width == target_width and height == target_height:
            return image_path
        if width < target_width or height < target_height:
            raise ValueError(f"Cannot crop {width}x{height} image to larger target {target_width}x{target_height}: {image_path}")
        left = (width - target_width) // 2
        top = (height - target_height) // 2
        cropped = image.crop((left, top, left + target_width, top + target_height))
        out = image_path.with_name(f"{image_path.stem}_crop_{target_width}x{target_height}{image_path.suffix}")
        cropped.save(out)
        return out


def resize_or_crop_to(image_path: Path, target_size: str) -> Path:
    target_width, target_height = parse_size(target_size)
    with Image.open(image_path) as image:
        image = image.convert("RGB")
        width, height = image.size
        if width == target_width and height == target_height:
            return image_path
        source_ratio = width / height
        target_ratio = target_width / target_height
        if abs(source_ratio - target_ratio) < 0.002:
            final = image.resize((target_width, target_height), Image.Resampling.LANCZOS)
        else:
            scale = max(target_width / width, target_height / height)
            resized = image.resize((math.ceil(width * scale), math.ceil(height * scale)), Image.Resampling.LANCZOS)
            left = (resized.width - target_width) // 2
            top = (resized.height - target_height) // 2
            final = resized.crop((left, top, left + target_width, top + target_height))
        out = image_path.with_name(f"{image_path.stem}_final_{target_width}x{target_height}{image_path.suffix}")
        final.save(out)
        return out


def inspect_image(path: Path) -> dict:
    with Image.open(path) as image:
        return {"path": str(path), "width": image.width, "height": image.height, "mode": image.mode, "format": image.format}


def apply_text_mode(prompt: str, text_mode: str) -> str:
    if text_mode not in TEXT_MODE_RULES:
        raise ValueError(f"Unsupported text mode: {text_mode}. Use one of: {', '.join(TEXT_MODE_RULES)}")
    return prompt + TEXT_MODE_RULES[text_mode]


def update_workbook(workbook_path: Path, product_id: str, node: str, image_outputs: list[str], metadata_path: str, node_only: bool) -> None:
    from openpyxl import load_workbook

    wb = load_workbook(workbook_path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    header_index = {header: i + 1 for i, header in enumerate(headers) if header}

    required = ["大建商品id", "当前工作节点", "处理备注"]
    missing = [header for header in required if header not in header_index]
    if missing:
        raise ValueError(f"Workbook missing required columns: {missing}")

    if not node_only:
        for header in ["A+成图文件", "A+成图元数据"]:
            if header not in header_index:
                ws.cell(1, ws.max_column + 1).value = header
                header_index[header] = ws.max_column

    target_row = None
    for row_no in range(2, ws.max_row + 1):
        if str(ws.cell(row_no, header_index["大建商品id"]).value).strip() == str(product_id).strip():
            target_row = row_no
            break
    if target_row is None:
        raise ValueError(f"Product ID {product_id} not found in workbook")

    ws.cell(target_row, header_index["当前工作节点"]).value = node
    ws.cell(target_row, header_index["处理备注"]).value = "A+成图完成"
    if not node_only:
        ws.cell(target_row, header_index["A+成图文件"]).value = "; ".join(image_outputs)
        ws.cell(target_row, header_index["A+成图元数据"]).value = metadata_path
    wb.save(workbook_path)


def default_output(product_dir: Path, image_no: int) -> Path:
    return product_dir / "new a plus" / f"a_plus_{image_no:02d}.png"


def main():
    parser = argparse.ArgumentParser(description="Generate one Amazon A+ image from Scriptwriter JSON with required reference images.")
    parser.add_argument("--product-dir", required=True)
    parser.add_argument("--image-no", type=int, required=True)
    parser.add_argument("--script-json")
    parser.add_argument("--output")
    parser.add_argument("--api-mode", default="generations", choices=["generations", "edits"], help="Use /v1/images/generations by default; edits keeps the older multipart async path.")
    parser.add_argument("--size", help="Legacy edits mode output size. Generations mode uses --aspect-ratio instead.")
    parser.add_argument("--aspect-ratio", default="97:60", help="Generations mode aspect ratio; default matches 1940x1200 Amazon A+ output.")
    parser.add_argument("--quality", default="high", choices=["low", "medium", "high", "auto"], help="Generations mode quality.")
    parser.add_argument("--response-format", default="url", choices=["url", "b64_json"], help="Generations mode response format.")
    parser.add_argument("--reference-max-side", type=int, default=1200, help="Resize reference images before base64 embedding to keep JSON requests stable.")
    parser.add_argument("--reference-quality", type=int, default=88, help="JPEG quality for embedded reference images.")
    parser.add_argument("--final-size", default="1940x1200", help="Final usable image size after local resize/crop; use 'none' to keep raw output only.")
    parser.add_argument("--text-mode", default="minimal", choices=sorted(TEXT_MODE_RULES.keys()), help="Append controlled A+ text rendering guidance to the Scriptwriter prompt.")
    parser.add_argument("--workbook", help="Optional workbook path to update on success.")
    parser.add_argument("--workbook-node", default="9-A Plus Image", help="Workbook 当前工作节点 value after success.")
    parser.add_argument("--no-workbook-columns", action="store_true", help="Only update 当前工作节点 and 处理备注 when updating workbook.")
    parser.add_argument("--n", type=int, default=1)
    parser.add_argument("--timeout", type=int, default=300)
    parser.add_argument("--config", default=str(DEFAULT_CONFIG))
    args = parser.parse_args()

    product_dir = Path(args.product_dir).expanduser()
    script_json = Path(args.script_json).expanduser() if args.script_json else product_dir / "image analysis" / "gpt_image_scripts.json"
    output = Path(args.output).expanduser() if args.output else default_output(product_dir, args.image_no)

    config = load_config(Path(args.config).expanduser())
    source, script = load_script(script_json, args.image_no)
    ref_paths = validate_refs(script)
    prompt = apply_text_mode(script["prompt"], args.text_mode)

    task_id = None
    submit = None
    last_result = None
    requested = {}
    if args.api_mode == "generations":
        requested = {"aspect_ratio": args.aspect_ratio, "quality": args.quality}
        submit = submit_generation(
            config,
            prompt,
            ref_paths,
            args.aspect_ratio,
            args.quality,
            args.n,
            args.response_format,
            args.reference_max_side,
            args.reference_quality,
            args.timeout,
        )
        task_id = submit.get("task_id")
        print(json.dumps({"status": "submitted", "api_mode": "generations", "task_id": task_id, **requested, "references": [str(p) for p in ref_paths]}, ensure_ascii=False))
        last_result = submit
        if task_id:
            start = time.time()
            while time.time() - start < args.timeout:
                last_result = query_task(config, task_id)
                status = last_result.get("data", {}).get("status", "UNKNOWN")
                print(json.dumps({"status": status, "task_id": task_id}, ensure_ascii=False), flush=True)
                if status == "SUCCESS":
                    break
                if status == "FAILURE":
                    raise RuntimeError(f"Task failed: {json.dumps(last_result, ensure_ascii=False)[:1000]}")
                time.sleep(3)
            else:
                raise TimeoutError(f"Timed out waiting for task {task_id}. Last result: {json.dumps(last_result, ensure_ascii=False)[:1000]}")
    else:
        size = normalize_size(args.size or script.get("output_size") or config.get("default_size", "2048x1152"))
        requested = {"size": size}
        submit = submit_edit(config, prompt, ref_paths, size, args.n)
        task_id = submit["task_id"]
        print(json.dumps({"status": "submitted", "api_mode": "edits", "task_id": task_id, "size": size, "references": [str(p) for p in ref_paths]}, ensure_ascii=False))
        start = time.time()
        while time.time() - start < args.timeout:
            last_result = query_task(config, task_id)
            status = last_result.get("data", {}).get("status", "UNKNOWN")
            print(json.dumps({"status": status, "task_id": task_id}, ensure_ascii=False), flush=True)
            if status == "SUCCESS":
                break
            if status == "FAILURE":
                raise RuntimeError(f"Task failed: {json.dumps(last_result, ensure_ascii=False)[:1000]}")
            time.sleep(3)
        else:
            raise TimeoutError(f"Timed out waiting for task {task_id}. Last result: {json.dumps(last_result, ensure_ascii=False)[:1000]}")

    urls = extract_urls(last_result)
    b64_images = extract_b64_images(last_result)
    if not urls and not b64_images:
        raise RuntimeError(f"No image URL or b64_json found: {json.dumps(last_result, ensure_ascii=False)[:1000]}")
    outputs = []
    if urls:
        for i, url in enumerate(urls, 1):
            out = output if len(urls) == 1 else output.with_name(f"{output.stem}_{i}{output.suffix}")
            download(url, out)
            outputs.append(str(out))
    else:
        output.parent.mkdir(parents=True, exist_ok=True)
        for i, b64_image in enumerate(b64_images, 1):
            out = output if len(b64_images) == 1 else output.with_name(f"{output.stem}_{i}{output.suffix}")
            out.write_bytes(base64.b64decode(b64_image))
            outputs.append(str(out))

    output_info = [inspect_image(Path(out)) for out in outputs]
    final_outputs = []
    if args.final_size and args.final_size.lower() != "none":
        for out in outputs:
            if args.api_mode == "edits":
                final_outputs.append(str(center_crop_to(Path(out), args.final_size)))
            else:
                final_outputs.append(str(resize_or_crop_to(Path(out), args.final_size)))
    final_output_info = [inspect_image(Path(out)) for out in final_outputs]
    meta = {
        "product_id": source.get("product_id"),
        "image_no": args.image_no,
        "api_mode": args.api_mode,
        "task_id": task_id,
        "request": requested,
        "final_size": None if args.final_size.lower() == "none" else args.final_size,
        "text_mode": args.text_mode,
        "script_json": str(script_json),
        "outputs": outputs,
        "output_info": output_info,
        "final_outputs": final_outputs,
        "final_output_info": final_output_info,
        "urls": urls,
        "reference_images": script.get("reference_images", []),
        "prompt": prompt,
        "provider_submit": submit,
        "provider_result": last_result,
    }
    meta_path = output.with_suffix(".metadata.json")
    meta_path.write_text(json.dumps(meta, ensure_ascii=False, indent=2), encoding="utf-8")
    if args.workbook:
        update_workbook(Path(args.workbook).expanduser(), source.get("product_id"), args.workbook_node, final_outputs or outputs, str(meta_path), args.no_workbook_columns)
    print(json.dumps({"status": "downloaded", "outputs": outputs, "final_outputs": final_outputs, "metadata": str(meta_path)}, ensure_ascii=False))


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print(json.dumps({"error": str(exc)}, ensure_ascii=False), file=sys.stderr)
        raise
